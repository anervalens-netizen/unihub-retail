#!/usr/bin/env python3
"""Compara strict grila veche din mai 2026 cu grila salariala noua.

Aceeasi cohorta, aceleasi vanzari, aceleasi targete si aceleasi zile lucrate
vin din arhiva grilelor vechi. Se schimba numai formula salariala.
"""

from __future__ import annotations

import asyncio
import itertools
import os
from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from statistics import median
from typing import Any
from zipfile import ZipFile

import asyncpg
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from generate_salary_grid_simulation import (
    ACCESSORY_COMMISSION_GATE,
    BASE_SALARY,
    COMMISSION_SCENARIOS,
    OLD_GRID_MAY_ARCHIVE,
    _number,
    _old_grid_comparable_salary,
    _old_grid_target_bonus,
    _qualitative_bonus,
    _qualitative_points,
    _target_bonus,
    build_simulation,
    candidate_agent_codes,
    fetch_data,
)

BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_DIR = BACKEND_DIR.parent
WORKSPACE_DIR = REPO_DIR.parent
DEFAULT_OUTPUT = WORKSPACE_DIR / "docs" / "comparatie_grila_veche_vs_noua_mai_2026.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")

load_dotenv(REPO_DIR / ".env")


@dataclass
class OldGridRow:
    registry_key: str
    company: str
    store: str
    site_code: str
    slot: int
    old_name: str
    old_base: float
    meal_tickets: float
    overtime: float
    total_salary: float
    sales: float
    target: float
    worked_days: int
    old_main_commission_total: float
    old_other_location_commission: float
    old_sim_bonus: float
    old_epay_bonus: float
    other_location_rows: list[tuple[float, float]]
    agent_code: str | None = None
    match_method: str = "nemapat"

    @property
    def target_pct(self) -> float | None:
        return self.sales * 100 / self.target if self.target > 0 else None

    @property
    def old_accessory_commission(self) -> float:
        return (
            self.sales * 0.03
            if self.target_pct is not None
            and self.target_pct >= ACCESSORY_COMMISSION_GATE
            else 0.0
        )

    @property
    def old_target_bonus(self) -> float:
        return _old_grid_target_bonus(self.target_pct)

    @property
    def old_comparable_salary(self) -> float:
        return (
            self.old_base
            + self.old_main_commission_total
            + self.old_other_location_commission
            + self.old_sim_bonus
            + self.old_epay_bonus
        )

    @property
    def old_cached_comparable_salary(self) -> float:
        return _old_grid_comparable_salary(
            self.total_salary,
            self.meal_tickets,
            self.overtime,
        )


def _read_other_location_rows(
    sheet: Any,
    *,
    target_column: int,
    realized_column: int,
) -> list[tuple[float, float]]:
    rows: list[tuple[float, float]] = []
    for row_number in range(5, 36):
        target = _number(sheet.cell(row_number, target_column).value)
        realized = _number(sheet.cell(row_number, realized_column).value)
        if target > 0 or realized > 0:
            rows.append((target, realized))
    return rows


def load_old_grid_rows(
    grid_registry: list[dict[str, Any]],
) -> list[OldGridRow]:
    registry_by_key = {
        row["registry_key"]: row["site_code"] for row in grid_registry
    }
    rows: list[OldGridRow] = []
    with ZipFile(OLD_GRID_MAY_ARCHIVE) as archive:
        for archive_name in archive.namelist():
            if not archive_name.lower().endswith(".xlsx"):
                continue
            registry_key = archive_name[:-5]
            site_code = registry_by_key.get(registry_key)
            if not site_code:
                continue
            company, store = registry_key.split("/", 1)
            workbook = load_workbook(
                BytesIO(archive.read(archive_name)),
                data_only=True,
                read_only=True,
            )
            sheet = workbook["Grila"]
            for slot, offset in ((1, 0), (2, 14)):
                old_name = str(sheet[f"D{2 + offset}"].value or "").strip()
                if not old_name:
                    continue
                if slot == 1:
                    other_rows = _read_other_location_rows(
                        sheet,
                        target_column=18,
                        realized_column=19,
                    )
                else:
                    other_rows = _read_other_location_rows(
                        sheet,
                        target_column=23,
                        realized_column=24,
                    )
                rows.append(
                    OldGridRow(
                        registry_key=registry_key,
                        company=company,
                        store=store,
                        site_code=site_code,
                        slot=slot,
                        old_name=old_name,
                        old_base=_number(sheet[f"D{3 + offset}"].value),
                        meal_tickets=_number(sheet[f"D{4 + offset}"].value),
                        overtime=_number(sheet[f"G{10 + offset}"].value),
                        total_salary=_number(sheet[f"D{6 + offset}"].value),
                        sales=_number(sheet[f"E{8 + offset}"].value),
                        target=_number(sheet[f"D{8 + offset}"].value),
                        worked_days=int(_number(sheet[f"D{9 + offset}"].value)),
                        old_main_commission_total=_number(
                            sheet[f"G{8 + offset}"].value
                        ),
                        old_other_location_commission=_number(
                            sheet[f"G{11 + offset}"].value
                        ),
                        old_sim_bonus=_number(sheet[f"G{12 + offset}"].value),
                        old_epay_bonus=(
                            _number(sheet[f"G{13 + offset}"].value)
                            + _number(sheet[f"G{14 + offset}"].value)
                        ),
                        other_location_rows=other_rows,
                    )
                )
    return rows


def _assignment_cost(old_row: OldGridRow, retail_row: dict[str, Any]) -> float:
    sales_scale = max(abs(old_row.sales), 1_000.0)
    sales_cost = abs(float(retail_row["total_sales"]) - old_row.sales) / sales_scale
    days_cost = abs(int(retail_row["working_days"]) - old_row.worked_days)
    return sales_cost * 10 + days_cost * 0.5


def match_old_rows_to_retail(
    old_rows: list[OldGridRow],
    retail_rows: list[dict[str, Any]],
) -> None:
    retail_by_site: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in retail_rows:
        if row["import_month"] == "2026-05":
            retail_by_site[row["site_code"]].append(row)

    old_by_site: dict[str, list[OldGridRow]] = defaultdict(list)
    for row in old_rows:
        old_by_site[row.site_code].append(row)

    for site_code, site_old_rows in old_by_site.items():
        candidates = retail_by_site.get(site_code, [])
        used_agents: set[str] = set()

        for old_row in site_old_rows:
            matches = [
                row
                for row in candidates
                if row["agent"] not in used_agents
                and row["agent"] in candidate_agent_codes(old_row.old_name)
            ]
            if len(matches) == 1:
                old_row.agent_code = matches[0]["agent"]
                old_row.match_method = "cod_nume"
                used_agents.add(matches[0]["agent"])

        remaining_old = [row for row in site_old_rows if not row.agent_code]
        remaining_retail = [
            row for row in candidates if row["agent"] not in used_agents
        ]
        if not remaining_old or not remaining_retail:
            continue

        pair_count = min(len(remaining_old), len(remaining_retail))
        best_pairs: list[tuple[OldGridRow, dict[str, Any]]] = []
        best_cost = float("inf")
        for retail_subset in itertools.combinations(remaining_retail, pair_count):
            for retail_order in itertools.permutations(retail_subset):
                pairs = list(zip(remaining_old[:pair_count], retail_order, strict=True))
                cost = sum(_assignment_cost(old, retail) for old, retail in pairs)
                if cost < best_cost:
                    best_cost = cost
                    best_pairs = pairs
        for old_row, retail_row in best_pairs:
            old_row.agent_code = retail_row["agent"]
            old_row.match_method = "vanzari_si_zile"


def _new_other_location_commission(
    row: OldGridRow,
    rate: float,
) -> float:
    return sum(
        realized * rate
        for target, realized in row.other_location_rows
        if target > 0 and realized * 100 / target >= ACCESSORY_COMMISSION_GATE
    )


def build_comparison(
    old_rows: list[OldGridRow],
    retail_simulation: pd.DataFrame,
) -> pd.DataFrame:
    simulation_by_agent = {
        row["Agent cod"]: row
        for row in retail_simulation[
            retail_simulation["Luna"] == "2026-05"
        ].to_dict("records")
    }
    rows_by_site: dict[str, list[OldGridRow]] = defaultdict(list)
    for row in old_rows:
        rows_by_site[row.site_code].append(row)

    output: list[dict[str, Any]] = []
    for old_row in old_rows:
        retail = simulation_by_agent.get(old_row.agent_code or "", {})
        daily_average = (
            old_row.sales / old_row.worked_days
            if old_row.worked_days > 0
            else None
        )
        peer_daily = [
            peer.sales / peer.worked_days
            for peer in rows_by_site[old_row.site_code]
            if peer is not old_row and peer.worked_days > 0
        ]
        daily_reference = (
            median(peer_daily)
            if peer_daily
            else _number(retail.get("Reper coleg")) or None
        )
        daily_vs_colleague_pct = (
            daily_average * 100 / daily_reference
            if daily_average is not None and daily_reference
            else None
        )
        points = _qualitative_points(
            daily_vs_colleague_pct=daily_vs_colleague_pct,
            receipt_2plus_pct=retail.get("Bonuri 2+ %"),
            focus_pct=retail.get("Focus %"),
            average_item_value=retail.get("Valoare medie produs"),
            premium_glass_pct=retail.get("Folii premium %"),
        )
        qualitative_bonus = _qualitative_bonus(points)
        target_bonus = _target_bonus(old_row.target_pct)

        result: dict[str, Any] = {
            "Companie": old_row.company,
            "Magazin": old_row.store,
            "Site": old_row.site_code,
            "Slot": old_row.slot,
            "Agent grila": old_row.old_name,
            "Agent cod Retail": old_row.agent_code,
            "Metoda mapare": old_row.match_method,
            "Salariu baza vechi": old_row.old_base,
            "Zile lucrate": old_row.worked_days,
            "Vanzari accesorii": round(old_row.sales, 2),
            "Target agent": round(old_row.target, 2),
            "Realizare target %": (
                round(old_row.target_pct, 2)
                if old_row.target_pct is not None
                else None
            ),
            "Comision accesorii vechi 3%": round(
                old_row.old_accessory_commission,
                2,
            ),
            "Bonus target vechi": old_row.old_target_bonus,
            "Comision alte locatii vechi": old_row.old_other_location_commission,
            "Bonus SIM": old_row.old_sim_bonus,
            "Bonus ePay": old_row.old_epay_bonus,
            "Ore suplimentare excluse": old_row.overtime,
            "Bonuri excluse": old_row.meal_tickets,
            "Grila veche comparabila": round(old_row.old_comparable_salary, 2),
            "Abatere cache total vechi": round(
                old_row.old_comparable_salary
                - old_row.old_cached_comparable_salary,
                2,
            ),
            "Observatie audit": (
                "Totalul cache din sheet nu include toate componentele"
                if abs(
                    old_row.old_comparable_salary
                    - old_row.old_cached_comparable_salary
                )
                > 0.05
                else ""
            ),
            "Medie zilnica": (
                round(daily_average, 2) if daily_average is not None else None
            ),
            "Reper coleg": (
                round(daily_reference, 2)
                if daily_reference is not None
                else None
            ),
            "Medie vs coleg %": (
                round(daily_vs_colleague_pct, 2)
                if daily_vs_colleague_pct is not None
                else None
            ),
            "Bonuri 2+ %": retail.get("Bonuri 2+ %"),
            "Focus %": retail.get("Focus %"),
            "Valoare medie produs": retail.get("Valoare medie produs"),
            "Folii premium %": retail.get("Folii premium %"),
            "Pct medie zilnica": points[0],
            "Pct bonuri 2+": points[1],
            "Pct focus": points[2],
            "Pct valoare": points[3],
            "Pct folii": points[4],
            "Puncte calitative": sum(points),
            "Bonus calitativ nou": qualitative_bonus,
            "Bonus target nou": target_bonus,
        }

        for scenario, rate in COMMISSION_SCENARIOS:
            main_commission = (
                old_row.sales * rate
                if old_row.target_pct is not None
                and old_row.target_pct >= ACCESSORY_COMMISSION_GATE
                else 0.0
            )
            other_commission = _new_other_location_commission(old_row, rate)
            new_salary = (
                BASE_SALARY
                + main_commission
                + other_commission
                + target_bonus
                + qualitative_bonus
                + old_row.old_sim_bonus
                + old_row.old_epay_bonus
            )
            result[f"Comision nou {scenario}"] = round(
                main_commission + other_commission,
                2,
            )
            result[f"Grila noua {scenario}"] = round(new_salary, 2)
            result[f"Diferenta {scenario}"] = round(
                new_salary - old_row.old_comparable_salary,
                2,
            )
        output.append(result)

    return pd.DataFrame(output).sort_values(
        ["Companie", "Magazin", "Slot"],
    )


def build_summary(comparison: pd.DataFrame) -> pd.DataFrame:
    old_total = comparison["Grila veche comparabila"].sum()
    agent_count = len(comparison)
    base_increase = (BASE_SALARY - comparison["Salariu baza vechi"]).sum()
    old_commission = (
        comparison["Comision accesorii vechi 3%"].sum()
        + comparison["Comision alte locatii vechi"].sum()
    )
    target_bonus_delta = (
        comparison["Bonus target nou"] - comparison["Bonus target vechi"]
    ).sum()
    qualitative_bonus = comparison["Bonus calitativ nou"].sum()
    rows: list[dict[str, Any]] = []
    for scenario, rate in COMMISSION_SCENARIOS:
        new_total = comparison[f"Grila noua {scenario}"].sum()
        delta = comparison[f"Diferenta {scenario}"]
        rows.append(
            {
                "Scenariu comision": scenario,
                "Procent comision": rate * 100,
                "Agenti": agent_count,
                "Grila veche comparabila": round(old_total, 2),
                "Grila noua comparabila": round(new_total, 2),
                "Diferenta buget": round(new_total - old_total, 2),
                "Diferenta buget %": (
                    (new_total - old_total) * 100 / old_total
                    if old_total
                    else None
                ),
                "Medie veche per agent": old_total / agent_count,
                "Medie noua per agent": new_total / agent_count,
                "Diferenta medie per agent": (new_total - old_total)
                / agent_count,
                "Agenti crestere": int(delta.gt(0.05).sum()),
                "Agenti scadere": int(delta.lt(-0.05).sum()),
                "Agenti fara diferenta": int(
                    delta.between(-0.05, 0.05).sum()
                ),
                "Crestere din salariul de baza": base_increase,
                "Diferenta din comision": (
                    comparison[f"Comision nou {scenario}"].sum()
                    - old_commission
                ),
                "Diferenta din bonus target": target_bonus_delta,
                "Bonus calitativ nou": qualitative_bonus,
            }
        )
    return pd.DataFrame(rows)


def rules_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Perioada", "Mai 2026"),
            (
                "Sursa principala",
                "arhiva exacta Grile - Mai 2026.zip: agenti, baza, vanzari, target, zile, SIM si ePay",
            ),
            (
                "Principiu",
                "aceleasi vanzari, targete si zile pentru grila veche si cea noua",
            ),
            ("Comparabil", "bonurile si plata orelor suplimentare sunt excluse"),
            (
                "Grila veche",
                "suma componentelor din sheet, fara bonuri si ore suplimentare; evita totalurile cache neactualizate",
            ),
            ("Baza noua", "2.700 lei"),
            (
                "Comision nou",
                "2,5% / 2,7% / 3,0%, numai de la minimum 80% din targetul vechi salvat",
            ),
            (
                "Bonus target nou",
                "200 lei la 100%; 300 lei la 110%; 400 lei la 120%",
            ),
            (
                "Criterii calitative",
                "date Retail mai 2026; medie zilnica recalculata din vanzarile si zilele grilei",
            ),
            (
                "Bonus calitativ",
                "100 / 200 / 300 lei; orice criteriu cu 0 puncte elimina bonusul",
            ),
            ("SIM", "aceeasi valoare din grila veche; regula ramane 3 lei/SIM"),
            (
                "ePay",
                "aceeasi valoare din grila veche; nu modifica diferenta intre grile",
            ),
            (
                "Vanzari alte locatii",
                "recalculate cu procentul scenariului si prag de minimum 80%",
            ),
        ],
        columns=["Regula", "Aplicare"],
    )


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    currency_words = (
        "salariu",
        "grila",
        "diferenta",
        "comision",
        "bonus",
        "vanzari",
        "target agent",
        "medie veche",
        "medie noua",
        "reper coleg",
        "medie zilnica",
    )
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            header = str(column_cells[0].value or "")
            lower_header = header.lower()
            if header.endswith("%") or header == "Procent comision":
                for cell in column_cells[1:]:
                    cell.number_format = '0.00"%"'
            elif any(word in lower_header for word in currency_words):
                for cell in column_cells[1:]:
                    cell.number_format = '#,##0.00 "lei"'
            values = [str(cell.value or "") for cell in column_cells[:300]]
            width = min(max(max(map(len, values), default=0) + 2, 10), 45)
            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = width

    detail = workbook["Comparatie agenti"]
    headers = {cell.value: cell.column for cell in detail[1]}
    mapping_column = headers["Metoda mapare"]
    for row_number in range(2, detail.max_row + 1):
        if detail.cell(row_number, mapping_column).value != "cod_nume":
            detail.cell(row_number, mapping_column).fill = WARNING_FILL
    for scenario, _rate in COMMISSION_SCENARIOS:
        delta_column = get_column_letter(headers[f"Diferenta {scenario}"])
        detail.conditional_formatting.add(
            f"{delta_column}2:{delta_column}{detail.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )
    workbook.save(path)


async def generate(output: Path) -> dict[str, Any]:
    if not OLD_GRID_MAY_ARCHIVE.exists():
        raise FileNotFoundError(OLD_GRID_MAY_ARCHIVE)
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL lipseste din .env")
    connection = await asyncpg.connect(database_url)
    try:
        data = await fetch_data(connection)
    finally:
        await connection.close()

    old_rows = load_old_grid_rows(data["grid_registry"])
    match_old_rows_to_retail(old_rows, data["base"])
    retail_simulation = build_simulation(data)
    comparison = build_comparison(old_rows, retail_simulation)
    summary = build_summary(comparison)

    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Sumar", index=False)
        comparison.to_excel(writer, sheet_name="Comparatie agenti", index=False)
        rules_frame().to_excel(writer, sheet_name="Reguli", index=False)
    format_workbook(output)

    return {
        "output": str(output),
        "grid_rows": len(old_rows),
        "mapped_rows": sum(row.agent_code is not None for row in old_rows),
        "summary": summary.to_dict("records"),
    }


def main() -> None:
    result = asyncio.run(generate(DEFAULT_OUTPUT))
    print(f"Generat: {result['output']}")
    print(
        f"Agenti grila={result['grid_rows']}, mapati Retail={result['mapped_rows']}"
    )
    for row in result["summary"]:
        print(
            f"{row['Scenariu comision']}: "
            f"veche={row['Grila veche comparabila']:.2f}, "
            f"noua={row['Grila noua comparabila']:.2f}, "
            f"dif={row['Diferenta buget']:+.2f} "
            f"({row['Diferenta buget %']:+.2f}%)"
        )


if __name__ == "__main__":
    main()
