#!/usr/bin/env python3
"""AC-13 destructive scale authority; accepts only disposable loopback test DBs."""
from __future__ import annotations

import argparse
import ast
import asyncio
import hashlib
import inspect
import io
import json
import math
import os
import platform
import shutil
import sys
import tempfile
import time
import uuid
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Awaitable, Callable
from urllib.parse import unquote, urlparse, urlunparse

ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / "backend"
asyncpg: Any = None
SEED = 20260812
UPLOAD_BYTES = 33_554_432
WARMUPS = 5
MEASUREMENTS = 30
PROFILES = {
    "2x": {"rows": 3_651_714, "stores": 200, "agents": 770, "months": 36},
    "5x": {"rows": 9_129_285, "stores": 200, "agents": 1_925, "months": 60},
}
LIMITS_MS = {"usual": 500, "complex": 1500, "target": 2000, "poll": 250, "upload": 1000, "reserve": 500}
EXCEPTION_STORES = (
    "AFICOTRO", "AUCHMIL2", "AUCHMILI", "AUCHTRIC", "CCTCIT", "CJIULMALL",
    "CJPPOL", "CLUJCFPOL", "CORALEX", "COTROCENI", "CRFFEER", "CTAUCH",
    "CTCITYPRK", "CTCORA", "CTCRFTOM", "CTVIVO", "MC-MEGAMALL", "MCRFBAL",
    "MEGAMALL", "PRKLK", "PROM", "PROMEN", "SUNPLZ", "TMACUH", "TMSHOPCITY", "UNIRII",
)
REQUIRED_B = (
    "services/dashboard/query_stores.py", "services/dashboard/query_agents.py",
    "services/dashboard/query_managers.py", "services/reporting_refresh_month.py",
)
DEPENDENCY_GLOBS = (
    "db/migrations/*.sql", "services/dashboard/query_*.py", "services/reporting_refresh*.py",
)
FLOW_KINDS = {
    "dashboard_store": "complex", "dashboard_agent": "complex", "dashboard_regional": "complex",
    "dashboard_history": "usual", "dashboard_all": "complex", "pnl_annual": "usual",
    "salary_agents": "usual", "salary_records": "usual", "target_context": "usual",
    "target_calculate": "target", "campaign_dense": "complex", "import_accept": "upload",
    "export_reserve": "reserve", "export_status": "poll", "export_resumable": "poll",
}
VOLATILE_KEYS = {"id", "scenario_id", "operation_id", "job_id", "created_at", "updated_at", "generated_at", "started_at", "finished_at"}
def safe_dsn(value: str) -> str:
    parsed = urlparse(value)
    dbname = unquote(parsed.path.lstrip("/"))
    valid = parsed.scheme in {"postgres", "postgresql"} and parsed.hostname in {"127.0.0.1", "localhost", "::1"}
    valid = valid and parsed.port is not None and parsed.port != 5432
    valid = valid and (dbname.startswith("test_") or dbname.endswith("_test"))
    if not valid:
        raise RuntimeError("AC-13 refuses non-loopback, port 5432, or non-test PostgreSQL DSN")
    if os.getenv("UNIHUB_TEST_DATABASE") != "1" or os.getenv("UNIHUB_RUNNING_TESTS") != "1":
        raise RuntimeError("AC-13 requires both destructive-test markers")
    return value
def database_dsn(admin_dsn: str, name: str) -> str:
    parsed = urlparse(admin_dsn)
    return safe_dsn(urlunparse(parsed._replace(path=f"/{name}")))
def canonical(value: Any) -> Any:
    if hasattr(value, "model_dump"):
        value = value.model_dump()
    if isinstance(value, dict):
        return {str(k): canonical(v) for k, v in sorted(value.items()) if str(k) not in VOLATILE_KEYS}
    if isinstance(value, (list, tuple, set)):
        return [canonical(item) for item in value]
    if isinstance(value, (date, datetime, Decimal, Path, uuid.UUID)):
        return str(value)
    if isinstance(value, bytes):
        return {"bytes": len(value), "sha256": hashlib.sha256(value).hexdigest()}
    return value
def digest(value: Any) -> str:
    material = json.dumps(canonical(value), sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha256(material.encode()).hexdigest()
def nearest_rank(values: list[float], percentile: float) -> float:
    return sorted(values)[max(0, math.ceil(percentile * len(values)) - 1)]
def assert_static_authority(source: str) -> None:
    tree = ast.parse(source)
    forbidden_imports = ("salary_exports", "import_salary", "export_process", "worker")
    imports = [node.module or "" for node in ast.walk(tree) if isinstance(node, ast.ImportFrom)]
    imports += [alias.name for node in ast.walk(tree) if isinstance(node, ast.Import) for alias in node.names]
    if any(token in name for name in imports for token in forbidden_imports):
        raise RuntimeError("AC-13 forbidden authority import")
    forbidden_calls = {
        "reserve_salary", "finalize", "save_final_targets", "export_excel", "build_target_excel",
        "promote_sales_generation", "rollback_sales_generation", "import_promo_actuals", "process_promo_actuals",
        "claim_download_owned", "download", "run_export_renderer_process", "get_pool",
    }
    calls = [node.func.attr for node in ast.walk(tree) if isinstance(node, ast.Call) and isinstance(node.func, ast.Attribute)]
    if forbidden_calls.intersection(calls):
        raise RuntimeError("AC-13 forbidden salary/Target/download call")
    required = {"get_performance_detail", "get_monthly_history", "get_dashboard_all", "annual", "get_agents_summary", "get_records", "get_context", "calculate", "get_promotions_incentives", "import_sales", "reserve", "status", "resumable"}
    if not required.issubset(set(calls)):
        raise RuntimeError("AC-13 real callable surface is incomplete")
def fail_closed_b() -> None:
    missing = [item for item in REQUIRED_B if not (BACKEND / item).is_file()]
    if missing:
        raise RuntimeError(f"AC-13 requires Release B callables: {', '.join(missing)}")
class Acquire:
    def __init__(self, pool: Any, wrap: Callable[[Any], Any]):
        self.pool, self.wrap, self.raw = pool, wrap, None

    async def get(self) -> Any:
        self.raw = await self.pool.acquire()
        return self.wrap(self.raw)

    def __await__(self):
        return self.get().__await__()

    async def __aenter__(self) -> Any:
        return await self.get()

    async def __aexit__(self, *_args: Any) -> None:
        await self.pool.release(self.raw)
class RecordingConnection:
    def __init__(self, raw: Any, records: list[tuple[str, tuple[Any, ...]]]):
        self.raw, self.records = raw, records

    async def _run(self, method: str, sql: str, *args: Any, **kwargs: Any) -> Any:
        self.records.append((sql, args))
        return await getattr(self.raw, method)(sql, *args, **kwargs)

    async def fetch(self, sql: str, *args: Any, **kwargs: Any) -> Any:
        return await self._run("fetch", sql, *args, **kwargs)

    async def fetchrow(self, sql: str, *args: Any, **kwargs: Any) -> Any:
        return await self._run("fetchrow", sql, *args, **kwargs)

    async def fetchval(self, sql: str, *args: Any, **kwargs: Any) -> Any:
        return await self._run("fetchval", sql, *args, **kwargs)

    async def execute(self, sql: str, *args: Any, **kwargs: Any) -> Any:
        return await self._run("execute", sql, *args, **kwargs)

    def __getattr__(self, name: str) -> Any:
        return getattr(self.raw, name)
class RecordingPool:
    def __init__(self, pool: Any):
        self.pool, self.records = pool, []

    def acquire(self) -> Acquire:
        return Acquire(self.pool, lambda raw: RecordingConnection(raw, self.records))

    async def release(self, conn: Any) -> None:
        await self.pool.release(getattr(conn, "raw", conn))

    def __getattr__(self, name: str) -> Any:
        return getattr(self.pool, name)
class BoundPool:
    def __init__(self, connection: Any):
        self.connection = connection

    def acquire(self) -> Any:
        connection = self.connection
        class Context:
            def __await__(self): return self.get().__await__()
            async def get(self): return connection
            async def __aenter__(self): return connection
            async def __aexit__(self, *_args): return None
        return Context()

    async def release(self, _connection: Any) -> None:
        return None
class FakeArq:
    def __init__(self, error: BaseException | None = None):
        self.error, self.calls = error, []

    async def enqueue_job(self, *args: Any, **kwargs: Any) -> Any:
        self.calls.append((args, kwargs))
        if self.error:
            raise self.error
        return SimpleNamespace(job_id=kwargs.get("_job_id", "synthetic-job"))

    async def delete(self, *_args: Any) -> int:
        return 1
@dataclass(frozen=True)
class Flow:
    name: str
    call: Callable[[], Awaitable[Any]]
def prepare_application_imports() -> None:
    sys.path.insert(0, str(BACKEND)) if str(BACKEND) not in sys.path else None
    import env_loader
    env_loader.load_repository_env = lambda override=False: False


def load_asyncpg() -> None:
    global asyncpg
    import asyncpg as driver
    asyncpg = driver
async def bootstrap(dsn: str) -> list[str]:
    prepare_application_imports()
    from db.migration_runner import run_migrations
    return await run_migrations(dsn)


async def provision_test_cluster(admin_dsn: str) -> None:
    connection = await asyncpg.connect(admin_dsn)
    try:
        await connection.execute("DO $$ BEGIN IF NOT EXISTS(SELECT 1 FROM pg_roles WHERE rolname='unihub_salary_export') THEN CREATE ROLE unihub_salary_export NOLOGIN NOSUPERUSER NOCREATEDB NOCREATEROLE NOINHERIT NOBYPASSRLS NOREPLICATION; END IF; END $$")
    finally:
        await connection.close()
STORE_SQL = """
INSERT INTO stores(site_code,locatie,firma,regional,asm,is_active,first_seen_month,last_seen_month)
SELECT CASE WHEN n<=26 THEN ($1::text[])[n] ELSE 'S'||lpad(n::text,3,'0') END,
       'Scale Store '||lpad(n::text,3,'0'), CASE WHEN n%2=0 THEN 'Mobicell' ELSE 'Mobiup' END,
       'REG-'||lpad(((n-1)%10+1)::text,2,'0'), 'ASM-'||lpad(((n-1)%20+1)::text,2,'0'), true, $2, '2026-07'
FROM generate_series(1,200) n;
INSERT INTO store_org_assignments(site_code,regional,asm,valid_from_month,is_current,source)
SELECT site_code,regional,asm,$1,true,'scale-authority' FROM stores;
INSERT INTO focus_products(item_code,item_name)
SELECT 'ITEM-'||lpad(n::text,4,'0'),'Scale Product '||n FROM generate_series(0,1999)n;
"""
SALES_SQL = """
WITH months AS (SELECT row_number() OVER() n,to_char(d,'YYYY-MM') m FROM generate_series(date '2026-07-01'-($2::int-1)*interval '1 month',date '2026-07-01',interval '1 month')d),
snaps AS (INSERT INTO import_snapshots(import_month,filename,is_month_final,rows_in_file,rows_imported,status,finished_at)
 SELECT m,'scale-'||m||'.xlsx',true,0,0,'completed',now() FROM months RETURNING id,import_month),
g AS (SELECT n,CASE WHEN n<=500000 THEN '2026-07' ELSE to_char(date '2026-07-01'-((n-500001)%$2)*interval '1 month','YYYY-MM') END m FROM generate_series(1,$1::bigint)n),
mapped AS (SELECT g.n,g.m,CASE WHEN ((identity-1+$5)%200)+1<=26 THEN ($4::text[])[((identity-1+$5)%200)+1] ELSE 'S'||lpad((((identity-1+$5)%200)+1)::text,3,'0') END site,identity receipt FROM
 (SELECT g.*,CASE WHEN g.n<=500000 THEN (g.n+1)/2 ELSE g.n END identity FROM g)g)
INSERT INTO sales_transactions(import_month,sale_date,site_code,bon_nr,item_code,item_name,brand,category,subcategory,quantity,unit_price,total_value,agent,is_cartela,is_return,snapshot_id)
SELECT x.m,CASE WHEN x.n<=500000 THEN date '2026-07-28' ELSE to_date(x.m||'-01','YYYY-MM-DD')+((x.receipt-1)%28)::int END,x.site,
 CASE WHEN x.n<=500000 THEN 'PAIR-'||lpad(x.receipt::text,9,'0') ELSE 'BON-'||x.n END,
 CASE WHEN x.n<=500000 AND x.n%2=1 THEN 'ITEM-'||lpad(((x.receipt-1+$5)%24)::text,4,'0') ELSE 'ITEM-'||lpad(((x.n-1+$5)%2000)::text,4,'0') END,
 'Scale Product '||((x.n-1)%2000),'SYNTH','Accesorii','Scale',1,10+(x.n%90),10+(x.n%90),
 'AGENT-'||lpad(((x.receipt-1+$5)%$3)::text,4,'0'),false,false,s.id FROM mapped x JOIN snaps s ON s.import_month=x.m;
UPDATE import_snapshots i SET rows_in_file=q.c,rows_imported=q.c FROM (SELECT import_month,count(*)::int c FROM sales_transactions GROUP BY import_month)q WHERE q.import_month=i.import_month;
"""
AUX_SQL = """
INSERT INTO store_targets(import_month,site_code,target_value,source_file)
SELECT m,site_code,200000,'scale-authority' FROM stores CROSS JOIN unnest(ARRAY['2026-07','2026-08'])m;
INSERT INTO salary_private.people(person_id,cnp,normalized_name,identity_source)
SELECT 'sp1_'||encode(digest('scale-person-'||n,'sha256'),'hex'),NULL,'scale agent '||n,'name' FROM generate_series(0,$1-1)n;
INSERT INTO salary_records(year,month,full_name,cnp,total_salary,company_name,site_code,locatie,person_id)
SELECT extract(year from d)::int,extract(month from d)::int,'Scale Agent '||a,NULL,3000+(a%500),
 CASE WHEN a%2=0 THEN 'Mobicell' ELSE 'Mobiup' END,(SELECT site_code FROM stores ORDER BY site_code OFFSET (a%200) LIMIT 1),
 'Scale Store','sp1_'||encode(digest('scale-person-'||a,'sha256'),'hex')
FROM generate_series(0,$1-1)a CROSS JOIN generate_series(date '2026-01-01',date '2026-07-01',interval '1 month')d;
INSERT INTO store_pnl_site_links(company_name,source_site_code,source_location_name,site_code,match_method,confidence,reviewed)
SELECT firma,site_code,locatie,site_code,'exact_code',1,true FROM stores;
INSERT INTO store_pnl_monthly(company_name,period,source_site_code,source_location_name,category_code,category_name,amount,data_kind,source_file,source_sha256)
SELECT s.firma,d,s.site_code,s.locatie,c,c,1000+(extract(month from d)*10),'actual','scale-pnl',repeat('a',64)
FROM stores s CROSS JOIN generate_series(date '2024-01-01',date '2026-07-01',interval '1 month')d CROSS JOIN unnest(ARRAY['v11','c11','c4','c5','c6'])c;
INSERT INTO incentive_campaigns(month,title,description) VALUES('2026-07','Scale incentive','synthetic');
INSERT INTO incentive_products(campaign_id,item_code,item_name,reward_value,valid_from,valid_to,category,subcategory,source_file)
SELECT id,'ITEM-'||lpad(n::text,4,'0'),'Scale Product',10,date '2026-07-01',date '2026-07-31','Accesorii','Scale','scale' FROM incentive_campaigns CROSS JOIN generate_series(0,23)n WHERE month='2026-07';
"""
FORECAST_SQL = """
WITH run AS (INSERT INTO ai_forecast_runs(forecast_month,source_month,model_name,model_mode,variant,status,metric,horizon,metadata)
 VALUES('2026-08','2026-07','scale-authority','offline','exact','completed','sales_value','current_month','{}') RETURNING id)
INSERT INTO ai_forecast_store_month(run_id,site_code,forecast_sales) SELECT run.id,s.site_code,220000 FROM run CROSS JOIN stores s;
"""


async def execute_script(conn: Any, sql: str, arguments: list[tuple[Any, ...]]) -> None:
    statements = [item.strip() for item in sql.split(";") if item.strip()]
    if len(statements) != len(arguments): raise RuntimeError("seed statement/argument contract drifted")
    for statement, args in zip(statements, arguments): await conn.execute(statement, *args)
async def seed_database(pool: asyncpg.Pool, profile: dict[str, int]) -> dict[str, int]:
    async with pool.acquire() as conn:
        before = await conn.fetchval("SELECT count(*) FROM sales_transactions")
        if before != 0:
            raise RuntimeError("AC-13 disposable database is not empty")
        first = 2026*12+6-(profile["months"]-1)
        first_month = f"{first//12:04d}-{first%12+1:02d}"
        await execute_script(conn, STORE_SQL, [(list(EXCEPTION_STORES), first_month), (first_month,), ()])
        await execute_script(conn, SALES_SQL, [(profile["rows"], profile["months"], profile["agents"], list(EXCEPTION_STORES), SEED), ()])
        await execute_script(conn, AUX_SQL, [(), (profile["agents"],), (profile["agents"],), (), (), (), ()])
        await conn.execute(FORECAST_SQL)
        from services.reporting_refresh import rebuild_reporting_all
        await rebuild_reporting_all(conn)
        await conn.execute("ANALYZE")
        row = await conn.fetchrow("SELECT count(*) rows,count(DISTINCT site_code) stores,count(DISTINCT agent) agents,count(DISTINCT import_month) months,count(DISTINCT item_code) products FROM sales_transactions")
        pairs = await conn.fetchval("SELECT count(*) FROM (SELECT sale_date,site_code,agent,bon_nr FROM sales_transactions WHERE import_month='2026-07' AND bon_nr LIKE 'PAIR-%' GROUP BY 1,2,3,4 HAVING sum(quantity)>=2)q")
        exceptions = await conn.fetchval("SELECT count(*) FROM stores WHERE site_code=ANY($1::text[])", list(EXCEPTION_STORES))
    result = {key: int(row[key]) for key in ("rows", "stores", "agents", "months", "products")}
    result.update(qualifying_receipts=int(pairs), target_exception_stores=int(exceptions))
    if result != {**profile, "products": 2_000, "qualifying_receipts": 250_000, "target_exception_stores": 26}:
        raise RuntimeError(f"AC-13 cardinality mismatch: {result}")
    return result
def write_campaign_config(path: Path) -> None:
    promotions = [{"key": f"scale-{n:02d}", "title": f"Scale {n:02d}", "rule_type": "selected_item_copurchase", "item_codes": [f"ITEM-{n:04d}"], "start_date": "2026-07-01", "end_date": "2026-07-31", "discount_rate": "0.20"} for n in range(24)]
    path.write_text(json.dumps({"promotions": promotions, "incentives": []}, sort_keys=True), encoding="utf-8")


async def dense_campaign_call(app: Any) -> Any:
    from services.request_deadline import RequestDeadline
    empty = os.environ["UNIHUB_HUB_SPECIALS_CONFIG"]
    os.environ["UNIHUB_HUB_SPECIALS_CONFIG"] = os.environ["UNIHUB_SCALE_DENSE_CAMPAIGN_CONFIG"]
    try:
        return await app.campaigns.get_promotions_incentives(date(2026, 7, 1), date(2026, 7, 31), None, None, None, None, None, view="all", deadline=RequestDeadline(2.5))
    finally:
        os.environ["UNIHUB_HUB_SPECIALS_CONFIG"] = empty
def exact_xlsx(size: int) -> bytes:
    from openpyxl import Workbook, load_workbook
    source = io.BytesIO(); workbook = Workbook(); workbook.properties.created = workbook.properties.modified = datetime(2026, 8, 12); workbook.active.append(["scale"]); workbook.save(source)
    entries = []
    with zipfile.ZipFile(io.BytesIO(source.getvalue())) as archive:
        entries = [(item.filename, archive.read(item.filename)) for item in archive.infolist()]
    def build(padding: int) -> bytes:
        out = io.BytesIO()
        with zipfile.ZipFile(out, "w", zipfile.ZIP_STORED) as archive:
            for name, content in entries: archive.writestr(zipfile.ZipInfo(name, (2026, 8, 12, 0, 0, 0)), content)
            archive.writestr(zipfile.ZipInfo("scale-padding.bin", (2026, 8, 12, 0, 0, 0)), b"0" * padding)
        return out.getvalue()
    empty = build(0); result = build(size - len(empty))
    if len(result) != size:
        raise RuntimeError("exact XLSX fixture size mismatch")
    load_workbook(io.BytesIO(result), read_only=True).close()
    return result
def load_app(pool: Any) -> SimpleNamespace:
    from repositories.campaigns import CampaignsRepository
    from repositories.dashboard import DashboardRepository
    from repositories.export_operations import ExportOperationsRepository
    from repositories.imports import ImportsRepository
    from repositories.salarii import SalariiRepository
    from repositories.store_pnl import StorePnlRepository
    from repositories.target_calculator import TargetCalculatorRepository
    from services.campaigns import CampaignsService
    from services.dashboard_service import DashboardService
    from services.export_operations import ExportOperationsService
    from services.imports import ImportsService
    from services.salarii import SalariiService
    from services.store_pnl import StorePnlService
    from services.target_calculator import TargetCalculatorService
    app = SimpleNamespace(
        dashboard=DashboardService(DashboardRepository(pool), pool), pnl=StorePnlService(StorePnlRepository(pool)),
        salary=SalariiService(SalariiRepository(pool)), target=TargetCalculatorService(TargetCalculatorRepository(pool)),
        campaigns=CampaignsService(CampaignsRepository(pool), pool), imports=ImportsService(ImportsRepository(pool), pool),
        exports=ExportOperationsService(pool), classes=(DashboardService, StorePnlService, SalariiService, TargetCalculatorService, CampaignsService, ImportsService, ExportOperationsService, CampaignsRepository, DashboardRepository, ExportOperationsRepository, ImportsRepository, SalariiRepository, StorePnlRepository, TargetCalculatorRepository),
    )
    app.runtime_sentinels = arm_runtime_sentinels(app)
    return app
def arm_runtime_sentinels(app: Any) -> list[str]:
    async def blocked(*_args: Any, **_kwargs: Any) -> None:
        raise RuntimeError("AC-13 runtime sentinel blocked forbidden mutation")
    boundaries = ((app.target, ("save_final_targets", "finalize", "export_excel")), (app.imports, ("promote_sales_generation", "import_promo_actuals", "process_promo_actuals")), (app.exports, ("reserve_salary", "download")))
    armed = []
    for service, names in boundaries:
        for name in names:
            if hasattr(service, name): setattr(service, name, blocked); armed.append(f"{type(service).__name__}.{name}")
    if len(armed) != 8: raise RuntimeError("AC-13 runtime sentinel surface drifted")
    return armed
async def target_call(pool: Any) -> Any:
    async with pool.acquire() as conn:
        tx = conn.transaction(); await tx.start()
        try:
            app = load_app(BoundPool(conn))
            return await app.target.calculate({"target_month": "2026-08", "cohort_month": "2026-07", "total_target": 40_000_000, "min_floor": 35_000, "seasonality_years": 3})
        finally:
            await tx.rollback()
def export_request() -> dict[str, Any]:
    return {"export_mode": "table", "dataset": "stores", "months": ["2026-07"], "daily_metrics": ["total_sales"], "dimensions": ["site_code", "locatie", "firma", "regional", "asm"], "metrics": ["total_sales"]}
async def export_call(pool: Any, owner: str, operation: str) -> Any:
    async with pool.acquire() as conn:
        tx = conn.transaction(); await tx.start()
        try:
            service = load_app(BoundPool(conn)).exports
            reserved = await service.reserve(export_request(), requested_by_sub=owner)
            if operation == "reserve": return reserved
            if operation == "status": return await service.status(reserved.id, requested_by_sub=owner)
            return await service.resumable(requested_by_sub=owner)
        finally:
            await tx.rollback()
async def import_call(app: Any, fixture: bytes) -> Any:
    from fastapi import UploadFile
    return await app.imports.import_sales(UploadFile(filename="scale.xlsx", file=io.BytesIO(fixture)), requested_by_sub="scale-authority")
def flows(pool: Any, fixture: bytes) -> list[Flow]:
    from services.request_deadline import RequestDeadline
    app = load_app(pool)
    deadline = lambda: RequestDeadline(2.5)
    return [
        Flow("dashboard_store", lambda: app.dashboard.get_performance_detail("2026-07", "store", "AFICOTRO", None, None, None, None, None, deadline=deadline())),
        Flow("dashboard_agent", lambda: app.dashboard.get_performance_detail("2026-07", "agent", "AGENT-0000", None, None, None, None, None, deadline=deadline())),
        Flow("dashboard_regional", lambda: app.dashboard.get_performance_detail("2026-07", "regional", "REG-01", None, None, None, None, None, deadline=deadline())),
        Flow("dashboard_history", lambda: app.dashboard.get_monthly_history("2026-07", 12, None, None, None, None, None, deadline=deadline())),
        Flow("dashboard_all", lambda: app.dashboard.get_dashboard_all("2026-07", None, None, None, None, None, deadline=deadline())),
        Flow("pnl_annual", lambda: app.pnl.annual(None, None)),
        Flow("salary_agents", lambda: app.salary.get_agents_summary(None, None, None, None, None, 2026, 7, 100, 0)),
        Flow("salary_records", lambda: app.salary.get_records(None, 2026, 7, None, 100, 0)),
        Flow("target_context", lambda: app.target.get_context()), Flow("target_calculate", lambda: target_call(pool)),
        Flow("campaign_dense", lambda: dense_campaign_call(app)),
        Flow("import_accept", lambda: import_call(app, fixture)),
        Flow("export_reserve", lambda: export_call(pool, "scale-reserve", "reserve")),
        Flow("export_status", lambda: export_call(pool, "scale-status", "status")),
        Flow("export_resumable", lambda: export_call(pool, "scale-resumable", "resumable")),
    ]
async def measure(flow: Flow) -> dict[str, Any]:
    for _ in range(WARMUPS): await flow.call()
    latencies, hashes = [], []
    for _ in range(MEASUREMENTS):
        started = time.perf_counter(); value = await flow.call()
        latencies.append((time.perf_counter() - started) * 1000); hashes.append(digest(value))
    p95 = nearest_rank(latencies, .95)
    if p95 >= LIMITS_MS[FLOW_KINDS[flow.name]]:
        raise RuntimeError(f"{flow.name} p95 {p95:.3f}ms exceeds gate")
    if len(set(hashes)) != 1:
        raise RuntimeError(f"{flow.name} business response is nondeterministic")
    return {"warmups": WARMUPS, "measurements": MEASUREMENTS, "latencies_ms": latencies, "p50_ms": nearest_rank(latencies, .50), "p95_ms": p95, "business_sha256": hashes[0]}
def explainable(records: list[tuple[str, tuple[Any, ...]]]) -> tuple[str, tuple[Any, ...]]:
    candidates = []
    for sql, args in records:
        upper = " ".join(sql.upper().split())
        if upper.startswith(("SELECT", "WITH")) and not any(token in upper for token in (" INSERT ", " UPDATE ", " DELETE ", "PG_ADVISORY", " FOR UPDATE")):
            candidates.append((sql, args))
    if not candidates:
        raise RuntimeError("real callable emitted no explainable SQL")
    return max(candidates, key=lambda item: len(item[0]))
def plan_nodes(plan: dict[str, Any]) -> list[dict[str, Any]]:
    return [plan] + [item for child in plan.get("Plans", []) for item in plan_nodes(child)]
async def explain_flow(pool: Any, flow_name: str, fixture: bytes) -> dict[str, Any]:
    recorder = RecordingPool(pool)
    selected = next(flow for flow in flows(recorder, fixture) if flow.name == flow_name)
    await selected.call(); sql, args = explainable(recorder.records)
    async with pool.acquire() as conn:
        payload = await conn.fetchval("EXPLAIN (ANALYZE, BUFFERS, FORMAT JSON) " + sql, *args)
    if isinstance(payload, str): payload = json.loads(payload)
    root = payload[0]["Plan"]; rows = max(float(root.get("Actual Rows", 0)), 1.0)
    bad = [node for node in plan_nodes(root) if node.get("Node Type") == "Seq Scan" and float(node.get("Plan Rows", 0)) > 100_000]
    if bad:
        raise RuntimeError(f"{flow_name} has prohibited large sequential scan")
    blocks = sum(float(root.get(key, 0)) for key in ("Shared Hit Blocks", "Shared Read Blocks"))
    return {"sql": sql, "args": canonical(args), "plan": payload, "sql_sha256": hashlib.sha256(sql.encode()).hexdigest(), "execution_ms_per_control_row": float(root.get("Actual Total Time", 0))/rows, "shared_blocks_per_control_row": blocks/rows}
async def queue_faults() -> dict[str, str]:
    import services.jobs as jobs
    original = jobs.get_arq_pool; outcomes = {}
    async def none(): return None
    async def connection_error(): raise ConnectionError("injected adapter failure")
    for name, adapter in (("none", none), ("connection_error", connection_error), ("uncertain", lambda: asyncio.sleep(0, result=FakeArq(ConnectionError("injected uncertain"))))):
        jobs.get_arq_pool = adapter
        try: await jobs.enqueue_complex_export(1)
        except Exception as exc: outcomes[name] = type(exc).__name__
        else: outcomes[name] = "unexpected-success"
    jobs.get_arq_pool = original
    if outcomes != {"none": "JobQueueUnavailableError", "connection_error": "ConnectionError", "uncertain": "JobPublishUncertainError"}:
        raise RuntimeError(f"queue fault contract mismatch: {outcomes}")
    return outcomes
async def lock_contention(pool: Any) -> str:
    from repositories.export_operations import ExportOperationsRepository
    async with pool.acquire() as holder:
        await holder.execute("SELECT pg_advisory_lock(hashtextextended($1,0))", "unihub:exports:active-capacity")
        try:
            await ExportOperationsRepository(pool).reserve(kind="daily_metrics", request_payload=export_request(), request_sha256=digest(export_request()), requested_by_sub="scale-lock")
        except Exception as exc:
            outcome = type(exc).__name__
        else:
            outcome = "unexpected-success"
        finally:
            await holder.execute("SELECT pg_advisory_unlock(hashtextextended($1,0))", "unihub:exports:active-capacity")
    if outcome == "unexpected-success": raise RuntimeError("advisory lock contention did not fail closed")
    return outcome
async def concurrency_gate(pool: Any, fixture: bytes, reference: str) -> dict[str, Any]:
    from services.request_deadline import RequestDeadline
    app = load_app(pool)
    reads = [app.dashboard.get_performance_detail("2026-07", "store", "AFICOTRO", None, None, None, None, None, deadline=RequestDeadline(2.5)) for _ in range(8)]
    exports = [export_call(pool, f"scale-concurrent-{n}", "reserve") for n in range(3)]
    results = await asyncio.gather(*reads, *exports, import_call(app, fixture))
    if any(digest(value) != reference for value in results[:8]): raise RuntimeError("concurrent dashboard hash mismatch")
    return {"dashboard_reads": 8, "distinct_owner_reserves": 3, "imports": 1, "hashes": [digest(value) for value in results]}
def dependencies(app: Any) -> list[dict[str, str]]:
    paths = {Path(inspect.getsourcefile(item) or "") for item in app.classes}
    paths.update({BACKEND / "services/reporting_refresh.py", BACKEND / "services/jobs.py", BACKEND / "services/sales_artifacts.py", BACKEND / "db/schema_v2.sql", BACKEND / "db/migrations/manifest.json", BACKEND / "db/migration_runner.py"})
    paths.update((BACKEND / item) for item in REQUIRED_B)
    for pattern in DEPENDENCY_GLOBS: paths.update(BACKEND.glob(pattern))
    loaded = [Path(raw).resolve() for module in tuple(sys.modules.values()) if (raw := getattr(module, "__file__", None))]
    paths.update(path for path in loaded if path.suffix == ".py" and path.is_relative_to(BACKEND) and "tests" not in path.parts)
    return [{"path": str(path.relative_to(ROOT)), "sha256": hashlib.sha256(path.read_bytes()).hexdigest()} for path in sorted(paths) if path.is_file()]
async def environment(pool: Any) -> dict[str, Any]:
    async with pool.acquire() as conn:
        pg = await conn.fetchrow("SELECT version() version,current_setting('shared_buffers') shared_buffers,current_setting('work_mem') work_mem,current_setting('max_connections') max_connections")
    memory = next(int(line.split()[1])*1024 for line in Path("/proc/meminfo").read_text().splitlines() if line.startswith("MemAvailable:"))
    return {"cpu": platform.processor(), "cores": os.cpu_count(), "python": sys.version, "ram_available_bytes": memory, "disk_available_bytes": shutil.disk_usage("/").free, "postgresql": canonical(dict(pg)), "config": {"pool_max": 24, "dashboard_deadline_ms": 2500, "statement_timeout_ms": 1_800_000, "lock_timeout_ms": 1500}}
PLAN_FLOWS = ("dashboard_store", "dashboard_agent", "dashboard_regional", "dashboard_history", "pnl_annual", "salary_agents", "salary_records", "target_calculate", "campaign_dense")
async def run_profile(admin_dsn: str, name: str, evidence_dir: Path) -> dict[str, Any]:
    dbname = f"test_retail_scale_{name}_{uuid.uuid4().hex[:10]}"; dsn = database_dsn(admin_dsn, dbname)
    admin = await asyncpg.connect(admin_dsn)
    await admin.execute(f'CREATE DATABASE "{dbname}"')
    pool = None; spool = None
    try:
        await bootstrap(dsn)
        pool = await asyncpg.create_pool(dsn, min_size=2, max_size=24, command_timeout=1800, server_settings={"statement_timeout": "1800000", "lock_timeout": "1500", "idle_in_transaction_session_timeout": "1800000", "application_name": "unihub-retail-ac13"})
        async with pool.acquire() as conn: migrations = [dict(row) for row in await conn.fetch("SELECT filename,checksum FROM schema_migrations ORDER BY filename")]
        if len(migrations) != 69 or not migrations[-1]["filename"].startswith("069_"): raise RuntimeError("production migration bootstrap is not exactly 001..069")
        cardinality = await seed_database(pool, PROFILES[name]); config = evidence_dir / f"campaigns-{name}.json"; empty_config = evidence_dir / f"campaigns-empty-{name}.json"; write_campaign_config(config); empty_config.write_text('{"incentives":[],"promotions":[]}', encoding="utf-8")
        spool = Path(tempfile.mkdtemp(prefix=f"unihub-ac13-{name}-")); os.environ.update({"DATABASE_URL": dsn, "UNIHUB_HUB_SPECIALS_CONFIG": str(empty_config), "UNIHUB_SCALE_DENSE_CAMPAIGN_CONFIG": str(config), "SALES_IMPORT_SPOOL_DIR": str(spool), "MAX_SALES_UPLOAD_BYTES": str(UPLOAD_BYTES)})
        fixture = exact_xlsx(UPLOAD_BYTES); app = load_app(pool)
        import services.imports as imports_module
        import services.jobs as jobs
        from services.jobs import JobResult, JobStatus
        fake = FakeArq(); original_pool, original_status = jobs.get_arq_pool, imports_module.get_job_status
        async def fake_pool(): return fake
        async def fake_status(job_id: str): return JobResult(job_id=job_id, status=JobStatus.QUEUED)
        jobs.get_arq_pool, imports_module.get_job_status = fake_pool, fake_status
        try:
            from fastapi import HTTPException, UploadFile
            try: await app.imports.import_sales(UploadFile(filename="too-large.xlsx", file=io.BytesIO(fixture+b"x")))
            except HTTPException as exc:
                if exc.status_code != 413: raise
            else: raise RuntimeError("max upload +1 was accepted")
            measured = {flow.name: await measure(flow) for flow in flows(pool, fixture)}
            plans = {flow: await explain_flow(pool, flow, fixture) for flow in PLAN_FLOWS}
            concurrent = await concurrency_gate(pool, fixture, measured["dashboard_store"]["business_sha256"])
            faults = await queue_faults(); contention = await lock_contention(pool)
        finally:
            jobs.get_arq_pool, imports_module.get_job_status = original_pool, original_status
        return {"profile": name, "seed": SEED, "cardinality": cardinality, "migration_manifest": migrations, "upload": {"accepted_bytes": len(fixture), "rejected_bytes": len(fixture)+1, "sha256": hashlib.sha256(fixture).hexdigest(), "content_address_replays": MEASUREMENTS+WARMUPS}, "measurements": measured, "plans": plans, "concurrency": concurrent, "queue_faults": faults, "lock_contention": contention, "runtime_sentinels": app.runtime_sentinels, "environment": await environment(pool), "dependencies": dependencies(app)}
    finally:
        if pool: await pool.close()
        if spool: shutil.rmtree(spool, ignore_errors=True)
        await admin.execute("SELECT pg_terminate_backend(pid) FROM pg_stat_activity WHERE datname=$1", dbname)
        await admin.execute(f'DROP DATABASE IF EXISTS "{dbname}"')
        await admin.close()
def compare_profiles(results: dict[str, Any]) -> dict[str, Any]:
    ratios = {}
    for flow in PLAN_FLOWS:
        small, large = results["2x"]["plans"][flow], results["5x"]["plans"][flow]
        ratios[flow] = {key: large[key] / max(small[key], 1e-12) for key in ("execution_ms_per_control_row", "shared_blocks_per_control_row")}
        if any(value > 2.5 for value in ratios[flow].values()): raise RuntimeError(f"{flow} 5x/2x plan ratio exceeds 2.5")
    latency = {flow: results["5x"]["measurements"][flow]["p95_ms"] / max(results["2x"]["measurements"][flow]["p95_ms"], 1e-12) for flow in FLOW_KINDS}
    if any(value > 2.5 for value in latency.values()): raise RuntimeError("5x/2x latency ratio exceeds 2.5")
    return {"plan_ratios": ratios, "latency_p95_ratios": latency}
def verify_profile_evidence(expected: dict[str, int], item: dict[str, Any]) -> None:
    if item.get("cardinality") != {**expected, "products": 2_000, "qualifying_receipts": 250_000, "target_exception_stores": 26}: raise RuntimeError("invalid profile cardinality evidence")
    if set(item.get("measurements", {})) != set(FLOW_KINDS): raise RuntimeError("incomplete callable evidence")
    if any(len(value.get("latencies_ms", [])) != MEASUREMENTS for value in item["measurements"].values()): raise RuntimeError("manual/fake measurement evidence")
    if set(item.get("plans", {})) != set(PLAN_FLOWS) or len(item.get("dependencies", [])) < 75: raise RuntimeError("incomplete real-plan authority")
    if item.get("queue_faults") != {"none": "JobQueueUnavailableError", "connection_error": "ConnectionError", "uncertain": "JobPublishUncertainError"}: raise RuntimeError("invalid queue-fault evidence")
    if item.get("upload", {}).get("accepted_bytes") != UPLOAD_BYTES or item.get("upload", {}).get("rejected_bytes") != UPLOAD_BYTES+1: raise RuntimeError("invalid upload-boundary evidence")
    if len(item.get("runtime_sentinels", [])) != 8 or item.get("concurrency", {}).get("dashboard_reads") != 8: raise RuntimeError("incomplete runtime/concurrency evidence")
    if len(item.get("migration_manifest", [])) != 69 or not item["migration_manifest"][-1]["filename"].startswith("069_"): raise RuntimeError("incomplete migration evidence")


def verify_evidence(payload: dict[str, Any]) -> None:
    if payload.get("contract") != "AC-13-real-scale-v1" or payload.get("seed") != SEED: raise RuntimeError("invalid evidence contract")
    profiles = payload.get("profiles")
    if not isinstance(profiles, dict) or set(profiles) != set(PROFILES): raise RuntimeError("both exact profiles are required")
    for name, expected in PROFILES.items(): verify_profile_evidence(expected, profiles[name])
    claimed = payload.get("evidence_sha256"); material = {k: v for k, v in payload.items() if k != "evidence_sha256"}
    if claimed != digest(material): raise RuntimeError("evidence digest mismatch")
def self_test() -> None:
    os.environ.update({"UNIHUB_TEST_DATABASE": "1", "UNIHUB_RUNNING_TESTS": "1"})
    safe_dsn("postgresql://127.0.0.1:15432/test_scale")
    for bad in ("postgresql://server:15432/test_scale", "postgresql://127.0.0.1:5432/test_scale", "postgresql://127.0.0.1:15432/unihub"):
        try: safe_dsn(bad)
        except RuntimeError: pass
        else: raise RuntimeError("live DSN self-test failed")
    assert_static_authority(Path(__file__).read_text(encoding="utf-8"))
    try: assert_static_authority("service.reserve_salary()")
    except RuntimeError: pass
    else: raise RuntimeError("forbidden call self-test failed")
    try: verify_evidence({"result": "PASS"})
    except RuntimeError: pass
    else: raise RuntimeError("manual PASS self-test failed")
    print("AC-13 authority self-test: PASS")
async def async_main(args: argparse.Namespace) -> None:
    load_asyncpg()
    fail_closed_b(); assert_static_authority(Path(__file__).read_text(encoding="utf-8"))
    if args.seed != SEED or args.exact_max_upload != UPLOAD_BYTES or args.profiles.split(",") != ["2x", "5x"]: raise RuntimeError("AC-13 exact seed/profiles/upload contract required")
    admin_dsn = safe_dsn(os.environ.get("UNIHUB_SCALE_ADMIN_DSN", "")); args.evidence.mkdir(parents=True, exist_ok=True); await provision_test_cluster(admin_dsn)
    results = {name: await run_profile(admin_dsn, name, args.evidence) for name in PROFILES}
    payload = {"contract": "AC-13-real-scale-v1", "seed": SEED, "profiles": results, "comparison": compare_profiles(results), "result": "PASS"}
    payload["evidence_sha256"] = digest(payload); verify_evidence(payload)
    target = args.evidence / "ac-13-scale-evidence.json"; temporary = target.with_suffix(".tmp")
    temporary.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str)+"\n", encoding="utf-8"); temporary.replace(target)
    print(json.dumps({"result": "PASS", "evidence": str(target), "sha256": payload["evidence_sha256"]}, sort_keys=True))
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(); parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--seed", type=int, default=SEED); parser.add_argument("--profiles", default="2x,5x")
    parser.add_argument("--exact-max-upload", type=int, default=UPLOAD_BYTES); parser.add_argument("--evidence", type=Path, default=ROOT/"evidence"/"ac-13")
    return parser.parse_args()
def main() -> int:
    args = parse_args()
    try:
        if args.self_test: self_test()
        else: asyncio.run(async_main(args))
    except Exception as exc:
        print(json.dumps({"result": "FAIL", "error_type": type(exc).__name__, "error": str(exc)}), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
