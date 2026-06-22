#!/usr/bin/env python3
"""Genereaza simularea grilei salariale propuse pentru 2025-12..2026-05.

Surse:
- KPI si targete: PostgreSQL Retail;
- salarii istorice: fisierele HR din /opt/Mobiup/docs/comisioane;
- reguli: /opt/Mobiup/docs/grila 2026.docx.

Salariul istoric comparabil este TOTAL SALARIU minus ore suplimentare, fara
bonuri de masa. Randurile asociate aceluiasi agent se agrega inainte de pragul
de 2.400 lei. ePay este zero deoarece nu exista sursa per agent.
"""

from __future__ import annotations

import argparse
import asyncio
import itertools
import os
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import asyncpg
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_DIR = BACKEND_DIR.parent
WORKSPACE_DIR = REPO_DIR.parent
DOCS_DIR = WORKSPACE_DIR / "docs"
COMISIOANE_DIR = DOCS_DIR / "comisioane"
DEFAULT_OUTPUT = DOCS_DIR / "simulare_grila_noua_2025-12_2026-05.xlsx"
DEFAULT_COMPARISON_OUTPUT = (
    DOCS_DIR / "comparatie_salarii_istorice_vs_grila_noua_2025-12_2026-05.xlsx"
)
OLD_GRID_MAY_ARCHIVE = (
    WORKSPACE_DIR
    / "grile-salarii"
    / "outputs"
    / "archive"
    / "Mai 2026"
    / "Grile - Mai 2026.zip"
)
PERIODS = ("2025-12", "2026-01", "2026-02", "2026-03", "2026-04", "2026-05")

sys.path.insert(0, str(BACKEND_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from import_salary_records import (  # noqa: E402
    LOCATION_ALIASES,
    company_key,
    format_cnp,
    normalize_text,
)
from services.grile_agent_targets import candidate_agent_codes  # noqa: E402

load_dotenv(REPO_DIR / ".env")

BASE_SALARY = 2700.0
MEAL_TICKETS = 480.0
ACCESSORY_COMMISSION_GATE = 80.0
FULL_MONTH_SALARY_FLOOR = 2400.0
COMMISSION_SCENARIOS = (
    ("2,5%", 0.025),
    ("2,7%", 0.027),
    ("3,0%", 0.030),
)


def commission_column(scenario: str) -> str:
    return f"Comision accesorii {scenario}"


def simulated_without_meals_column(scenario: str) -> str:
    return f"Simulat {scenario} fara bonuri"


def simulated_with_meals_column(scenario: str) -> str:
    return f"Simulat {scenario} cu bonuri"


def difference_column(scenario: str) -> str:
    return f"Diferenta {scenario} vs istoric"


def difference_pct_column(scenario: str) -> str:
    return f"Diferenta {scenario} %"

EXTRA_LOCATION_ALIASES: dict[tuple[str, str], str | None] = {
    ("Mobiup", "BAIA MARE VIVO"): "BMAREVIVO",
    ("Mobiup", "BRASOV CARREFOUR"): "CRFBV",
    ("Mobiup", "ORHIDEEA"): "CRFORH1",
    ("Mobiup", "SIBIU CITY"): "SBPROMEN",
    ("Mobicell", "AFI PLOIESTI"): "PLAFIPL",
    ("Mobicell", "IASI MOLDOVA"): "MOLDMLL",
}

LOCAL_SALARY_FILES: dict[tuple[int, int, str], str] = {
    (2026, 3, "Mobiup"): "MOBIUP COMISIOANE AGENTI MARTIE.xls",
    (2026, 3, "Mobicell"): "COMISIOANE AGENTI Mobicell martie.xls",
    (2026, 4, "Mobiup"): "MOBIUP COMISIOANE AGENTI APRILIE.xls",
    (2026, 4, "Mobicell"): "COMISIOANE AGENTI Mobicell aprilie.xls",
    (2026, 5, "Mobiup"): "MOBI COMISIOANE AGENTI MAI.xls",
    (2026, 5, "Mobicell"): "COMISIOANE AGENTI Mobicell mai.xls",
}

ARCHIVE_MONTH_WORDS = {
    (2025, 12): "DECEMBRIE",
    (2026, 1): "IANUARIE",
    (2026, 2): "FEBRUARIE",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
OK_FILL = PatternFill("solid", fgColor="D9EAD3")


@dataclass
class SalarySourceRow:
    period: str
    company: str
    full_name: str
    normalized_name: str
    cnp: str
    location: str
    normalized_location: str
    site_code: str | None
    total_salary: float
    overtime: float
    adjusted_salary: float
    meal_tickets: float
    source_file: str
    agent: str | None = None
    match_method: str = "unmatched"
    match_confidence: float = 0.0


def _number(value: Any) -> float:
    if value is None or pd.isna(value):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _normalized_company(value: str) -> str:
    return "Mobicell" if company_key(value) == "MOBICELL" else "Mobiup"


def _parse_salary_dataframe(
    df: pd.DataFrame,
    *,
    year: int,
    month: int,
    company: str,
    source_file: str,
) -> list[SalarySourceRow]:
    normalized_columns = {normalize_text(column): column for column in df.columns}
    name_column = normalized_columns.get("NUME PRENUME")
    location_column = normalized_columns.get("DENUMIRE LOCATIE")
    cnp_column = normalized_columns.get("CNP")
    total_column = normalized_columns.get("TOTAL SALARIU")
    overtime_column = next(
        (column for column in df.columns if "ORE SUPLIM" in normalize_text(column)),
        None,
    )
    meal_column = next(
        (
            column
            for column in df.columns
            if normalize_text(column).startswith("BONURI MASA")
        ),
        None,
    )
    missing = [
        label
        for label, column in (
            ("Nume Prenume", name_column),
            ("Denumire locatie", location_column),
            ("CNP", cnp_column),
            ("TOTAL SALARIU", total_column),
        )
        if column is None
    ]
    if missing:
        raise ValueError(f"{source_file}: coloane lipsa: {', '.join(missing)}")

    rows: list[SalarySourceRow] = []
    for _, row in df.iterrows():
        full_name = str(row[name_column]).strip() if pd.notna(row[name_column]) else ""
        if not full_name or normalize_text(full_name).startswith("TOTAL"):
            continue
        location = (
            str(row[location_column]).strip()
            if pd.notna(row[location_column])
            else ""
        )
        total_salary = _number(row[total_column])
        overtime = _number(row[overtime_column]) if overtime_column else 0.0
        meal_tickets = _number(row[meal_column]) if meal_column else 0.0
        rows.append(
            SalarySourceRow(
                period=f"{year}-{month:02d}",
                company=company,
                full_name=full_name,
                normalized_name=normalize_text(full_name),
                cnp=format_cnp(row[cnp_column]),
                location=location,
                normalized_location=normalize_text(location),
                site_code=None,
                total_salary=total_salary,
                overtime=overtime,
                adjusted_salary=total_salary - overtime,
                meal_tickets=meal_tickets,
                source_file=source_file,
            )
        )
    return rows


def load_salary_source_rows() -> list[SalarySourceRow]:
    rows: list[SalarySourceRow] = []
    archive_path = COMISIOANE_DIR / "salarii-istoric.zip"
    with ZipFile(archive_path) as archive:
        for (year, month), month_word in ARCHIVE_MONTH_WORDS.items():
            for company, marker in (
                ("Mobiup", "MOBI COMISIOANE"),
                ("Mobicell", "MOBICELL"),
            ):
                matches = [
                    name
                    for name in archive.namelist()
                    if str(year) in name
                    and month_word in name.upper()
                    and marker in name.upper()
                    and name.lower().endswith((".xls", ".xlsx"))
                ]
                if len(matches) != 1:
                    raise ValueError(
                        f"Arhiva salarii: {year}-{month:02d} {company}, "
                        f"asteptat un fisier, gasit {matches}"
                    )
                content = archive.read(matches[0])
                rows.extend(
                    _parse_salary_dataframe(
                        pd.read_excel(BytesIO(content), sheet_name=0),
                        year=year,
                        month=month,
                        company=company,
                        source_file=matches[0],
                    )
                )

    for (year, month, company), file_name in LOCAL_SALARY_FILES.items():
        path = COMISIOANE_DIR / file_name
        rows.extend(
            _parse_salary_dataframe(
                pd.read_excel(path, sheet_name=0),
                year=year,
                month=month,
                company=company,
                source_file=file_name,
            )
        )
    return rows


async def fetch_data(conn: asyncpg.Connection) -> dict[str, list[dict[str, Any]]]:
    base_rows = await conn.fetch(
        """
        WITH location_days AS (
            SELECT import_month, site_code, COUNT(DISTINCT sale_date)::INT AS location_days
            FROM reporting_agent_day
            WHERE import_month = ANY($1::TEXT[])
            GROUP BY import_month, site_code
        ),
        agent_days AS (
            SELECT import_month, agent, COUNT(DISTINCT sale_date)::INT AS agent_days
            FROM reporting_agent_day
            WHERE import_month = ANY($1::TEXT[])
            GROUP BY import_month, agent
        )
        SELECT
            ram.import_month,
            ram.site_code,
            ram.locatie,
            ram.firma,
            ram.regional,
            ram.asm,
            ram.agent,
            ram.total_sales,
            ram.total_quantity,
            ram.focus_quantity,
            ram.receipt_count,
            ram.receipt_2plus_count,
            ram.working_days,
            COALESCE(ld.location_days, 0) AS location_days,
            COALESCE(ad.agent_days, 0) AS agent_days,
            COALESCE(st.target_value, 0) AS store_target
        FROM reporting_agent_month ram
        LEFT JOIN location_days ld
          ON ld.import_month = ram.import_month
         AND ld.site_code = ram.site_code
        LEFT JOIN agent_days ad
          ON ad.import_month = ram.import_month
         AND ad.agent = ram.agent
        LEFT JOIN store_targets st
          ON st.import_month = ram.import_month
         AND st.site_code = ram.site_code
        WHERE ram.import_month = ANY($1::TEXT[])
          AND ram.agent IS NOT NULL
          AND TRIM(ram.agent) NOT IN ('', '-')
          AND ram.agent NOT ILIKE 'TR%'
        ORDER BY ram.import_month, ram.agent, ram.total_sales DESC
        """,
        list(PERIODS),
    )
    history_rows = await conn.fetch(
        """
        SELECT import_month, site_code, asm, agent, total_sales, working_days
        FROM reporting_agent_month
        WHERE import_month BETWEEN '2025-09' AND '2026-05'
          AND agent IS NOT NULL
          AND TRIM(agent) NOT IN ('', '-')
          AND agent NOT ILIKE 'TR%'
        """
    )
    premium_rows = await conn.fetch(
        """
        WITH premium_lines AS (
            SELECT DISTINCT
                st.id,
                st.import_month,
                st.agent,
                st.quantity,
                pgm.is_premium_glass AS is_premium
            FROM sales_transactions st
            JOIN premium_glass_item_models pgm ON pgm.item_code = st.item_code
            WHERE st.import_month = ANY($1::TEXT[])
              AND LOWER(TRIM(COALESCE(st.category, ''))) = 'folii sticla'
              AND st.quantity > 0
              AND st.agent IS NOT NULL
              AND TRIM(st.agent) NOT IN ('', '-')
              AND st.agent NOT ILIKE 'TR%'
        )
        SELECT
            import_month,
            agent,
            COALESCE(SUM(quantity), 0)::INT AS glass_qty,
            COALESCE(SUM(quantity) FILTER (WHERE is_premium), 0)::INT AS premium_glass_qty
        FROM premium_lines
        GROUP BY import_month, agent
        """,
        list(PERIODS),
    )
    sim_rows = await conn.fetch(
        """
        SELECT
            import_month,
            agent,
            GREATEST(
                COALESCE(SUM(quantity) FILTER (
                    WHERE item_code ~* '^(SIM|SPN)'
                       OR item_name ILIKE 'CARTELA %'
                       OR item_name ILIKE 'SIM %'
                ), 0),
                0
            )::INT AS sim_qty
        FROM sales_transactions
        WHERE import_month = ANY($1::TEXT[])
          AND agent IS NOT NULL
          AND TRIM(agent) NOT IN ('', '-')
          AND agent NOT ILIKE 'TR%'
          AND item_name NOT ILIKE 'PACHET%'
        GROUP BY import_month, agent
        """,
        list(PERIODS),
    )
    stores = await conn.fetch(
        "SELECT site_code, locatie, firma FROM stores ORDER BY site_code"
    )
    known_names = await conn.fetch(
        """
        SELECT site_code, agent, source_agent_name
        FROM agent_targets
        WHERE source_agent_name IS NOT NULL
          AND TRIM(source_agent_name) != ''
        """
    )
    grid_registry = await conn.fetch(
        """
        SELECT site_code, registry_key
        FROM grile_sheets
        WHERE is_active = TRUE
        """
    )
    return {
        "base": [dict(row) for row in base_rows],
        "history": [dict(row) for row in history_rows],
        "premium": [dict(row) for row in premium_rows],
        "sim": [dict(row) for row in sim_rows],
        "stores": [dict(row) for row in stores],
        "known_names": [dict(row) for row in known_names],
        "grid_registry": [dict(row) for row in grid_registry],
    }


def _month_index(period: str) -> int:
    year, month = map(int, period.split("-"))
    return year * 12 + month


def _median(values: list[float]) -> float | None:
    clean = sorted(value for value in values if value > 0)
    if not clean:
        return None
    middle = len(clean) // 2
    if len(clean) % 2:
        return clean[middle]
    return (clean[middle - 1] + clean[middle]) / 2


def _qualitative_points(
    *,
    daily_vs_colleague_pct: float | None,
    receipt_2plus_pct: float | None,
    focus_pct: float | None,
    average_item_value: float | None,
    premium_glass_pct: float | None,
) -> tuple[int, int, int, int, int]:
    if daily_vs_colleague_pct is None or daily_vs_colleague_pct < 90:
        daily_points = 0
    elif daily_vs_colleague_pct <= 110:
        daily_points = 1
    else:
        daily_points = 3

    if receipt_2plus_pct is None or receipt_2plus_pct < 25:
        receipt_points = 0
    elif receipt_2plus_pct < 30:
        receipt_points = 1
    elif receipt_2plus_pct < 35:
        receipt_points = 2
    else:
        receipt_points = 3

    if focus_pct is None or focus_pct < 6:
        focus_points = 0
    elif focus_pct < 7:
        focus_points = 1
    elif focus_pct < 8:
        focus_points = 2
    else:
        focus_points = 3

    if average_item_value is None or average_item_value < 80:
        value_points = 0
    elif average_item_value < 90:
        value_points = 1
    elif average_item_value < 100:
        value_points = 2
    else:
        value_points = 3

    if premium_glass_pct is None or premium_glass_pct < 30:
        premium_points = 0
    elif premium_glass_pct <= 40:
        premium_points = 1
    elif premium_glass_pct <= 50:
        premium_points = 2
    else:
        premium_points = 3

    return (
        daily_points,
        receipt_points,
        focus_points,
        value_points,
        premium_points,
    )


def _qualitative_bonus(points: tuple[int, int, int, int, int]) -> float:
    if any(point == 0 for point in points):
        return 0.0
    total = sum(points)
    if total == 15:
        return 300.0
    if total >= 13:
        return 200.0
    if total >= 11:
        return 100.0
    return 0.0


def _target_bonus(target_pct: float | None) -> float:
    if target_pct is None or target_pct < 100:
        return 0.0
    if target_pct < 110:
        return 200.0
    if target_pct < 120:
        return 300.0
    return 400.0


def build_simulation(data: dict[str, list[dict[str, Any]]]) -> pd.DataFrame:
    base = data["base"]
    history = data["history"]
    premium_map = {
        (row["import_month"], row["agent"]): row for row in data["premium"]
    }
    sim_map = {(row["import_month"], row["agent"]): row for row in data["sim"]}

    site_month_daily: dict[tuple[str, str], dict[str, float]] = defaultdict(dict)
    for row in base:
        if row["working_days"]:
            site_month_daily[(row["import_month"], row["site_code"])][row["agent"]] = (
                float(row["total_sales"]) / int(row["working_days"])
            )

    history_by_site: dict[str, list[dict[str, Any]]] = defaultdict(list)
    history_by_asm: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in history:
        history_by_site[row["site_code"]].append(row)
        history_by_asm[row["asm"]].append(row)

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in base:
        grouped[(row["import_month"], row["agent"])].append(row)

    output: list[dict[str, Any]] = []
    for (period, agent), rows in sorted(grouped.items()):
        primary = max(rows, key=lambda item: float(item["total_sales"]))
        total_sales = sum(float(row["total_sales"]) for row in rows)
        total_quantity = sum(int(row["total_quantity"]) for row in rows)
        focus_quantity = sum(int(row["focus_quantity"]) for row in rows)
        receipt_count = sum(int(row["receipt_count"]) for row in rows)
        receipt_2plus_count = sum(int(row["receipt_2plus_count"]) for row in rows)
        working_days = max(int(row["agent_days"]) for row in rows)
        target_value = sum(
            float(row["store_target"]) * int(row["working_days"]) / int(row["location_days"])
            if int(row["location_days"]) > 0
            else 0.0
            for row in rows
        )
        target_pct = total_sales * 100 / target_value if target_value > 0 else None
        daily_average = total_sales / working_days if working_days > 0 else None

        peers = [
            value
            for peer_agent, value in site_month_daily[
                (period, primary["site_code"])
            ].items()
            if peer_agent != agent
        ]
        daily_reference = _median(peers)
        daily_reference_type = "colegi"
        period_idx = _month_index(period)
        if daily_reference is None:
            location_history = [
                row
                for row in history_by_site[primary["site_code"]]
                if period_idx - 3 <= _month_index(row["import_month"]) <= period_idx - 1
            ]
            history_days = sum(int(row["working_days"]) for row in location_history)
            if history_days > 0:
                daily_reference = (
                    sum(float(row["total_sales"]) for row in location_history)
                    / history_days
                )
                daily_reference_type = "istoric_locatie"
        if daily_reference is None:
            manager_history = [
                row
                for row in history_by_asm[primary["asm"]]
                if period_idx - 3 <= _month_index(row["import_month"]) <= period_idx - 1
            ]
            history_days = sum(int(row["working_days"]) for row in manager_history)
            if history_days > 0:
                daily_reference = (
                    sum(float(row["total_sales"]) for row in manager_history)
                    / history_days
                )
                daily_reference_type = "media_manager"

        daily_vs_colleague_pct = (
            daily_average * 100 / daily_reference
            if daily_average is not None and daily_reference
            else None
        )
        receipt_2plus_pct = (
            receipt_2plus_count * 100 / receipt_count if receipt_count > 0 else None
        )
        focus_pct = (
            focus_quantity * 100 / total_quantity if total_quantity > 0 else None
        )
        average_item_value = (
            total_sales / total_quantity if total_quantity > 0 else None
        )
        premium = premium_map.get((period, agent), {})
        glass_qty = int(premium.get("glass_qty") or 0)
        premium_glass_qty = int(premium.get("premium_glass_qty") or 0)
        premium_glass_pct = (
            premium_glass_qty * 100 / glass_qty if glass_qty > 0 else 0.0
        )
        sim_qty = int(sim_map.get((period, agent), {}).get("sim_qty") or 0)

        qualitative_points = _qualitative_points(
            daily_vs_colleague_pct=daily_vs_colleague_pct,
            receipt_2plus_pct=receipt_2plus_pct,
            focus_pct=focus_pct,
            average_item_value=average_item_value,
            premium_glass_pct=premium_glass_pct,
        )
        commissionable_sales = (
            total_sales
            if target_pct is not None and target_pct >= ACCESSORY_COMMISSION_GATE
            else 0.0
        )
        target_bonus = _target_bonus(target_pct)
        qualitative_bonus = _qualitative_bonus(qualitative_points)
        sim_bonus = sim_qty * 3.0
        epay_bonus = 0.0
        fixed_salary_components = (
            BASE_SALARY
            + target_bonus
            + qualitative_bonus
            + sim_bonus
            + epay_bonus
        )

        output_row = {
                "Luna": period,
                "Agent cod": agent,
                "Firma": primary["firma"],
                "ASM": primary["asm"],
                "Site principal": primary["site_code"],
                "Locatie principala": primary["locatie"],
                "Site-uri": ", ".join(sorted({row["site_code"] for row in rows})),
                "Locatii": " / ".join(sorted({row["locatie"] for row in rows})),
                "Nr magazine": len(rows),
                "Vanzari accesorii": round(total_sales, 2),
                "Target agent": round(target_value, 2),
                "Realizare target %": round(target_pct, 2) if target_pct is not None else None,
                "Zile vanzare": working_days,
                "Bonuri": receipt_count,
                "Produse": total_quantity,
                "Medie zilnica": round(daily_average, 2) if daily_average is not None else None,
                "Reper coleg": round(daily_reference, 2) if daily_reference is not None else None,
                "Tip reper": daily_reference_type,
                "Medie vs coleg %": round(daily_vs_colleague_pct, 2) if daily_vs_colleague_pct is not None else None,
                "Bonuri 2+ %": round(receipt_2plus_pct, 2) if receipt_2plus_pct is not None else None,
                "Focus %": round(focus_pct, 2) if focus_pct is not None else None,
                "Valoare medie produs": round(average_item_value, 2) if average_item_value is not None else None,
                "Folii eligibile": glass_qty,
                "Folii premium": premium_glass_qty,
                "Folii premium %": round(premium_glass_pct, 2),
                "Pct medie zilnica": qualitative_points[0],
                "Pct bonuri 2+": qualitative_points[1],
                "Pct focus": qualitative_points[2],
                "Pct valoare": qualitative_points[3],
                "Pct folii": qualitative_points[4],
                "Puncte calitative": sum(qualitative_points),
                "Are criteriu zero": "DA" if any(point == 0 for point in qualitative_points) else "NU",
                "SIM": sim_qty,
                "Salariu baza": BASE_SALARY,
                "Bonus target": target_bonus,
                "Bonus calitativ": qualitative_bonus,
                "Bonus SIM": sim_bonus,
                "Bonus ePay": epay_bonus,
                "Bonuri masa propuse": MEAL_TICKETS,
        }
        for scenario, rate in COMMISSION_SCENARIOS:
            accessory_commission = commissionable_sales * rate
            simulated_without_meals = fixed_salary_components + accessory_commission
            output_row[commission_column(scenario)] = round(accessory_commission, 2)
            output_row[simulated_without_meals_column(scenario)] = round(
                simulated_without_meals, 2
            )
            output_row[simulated_with_meals_column(scenario)] = round(
                simulated_without_meals + MEAL_TICKETS, 2
            )
        output.append(output_row)
    return pd.DataFrame(output)


def assign_salary_sites(
    salary_rows: list[SalarySourceRow],
    stores: list[dict[str, Any]],
) -> None:
    store_map: dict[tuple[str, str], str | None] = {}
    for store in stores:
        key = (company_key(store["firma"]), normalize_text(store["locatie"]))
        store_map[key] = store["site_code"] if key not in store_map else None

    aliases = {**LOCATION_ALIASES, **EXTRA_LOCATION_ALIASES}
    for row in salary_rows:
        alias_key = (row.company, row.normalized_location)
        if alias_key in aliases:
            row.site_code = aliases[alias_key]
            continue
        row.site_code = store_map.get(
            (company_key(row.company), row.normalized_location)
        )


def _name_code_score(full_name: str, agent_code: str) -> float:
    if agent_code in candidate_agent_codes(full_name):
        return 1.0
    name_compact = normalize_text(full_name).replace(" ", "")
    code = normalize_text(agent_code).replace(" ", "")
    scores = [SequenceMatcher(None, name_compact, code).ratio()]
    tokens = [token for token in normalize_text(full_name).split() if token]
    for first, second in itertools.permutations(tokens, 2):
        variants = (
            first + second[:1],
            first[:1] + second,
            first[:8] + second[:1],
            first + second[:2],
        )
        scores.extend(SequenceMatcher(None, variant, code).ratio() for variant in variants)
    return max(scores)


def _best_group_assignment(
    rows: list[SalarySourceRow],
    agents: list[str],
) -> tuple[list[tuple[SalarySourceRow, str, float]], float]:
    if not rows or not agents:
        return [], 0.0
    pair_count = min(len(rows), len(agents))
    best_pairs: list[tuple[SalarySourceRow, str, float]] = []
    best_total = -1.0
    second_total = -1.0
    for row_subset in itertools.combinations(rows, pair_count):
        for agent_perm in itertools.permutations(agents, pair_count):
            pairs = [
                (row, agent, _name_code_score(row.full_name, agent))
                for row, agent in zip(row_subset, agent_perm, strict=True)
            ]
            total = sum(score for _, _, score in pairs)
            if total > best_total:
                second_total = best_total
                best_total = total
                best_pairs = pairs
            elif total > second_total:
                second_total = total
    return best_pairs, best_total - max(second_total, 0.0)


def match_salary_rows(
    salary_rows: list[SalarySourceRow],
    simulation: pd.DataFrame,
    known_names: list[dict[str, Any]],
) -> None:
    sales_agents_by_group: dict[tuple[str, str], set[str]] = defaultdict(set)
    sales_agents_by_month: dict[str, set[str]] = defaultdict(set)
    for row in simulation.to_dict("records"):
        period = row["Luna"]
        agent = row["Agent cod"]
        sales_agents_by_month[period].add(agent)
        for site_code in str(row["Site-uri"]).split(", "):
            sales_agents_by_group[(period, site_code)].add(agent)

    known_map: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in known_names:
        known_map[
            (row["site_code"], normalize_text(row["source_agent_name"]))
        ].add(row["agent"])

    # 1. Asocieri deterministe: cod generat din nume sau mapare existenta Grile.
    for row in salary_rows:
        scope = (
            sales_agents_by_group.get((row.period, row.site_code), set())
            if row.site_code
            else sales_agents_by_month[row.period]
        )
        generated = candidate_agent_codes(row.full_name) & scope
        if len(generated) == 1:
            row.agent = next(iter(generated))
            row.match_method = "cod_nume"
            row.match_confidence = 1.0
            continue
        if row.site_code:
            known = (
                known_map.get((row.site_code, row.normalized_name), set()) & scope
            )
            if len(known) == 1:
                row.agent = next(iter(known))
                row.match_method = "mapare_grile"
                row.match_confidence = 1.0

    # 2. Invata numele complete rezolvate sigur si aplica aceeasi identitate in alte luni.
    for _ in range(3):
        learned: dict[str, set[str]] = defaultdict(set)
        for row in salary_rows:
            if row.agent and row.match_confidence >= 0.95:
                learned[row.normalized_name].add(row.agent)
        changed = False
        for row in salary_rows:
            if row.agent:
                continue
            candidates = learned.get(row.normalized_name, set())
            scope = sales_agents_by_month[row.period]
            matches = candidates & scope
            if len(matches) == 1:
                row.agent = next(iter(matches))
                row.match_method = "nume_alta_luna"
                row.match_confidence = 0.98
                changed = True
        if not changed:
            break

    # 3. Asignare bipartita in acelasi magazin/luna.
    grouped_rows: dict[tuple[str, str], list[SalarySourceRow]] = defaultdict(list)
    for row in salary_rows:
        if row.site_code:
            grouped_rows[(row.period, row.site_code)].append(row)

    for group_key, group_rows in grouped_rows.items():
        sales_agents = sorted(sales_agents_by_group.get(group_key, set()))
        used_agents = {row.agent for row in group_rows if row.agent}
        remaining_rows = [row for row in group_rows if not row.agent]
        remaining_agents = [agent for agent in sales_agents if agent not in used_agents]
        pairs, assignment_gap = _best_group_assignment(
            remaining_rows, remaining_agents
        )
        equal_group_size = len(remaining_rows) == len(remaining_agents)
        for row, agent, score in pairs:
            if score >= 0.72:
                accepted = True
                method = "similaritate_nume"
                confidence = min(0.95, score)
            elif equal_group_size and score >= 0.48 and assignment_gap >= 0.08:
                accepted = True
                method = "asignare_magazin"
                confidence = min(0.89, score + 0.20)
            elif (
                len(remaining_rows) == 1
                and len(remaining_agents) == 1
                and score >= 0.40
            ):
                accepted = True
                method = "unic_in_magazin"
                confidence = min(0.82, score + 0.18)
            else:
                accepted = False
                method = "unmatched"
                confidence = score
            if accepted:
                row.agent = agent
                row.match_method = method
                row.match_confidence = confidence

    # 4. Refoloseste asocierile de nume rezultate din magazin.
    learned = defaultdict(set)
    for row in salary_rows:
        if row.agent and row.match_confidence >= 0.65:
            learned[row.normalized_name].add(row.agent)
    for row in salary_rows:
        if row.agent:
            continue
        matches = learned.get(row.normalized_name, set()) & sales_agents_by_month[
            row.period
        ]
        if len(matches) == 1:
            row.agent = next(iter(matches))
            row.match_method = "nume_repetat"
            row.match_confidence = 0.75


def build_salary_frames(
    salary_rows: list[SalarySourceRow],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    source_df = pd.DataFrame(
        [
            {
                "Luna": row.period,
                "Companie HR": row.company,
                "Nume HR": row.full_name,
                "CNP": row.cnp,
                "Locatie HR": row.location,
                "Site mapat": row.site_code,
                "Agent cod": row.agent,
                "Metoda mapare": row.match_method,
                "Incredere mapare": round(row.match_confidence, 3),
                "Total salariu HR": round(row.total_salary, 2),
                "Ore suplimentare": round(row.overtime, 2),
                "Salariu ajustat": round(row.adjusted_salary, 2),
                "Bonuri masa istorice": round(row.meal_tickets, 2),
                "Fisier sursa": row.source_file,
            }
            for row in salary_rows
        ]
    )
    matched = source_df[source_df["Agent cod"].notna()].copy()
    if matched.empty:
        return source_df, pd.DataFrame()

    aggregate = (
        matched.groupby(["Luna", "Agent cod"], as_index=False)
        .agg(
            {
                "Nume HR": lambda values: " / ".join(sorted(set(values))),
                "Companie HR": lambda values: " / ".join(sorted(set(values))),
                "Locatie HR": lambda values: " / ".join(sorted(set(values))),
                "Total salariu HR": "sum",
                "Ore suplimentare": "sum",
                "Salariu ajustat": "sum",
                "Bonuri masa istorice": "sum",
                "Incredere mapare": "min",
                "Metoda mapare": lambda values: " / ".join(sorted(set(values))),
            }
        )
        .rename(
            columns={
                "Nume HR": "Nume istoric",
                "Companie HR": "Companie istorica",
                "Locatie HR": "Locatie istorica",
                "Incredere mapare": "Incredere minima",
                "Metoda mapare": "Metode mapare",
            }
        )
    )
    aggregate["Eligibil luna intreaga"] = aggregate["Salariu ajustat"].ge(
        FULL_MONTH_SALARY_FLOOR
    )
    aggregate["Mapare sigura"] = aggregate["Incredere minima"].ge(0.65)
    return source_df, aggregate


def merge_simulation_with_salary(
    simulation: pd.DataFrame,
    salary_aggregate: pd.DataFrame,
) -> pd.DataFrame:
    if salary_aggregate.empty:
        result = simulation.copy()
        result["Status comparatie"] = "fara_salariu_mapat"
        return result
    result = simulation.merge(
        salary_aggregate,
        on=["Luna", "Agent cod"],
        how="left",
        validate="one_to_one",
    )
    result["Status comparatie"] = "fara_salariu_mapat"
    mapped = result["Salariu ajustat"].notna()
    eligible_full_month = result["Eligibil luna intreaga"].eq(True)
    safe_mapping = result["Mapare sigura"].eq(True)
    result.loc[mapped, "Status comparatie"] = "salariu_sub_2400"
    uncertain = (
        mapped
        & eligible_full_month
        & ~safe_mapping
    )
    result.loc[uncertain, "Status comparatie"] = "mapare_incerta"
    comparable = (
        mapped
        & eligible_full_month
        & safe_mapping
    )
    result.loc[comparable, "Status comparatie"] = "comparabil"
    for scenario, _rate in COMMISSION_SCENARIOS:
        result[difference_column(scenario)] = (
            result[simulated_without_meals_column(scenario)]
            - result["Salariu ajustat"]
        )
        result[difference_pct_column(scenario)] = (
            result[difference_column(scenario)]
            * 100
            / result["Salariu ajustat"]
        )
    return result


def build_summaries(combined: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    comparable = combined[combined["Status comparatie"] == "comparabil"].copy()
    monthly_base = (
        comparable.groupby("Luna", as_index=False)
        .agg(
            Agenti=("Agent cod", "nunique"),
            Total_salariu_HR_inainte_de_ore_suplimentare=("Total salariu HR", "sum"),
            Ore_suplimentare_excluse=("Ore suplimentare", "sum"),
            Istoric_fara_bonuri=("Salariu ajustat", "sum"),
            Bonus_target=("Bonus target", "sum"),
            Bonus_calitativ=("Bonus calitativ", "sum"),
            Bonus_SIM=("Bonus SIM", "sum"),
        )
    )
    monthly_frames: list[pd.DataFrame] = []
    manager_frames: list[pd.DataFrame] = []
    for scenario, rate in COMMISSION_SCENARIOS:
        scenario_monthly = monthly_base.copy()
        scenario_monthly.insert(1, "Scenariu comision", scenario)
        scenario_monthly.insert(2, "Procent comision", rate * 100)
        monthly_values = (
            comparable.groupby("Luna", as_index=False)
            .agg(
                Simulat_fara_bonuri=(
                    simulated_without_meals_column(scenario),
                    "sum",
                ),
                Comision_accesorii=(commission_column(scenario), "sum"),
            )
        )
        scenario_monthly = scenario_monthly.merge(
            monthly_values, on="Luna", how="left", validate="one_to_one"
        )
        scenario_monthly["Diferenta"] = (
            scenario_monthly["Simulat_fara_bonuri"]
            - scenario_monthly["Istoric_fara_bonuri"]
        )
        scenario_monthly["Diferenta %"] = (
            scenario_monthly["Diferenta"]
            * 100
            / scenario_monthly["Istoric_fara_bonuri"]
        )
        scenario_monthly["Salariu_mediu_istoric"] = (
            scenario_monthly["Istoric_fara_bonuri"] / scenario_monthly["Agenti"]
        )
        scenario_monthly["Salariu_mediu_simulat"] = (
            scenario_monthly["Simulat_fara_bonuri"] / scenario_monthly["Agenti"]
        )
        scenario_monthly["Diferenta_medie_per_agent"] = (
            scenario_monthly["Salariu_mediu_simulat"]
            - scenario_monthly["Salariu_mediu_istoric"]
        )
        scenario_monthly["Bonuri masa propuse"] = (
            scenario_monthly["Agenti"] * MEAL_TICKETS
        )
        monthly_frames.append(scenario_monthly)

        manager = (
            comparable.groupby(["Luna", "ASM"], as_index=False)
            .agg(
                Agenti=("Agent cod", "nunique"),
                Istoric_fara_bonuri=("Salariu ajustat", "sum"),
                Simulat_fara_bonuri=(
                    simulated_without_meals_column(scenario),
                    "sum",
                ),
            )
        )
        manager.insert(2, "Scenariu comision", scenario)
        manager.insert(3, "Procent comision", rate * 100)
        manager["Diferenta"] = (
            manager["Simulat_fara_bonuri"] - manager["Istoric_fara_bonuri"]
        )
        manager["Diferenta %"] = (
            manager["Diferenta"] * 100 / manager["Istoric_fara_bonuri"]
        )
        manager_frames.append(manager)

    monthly = pd.concat(monthly_frames, ignore_index=True).sort_values(
        ["Luna", "Procent comision"]
    )
    manager = pd.concat(manager_frames, ignore_index=True).sort_values(
        ["Luna", "ASM", "Procent comision"]
    )
    return monthly, manager


def build_audit(
    salary_source: pd.DataFrame,
    combined: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    for period in PERIODS:
        salary_period = salary_source[salary_source["Luna"] == period]
        combined_period = combined[combined["Luna"] == period]
        rows.append(
            {
                "Luna": period,
                "Randuri HR": len(salary_period),
                "Randuri HR mapate": int(salary_period["Agent cod"].notna().sum()),
                "Randuri HR nemapate": int(salary_period["Agent cod"].isna().sum()),
                "Mapari incredere < 0.65": int(
                    (
                        salary_period["Agent cod"].notna()
                        & salary_period["Incredere mapare"].lt(0.65)
                    ).sum()
                ),
                "Agenti vanzari": int(combined_period["Agent cod"].nunique()),
                "Agenti comparabili": int(
                    combined_period.loc[
                        combined_period["Status comparatie"] == "comparabil",
                        "Agent cod",
                    ].nunique()
                ),
                "Agenti salariu sub 2400": int(
                    combined_period.loc[
                        combined_period["Status comparatie"] == "salariu_sub_2400",
                        "Agent cod",
                    ].nunique()
                ),
                "Agenti cu mapare incerta": int(
                    combined_period.loc[
                        combined_period["Status comparatie"] == "mapare_incerta",
                        "Agent cod",
                    ].nunique()
                ),
                "Agenti fara salariu mapat": int(
                    combined_period.loc[
                        combined_period["Status comparatie"] == "fara_salariu_mapat",
                        "Agent cod",
                    ].nunique()
                ),
            }
        )
    return pd.DataFrame(rows)


def _old_grid_target_bonus(target_pct: float | None) -> float:
    if target_pct is None or target_pct < 100:
        return 0.0
    if target_pct < 120:
        return 200.0
    return 400.0


def _old_grid_comparable_salary(
    total_salary: float,
    meal_tickets: float,
    overtime: float,
) -> float:
    return total_salary - meal_tickets - overtime


def build_old_grid_may_control(
    simulation: pd.DataFrame,
    combined: pd.DataFrame,
    grid_registry: list[dict[str, Any]],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Compara exact arhiva grilei vechi din mai cu simularea noua.

    Arhive locale complete pentru grila veche exista doar pentru mai 2026.
    Controlul este separat de salariile HR, deoarece payroll-ul poate include
    incentive, prime si reglari care nu fac parte din formula standard.
    """

    detail_columns = [
        "Luna",
        "Companie",
        "Magazin grila veche",
        "Site",
        "Nume grila veche",
        "Agent cod",
        "Salariu baza vechi",
        "Bonuri vechi",
        "Ore suplimentare vechi",
        "Vanzari vechi",
        "Target vechi",
        "Realizare target vechi %",
        "Comision accesorii vechi 3%",
        "Bonus target vechi",
        "Bonus SIM vechi",
        "Bonus ePay vechi",
        "Comision alte locatii vechi",
        "Salariu total vechi",
        "Grila veche comparabila",
        "Vanzari simulare",
        "Target simulare",
        "Realizare target simulare %",
        "Comision accesorii nou 3%",
        "Bonus target nou",
        "Bonus calitativ nou",
        "Bonus SIM nou",
        "Bonus ePay nou",
        "Grila noua 3% comparabila",
        "Diferenta nou minus vechi",
        "Salariu HR efectiv ajustat",
        "Diferenta HR minus grila veche",
        "Diferenta HR minus grila noua",
        "Status salariu HR",
    ]
    summary_columns = [
        "Luna",
        "Potriviri grila veche",
        "Potriviri cu salariu HR comparabil",
        "Agenti cu crestere",
        "Agenti cu scadere",
        "Agenti fara diferenta",
        "Grila veche comparabila",
        "Grila noua 3% comparabila",
        "Diferenta nou minus vechi",
        "Diferenta nou minus vechi %",
        "Medie grila veche",
        "Medie grila noua",
        "Diferenta medie per agent",
        "Salarii HR efective ajustate",
        "Diferenta HR minus grila veche",
        "Diferenta HR minus grila noua",
    ]
    if not OLD_GRID_MAY_ARCHIVE.exists():
        return pd.DataFrame(columns=detail_columns), pd.DataFrame(
            columns=summary_columns
        )

    may_simulation = simulation[simulation["Luna"] == "2026-05"].copy()
    simulation_by_site: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in may_simulation.to_dict("records"):
        for site_code in str(row["Site-uri"]).split(", "):
            simulation_by_site[site_code].append(row)

    may_combined = combined[combined["Luna"] == "2026-05"].set_index("Agent cod")
    registry_by_key = {
        row["registry_key"]: row["site_code"] for row in grid_registry
    }
    details: list[dict[str, Any]] = []

    with ZipFile(OLD_GRID_MAY_ARCHIVE) as archive:
        for archive_name in archive.namelist():
            if not archive_name.lower().endswith(".xlsx"):
                continue
            registry_key = archive_name[:-5]
            site_code = registry_by_key.get(registry_key)
            if not site_code:
                continue
            company, store = registry_key.split("/", 1)
            workbook = load_workbook(
                BytesIO(archive.read(archive_name)),
                data_only=True,
                read_only=True,
            )
            sheet = workbook["Grila"]
            for offset in (0, 14):
                old_name = str(sheet[f"D{2 + offset}"].value or "").strip()
                if not old_name:
                    continue
                name_candidates = candidate_agent_codes(old_name)
                matched_agents = [
                    row
                    for row in simulation_by_site.get(site_code, [])
                    if row["Agent cod"] in name_candidates
                ]
                if len(matched_agents) != 1:
                    continue

                new_row = matched_agents[0]
                old_total = _number(sheet[f"D{6 + offset}"].value)
                old_base = _number(sheet[f"D{3 + offset}"].value)
                old_meals = _number(sheet[f"D{4 + offset}"].value)
                old_overtime = _number(sheet[f"G{10 + offset}"].value)
                old_other_location = _number(sheet[f"G{11 + offset}"].value)
                old_sales = _number(sheet[f"E{8 + offset}"].value)
                old_target = _number(sheet[f"D{8 + offset}"].value)
                old_target_pct = (
                    old_sales * 100 / old_target if old_target > 0 else None
                )
                old_commission = (
                    old_sales * 0.03
                    if old_target_pct is not None
                    and old_target_pct >= ACCESSORY_COMMISSION_GATE
                    else 0.0
                )
                old_target_bonus = _old_grid_target_bonus(old_target_pct)
                old_sim = _number(sheet[f"G{12 + offset}"].value)
                old_epay = _number(sheet[f"G{13 + offset}"].value) + _number(
                    sheet[f"G{14 + offset}"].value
                )
                old_comparable = _old_grid_comparable_salary(
                    old_total,
                    old_meals,
                    old_overtime,
                )

                agent_code = new_row["Agent cod"]
                hr_row = (
                    may_combined.loc[agent_code]
                    if agent_code in may_combined.index
                    else None
                )
                hr_status = (
                    str(hr_row["Status comparatie"]) if hr_row is not None else ""
                )
                hr_adjusted = (
                    _number(hr_row["Salariu ajustat"])
                    if hr_row is not None and hr_status == "comparabil"
                    else None
                )
                new_comparable = _number(
                    new_row[simulated_without_meals_column("3,0%")]
                )
                details.append(
                    {
                        "Luna": "2026-05",
                        "Companie": company,
                        "Magazin grila veche": store,
                        "Site": site_code,
                        "Nume grila veche": old_name,
                        "Agent cod": agent_code,
                        "Salariu baza vechi": old_base,
                        "Bonuri vechi": old_meals,
                        "Ore suplimentare vechi": old_overtime,
                        "Vanzari vechi": round(old_sales, 2),
                        "Target vechi": round(old_target, 2),
                        "Realizare target vechi %": (
                            round(old_target_pct, 2)
                            if old_target_pct is not None
                            else None
                        ),
                        "Comision accesorii vechi 3%": round(old_commission, 2),
                        "Bonus target vechi": old_target_bonus,
                        "Bonus SIM vechi": old_sim,
                        "Bonus ePay vechi": old_epay,
                        "Comision alte locatii vechi": old_other_location,
                        "Salariu total vechi": round(old_total, 2),
                        "Grila veche comparabila": round(old_comparable, 2),
                        "Vanzari simulare": new_row["Vanzari accesorii"],
                        "Target simulare": new_row["Target agent"],
                        "Realizare target simulare %": new_row[
                            "Realizare target %"
                        ],
                        "Comision accesorii nou 3%": new_row[
                            commission_column("3,0%")
                        ],
                        "Bonus target nou": new_row["Bonus target"],
                        "Bonus calitativ nou": new_row["Bonus calitativ"],
                        "Bonus SIM nou": new_row["Bonus SIM"],
                        "Bonus ePay nou": new_row["Bonus ePay"],
                        "Grila noua 3% comparabila": round(new_comparable, 2),
                        "Diferenta nou minus vechi": round(
                            new_comparable - old_comparable,
                            2,
                        ),
                        "Salariu HR efectiv ajustat": (
                            round(hr_adjusted, 2)
                            if hr_adjusted is not None
                            else None
                        ),
                        "Diferenta HR minus grila veche": (
                            round(hr_adjusted - old_comparable, 2)
                            if hr_adjusted is not None
                            else None
                        ),
                        "Diferenta HR minus grila noua": (
                            round(hr_adjusted - new_comparable, 2)
                            if hr_adjusted is not None
                            else None
                        ),
                        "Status salariu HR": hr_status,
                    }
                )

    detail = pd.DataFrame(details, columns=detail_columns)
    if detail.empty:
        return detail, pd.DataFrame(columns=summary_columns)

    detail = detail.sort_values(["Companie", "Magazin grila veche", "Agent cod"])
    hr_comparable = detail["Salariu HR efectiv ajustat"].notna()
    old_total = detail["Grila veche comparabila"].sum()
    new_total = detail["Grila noua 3% comparabila"].sum()
    count = len(detail)
    delta = detail["Diferenta nou minus vechi"]
    summary = pd.DataFrame(
        [
            {
                "Luna": "2026-05",
                "Potriviri grila veche": count,
                "Potriviri cu salariu HR comparabil": int(hr_comparable.sum()),
                "Agenti cu crestere": int(delta.gt(0.05).sum()),
                "Agenti cu scadere": int(delta.lt(-0.05).sum()),
                "Agenti fara diferenta": int(delta.between(-0.05, 0.05).sum()),
                "Grila veche comparabila": round(old_total, 2),
                "Grila noua 3% comparabila": round(new_total, 2),
                "Diferenta nou minus vechi": round(new_total - old_total, 2),
                "Diferenta nou minus vechi %": (
                    (new_total - old_total) * 100 / old_total
                    if old_total
                    else None
                ),
                "Medie grila veche": old_total / count,
                "Medie grila noua": new_total / count,
                "Diferenta medie per agent": (new_total - old_total) / count,
                "Salarii HR efective ajustate": round(
                    detail.loc[
                        hr_comparable,
                        "Salariu HR efectiv ajustat",
                    ].sum(),
                    2,
                ),
                "Diferenta HR minus grila veche": round(
                    detail.loc[
                        hr_comparable,
                        "Diferenta HR minus grila veche",
                    ].sum(),
                    2,
                ),
                "Diferenta HR minus grila noua": round(
                    detail.loc[
                        hr_comparable,
                        "Diferenta HR minus grila noua",
                    ].sum(),
                    2,
                ),
            }
        ],
        columns=summary_columns,
    )
    return detail, summary


def comparison_notes_frame(
    old_grid_summary: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[tuple[str, str]] = [
        (
            "Comparatia principala",
            "grila noua versus salariile HR efectiv platite; nu versus formula grilei vechi",
        ),
        (
            "Salarii HR",
            "pot include incentive, prime si reglari care nu apar in formula standard a grilei",
        ),
        (
            "Control grila veche",
            "comparatie exacta disponibila pentru mai 2026 din arhiva locala a grilelor",
        ),
        (
            "Comparabilitate",
            "atat grila veche, cat si grila noua exclud bonurile si plata orelor suplimentare",
        ),
        (
            "Target",
            "grila noua foloseste formula Evaluare noua; targetul vechi era cel salvat manual in grila",
        ),
    ]
    if not old_grid_summary.empty:
        row = old_grid_summary.iloc[0]
        rows.extend(
            [
                (
                    "Mai 2026 - grila veche",
                    f"{row['Grila veche comparabila']:.2f} lei pentru "
                    f"{int(row['Potriviri grila veche'])} agenti potriviti",
                ),
                (
                    "Mai 2026 - grila noua 3%",
                    f"{row['Grila noua 3% comparabila']:.2f} lei; "
                    f"diferenta {row['Diferenta nou minus vechi']:+.2f} lei",
                ),
            ]
        )
    return pd.DataFrame(rows, columns=["Subiect", "Clarificare"])


def rules_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Perioada", "2025-12..2026-05, numai luni inchise"),
            ("Salariu baza", "2.700 lei net"),
            ("Bonuri masa", "480 lei separat"),
            (
                "Comision accesorii",
                "scenarii 2,5% / 2,7% / 3,0% din vanzari, numai de la 80% target",
            ),
            ("Target agent", "target magazin / zile vanzare magazin x zile vanzare agent"),
            ("Bonus target", "100%=200; 110%=300 cumulativ; 120%=400 cumulativ"),
            ("Bonus calitativ", "11-12=100; 13-14=200; 15=300"),
            ("Regula eliminatorie", "orice criteriu cu 0 puncte => bonus calitativ 0"),
            ("Medie zilnica", "<90%=0; 90-110%=1; >110%=3; nu exista 2 puncte"),
            (
                "Reper medie zilnica",
                "mediana colegilor din magazin/luna; fallback istoric locatie 3 luni, apoi manager",
            ),
            ("Bonuri 2+", "<25=0; 25-<30=1; 30-<35=2; >=35=3"),
            ("Focus", "<6=0; 6-<7=1; 7-<8=2; >=8=3"),
            ("Valoare medie", "<80=0; 80-<90=1; 90-<100=2; >=100=3"),
            ("Folii premium", "<30=0; 30-40=1; >40-50=2; >50=3"),
            ("SIM", "3 lei per SIM; coduri SIM/SPN sau denumire Cartela/SIM, fara pachete telefon"),
            ("ePay", "0 lei in simulare; se adauga manual"),
            ("Salariu istoric", "TOTAL SALARIU minus ore suplimentare; fara bonuri masa"),
            ("Prag luna intreaga", "salariul istoric ajustat agregat per agent >= 2.400 lei"),
            ("Comparatie", "numai agentii mapati si eligibili dupa pragul de 2.400 lei"),
        ],
        columns=["Regula", "Interpretare aplicata"],
    )


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    currency_headers = {
        "Vanzari accesorii",
        "Target agent",
        "Medie zilnica",
        "Reper coleg",
        "Valoare medie produs",
        "Salariu baza",
        "Bonus target",
        "Bonus calitativ",
        "Bonus SIM",
        "Bonus ePay",
        "Simulat fara bonuri",
        "Bonuri masa propuse",
        "Simulat cu bonuri",
        "Total salariu HR",
        "Ore suplimentare",
        "Salariu ajustat",
        "Bonuri masa istorice",
        "Istoric_fara_bonuri",
        "Simulat_fara_bonuri",
        "Comision_accesorii",
        "Bonus_target",
        "Bonus_calitativ",
        "Bonus_SIM",
        "Diferenta",
        "Total_salariu_HR_inainte_de_ore_suplimentare",
        "Ore_suplimentare_excluse",
        "Salariu_mediu_istoric",
        "Salariu_mediu_simulat",
        "Diferenta_medie_per_agent",
        "Salariu_mediu_simulat_fara_bonuri",
        "Salariu_mediu_simulat_cu_bonuri",
    }
    percent_headers = {
        "Realizare target %",
        "Medie vs coleg %",
        "Bonuri 2+ %",
        "Focus %",
        "Folii premium %",
        "Diferenta %",
        "Procent comision",
    }
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        headers = {cell.value: cell.column for cell in sheet[1]}
        for header, column in headers.items():
            is_scenario_currency = isinstance(header, str) and (
                header.startswith("Comision accesorii ")
                or header.startswith("Simulat ")
                or header.startswith("Diferenta ")
            ) and not header.endswith(" %")
            if header in currency_headers or is_scenario_currency:
                for cell in sheet[get_column_letter(column)][1:]:
                    cell.number_format = '#,##0.00 "lei"'
            elif header in percent_headers or (
                isinstance(header, str) and header.endswith(" %")
            ):
                for cell in sheet[get_column_letter(column)][1:]:
                    cell.number_format = '0.00"%"'
        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in column_cells[:300]]
            width = min(max(max(map(len, values), default=0) + 2, 10), 42)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        if sheet.title == "Comparatie eligibili":
            status_col = headers.get("Status comparatie")
            difference_col = headers.get(difference_column("2,5%"))
            if status_col:
                for cell in sheet[get_column_letter(status_col)][1:]:
                    cell.fill = OK_FILL if cell.value == "comparabil" else WARNING_FILL
            if difference_col and sheet.max_row > 1:
                column = get_column_letter(difference_col)
                sheet.conditional_formatting.add(
                    f"{column}2:{column}{sheet.max_row}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="63BE7B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="F8696B",
                    ),
                )
        if sheet.title == "Salarii sursa":
            agent_col = headers.get("Agent cod")
            confidence_col = headers.get("Incredere mapare")
            if agent_col:
                for row in range(2, sheet.max_row + 1):
                    if not sheet.cell(row, agent_col).value:
                        for cell in sheet[row]:
                            cell.fill = ERROR_FILL
            if confidence_col:
                for row in range(2, sheet.max_row + 1):
                    value = sheet.cell(row, confidence_col).value
                    if value is not None and value < 0.65:
                        sheet.cell(row, confidence_col).fill = WARNING_FILL

    summary_sheet = next(
        (
            name
            for name in ("Total scenarii", "Sumar lunar", "Sumar simulare")
            if name in workbook.sheetnames
        ),
        None,
    )
    if summary_sheet:
        sheet = workbook[summary_sheet]
        if sheet.max_row > 1:
            summary_headers = {
                cell.value: cell.column for cell in sheet[1]
            }
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            if "Istoric_fara_bonuri" in summary_headers:
                first_column = summary_headers["Istoric_fara_bonuri"]
                last_column = summary_headers["Simulat_fara_bonuri"]
                chart.title = "Cost istoric vs simulare"
            else:
                first_column = summary_headers["Simulat_fara_bonuri"]
                last_column = first_column
                chart.title = "Cost simulat fara bonuri"
            chart.y_axis.title = "Lei"
            chart.x_axis.title = "Luna"
            data = Reference(
                sheet,
                min_col=first_column,
                max_col=last_column,
                min_row=1,
                max_row=sheet.max_row,
            )
            categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 15
            sheet.add_chart(chart, "M2")

    workbook.save(path)


def build_simulation_summary(simulation: pd.DataFrame) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for scenario, rate in COMMISSION_SCENARIOS:
        summary = (
            simulation.groupby("Luna", as_index=False)
            .agg(
                Agenti=("Agent cod", "nunique"),
                Vanzari_accesorii=("Vanzari accesorii", "sum"),
                Target_agenti=("Target agent", "sum"),
                Simulat_fara_bonuri=(
                    simulated_without_meals_column(scenario),
                    "sum",
                ),
                Comision_accesorii=(commission_column(scenario), "sum"),
                Bonus_target=("Bonus target", "sum"),
                Bonus_calitativ=("Bonus calitativ", "sum"),
                Bonus_SIM=("Bonus SIM", "sum"),
            )
        )
        summary.insert(1, "Scenariu comision", scenario)
        summary.insert(2, "Procent comision", rate * 100)
        summary["Bonuri_masa_propuse"] = summary["Agenti"] * MEAL_TICKETS
        summary["Simulat_cu_bonuri"] = (
            summary["Simulat_fara_bonuri"] + summary["Bonuri_masa_propuse"]
        )
        summary["Salariu_mediu_simulat_fara_bonuri"] = (
            summary["Simulat_fara_bonuri"] / summary["Agenti"]
        )
        summary["Salariu_mediu_simulat_cu_bonuri"] = (
            summary["Simulat_cu_bonuri"] / summary["Agenti"]
        )
        frames.append(summary)
    return pd.concat(frames, ignore_index=True).sort_values(
        ["Luna", "Procent comision"]
    )


def build_total_scenario_summary(
    simulation_summary: pd.DataFrame,
    comparison_summary: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    simulation_totals = (
        simulation_summary.groupby(
            ["Scenariu comision", "Procent comision"], as_index=False
        )
        .agg(
            Agent_luni=("Agenti", "sum"),
            Simulat_fara_bonuri=("Simulat_fara_bonuri", "sum"),
            Comision_accesorii=("Comision_accesorii", "sum"),
            Bonuri_masa_propuse=("Bonuri_masa_propuse", "sum"),
            Simulat_cu_bonuri=("Simulat_cu_bonuri", "sum"),
        )
    )
    simulation_totals["Salariu_mediu_simulat_fara_bonuri"] = (
        simulation_totals["Simulat_fara_bonuri"]
        / simulation_totals["Agent_luni"]
    )
    simulation_totals["Salariu_mediu_simulat_cu_bonuri"] = (
        simulation_totals["Simulat_cu_bonuri"]
        / simulation_totals["Agent_luni"]
    )

    comparison_totals = (
        comparison_summary.groupby(
            ["Scenariu comision", "Procent comision"], as_index=False
        )
        .agg(
            Agent_luni=("Agenti", "sum"),
            Total_salariu_HR_inainte_de_ore_suplimentare=(
                "Total_salariu_HR_inainte_de_ore_suplimentare",
                "sum",
            ),
            Ore_suplimentare_excluse=("Ore_suplimentare_excluse", "sum"),
            Istoric_fara_bonuri=("Istoric_fara_bonuri", "sum"),
            Simulat_fara_bonuri=("Simulat_fara_bonuri", "sum"),
            Comision_accesorii=("Comision_accesorii", "sum"),
        )
    )
    comparison_totals["Diferenta"] = (
        comparison_totals["Simulat_fara_bonuri"]
        - comparison_totals["Istoric_fara_bonuri"]
    )
    comparison_totals["Diferenta %"] = (
        comparison_totals["Diferenta"]
        * 100
        / comparison_totals["Istoric_fara_bonuri"]
    )
    comparison_totals["Salariu_mediu_istoric"] = (
        comparison_totals["Istoric_fara_bonuri"]
        / comparison_totals["Agent_luni"]
    )
    comparison_totals["Salariu_mediu_simulat"] = (
        comparison_totals["Simulat_fara_bonuri"]
        / comparison_totals["Agent_luni"]
    )
    comparison_totals["Diferenta_medie_per_agent"] = (
        comparison_totals["Salariu_mediu_simulat"]
        - comparison_totals["Salariu_mediu_istoric"]
    )
    return simulation_totals, comparison_totals


async def generate(output: Path, comparison_output: Path) -> dict[str, Any]:
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL lipseste din .env")
    conn = await asyncpg.connect(database_url)
    try:
        data = await fetch_data(conn)
    finally:
        await conn.close()

    simulation = build_simulation(data)
    salary_rows = load_salary_source_rows()
    assign_salary_sites(salary_rows, data["stores"])
    match_salary_rows(salary_rows, simulation, data["known_names"])
    salary_source, salary_aggregate = build_salary_frames(salary_rows)
    combined = merge_simulation_with_salary(simulation, salary_aggregate)
    monthly_summary, manager_summary = build_summaries(combined)
    simulation_summary = build_simulation_summary(simulation)
    simulation_totals, comparison_totals = build_total_scenario_summary(
        simulation_summary, monthly_summary
    )
    audit = build_audit(salary_source, combined)
    old_grid_detail, old_grid_summary = build_old_grid_may_control(
        simulation,
        combined,
        data["grid_registry"],
    )
    comparison_notes = comparison_notes_frame(old_grid_summary)

    comparable = combined[combined["Status comparatie"] == "comparabil"].copy()
    excluded = combined[combined["Status comparatie"] != "comparabil"].copy()
    salary_unmatched = salary_source[
        salary_source["Agent cod"].isna()
        | salary_source["Incredere mapare"].lt(0.65)
    ].copy()

    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        simulation_totals.to_excel(writer, sheet_name="Total scenarii", index=False)
        simulation_summary.to_excel(writer, sheet_name="Sumar simulare", index=False)
        simulation.to_excel(writer, sheet_name="Simulare completa", index=False)
        rules_frame().to_excel(writer, sheet_name="Reguli", index=False)
    _format_workbook(output)

    comparison_output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(comparison_output, engine="openpyxl") as writer:
        comparison_notes.to_excel(writer, sheet_name="Citeste intai", index=False)
        comparison_totals.to_excel(writer, sheet_name="Total scenarii", index=False)
        monthly_summary.to_excel(writer, sheet_name="Sumar lunar", index=False)
        manager_summary.to_excel(writer, sheet_name="Sumar ASM", index=False)
        old_grid_summary.to_excel(
            writer,
            sheet_name="Control grila veche Mai",
            index=False,
        )
        old_grid_detail.to_excel(
            writer,
            sheet_name="Detaliu grila veche Mai",
            index=False,
        )
        comparable.to_excel(writer, sheet_name="Comparatie eligibili", index=False)
        salary_source.to_excel(writer, sheet_name="Salarii sursa", index=False)
        excluded.to_excel(writer, sheet_name="Exclusi comparatie", index=False)
        salary_unmatched.to_excel(writer, sheet_name="Mapari de verificat", index=False)
        audit.to_excel(writer, sheet_name="Audit acoperire", index=False)
        rules_frame().to_excel(writer, sheet_name="Reguli", index=False)
    _format_workbook(comparison_output)

    return {
        "output": str(output),
        "comparison_output": str(comparison_output),
        "simulation_rows": len(simulation),
        "salary_rows": len(salary_source),
        "mapped_salary_rows": int(salary_source["Agent cod"].notna().sum()),
        "comparable_rows": len(comparable),
        "unmatched_or_low_confidence_salary_rows": len(salary_unmatched),
        "monthly_summary": monthly_summary.to_dict("records"),
        "comparison_totals": comparison_totals.to_dict("records"),
        "old_grid_summary": old_grid_summary.to_dict("records"),
        "match_methods": Counter(
            row.match_method for row in salary_rows
        ),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--comparison-output",
        type=Path,
        default=DEFAULT_COMPARISON_OUTPUT,
    )
    args = parser.parse_args()
    result = asyncio.run(generate(args.output, args.comparison_output))
    print(f"Generat simulare: {result['output']}")
    print(f"Generat comparatie: {result['comparison_output']}")
    print(
        "Randuri simulare={simulation_rows}, HR={salary_rows}, HR mapate={mapped_salary_rows}, "
        "comparabile={comparable_rows}, mapari de verificat={unmatched_or_low_confidence_salary_rows}".format(
            **result
        )
    )
    print("Metode mapare:", dict(result["match_methods"]))
    for row in result["comparison_totals"]:
        print(
            f"{row['Scenariu comision']}: agent-luni={row['Agent_luni']}, "
            f"istoric={row['Istoric_fara_bonuri']:.2f}, "
            f"simulat={row['Simulat_fara_bonuri']:.2f}, "
            f"dif={row['Diferenta']:.2f} ({row['Diferenta %']:.2f}%)"
        )
    for row in result["old_grid_summary"]:
        print(
            "Control grila veche mai 2026: "
            f"potriviri={row['Potriviri grila veche']}, "
            f"veche={row['Grila veche comparabila']:.2f}, "
            f"noua 3%={row['Grila noua 3% comparabila']:.2f}, "
            f"dif={row['Diferenta nou minus vechi']:.2f}"
        )


if __name__ == "__main__":
    main()
