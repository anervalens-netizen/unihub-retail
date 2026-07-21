from __future__ import annotations

import asyncio
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from services.spreadsheet_safety import append_openpyxl_row

from repositories.exports import ExportsRepository
from services.dashboard.queries import (
    _compute_dashboard_promotion_result,
    _get_store_incentive_multipliers,
    _load_dashboard_campaign_context,
)
from services.incentive_db import get_incentive_campaign
from services.promo_copurchase import promo_actuals_cutoff_date
from services.dashboard_specials import (
    load_promotion_rule_products,
    load_special_cards_config,
    parse_promotion_definitions,
)


@dataclass(frozen=True)
class ColumnDef:
    key: str
    label: str
    type: str
    group: str


DATASETS = {
    "agents": {
        "label": "Agenti",
        "description": "Rand pe agent si magazin, folosind structura curenta.",
        "dimensions": ["agent", "site_code", "locatie", "firma", "regional", "asm"],
    },
    "stores": {
        "label": "Magazine",
        "description": "Rand pe magazin, folosind structura curenta.",
        "dimensions": ["site_code", "locatie", "firma", "regional", "asm"],
    },
    "regionals": {
        "label": "RM",
        "description": "Rand pe regional manager.",
        "dimensions": ["regional"],
    },
    "asms": {
        "label": "ASM",
        "description": "Rand pe ASM.",
        "dimensions": ["regional", "asm"],
    },
    "incentive_products": {
        "label": "Incentive pe produs",
        "description": "Produse incentive, categorii, excluderi promo si plata calculata pe pragul fiecarui magazin.",
        "dimensions": [],
    },
}

DIMENSIONS = {
    "agent": ColumnDef("agent", "Agent", "text", "Identificare"),
    "site_code": ColumnDef("site_code", "Cod magazin", "text", "Identificare"),
    "locatie": ColumnDef("locatie", "Magazin", "text", "Identificare"),
    "firma": ColumnDef("firma", "Firma", "text", "Identificare"),
    "regional": ColumnDef("regional", "RM", "text", "Identificare"),
    "asm": ColumnDef("asm", "ASM", "text", "Identificare"),
}

METRICS = {
    "total_sales": ColumnDef("total_sales", "Vanzari", "currency", "Vanzari"),
    "total_quantity": ColumnDef("total_quantity", "Cantitate", "integer", "Vanzari"),
    "total_receipts": ColumnDef("total_receipts", "Bonuri", "integer", "Vanzari"),
    "avg_receipt_value": ColumnDef("avg_receipt_value", "Val. medie bon", "currency", "KPI"),
    "proc_bon2acc": ColumnDef("proc_bon2acc", "Bon2Acc %", "percent", "KPI"),
    "prc_focus_acc_qty": ColumnDef("prc_focus_acc_qty", "Focus %", "percent", "KPI"),
    "target": ColumnDef("target", "Target", "currency", "Target"),
    "target_progress_pct": ColumnDef("target_progress_pct", "Realizare target %", "percent", "Target"),
    "working_days": ColumnDef("working_days", "Zile active", "integer", "KPI"),
    "daily_average": ColumnDef("daily_average", "Medie zilnica", "currency", "KPI"),
    "store_count": ColumnDef("store_count", "Magazine active", "integer", "Identificare"),
    "agent_count": ColumnDef("agent_count", "Agenti activi", "integer", "Identificare"),
    "incentive_sales": ColumnDef("incentive_sales", "Incentive vanzari", "currency", "Campanii"),
    "incentive_quantity": ColumnDef("incentive_quantity", "Incentive cantitate", "integer", "Campanii"),
    "incentive_bonus": ColumnDef("incentive_bonus", "Incentive bonus", "currency", "Campanii"),
    "promo_sales": ColumnDef("promo_sales", "Promo vanzari", "currency", "Campanii"),
    "promo_quantity": ColumnDef("promo_quantity", "Promo cantitate", "integer", "Campanii"),
}

EVOLUTION_METRICS = {
    key: METRICS[key]
    for key in [
        "total_sales",
        "total_quantity",
        "total_receipts",
        "avg_receipt_value",
        "proc_bon2acc",
        "prc_focus_acc_qty",
        "target",
        "target_progress_pct",
        "incentive_sales",
        "incentive_quantity",
        "incentive_bonus",
        "promo_sales",
        "promo_quantity",
    ]
}
DAILY_EVOLUTION_METRICS = {
    key: METRICS[key]
    for key in [
        "total_sales",
        "total_quantity",
        "total_receipts",
        "avg_receipt_value",
        "proc_bon2acc",
        "prc_focus_acc_qty",
        "incentive_sales",
        "incentive_quantity",
        "incentive_bonus",
        "promo_sales",
        "promo_quantity",
    ]
}

DEFAULT_METRICS = [
    "total_sales",
    "total_quantity",
    "total_receipts",
    "target",
    "target_progress_pct",
    "proc_bon2acc",
    "prc_focus_acc_qty",
    "daily_average",
]

COMPARISON_LEVELS = {
    "general": {
        "label": "General",
        "sheet": "General",
        "dimensions": [],
    },
    "asms": {
        "label": "ASM",
        "sheet": "ASM",
        "dimensions": ["asm"],
    },
    "stores": {
        "label": "Magazine",
        "sheet": "Magazine",
        "dimensions": ["site_code", "locatie", "asm"],
    },
    "agents": {
        "label": "Agenti",
        "sheet": "Agenti",
        "dimensions": ["agent", "site_code", "locatie", "asm"],
    },
}


class ExportValidationError(ValueError):
    pass


class ExportsService:
    def __init__(self, repo: ExportsRepository):
        self.repo = repo

    def catalog(self) -> dict[str, Any]:
        return {
            "datasets": [
                {
                    "key": key,
                    "label": value["label"],
                    "description": value["description"],
                    "dimensions": [self._column_payload(DIMENSIONS[item]) for item in value["dimensions"]],
                }
                for key, value in DATASETS.items()
            ],
            "metrics": [self._column_payload(item) for item in METRICS.values()],
            "monthly_metrics": [self._column_payload(item) for item in EVOLUTION_METRICS.values()],
            "daily_metrics": [self._column_payload(item) for item in DAILY_EVOLUTION_METRICS.values()],
            "comparison_levels": [
                {"key": key, "label": value["label"]}
                for key, value in COMPARISON_LEVELS.items()
            ],
        }

    async def preview(self, request: dict[str, Any]) -> dict[str, Any]:
        if request.get("export_mode") == "daily_comparison":
            return await self._preview_daily_comparison(request)
        result = await self.build_report(request)
        limit = int(request.get("preview_limit") or 100)
        return {
            "columns": result["columns"],
            "rows": result["rows"][:limit],
            "total_rows": len(result["rows"]),
            "truncated": len(result["rows"]) > limit,
        }

    async def build_report(self, request: dict[str, Any]) -> dict[str, Any]:
        dataset = str(request.get("dataset") or "")
        if dataset not in DATASETS:
            raise ExportValidationError("Dataset invalid.")
        months = sorted({str(item) for item in request.get("months", []) if item})
        if not months:
            raise ExportValidationError("Selecteaza cel putin o luna.")
        if len(months) > 144:
            raise ExportValidationError("Selectia poate contine maxim 144 luni.")
        selected_days = self._selected_days(request)
        include_closed_stores = bool(request.get("include_closed_stores", False))

        if dataset == "incentive_products":
            return await self._build_incentive_products_report(
                months=months,
                filters=self._normalize_filters(request.get("filters") or {}),
                include_closed_stores=include_closed_stores,
                selected_days=selected_days,
            )

        dimensions = self._valid_keys(
            request.get("dimensions"),
            set(DATASETS[dataset]["dimensions"]),
            list(DATASETS[dataset]["dimensions"]),
            "dimensiuni",
        )
        metrics = self._valid_keys(
            request.get("metrics"),
            set(METRICS),
            DEFAULT_METRICS,
            "metrici",
        )
        monthly_metrics = self._valid_keys(
            request.get("monthly_metrics"),
            set(EVOLUTION_METRICS),
            [],
            "metrici lunare",
        )
        daily_metrics = self._valid_keys(
            request.get("daily_metrics"),
            set(DAILY_EVOLUTION_METRICS),
            [],
            "metrici zilnice",
        )
        if daily_metrics and len(months) > 3:
            raise ExportValidationError("Evolutia zilnica este limitata la maxim 3 luni selectate.")
        if daily_metrics and len(daily_metrics) * 31 * len(months) > 220:
            raise ExportValidationError("Prea multe coloane zilnice. Redu lunile sau metricile zilnice.")

        filters = self._normalize_filters(request.get("filters") or {})
        campaign_codes_by_month = self._campaign_codes_by_month(months)
        campaign_exclusions_by_month = await self._campaign_exclusions_by_month(
            months, filters, selected_days
        )

        total_records = await self.repo.fetch_report_rows(
            dataset=dataset,
            months=months,
            filters=filters,
            include_closed_stores=include_closed_stores,
            campaign_codes_by_month=campaign_codes_by_month,
            campaign_exclusions_by_month=campaign_exclusions_by_month,
            selected_days=selected_days,
        )
        rows_by_key: dict[tuple[Any, ...], dict[str, Any]] = {}
        for record in total_records:
            row = self._base_row(record, dimensions, metrics)
            rows_by_key[self._row_key(record, dataset)] = row

        if monthly_metrics:
            monthly_records = await self.repo.fetch_report_rows(
                dataset=dataset,
                months=months,
                filters=filters,
                include_closed_stores=include_closed_stores,
                campaign_codes_by_month=campaign_codes_by_month,
                campaign_exclusions_by_month=campaign_exclusions_by_month,
                selected_days=selected_days,
                period="month",
            )
            self._attach_period_metrics(
                rows_by_key,
                monthly_records,
                dataset,
                monthly_metrics,
                period_prefix="month",
            )

        if daily_metrics:
            daily_records = await self.repo.fetch_report_rows(
                dataset=dataset,
                months=months,
                filters=filters,
                include_closed_stores=include_closed_stores,
                campaign_codes_by_month=campaign_codes_by_month,
                campaign_exclusions_by_month=campaign_exclusions_by_month,
                selected_days=selected_days,
                period="day",
            )
            self._attach_period_metrics(
                rows_by_key,
                daily_records,
                dataset,
                daily_metrics,
                period_prefix="day",
            )

        columns = self._build_columns(dataset, dimensions, metrics, months, monthly_metrics, rows_by_key, daily_metrics)
        rows = list(rows_by_key.values())
        rows.sort(key=lambda row: tuple(str(row.get(dim) or "") for dim in dimensions))
        return {
            "columns": columns,
            "rows": [self._public_row(row, columns) for row in rows],
        }

    async def _build_incentive_products_report(
        self,
        *,
        months: list[str],
        filters: dict[str, list[str]],
        include_closed_stores: bool,
        selected_days: list[int] | None,
    ) -> dict[str, Any]:
        columns = [
            {"key": "month", "label": "Luna", "type": "text", "group": "Perioada"},
            {"key": "category", "label": "Categorie", "type": "text", "group": "Produs"},
            {"key": "subcategory", "label": "Subcategorie", "type": "text", "group": "Produs"},
            {"key": "item_code", "label": "Cod produs", "type": "text", "group": "Produs"},
            {"key": "item_name", "label": "Produs", "type": "text", "group": "Produs"},
            {"key": "reward_value", "label": "Reward/unitate RON", "type": "currency", "group": "Incentive"},
            {"key": "positive_quantity", "label": "Vandute pozitive", "type": "integer", "group": "Vanzari"},
            {"key": "return_quantity", "label": "Retururi", "type": "integer", "group": "Vanzari"},
            {"key": "net_quantity", "label": "Vandute net", "type": "integer", "group": "Vanzari"},
            {"key": "promo_excluded_quantity", "label": "Excluse promo", "type": "integer", "group": "Incentive"},
            {"key": "eligible_quantity", "label": "Eligibile dupa promo", "type": "integer", "group": "Incentive"},
            {"key": "paid_quantity", "label": "Buc. platite >0", "type": "integer", "group": "Plata"},
            {"key": "paid_full_quantity", "label": "Buc. platite 100%", "type": "integer", "group": "Plata"},
            {"key": "paid_half_quantity", "label": "Buc. platite 50%", "type": "integer", "group": "Plata"},
            {"key": "unpaid_quantity", "label": "Neplatite", "type": "integer", "group": "Plata"},
            {"key": "qualified_ui_quantity", "label": "Calificate UI", "type": "integer", "group": "Plata"},
            {"key": "potential_value", "label": "Valoare potentiala RON", "type": "currency", "group": "Incentive"},
            {"key": "paid_value", "label": "RON platiti", "type": "currency", "group": "Plata"},
        ]
        pool = getattr(self.repo, "pool", None)
        if pool is None:
            raise ExportValidationError("Exportul incentive nu are conexiune la baza de date.")

        def csv_filter(key: str) -> str | None:
            values = [value for value in filters.get(key, []) if value]
            return ",".join(values) if values else None

        rows_by_key: dict[tuple[Any, ...], dict[str, Any]] = {}
        async with pool.acquire() as conn:
            for month in months:
                campaign = await get_incentive_campaign(conn, month)
                if campaign is None:
                    continue
                period_exclusions: dict[tuple[str, str, str], int] = {}
                periods = campaign.get("periods") or []
                if len(periods) <= 1:
                    month_exclusions = await self._campaign_exclusions_by_month(
                        [month], filters, selected_days
                    )
                    for (site_code_value, _agent, item_code_value), units in month_exclusions.get(month, {}).items():
                        key = (
                            periods[0]["valid_from"].isoformat() if periods else "",
                            site_code_value,
                            item_code_value,
                        )
                        period_exclusions[key] = period_exclusions.get(key, 0) + units
                else:
                    requested_days = set(selected_days or range(1, 32))
                    for period in periods:
                        period_days = [
                            day for day in requested_days
                            if period["valid_from"].day <= day <= period["valid_to"].day
                        ]
                        if not period_days:
                            continue
                        period_result = await self._campaign_exclusions_by_month(
                            [month], filters, sorted(period_days)
                        )
                        for (site_code_value, _agent, item_code_value), units in period_result.get(month, {}).items():
                            key = (period["valid_from"].isoformat(), site_code_value, item_code_value)
                            period_exclusions[key] = period_exclusions.get(key, 0) + units
                multipliers, achievements = await _get_store_incentive_multipliers(
                    conn,
                    month,
                    firma=csv_filter("firma"),
                    regional=csv_filter("regional"),
                    asm=csv_filter("asm"),
                    site_code=csv_filter("site_code"),
                    current_scope=True,
                    include_closed_stores=include_closed_stores,
                )
                records = await self.repo.fetch_incentive_product_rows(
                    month=month,
                    filters=filters,
                    include_closed_stores=include_closed_stores,
                    selected_days=selected_days,
                )
                for record in records:
                    reward = Decimal(record["reward_value"] or 0)
                    row_key = (
                        month,
                        record["category"],
                        record["subcategory"],
                        record["item_code"],
                        record["item_name"],
                        reward,
                    )
                    row = rows_by_key.setdefault(row_key, {
                        "month": month,
                        "category": record["category"],
                        "subcategory": record["subcategory"],
                        "item_code": record["item_code"],
                        "item_name": record["item_name"],
                        "reward_value": reward,
                        "positive_quantity": 0,
                        "return_quantity": 0,
                        "net_quantity": 0,
                        "promo_excluded_quantity": 0,
                        "eligible_quantity": 0,
                        "paid_quantity": 0,
                        "paid_full_quantity": 0,
                        "paid_half_quantity": 0,
                        "unpaid_quantity": 0,
                        "qualified_ui_quantity": 0,
                        "potential_value": Decimal(0),
                        "paid_value": Decimal(0),
                    })
                    net_quantity = int(record["net_quantity"] or 0)
                    # Exclusions cannot reduce a store-product below zero. This
                    # is the same clipping used by Focus for promo incentive.
                    excluded_quantity = min(
                        max(0, net_quantity),
                        int(period_exclusions.get((
                            record.get("valid_from").isoformat() if record.get("valid_from") else "",
                            record["site_code"],
                            record["item_code"],
                        ), 0)),
                    )
                    eligible_quantity = max(0, net_quantity - excluded_quantity)
                    multiplier = Decimal(str(multipliers.get(record["site_code"], 0)))
                    achievement = achievements.get(record["site_code"])
                    row["positive_quantity"] += int(record["positive_quantity"] or 0)
                    row["return_quantity"] += int(record["return_quantity"] or 0)
                    row["net_quantity"] += net_quantity
                    row["promo_excluded_quantity"] += excluded_quantity
                    row["eligible_quantity"] += eligible_quantity
                    row["potential_value"] += eligible_quantity * reward
                    row["paid_value"] += eligible_quantity * reward * multiplier
                    if multiplier > 0:
                        row["paid_quantity"] += eligible_quantity
                    else:
                        row["unpaid_quantity"] += eligible_quantity
                    if multiplier == 1:
                        row["paid_full_quantity"] += eligible_quantity
                    elif multiplier > 0:
                        row["paid_half_quantity"] += eligible_quantity
                    if achievement is not None and achievement >= 0.9:
                        row["qualified_ui_quantity"] += eligible_quantity

        rows = [self._public_row(row, columns) for row in rows_by_key.values()]
        rows.sort(key=lambda row: (
            str(row["month"]), str(row["category"]), str(row["subcategory"]), str(row["item_code"])
        ))
        return {"columns": columns, "rows": rows}

    async def build_xlsx(self, request: dict[str, Any]) -> tuple[bytes, str]:
        if request.get("export_mode") == "daily_comparison":
            return await self._build_daily_comparison_xlsx(request)

        result = await self.build_report(request)
        selected_days = self._selected_days(request)
        daily_rows: list[Any] | None = None
        if request.get("daily_metrics"):
            filters = self._normalize_filters(request.get("filters") or {})
            daily_rows = await self.repo.fetch_daily_evolution_rows(
                months=request["months"],
                filters=filters,
                include_closed_stores=bool(request.get("include_closed_stores", False)),
                campaign_codes_by_month=self._campaign_codes_by_month(request["months"]),
                campaign_exclusions_by_month=await self._campaign_exclusions_by_month(
                    request["months"],
                    filters,
                    selected_days,
                ),
                selected_days=selected_days,
            )
        return await asyncio.to_thread(
            self._render_table_xlsx,
            request,
            result,
            selected_days,
            daily_rows,
        )

    def _render_table_xlsx(
        self,
        request: dict[str, Any],
        result: dict[str, Any],
        selected_days: list[int] | None,
        daily_rows: list[Any] | None,
    ) -> tuple[bytes, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Raport"

        columns = result["columns"]
        rows = result["rows"]
        append_openpyxl_row(ws, [column["label"] for column in columns])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1f2937")
            cell.fill = PatternFill("solid", fgColor="EEF2FF")

        for row in rows:
            append_openpyxl_row(ws, [row.get(column["key"]) for column in columns])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for idx, column in enumerate(columns, start=1):
            letter = get_column_letter(idx)
            width = max(12, min(28, len(column["label"]) + 3))
            ws.column_dimensions[letter].width = width
            number_format = self._excel_number_format(column["type"])
            if number_format:
                for cell in ws[letter][1:]:
                    cell.number_format = number_format

        cfg = wb.create_sheet("Configuratie")
        append_openpyxl_row(cfg, ["Optiune", "Valoare"])
        append_openpyxl_row(cfg, ["Dataset", DATASETS[request["dataset"]]["label"]])
        append_openpyxl_row(cfg, ["Luni", ", ".join(request["months"])])
        append_openpyxl_row(cfg, ["Zile", ", ".join(str(day) for day in selected_days) if selected_days else "Toata luna"])
        append_openpyxl_row(cfg, ["Generat", datetime.now().strftime("%Y-%m-%d %H:%M")])
        append_openpyxl_row(cfg, ["Randuri", len(rows)])
        for cell in cfg[1]:
            cell.font = Font(bold=True)
        cfg.column_dimensions["A"].width = 24
        cfg.column_dimensions["B"].width = 64

        if request.get("daily_metrics") and daily_rows is not None:
            self._add_daily_evolution_sheet(
                wb,
                months=sorted(request["months"]),
                metrics=[metric for metric in request["daily_metrics"] if metric in DAILY_EVOLUTION_METRICS],
                records=daily_rows,
            )

        stream = BytesIO()
        wb.save(stream)
        filename = request.get("filename") or (
            f"export_retail_{request['dataset']}_{'_'.join(request['months'])}"
            f"{self._days_filename_suffix(selected_days)}.xlsx"
        )
        return stream.getvalue(), self._safe_filename(str(filename))

    async def _preview_daily_comparison(self, request: dict[str, Any]) -> dict[str, Any]:
        months, metrics, levels, filters, include_closed_stores, selected_days = self._daily_comparison_params(request)
        campaign_codes_by_month = self._campaign_codes_by_month(months)
        preview_level = "general" if "general" in levels else levels[0]
        records = await self.repo.fetch_daily_comparison_rows(
            level=preview_level,
            months=months,
            filters=filters,
            include_closed_stores=include_closed_stores,
            campaign_codes_by_month=campaign_codes_by_month,
            selected_days=selected_days,
        )
        table = self._daily_comparison_table(
            level=preview_level,
            months=months,
            metrics=metrics,
            records=records,
            selected_days=selected_days,
        )
        limit = int(request.get("preview_limit") or 100)
        return {
            "columns": table["columns"],
            "rows": table["rows"][:limit],
            "total_rows": len(table["rows"]),
            "truncated": len(table["rows"]) > limit,
        }

    async def _build_daily_comparison_xlsx(self, request: dict[str, Any]) -> tuple[bytes, str]:
        months, metrics, levels, filters, include_closed_stores, selected_days = self._daily_comparison_params(request)
        campaign_codes_by_month = self._campaign_codes_by_month(months)
        tables: list[tuple[str, dict[str, Any]]] = []
        for level in levels:
            records = await self.repo.fetch_daily_comparison_rows(
                level=level,
                months=months,
                filters=filters,
                include_closed_stores=include_closed_stores,
                campaign_codes_by_month=campaign_codes_by_month,
                selected_days=selected_days,
            )
            table = await asyncio.to_thread(
                self._daily_comparison_table,
                level=level,
                months=months,
                metrics=metrics,
                records=records,
                selected_days=selected_days,
            )
            tables.append((level, table))
        return await asyncio.to_thread(
            self._render_daily_comparison_xlsx,
            request,
            months,
            metrics,
            levels,
            include_closed_stores,
            selected_days,
            tables,
        )

    def _render_daily_comparison_xlsx(
        self,
        request: dict[str, Any],
        months: list[str],
        metrics: list[str],
        levels: list[str],
        include_closed_stores: bool,
        selected_days: list[int] | None,
        tables: list[tuple[str, dict[str, Any]]],
    ) -> tuple[bytes, str]:
        wb = Workbook()
        first_sheet = True
        total_rows = 0

        for level, table in tables:
            sheet_name = COMPARISON_LEVELS[level]["sheet"]
            ws = wb.active if first_sheet else wb.create_sheet(sheet_name)
            ws.title = sheet_name
            first_sheet = False
            self._write_table_sheet(ws, table["columns"], table["rows"], header_fill="DCFCE7")
            total_rows += len(table["rows"])
            self._add_daily_comparison_chart(
                ws,
                months=months,
                metric=metrics[0],
                max_row=len(table["rows"]) + 1,
                first_data_col=len(COMPARISON_LEVELS[level]["dimensions"]) + 2,
            )

        cfg = wb.create_sheet("Configuratie")
        append_openpyxl_row(cfg, ["Optiune", "Valoare"])
        append_openpyxl_row(cfg, ["Tip export", "Evolutie zilnica comparativa"])
        append_openpyxl_row(cfg, ["Luni", ", ".join(months)])
        append_openpyxl_row(cfg, ["Zile", ", ".join(str(day) for day in selected_days) if selected_days else "Toata luna"])
        append_openpyxl_row(cfg, ["Metrici zilnice", ", ".join(DAILY_EVOLUTION_METRICS[item].label for item in metrics)])
        append_openpyxl_row(cfg, ["Niveluri", ", ".join(str(COMPARISON_LEVELS[item]["label"]) for item in levels)])
        append_openpyxl_row(cfg, ["Include magazine inchise", "Da" if include_closed_stores else "Nu"])
        append_openpyxl_row(cfg, ["Generat", datetime.now().strftime("%Y-%m-%d %H:%M")])
        append_openpyxl_row(cfg, ["Randuri", total_rows])
        for cell in cfg[1]:
            cell.font = Font(bold=True)
        cfg.column_dimensions["A"].width = 28
        cfg.column_dimensions["B"].width = 72

        stream = BytesIO()
        wb.save(stream)
        filename = request.get("filename") or (
            f"export_retail_evolutie_zilnica_{'_'.join(months)}"
            f"{self._days_filename_suffix(selected_days)}.xlsx"
        )
        return stream.getvalue(), self._safe_filename(str(filename))

    def _base_row(self, record: Any, dimensions: list[str], metrics: list[str]) -> dict[str, Any]:
        row = {dimension: record[dimension] for dimension in dimensions}
        computed = self._compute_metrics(record)
        for metric in metrics:
            row[metric] = computed.get(metric)
        return row

    def _attach_period_metrics(
        self,
        rows_by_key: dict[tuple[Any, ...], dict[str, Any]],
        records: list[Any],
        dataset: str,
        metrics: list[str],
        *,
        period_prefix: str,
    ) -> None:
        for record in records:
            row = rows_by_key.get(self._row_key(record, dataset))
            if row is None:
                continue
            period_key = str(record["period_key"])
            computed = self._compute_metrics(record)
            for metric in metrics:
                row[f"{period_prefix}:{period_key}:{metric}"] = computed.get(metric)

    def _build_columns(
        self,
        dataset: str,
        dimensions: list[str],
        metrics: list[str],
        months: list[str],
        monthly_metrics: list[str],
        rows_by_key: dict[tuple[Any, ...], dict[str, Any]],
        daily_metrics: list[str],
    ) -> list[dict[str, str]]:
        columns: list[dict[str, str]] = []
        for key in dimensions:
            columns.append(self._column_payload(DIMENSIONS[key]))
        for key in metrics:
            columns.append(self._column_payload(METRICS[key]))
        for month in months:
            for metric in monthly_metrics:
                definition = EVOLUTION_METRICS[metric]
                columns.append({
                    "key": f"month:{month}:{metric}",
                    "label": f"{month} {definition.label}",
                    "type": definition.type,
                    "group": "Evolutie lunara",
                })
        day_keys = sorted({
            key.split(":")[1]
            for row in rows_by_key.values()
            for key in row
            if key.startswith("day:")
        })
        for day in day_keys:
            for metric in daily_metrics:
                definition = DAILY_EVOLUTION_METRICS[metric]
                columns.append({
                    "key": f"day:{day}:{metric}",
                    "label": f"{day} {definition.label}",
                    "type": definition.type,
                    "group": "Evolutie zilnica",
                })
        return columns

    def _compute_metrics(self, row: Any) -> dict[str, Any]:
        total_sales = Decimal(row["total_sales"] or 0)
        total_quantity = int(row["total_quantity"] or 0)
        total_receipts = int(row["total_receipts"] or 0)
        receipt_2plus_count = int(row["receipt_2plus_count"] or 0)
        focus_quantity = int(row["focus_quantity"] or 0)
        target = Decimal(row["target"] or 0)
        working_days = int(row["working_days"] or 0)
        return {
            "total_sales": total_sales,
            "total_quantity": total_quantity,
            "total_receipts": total_receipts,
            "avg_receipt_value": self._ratio(total_sales, total_receipts),
            "proc_bon2acc": self._pct(Decimal(receipt_2plus_count), Decimal(total_receipts)),
            "prc_focus_acc_qty": self._pct(Decimal(focus_quantity), Decimal(total_quantity)),
            "target": target,
            "target_progress_pct": self._pct(total_sales, target),
            "working_days": working_days,
            "daily_average": self._ratio(total_sales, working_days),
            "store_count": int(row["store_count"] or 0),
            "agent_count": int(row["agent_count"] or 0),
            "incentive_sales": Decimal(row["incentive_sales"] or 0),
            "incentive_quantity": int(row["incentive_quantity"] or 0),
            "incentive_bonus": Decimal(row["incentive_bonus"] or 0),
            "promo_sales": Decimal(row["promo_sales"] or 0),
            "promo_quantity": int(row["promo_quantity"] or 0),
        }

    def _campaign_codes_by_month(self, months: list[str]) -> dict[str, list[str]]:
        config, error = load_special_cards_config()
        if error is not None:
            return {}

        out: dict[str, list[str]] = {}
        for month in months:
            definitions, definitions_error = parse_promotion_definitions(config, month)
            if definitions_error is not None:
                continue
            codes: set[str] = set()
            for definition in definitions:
                products, products_error = load_promotion_rule_products(definition)
                if products_error is not None or products is None:
                    continue
                rule_type = definition.get("rule_type") or "selected_item_copurchase"
                if rule_type in {"same_model_screen_camera", "trigger_discounted"}:
                    codes.update(str(code) for code in products.get("discounted_codes", []))
                else:
                    codes.update(str(code) for code in products.get("item_codes", []))
            if codes:
                out[month] = sorted(codes)
        return out

    async def _campaign_exclusions_by_month(
        self,
        months: list[str],
        filters: dict[str, list[str]],
        selected_days: list[int] | None = None,
    ) -> dict[str, dict[tuple[str, str, str], int]]:
        pool = getattr(self.repo, "pool", None)
        if pool is None:
            return {}

        def csv_filter(key: str) -> str | None:
            values = [value for value in filters.get(key, []) if value]
            return ",".join(values) if values else None

        out: dict[str, dict[tuple[str, str, str], int]] = {}
        async with pool.acquire() as conn:
            for month in months:
                if selected_days:
                    config, config_error = load_special_cards_config()
                    definitions, definitions_error = parse_promotion_definitions(config, month)
                    if config_error is not None or definitions_error is not None:
                        continue
                    month_units: dict[tuple[str, str, str], int] = {}
                    for definition in definitions:
                        selected_dates: list[date] = []
                        year, month_number = (int(value) for value in month.split("-", 1))
                        for day_number in selected_days:
                            try:
                                selected_date = date(year, month_number, day_number)
                            except ValueError:
                                continue
                            if definition["start_date"] <= selected_date <= definition["end_date"]:
                                selected_dates.append(selected_date)
                        if not selected_dates:
                            continue
                        ranges: list[tuple[date, date]] = []
                        range_start = range_end = selected_dates[0]
                        for selected_date in selected_dates[1:]:
                            if selected_date == range_end + timedelta(days=1):
                                range_end = selected_date
                            else:
                                ranges.append((range_start, range_end))
                                range_start = range_end = selected_date
                        ranges.append((range_start, range_end))
                        for range_start, range_end in ranges:
                            scoped_definition = {
                                **definition,
                                "start_date": range_start,
                                "end_date": range_end,
                            }
                            # A POS report contains a cumulative actual through its
                            # cutoff. Use it only when the selected range fully
                            # contains that reported interval; otherwise the
                            # report cannot be split reliably by day.
                            if definition.get("actuals_source_file") or definition.get("actuals_file"):
                                cutoff_date = promo_actuals_cutoff_date(definition)
                                use_actuals = (
                                    range_start > cutoff_date
                                    if cutoff_date is not None
                                    else range_start <= definition["start_date"] and range_end >= definition["end_date"]
                                ) or (
                                    cutoff_date is not None
                                    and range_start <= definition["start_date"]
                                    and range_end >= cutoff_date
                                )
                                if not use_actuals:
                                    scoped_definition["actuals_source_file"] = None
                                    scoped_definition["actuals_file"] = None
                            result = await _compute_dashboard_promotion_result(
                                conn,
                                month=month,
                                definition=scoped_definition,
                                firma=csv_filter("firma"),
                                regional=csv_filter("regional"),
                                asm=csv_filter("asm"),
                                site_code=csv_filter("site_code"),
                                agent=csv_filter("agent"),
                            )
                            if result is None:
                                continue
                            for key, units in result.excluded_units.items():
                                month_units[key] = month_units.get(key, 0) + units
                    if month_units:
                        out[month] = month_units
                    continue
                context = await _load_dashboard_campaign_context(
                    conn,
                    month,
                    firma=csv_filter("firma"),
                    regional=csv_filter("regional"),
                    asm=csv_filter("asm"),
                    site_code=csv_filter("site_code"),
                    agent=csv_filter("agent"),
                )
                if context.promo_excluded_units:
                    out[month] = {
                        (str(site), str(agent), str(item)): int(units)
                        for (site, agent, item), units in context.promo_excluded_units.items()
                    }
        return out

    def _row_key(self, row: Any, dataset: str) -> tuple[Any, ...]:
        key_fields = {
            "agents": ["agent", "site_code"],
            "stores": ["site_code"],
            "regionals": ["regional"],
            "asms": ["regional", "asm"],
        }[dataset]
        return tuple(row[field] for field in key_fields)

    def _public_row(self, row: dict[str, Any], columns: list[dict[str, str]]) -> dict[str, Any]:
        return {column["key"]: self._json_value(row.get(column["key"])) for column in columns}

    def _daily_comparison_params(
        self,
        request: dict[str, Any],
    ) -> tuple[list[str], list[str], list[str], dict[str, list[str]], bool, list[int] | None]:
        months = sorted({str(item) for item in request.get("months", []) if item})
        if not months:
            raise ExportValidationError("Selecteaza cel putin o luna.")
        if len(months) > 6:
            raise ExportValidationError("Comparatia zilnica este limitata la maxim 6 luni.")

        metrics = self._valid_keys(
            request.get("daily_metrics"),
            set(DAILY_EVOLUTION_METRICS),
            ["total_sales"],
            "metrici zilnice",
        )
        if len(metrics) > 4:
            raise ExportValidationError("Comparatia zilnica poate contine maxim 4 metrici.")

        levels = self._valid_keys(
            request.get("comparison_levels"),
            set(COMPARISON_LEVELS),
            ["general", "asms", "stores", "agents"],
            "niveluri de comparatie",
        )
        filters = self._normalize_filters(request.get("filters") or {})
        include_closed_stores = bool(request.get("include_closed_stores", False))
        return months, metrics, levels, filters, include_closed_stores, self._selected_days(request)

    def _daily_comparison_table(
        self,
        *,
        level: str,
        months: list[str],
        metrics: list[str],
        records: list[Any],
        selected_days: list[int] | None = None,
    ) -> dict[str, Any]:
        dimensions = list(COMPARISON_LEVELS[level]["dimensions"])
        columns: list[dict[str, str]] = [
            self._column_payload(DIMENSIONS[dimension]) for dimension in dimensions
        ]
        columns.append({"key": "day_of_month", "label": "Zi", "type": "integer", "group": "Perioada"})

        for metric in metrics:
            definition = DAILY_EVOLUTION_METRICS[metric]
            for month in months:
                columns.append({
                    "key": f"{month}:{metric}",
                    "label": f"{month} {definition.label}",
                    "type": definition.type,
                    "group": "Evolutie zilnica",
                })
            if len(months) == 2:
                if definition.type == "percent":
                    columns.append({
                        "key": f"delta_pp:{metric}",
                        "label": f"Delta pp {months[1]} vs {months[0]} {definition.label}",
                        "type": "percent",
                        "group": "Comparatie",
                    })
                else:
                    columns.append({
                        "key": f"delta:{metric}",
                        "label": f"Delta {months[1]} vs {months[0]} {definition.label}",
                        "type": definition.type,
                        "group": "Comparatie",
                    })
                    columns.append({
                        "key": f"delta_pct:{metric}",
                        "label": f"Delta % {months[1]} vs {months[0]} {definition.label}",
                        "type": "percent",
                        "group": "Comparatie",
                    })

        values: dict[tuple[tuple[Any, ...], int, str], dict[str, Any]] = {}
        dimension_labels: dict[tuple[Any, ...], dict[str, Any]] = {}
        for record in records:
            dim_key = tuple(record[dimension] for dimension in dimensions)
            dimension_labels[dim_key] = {dimension: record[dimension] for dimension in dimensions}
            day = int(record["day_of_month"] or 0)
            if day <= 0:
                continue
            values[(dim_key, day, str(record["import_month"]))] = self._compute_metrics(record)

        if level == "general" and not dimension_labels:
            dimension_labels[()] = {}
        max_day = self._max_days_for_months(months)
        days = selected_days or list(range(1, max_day + 1))
        rows: list[dict[str, Any]] = []

        for dim_key in sorted(dimension_labels, key=lambda item: tuple(str(value or "") for value in item)):
            dim_values = dimension_labels[dim_key]
            for day in days:
                row: dict[str, Any] = {dimension: dim_values.get(dimension) for dimension in dimensions}
                row["day_of_month"] = day
                for metric in metrics:
                    month_values: list[Any] = []
                    for month in months:
                        value = values.get((dim_key, day, month), {}).get(metric)
                        month_values.append(value)
                        row[f"{month}:{metric}"] = self._json_value(value)
                    if len(months) == 2:
                        left, right = month_values
                        if left is not None and right is not None:
                            delta = Decimal(str(right)) - Decimal(str(left))
                            if DAILY_EVOLUTION_METRICS[metric].type == "percent":
                                row[f"delta_pp:{metric}"] = self._json_value(delta)
                            else:
                                row[f"delta:{metric}"] = self._json_value(delta)
                                row[f"delta_pct:{metric}"] = self._json_value(
                                    self._pct(delta, Decimal(str(left))) if Decimal(str(left)) != 0 else None
                                )
                        else:
                            if DAILY_EVOLUTION_METRICS[metric].type == "percent":
                                row[f"delta_pp:{metric}"] = None
                            else:
                                row[f"delta:{metric}"] = None
                                row[f"delta_pct:{metric}"] = None
                rows.append(row)

        return {
            "columns": columns,
            "rows": [self._public_row(row, columns) for row in rows],
            "max_day": max_day,
        }

    def _selected_days(self, request: dict[str, Any]) -> list[int] | None:
        raw_days = request.get("selected_days")
        if raw_days is None:
            return None
        try:
            days = sorted({int(day) for day in raw_days})
        except (TypeError, ValueError) as exc:
            raise ExportValidationError("Selectia zilelor este invalida.") from exc
        if not days:
            raise ExportValidationError("Selecteaza cel putin o zi.")
        invalid = [day for day in days if day < 1 or day > 31]
        if invalid:
            raise ExportValidationError("Zilele trebuie sa fie intre 1 si 31.")
        return None if days == list(range(1, 32)) else days

    def _days_filename_suffix(self, selected_days: list[int] | None) -> str:
        if not selected_days:
            return ""
        value = "-".join(str(day) for day in selected_days) if len(selected_days) <= 10 else f"{len(selected_days)}selectate"
        return f"_zile_{value}"

    def _max_days_for_months(self, months: list[str]) -> int:
        max_day = 0
        for month in months:
            try:
                year_value, month_value = month.split("-", 1)
                year = int(year_value)
                month_number = int(month_value)
                if month_number == 12:
                    next_month = datetime(year + 1, 1, 1)
                else:
                    next_month = datetime(year, month_number + 1, 1)
                current_month = datetime(year, month_number, 1)
                max_day = max(max_day, (next_month - current_month).days)
            except (TypeError, ValueError):
                continue
        return max_day or 31

    def _normalize_filters(self, filters: dict[str, Any]) -> dict[str, list[str]]:
        allowed = {"firma", "regional", "asm", "site_code", "agent"}
        normalized: dict[str, list[str]] = {}
        for key, value in filters.items():
            if key not in allowed:
                continue
            if isinstance(value, list):
                normalized[key] = [str(item) for item in value if str(item).strip()]
            elif value:
                normalized[key] = [str(value)]
        return normalized

    def _valid_keys(
        self,
        value: Any,
        allowed: set[str],
        default: list[str],
        label: str,
    ) -> list[str]:
        if not value:
            return default
        keys = [str(item) for item in value]
        invalid = [key for key in keys if key not in allowed]
        if invalid:
            raise ExportValidationError(f"Selectie invalida pentru {label}: {', '.join(invalid)}")
        return keys

    def _column_payload(self, definition: ColumnDef) -> dict[str, str]:
        return {
            "key": definition.key,
            "label": definition.label,
            "type": definition.type,
            "group": definition.group,
        }

    def _ratio(self, value: Decimal, base: int | Decimal) -> Decimal | None:
        base_decimal = Decimal(base)
        if base_decimal == 0:
            return None
        return (value / base_decimal).quantize(Decimal("0.01"))

    def _pct(self, value: Decimal, base: Decimal) -> Decimal | None:
        if base == 0:
            return None
        return ((value * Decimal(100)) / base).quantize(Decimal("0.01"))

    def _json_value(self, value: Any) -> Any:
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _safe_filename(self, value: str) -> str:
        safe = "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in value).strip("._")
        if not safe:
            safe = "export_retail"
        if not safe.lower().endswith(".xlsx"):
            safe += ".xlsx"
        return safe[:140]

    def _excel_number_format(self, column_type: str) -> str | None:
        if column_type == "currency":
            return '#,##0.00'
        if column_type == "percent":
            return '0.00'
        if column_type == "integer":
            return '0'
        return None

    def _write_table_sheet(
        self,
        ws: Any,
        columns: list[dict[str, str]],
        rows: list[dict[str, Any]],
        *,
        header_fill: str,
    ) -> None:
        append_openpyxl_row(ws, [column["label"] for column in columns])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1f2937")
            cell.fill = PatternFill("solid", fgColor=header_fill)

        for row in rows:
            append_openpyxl_row(ws, [row.get(column["key"]) for column in columns])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for idx, column in enumerate(columns, start=1):
            letter = get_column_letter(idx)
            ws.column_dimensions[letter].width = max(10, min(30, len(column["label"]) + 3))
            number_format = self._excel_number_format(column["type"])
            if number_format:
                for cell in ws[letter][1:]:
                    cell.number_format = number_format

    def _add_daily_comparison_chart(
        self,
        ws: Any,
        *,
        months: list[str],
        metric: str,
        max_row: int,
        first_data_col: int,
    ) -> None:
        if not months or max_row < 2:
            return
        chart = LineChart()
        chart.title = f"Comparatie zilnica - {DAILY_EVOLUTION_METRICS[metric].label}"
        chart.y_axis.title = DAILY_EVOLUTION_METRICS[metric].label
        self._configure_day_axis(chart)
        chart.visible_cells_only = True
        chart.style = 13
        data = Reference(
            ws,
            min_col=first_data_col,
            max_col=first_data_col + len(months) - 1,
            min_row=1,
            max_row=max_row,
        )
        categories = Reference(ws, min_col=first_data_col - 1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 20
        ws.add_chart(chart, f"{get_column_letter(first_data_col + len(months) + 3)}2")

    def _add_daily_evolution_sheet(
        self,
        wb: Workbook,
        *,
        months: list[str],
        metrics: list[str],
        records: list[Any],
    ) -> None:
        if not metrics:
            return
        ws = wb.create_sheet("Evolutie zilnica")
        values: dict[tuple[int, str], dict[str, Any]] = {}
        max_day = 31
        for record in records:
            day = int(record["day_of_month"] or 0)
            if day <= 0:
                continue
            max_day = max(max_day, day)
            values[(day, str(record["import_month"]))] = self._compute_metrics(record)

        headers = ["Zi"]
        for metric in metrics:
            definition = DAILY_EVOLUTION_METRICS[metric]
            for month in months:
                headers.append(f"{month} {definition.label}")
            if len(months) == 2:
                if definition.type == "percent":
                    headers.append(f"Delta pp {definition.label}")
                else:
                    headers.append(f"Delta {definition.label}")
                    headers.append(f"Delta % {definition.label}")
        append_openpyxl_row(ws, headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1f2937")
            cell.fill = PatternFill("solid", fgColor="DCFCE7")

        for day in range(1, max_day + 1):
            row: list[Any] = [day]
            for metric in metrics:
                month_values: list[Any] = []
                for month in months:
                    value = values.get((day, month), {}).get(metric)
                    month_values.append(value)
                    row.append(self._json_value(value))
                if len(months) == 2:
                    left = month_values[0]
                    right = month_values[1]
                    if left is not None and right is not None:
                        delta = Decimal(str(right)) - Decimal(str(left))
                        row.append(self._json_value(delta))
                        if DAILY_EVOLUTION_METRICS[metric].type != "percent":
                            row.append(self._json_value(self._pct(delta, Decimal(str(left))) if Decimal(str(left)) != 0 else None))
                    else:
                        row.append(None)
                        if DAILY_EVOLUTION_METRICS[metric].type != "percent":
                            row.append(None)
            append_openpyxl_row(ws, row)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions["A"].width = 8
        for idx, header in enumerate(headers[1:], start=2):
            letter = get_column_letter(idx)
            ws.column_dimensions[letter].width = max(14, min(26, len(header) + 2))
        if months:
            chart = LineChart()
            chart.title = f"Evolutie zilnica - {DAILY_EVOLUTION_METRICS[metrics[0]].label}"
            chart.y_axis.title = DAILY_EVOLUTION_METRICS[metrics[0]].label
            self._configure_day_axis(chart)
            data = Reference(ws, min_col=2, max_col=1 + len(months), min_row=1, max_row=max_day + 1)
            categories = Reference(ws, min_col=1, min_row=2, max_row=max_day + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 18
            ws.add_chart(chart, f"{get_column_letter(len(headers) + 2)}2")

    def _configure_day_axis(self, chart: LineChart) -> None:
        chart.x_axis.title = "Zi"
        chart.x_axis.delete = False
        chart.x_axis.axPos = "b"
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1
        chart.x_axis.majorTickMark = "out"
        chart.x_axis.noMultiLvlLbl = True
