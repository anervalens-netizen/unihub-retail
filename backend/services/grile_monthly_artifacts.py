"""Pure path and workbook artifact helpers for monthly Grile operations."""

from __future__ import annotations

import os
import re
import shutil
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.grile_monthly_integrity import (
    MonthlyIntegrityError,
    file_sha256,
    secure_directory,
    secure_file,
)
from services.grile_monthly_types import (
    AUDIT_HEADERS,
    HEADERS,
    ExtractedAgentRow,
    StoreEntry,
)
from services.spreadsheet_safety import TrustedFormula, append_openpyxl_row


FINAL_EXPORT_NAME_PREFIX = "Tabel Salarii -"
ARCHIVE_DIR_NAME = "archive"


def safe_filename(value: str) -> str:
    cleaned = value.replace("/", " - ").replace("\\", " - ")
    cleaned = re.sub(r'[<>:"|?*]', "_", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned.rstrip(". ") or "untitled"


def month_slug(month: str) -> str:
    slug = re.sub(r"[^0-9A-Za-zĂÂÎȘȚăâîșț]+", "-", month.strip())
    return slug.strip("-")


def build_final_export_path(outputs_dir: Path, month: str) -> Path:
    return outputs_dir / f"{FINAL_EXPORT_NAME_PREFIX} {month}.xlsx"


def build_archive_dir(outputs_dir: Path, month: str) -> Path:
    return outputs_dir / ARCHIVE_DIR_NAME / month


def build_archive_manifest_path(outputs_dir: Path, month: str) -> Path:
    return build_archive_dir(outputs_dir, month) / f"archive-manifest-{month_slug(month)}.json"


def build_archive_zip_path(outputs_dir: Path, month: str) -> Path:
    return build_archive_dir(outputs_dir, month) / f"Grile - {month}.zip"


def build_reset_report_path(outputs_dir: Path, next_month: str) -> Path:
    return outputs_dir / f"reset-report-{month_slug(next_month)}.json"


def build_reset_dry_run_report_path(outputs_dir: Path, next_month: str) -> Path:
    return outputs_dir / f"reset-dry-run-{month_slug(next_month)}.json"


def build_reset_backup_dir(
    outputs_dir: Path,
    closing_month: str,
    operation_id: int,
) -> Path:
    return outputs_dir / "reset-backups" / month_slug(closing_month) / str(operation_id)


def build_store_export_path(outputs_dir: Path, month: str, entry: StoreEntry) -> Path:
    return (
        build_archive_dir(outputs_dir, month)
        / safe_filename(entry.company)
        / f"{safe_filename(entry.store)}.xlsx"
    )


def build_manager_zip_path(outputs_dir: Path, month: str, manager: str) -> Path:
    return (
        build_archive_dir(outputs_dir, month)
        / "ASM"
        / f"Grile - {month} - {safe_filename(manager)}.zip"
    )


def resolve_output_path(month: str, only: str | None, output_dir: Path) -> Path:
    output_path = build_final_export_path(output_dir, month)
    if only:
        output_path = output_path.with_name(
            f"{output_path.stem} - TEST {safe_filename(only)}{output_path.suffix}"
        )
    return output_path


def validate_archive_manifest(
    manifest: dict[str, Any],
    expected_count: int,
) -> tuple[bool, list[str]]:
    errors = _archive_count_errors(manifest, expected_count)
    stores = manifest.get("stores")
    if not isinstance(stores, list) or len(stores) != expected_count:
        count = len(stores) if isinstance(stores, list) else "invalid"
        errors.append(f"stores count mismatch: {count} != {expected_count}")
    else:
        errors.extend(error for store in stores for error in _store_artifact_errors(store))
    zip_path = Path(str(manifest.get("zip_path", "")))
    if not zip_path.exists() or zip_path.stat().st_size == 0:
        errors.append(f"missing or empty archive zip: {zip_path}")
    return not errors, errors


def _archive_count_errors(manifest: dict[str, Any], expected_count: int) -> list[str]:
    errors: list[str] = []
    if manifest.get("registry_count") != expected_count:
        errors.append(f"registry_count mismatch: {manifest.get('registry_count')} != {expected_count}")
    if manifest.get("exported_count") != expected_count:
        errors.append(f"exported_count mismatch: {manifest.get('exported_count')} != {expected_count}")
    if manifest.get("error_count") != 0:
        errors.append(f"archive has {manifest.get('error_count')} export errors")
    return errors


def _store_artifact_errors(store: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    company = store.get("company", "?")
    name = store.get("store", "?")
    if store.get("status") != "OK":
        errors.append(f"{company}/{name} status is {store.get('status')}")
    xlsx_path = Path(str(store.get("xlsx_path", "")))
    if not xlsx_path.exists() or xlsx_path.stat().st_size == 0:
        errors.append(f"missing or empty export: {xlsx_path}")
    return errors


def make_output_row(
    row: ExtractedAgentRow,
    nr: int,
    metadata: dict[str, Any],
) -> list[Any]:
    excel_row = nr + 1
    return [
        nr,
        metadata.get("Manager", ""),
        row.store,
        row.agent,
        row.base_salary,
        row.sales_commission,
        metadata.get("Flip", ""),
        row.extra_location_commission,
        metadata.get("Incentive lunar", ""),
        row.extra_hours_pay,
        TrustedFormula(f"=SUM(E{excel_row}:J{excel_row},M{excel_row})"),
        TrustedFormula(f"=K{excel_row}-M{excel_row}"),
        row.bonuri,
        metadata.get("Data angajarii", ""),
        metadata.get("Data plecarii", ""),
        row.worked_hours,
        metadata.get("Zile CO luna in curs", ""),
    ]


def style_sheet(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 8,
        "B": 22,
        "C": 30,
        "D": 30,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 26,
        "I": 16,
        "J": 18,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 16,
        "O": 16,
        "P": 16,
        "Q": 16,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def build_workbook(
    rows: list[ExtractedAgentRow],
    output_path: Path,
    metadata_by_company_store: dict[tuple[str, str], dict[str, Any]],
    *,
    style: Callable[[Any], None] = style_sheet,
) -> None:
    workbook = _new_workbook()
    counters = {"Mobiup": 1, "Mobicell": 1}
    for row in rows:
        if row.status == "OK":
            worksheet = workbook[row.company]
            metadata = metadata_by_company_store.get((row.company, row.store), {})
            append_openpyxl_row(worksheet, make_output_row(row, counters[row.company], metadata))
            counters[row.company] += 1
    _append_audit_rows(workbook["Audit"], rows)
    for worksheet in workbook.worksheets:
        _finish_sheet(worksheet, style)
    secure_directory(output_path.parent)
    workbook.save(output_path)


def _new_workbook() -> Workbook:
    workbook = Workbook()
    workbook.active.title = "Mobiup"
    workbook.create_sheet("Mobicell")
    audit = workbook.create_sheet("Audit")
    for worksheet in (workbook["Mobiup"], workbook["Mobicell"]):
        append_openpyxl_row(worksheet, HEADERS)
    append_openpyxl_row(audit, AUDIT_HEADERS)
    return workbook


def _append_audit_rows(worksheet: Any, rows: list[ExtractedAgentRow]) -> None:
    for row in rows:
        append_openpyxl_row(
            worksheet,
            [
                row.company,
                row.store,
                row.slot,
                row.agent,
                row.sheet_id,
                row.sales_commission,
                row.extra_location_commission,
                row.extra_hours_pay,
                row.bonuri,
                row.worked_hours,
                f"https://docs.google.com/spreadsheets/d/{row.sheet_id}",
                row.status,
                row.error,
            ],
        )


def _finish_sheet(worksheet: Any, style: Callable[[Any], None]) -> None:
    style(worksheet)
    for row_cells in worksheet.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_index in range(1, worksheet.max_column + 1):
        column = get_column_letter(column_index)
        if worksheet.column_dimensions[column].width is None:
            worksheet.column_dimensions[column].width = 14


def validate_final_workbook(path: Path, *, expected_agents: int) -> None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            _validate_workbook_contents(workbook, expected_agents)
        finally:
            workbook.close()
    except MonthlyIntegrityError:
        raise
    except Exception as exc:
        raise MonthlyIntegrityError("workbook_invalid", "Workbook cannot be verified") from exc


def _validate_workbook_contents(workbook: Any, expected_agents: int) -> None:
    if set(workbook.sheetnames) != {"Mobiup", "Mobicell", "Audit"}:
        raise MonthlyIntegrityError("workbook_structure_invalid", "Workbook structure is invalid")
    agent_rows = sum(max(workbook[company].max_row - 1, 0) for company in ("Mobiup", "Mobicell"))
    if agent_rows != expected_agents:
        raise MonthlyIntegrityError("workbook_coverage_incomplete", "Workbook coverage is incomplete")
    audit = workbook["Audit"]
    statuses = [row[11].value for row in audit.iter_rows(min_row=2) if len(row) >= 12]
    if len(statuses) != expected_agents or any(status != "OK" for status in statuses):
        raise MonthlyIntegrityError("workbook_audit_invalid", "Workbook audit is invalid")


def staging_dir(outputs_dir: Path, operation: str, operation_id: int | None) -> Path:
    suffix = str(operation_id) if operation_id is not None else "direct"
    path = outputs_dir / ".staging" / f"{operation}-{suffix}"
    if path.exists():
        shutil.rmtree(path)
    secure_directory(path)
    return path


def promote_file(outputs_dir: Path, staged: Path, destination: Path) -> None:
    secure_directory(destination.parent)
    revision = _preserve_revision(outputs_dir, destination)
    try:
        os.replace(staged, destination)
        secure_file(destination)
    except Exception:
        if revision is not None and revision.exists() and not destination.exists():
            os.replace(revision, destination)
        raise


def _preserve_revision(outputs_dir: Path, destination: Path) -> Path | None:
    if not destination.exists():
        return None
    revision_dir = outputs_dir / ".revisions"
    secure_directory(revision_dir)
    revision = revision_dir / f"{destination.name}.{file_sha256(destination)[:16]}"
    if revision.exists():
        destination.unlink()
    else:
        os.replace(destination, revision)
    return revision
