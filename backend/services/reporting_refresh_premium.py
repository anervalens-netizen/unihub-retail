from __future__ import annotations

from pathlib import Path
from typing import Sequence

import asyncpg
from openpyxl import load_workbook

from retail_filters import retail_exclusion_clauses
from services.phone_models import extract_phone_model_keys, phone_model_metadata


_REPO_ROOT = Path(__file__).resolve().parents[2]
_PREMIUM_CAMERA_FILE = _REPO_ROOT / "data" / "folii premium camera.xlsx"


_BRAND_GROUP_SQL = """
CASE
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%apple%'
         OR LOWER(COALESCE(st.item_name, '')) LIKE '%iphone%'
    THEN 'Apple'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%samsung%' THEN 'Samsung'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%honor%' THEN 'Honor'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%xiaomi%' THEN 'Xiaomi'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%huawei%' THEN 'Huawei'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%motorola%' THEN 'Motorola'
    ELSE 'Altele'
END
""".strip()

_MONTH_INDEX_SQL = (
    "(split_part({alias}, '-', 1)::INT * 12 + split_part({alias}, '-', 2)::INT)"
)

_RETAIL_TRANSACTION_EXCLUSIONS_SQL = "\n          AND ".join(
    retail_exclusion_clauses(site_alias="st", store_alias="s")
)

_PREMIUM_GLASS_INSERT_SQL = """
WITH target_models(model_key, model_label, model_regex, exclude_regex) AS (
    VALUES
        ('iphone_15', 'iPhone 15', 'IPHONE 15', 'IPHONE 15 PRO|IPHONE 15 PLUS -'),
        ('iphone_15_pro', 'iPhone 15 Pro', 'IPHONE 15 PRO|15 PRO/', 'IPHONE 15 PRO MAX|15 PRO MAX'),
        ('iphone_15_pro_max', 'iPhone 15 Pro Max', 'IPHONE 15 PRO MAX|15 PRO MAX', NULL),
        ('iphone_16', 'iPhone 16', 'IPHONE 16|/16([[:space:]]|-|$)', 'IPHONE 16 PRO|16 PRO|IPHONE 15 PLUS/16 PLUS|IPHONE 16 PLUS -'),
        ('iphone_16_pro', 'iPhone 16 Pro', 'IPHONE 16 PRO|16 PRO/', 'IPHONE 16 PRO MAX|16 PRO MAX'),
        ('iphone_16_pro_max', 'iPhone 16 Pro Max', 'IPHONE 16 PRO MAX|16 PRO MAX', NULL),
        ('iphone_17', 'iPhone 17', 'IPHONE 17|PRO/17([[:space:]]|-|$)', 'IPHONE 17 PRO|IPHONE 17 AIR'),
        ('iphone_17_pro', 'iPhone 17 Pro', 'IPHONE 17 PRO|17 PRO/', 'IPHONE 17 PRO MAX|17 PRO MAX'),
        ('iphone_17_pro_max', 'iPhone 17 Pro Max', 'IPHONE 17 PRO MAX|17 PRO MAX', NULL),
        ('samsung_s26', 'Samsung S26', 'SAMSUNG GALAXY S26 5G|S26 5G', 'S26 PLUS|S26 ULTRA'),
        ('samsung_s26_ultra', 'Samsung S26 Ultra', 'SAMSUNG GALAXY S26 ULTRA|S26 ULTRA', NULL)
),
camera_products AS (
    SELECT *
    FROM UNNEST(
        $1::TEXT[],
        $2::TEXT[],
        $3::BOOLEAN[],
        $4::TEXT[],
        $5::TEXT[]
    ) AS t(item_code, item_name, is_premium_glass, model_key, model_label)
),
source_products AS (
    SELECT DISTINCT
        st.item_code,
        st.item_name,
        UPPER(COALESCE(st.item_name, '')) AS item_name_upper
    FROM sales_transactions st
    WHERE LOWER(TRIM(COALESCE(st.category, ''))) = 'folii sticla'
      AND st.item_code IS NOT NULL
      AND TRIM(st.item_code) != ''
),
screen_matches AS (
    SELECT DISTINCT ON (sp.item_code, tm.model_key)
        sp.item_code,
        sp.item_name,
        (sp.item_name_upper ~ '(SAPPHIRE|CERAMIC|CORNING)') AS is_premium_glass,
        tm.model_key,
        tm.model_label
    FROM source_products sp
    JOIN target_models tm
        ON sp.item_name_upper ~ tm.model_regex
       AND (tm.exclude_regex IS NULL OR sp.item_name_upper !~ tm.exclude_regex)
    WHERE NOT (sp.item_code = ANY($1::TEXT[]))
    ORDER BY sp.item_code, tm.model_key, sp.item_name
),
camera_matches AS (
    SELECT DISTINCT
        item_code,
        item_name,
        is_premium_glass,
        model_key,
        model_label
    FROM camera_products
)
INSERT INTO premium_glass_item_models (
    item_code,
    item_name,
    is_premium_glass,
    model_key,
    model_label
)
SELECT DISTINCT ON (sp.item_code, sp.model_key)
    sp.item_code,
    sp.item_name,
    sp.is_premium_glass,
    sp.model_key,
    sp.model_label
FROM (
    SELECT * FROM camera_matches
    UNION ALL
    SELECT * FROM screen_matches
) sp
ORDER BY sp.item_code, sp.model_key, sp.item_name;

""".strip()


def _load_premium_camera_rows(
    path: Path = _PREMIUM_CAMERA_FILE,
) -> tuple[list[str], list[str], list[bool], list[str], list[str]]:
    if not path.exists():
        return [], [], [], [], []

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Folii Camera"] if "Folii Camera" in workbook.sheetnames else workbook.active
    headers = [str(value or "").strip().lower() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {name: index for index, name in enumerate(headers)}
    code_idx = header_index.get("cod")
    premium_idx = header_index.get("premium")
    name_idx = header_index.get("itemname")
    if code_idx is None or premium_idx is None or name_idx is None:
        return [], [], [], [], []

    item_codes: list[str] = []
    item_names: list[str] = []
    premium_flags: list[bool] = []
    model_keys: list[str] = []
    model_labels: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = str(row[code_idx] or "").strip()
        item_name = str(row[name_idx] or "").strip()
        premium_value = str(row[premium_idx] or "").strip().lower()
        if not code or not item_name:
            continue
        is_premium = premium_value in {"da", "yes", "true", "1"}
        for model in sorted(extract_phone_model_keys(item_name)):
            metadata = phone_model_metadata(model)
            if metadata is None:
                continue
            model_key, model_label = metadata
            item_codes.append(code)
            item_names.append(item_name)
            premium_flags.append(is_premium)
            model_keys.append(model_key)
            model_labels.append(model_label)
    return item_codes, item_names, premium_flags, model_keys, model_labels


async def refresh_premium_glass_indicators(conn: asyncpg.Connection) -> None:
    await conn.execute("TRUNCATE premium_glass_item_models")
    await conn.execute(_PREMIUM_GLASS_INSERT_SQL, *_load_premium_camera_rows())
    await conn.execute("ANALYZE premium_glass_item_models")
