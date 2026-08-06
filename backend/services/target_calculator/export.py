"""XLSX presentation boundary for Target Calculator scenarios."""

from __future__ import annotations

import json
from collections import defaultdict
from collections.abc import Awaitable, Callable
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from business_clock import business_today
from services.spreadsheet_safety import append_openpyxl_row
from services.target_calculator.rules import percent_change
from services.target_calculator.seasonality import month_label_ro, shift_month


def _money(value: object) -> Decimal:
    return Decimal(str(value or 0))


def manager_allocation_analysis(scenario: dict[str, Any]) -> list[dict[str, Any]]:
    target_month = scenario["target_month"]
    previous_year_base_month = shift_month(target_month, -13)
    previous_year_target_month = shift_month(target_month, -12)
    previous_month = shift_month(target_month, -1)
    rows_by_manager: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in scenario["rows"]:
        rows_by_manager[row["regional"]].append(row)

    def period_value(row: dict[str, Any], month: str) -> Decimal:
        period = next(
            (item for item in row["history"] if item["month"] == month), None
        )
        return _money((period or {}).get("realized"))

    def build(manager: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
        target = sum((_money(row["proposed_target"]) for row in rows), Decimal("0"))
        previous = sum((period_value(row, previous_month) for row in rows), Decimal("0"))
        previous_year_base = sum(
            (period_value(row, previous_year_base_month) for row in rows), Decimal("0")
        )
        previous_year_target = sum(
            (period_value(row, previous_year_target_month) for row in rows), Decimal("0")
        )
        forecast_values = [
            (row.get("profitability") or {}).get("forecast_sales") for row in rows
        ]
        forecast = (
            sum((_money(value) for value in forecast_values if value is not None), Decimal("0"))
            if all(value is not None for value in forecast_values)
            else None
        )
        seasonality_pct = percent_change(float(previous_year_target), float(previous_year_base))
        seasonal_target = (
            previous * (Decimal("1") + Decimal(str(seasonality_pct)) / Decimal("100"))
            if seasonality_pct is not None
            else None
        )
        target_vs_previous_pct = percent_change(float(target), float(previous))
        target_vs_seasonal_pct = (
            percent_change(float(target), float(seasonal_target))
            if seasonal_target is not None
            else None
        )
        target_vs_forecast_pct = (
            percent_change(float(target), float(forecast)) if forecast is not None else None
        )
        if target_vs_forecast_pct is not None and target_vs_forecast_pct >= 5:
            signal = "Peste AI"
        elif target_vs_seasonal_pct is not None and round(target_vs_seasonal_pct, 1) >= 3:
            signal = "Peste sezonier"
        else:
            signal = "Echilibrat"
        return {
            "manager": manager,
            "store_count": len(rows),
            "target": float(target),
            "previous": float(previous),
            "previous_year_base": float(previous_year_base),
            "previous_year_target": float(previous_year_target),
            "forecast": float(forecast) if forecast is not None else None,
            "target_vs_previous_pct": target_vs_previous_pct,
            "seasonality_pct": seasonality_pct,
            "seasonality_deviation_pp": (
                target_vs_previous_pct - seasonality_pct
                if target_vs_previous_pct is not None and seasonality_pct is not None
                else None
            ),
            "seasonal_target": float(seasonal_target) if seasonal_target is not None else None,
            "target_vs_seasonal_pct": target_vs_seasonal_pct,
            "target_vs_previous_year_pct": percent_change(float(target), float(previous_year_target)),
            "target_vs_forecast_pct": target_vs_forecast_pct,
            "signal": signal,
        }

    managers = [
        build(manager, manager_rows) for manager, manager_rows in rows_by_manager.items()
    ]
    managers.sort(key=lambda item: (-item["target"], item["manager"]))
    network = build("TOTAL REȚEA", list(scenario["rows"]))
    network["signal"] = "Rețea"
    for item in [*managers, network]:
        item["target_share"] = item["target"] / network["target"] if network["target"] > 0 else 0
        item["previous_share"] = item["previous"] / network["previous"] if network["previous"] > 0 else 0
        item["previous_year_share"] = (
            item["previous_year_target"] / network["previous_year_target"]
            if network["previous_year_target"] > 0
            else 0
        )
        item["forecast_share"] = (
            item["forecast"] / network["forecast"]
            if item["forecast"] is not None and network["forecast"]
            else None
        )
        item["target_vs_previous_share_pp"] = (
            item["target_share"] - item["previous_share"]
        ) * 100
        item["target_vs_previous_year_share_pp"] = (
            item["target_share"] - item["previous_year_share"]
        ) * 100
        item["target_vs_forecast_share_pp"] = (
            (item["target_share"] - item["forecast_share"]) * 100
            if item["forecast_share"] is not None
            else None
        )
    return [*managers, network]


async def build_target_excel(
    scenario_id: int,
    load_scenario: Callable[[int], Awaitable[dict[str, Any]]],
) -> tuple[BytesIO, str]:
    scenario = await load_scenario(scenario_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Target + profitabilitate"
    target_month = scenario["target_month"]
    comparison_months = [
        shift_month(target_month, -13),
        shift_month(target_month, -12),
        shift_month(target_month, -1),
    ]
    headers = ["Firma", "Manager", "Nume locație", "Cod locație"]
    for month in comparison_months:
        headers.extend([f"Target {month}", f"Realizat {month}", f"% {month}"])
    headers.extend([
        "Pondere calcul",
        f"Calcul target {month_label_ro(target_month)}",
        "Propunere manager",
        "Cheltuieli salariale la 90% - P&L estimat",
        "Cheltuieli operaționale estimate",
        "Break-even vânzări brute",
        f"Forecast {month_label_ro(target_month)}",
    ])
    append_openpyxl_row(sheet, ["SUBTOTAL", *([""] * (len(headers) - 1))])
    append_openpyxl_row(sheet, headers)
    sorted_rows = sorted(
        scenario["rows"],
        key=lambda row: (-row["proposed_target"], row["locatie"], row["site_code"]),
    )
    for row in sorted_rows:
        history_by_month = {period["month"]: period for period in row["history"]}
        values: list[Any] = [
            row["firma"], row["regional"], row["locatie"], row["site_code"],
        ]
        for month in comparison_months:
            history = history_by_month.get(month) or {}
            attainment = history.get("attainment_pct")
            values.extend([
                history.get("target", 0),
                history.get("realized", 0),
                None if attainment is None else attainment / 100,
            ])
        profitability = row.get("profitability") or {}
        values.extend([
            row.get("normalized_weight", row["calculated_weight"]),
            row["proposed_target"],
            row["final_target"],
            profitability.get("salary_cost_at_90_pct"),
            profitability.get("operating_costs"),
            profitability.get("break_even_gross_sales"),
            profitability.get("forecast_sales"),
        ])
        append_openpyxl_row(sheet, values)
    last_row = sheet.max_row
    sheet.freeze_panes = "E3"
    sheet.auto_filter.ref = f"A2:T{last_row}"

    total_columns = ("E", "F", "H", "I", "K", "L", "N", "O", "P", "Q", "R", "S", "T")
    for column in total_columns:
        sheet[f"{column}1"] = f"=SUBTOTAL(109,{column}3:{column}{last_row})"
    for percentage_column, target_column, realized_column in (
        ("G", "E", "F"), ("J", "H", "I"), ("M", "K", "L")
    ):
        sheet[f"{percentage_column}1"] = f"=IF({target_column}1=0,0,{realized_column}1/{target_column}1)"

    comparison = workbook.create_sheet("Comparație manageri")
    manager_analysis = manager_allocation_analysis(scenario)
    append_openpyxl_row(comparison, ["1. Distribuția targetului", *([""] * 8)])
    comparison.merge_cells("A1:I1")
    append_openpyxl_row(comparison, [
        "Manager", "Nr. locații", "Pondere target calculat",
        f"Pondere realizat {comparison_months[2]}",
        f"Δ vs {comparison_months[2]} (pp)",
        f"Pondere realizat {comparison_months[1]}",
        f"Δ vs {comparison_months[1]} (pp)",
        f"Pondere forecast {target_month}", "Δ vs forecast (pp)",
    ])
    for item in manager_analysis:
        append_openpyxl_row(comparison, [
            item["manager"], item["store_count"], item["target_share"],
            item["previous_share"], item["target_vs_previous_share_pp"],
            item["previous_year_share"], item["target_vs_previous_year_share_pp"],
            item["forecast_share"], item["target_vs_forecast_share_pp"],
        ])
    distribution_total_row = comparison.max_row
    append_openpyxl_row(comparison, [])
    second_title_row = distribution_total_row + 2
    append_openpyxl_row(comparison, [
        "2. Target vs lună precedentă, an precedent, sezonalitate și forecast AI",
        *([""] * 12),
    ])
    comparison.merge_cells(
        start_row=second_title_row, start_column=1,
        end_row=second_title_row, end_column=13,
    )
    second_header_row = second_title_row + 1
    append_openpyxl_row(comparison, [
        "Manager", f"Target {target_month}", f"Realizat {comparison_months[2]}",
        f"Target vs {comparison_months[2]}", "Sezonalitate istorică",
        "Abatere sezonalitate (pp)", "Target sezonier estimat",
        "Gap vs target sezonier", f"Realizat {comparison_months[1]}",
        f"Target vs {comparison_months[1]}", f"Forecast AI {target_month}",
        "Target vs AI", "Semnal alocare",
    ])
    for item in manager_analysis:
        append_openpyxl_row(comparison, [
            item["manager"], item["target"], item["previous"],
            item["target_vs_previous_pct"], item["seasonality_pct"],
            item["seasonality_deviation_pp"], item["seasonal_target"],
            item["target_vs_seasonal_pct"], item["previous_year_target"],
            item["target_vs_previous_year_pct"], item["forecast"],
            item["target_vs_forecast_pct"], item["signal"],
        ])
    analysis_total_row = comparison.max_row

    summary = workbook.create_sheet("Rezumat calcul")
    append_openpyxl_row(summary, [
        "Regional", "Magazine", "Floor", "Target propus", "Target final", "Diferenta",
        "Luna curenta", "Forecast luna curenta", "% crestere propus vs luna curenta",
        "Baza anul trecut", "Target anul trecut", "Realizat baza anul trecut",
        "Realizat target anul trecut", "% crestere anul trecut",
    ])
    for row in scenario["regional_summary"]:
        append_openpyxl_row(summary, [
            row["regional"], row["store_count"], row["floor_total"],
            row["proposed_total"], row["final_total"], row["final_total"] - row["proposed_total"],
            row.get("current_month"), row.get("current_forecast_total"), row.get("proposed_growth_vs_current_pct"),
            row.get("last_year_base_month"), row.get("last_year_target_month"),
            row.get("last_year_base_total"), row.get("last_year_target_total"), row.get("last_year_growth_pct"),
        ])

    parameters = workbook.create_sheet("Parametri")
    append_openpyxl_row(parameters, ["Parametru", "Valoare"])
    for key, value in (
        ("Scenariu", scenario["id"]), ("Status", scenario["status"]),
        ("Luna target", scenario["target_month"]),
        ("Luna cohorta magazine active", scenario["cohort_month"]),
        ("Target total", scenario["total_target"]),
        ("Prag minim absolut", scenario["min_floor"]),
        ("Floor fata de luna precedenta", scenario["previous_month_floor_pct"]),
        ("Metoda", scenario["calculation_method"]),
    ):
        append_openpyxl_row(parameters, [key, value])
    for key, value in (scenario.get("calculation_params") or {}).items():
        append_openpyxl_row(parameters, [
            f"Parametru {key}",
            json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value,
        ])
    for item in scenario["source_months"]:
        append_openpyxl_row(parameters, [item["label"], item["month"]])
    for item in scenario["source_summary"]:
        if item["is_forecast"]:
            append_openpyxl_row(parameters, [
                f"Forecast {item['month']}",
                f"{item['forecast_factor']:.4f}x; importat {item['actual_realized']:.2f}; folosit {item['realized']:.2f}",
            ])
    for warning in scenario["warnings"]:
        append_openpyxl_row(parameters, ["Atentionare", warning])
    profitability_summary = scenario.get("profitability_summary") or {}
    append_openpyxl_row(parameters, ["Status surse profitabilitate", profitability_summary.get("status")])
    append_openpyxl_row(parameters, ["Luni P&L reale", ", ".join(profitability_summary.get("pnl_months") or [])])
    forecast_run = profitability_summary.get("forecast_run") or {}
    append_openpyxl_row(parameters, ["Forecast run", forecast_run.get("id")])
    append_openpyxl_row(parameters, ["Forecast model", forecast_run.get("model_name")])
    append_openpyxl_row(parameters, ["Forecast variant", forecast_run.get("variant")])

    navy_fill = PatternFill("solid", fgColor="17365D")
    subtotal_fill = PatternFill("solid", fgColor="D9E2F3")
    percentage_fill = PatternFill("solid", fgColor="F3F4F6")
    manager_fill = PatternFill("solid", fgColor="FFF8D9")
    break_even_fill = PatternFill("solid", fgColor="FFF7ED")
    forecast_fill = PatternFill("solid", fgColor="EFF8F1")
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    green_fill = PatternFill("solid", fgColor="E2F0D9")
    red_font = Font(color="9C0006", bold=True)
    amber_font = Font(color="C65911", bold=True)
    green_font = Font(color="00B050", bold=True)
    for cell in sheet[1]:
        cell.font = Font(color="111827", bold=True)
        cell.fill = subtotal_fill
    for cell in sheet[2]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 52
    for row_number in range(1, last_row + 1):
        for column in ("G", "J", "M", "N"):
            sheet[f"{column}{row_number}"].number_format = "0.0%"
    for row_number in range(3, last_row + 1):
        for column in ("G", "J", "M", "N"):
            sheet[f"{column}{row_number}"].fill = percentage_fill
        sheet[f"P{row_number}"].fill = manager_fill
        sheet[f"S{row_number}"].fill = break_even_fill
        sheet[f"T{row_number}"].fill = forecast_fill
        sheet[f"O{row_number}"].font = Font(color="111827", bold=True)
        sheet[f"S{row_number}"].font = Font(color="8A4B16", bold=True)
        sheet[f"T{row_number}"].font = Font(color="27633B", bold=True)
    for row_number in range(1, last_row + 1):
        for column in ("E", "F", "H", "I", "K", "L", "O", "P", "Q", "R", "S", "T"):
            sheet[f"{column}{row_number}"].number_format = '#,##0;[Red]-#,##0;-'
    for column in ("G", "J", "M"):
        data_range = f"{column}3:{column}{last_row}"
        sheet.conditional_formatting.add(data_range, CellIsRule(operator="lessThan", formula=["0.9"], font=red_font))
        sheet.conditional_formatting.add(data_range, CellIsRule(operator="between", formula=["0.9", "0.999999999"], font=amber_font))
        sheet.conditional_formatting.add(data_range, CellIsRule(operator="greaterThanOrEqual", formula=["1"], font=green_font))
    sheet.conditional_formatting.add(f"O3:O{last_row}", FormulaRule(formula=["AND(ISNUMBER($S3),$O3<$S3)"], fill=red_fill, font=red_font))
    sheet.conditional_formatting.add(f"P3:P{last_row}", FormulaRule(formula=["AND(ISNUMBER($P3),ISNUMBER($S3),$P3<$S3)"], fill=red_fill, font=red_font))
    sheet.conditional_formatting.add(f"T3:T{last_row}", FormulaRule(formula=["AND(ISNUMBER($T3),ISNUMBER($S3),$T3<$S3)"], fill=red_fill, font=red_font))
    sheet.conditional_formatting.add(f"T3:T{last_row}", FormulaRule(formula=["AND(ISNUMBER($T3),ISNUMBER($S3),$T3>=$S3)"], fill=green_fill, font=green_font))
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 29
    sheet.column_dimensions["D"].width = 15
    for column in ("E", "F", "H", "I", "K", "L", "O", "P", "Q", "R", "S", "T"):
        sheet.column_dimensions[column].width = 15
    for column in ("G", "J", "M", "N"):
        sheet.column_dimensions[column].width = 11
    sheet.column_dimensions["Q"].width = 20
    sheet.column_dimensions["R"].width = 19
    sheet.column_dimensions["S"].width = 18

    section_fill = PatternFill("solid", fgColor="17365D")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    total_fill = PatternFill("solid", fgColor="1F2937")
    signal_balanced_fill = PatternFill("solid", fgColor="C6EFCE")
    signal_seasonal_fill = PatternFill("solid", fgColor="FFF2CC")
    signal_ai_fill = PatternFill("solid", fgColor="F4CCCC")
    for row_number in (1, second_title_row):
        for cell in comparison[row_number]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = section_fill
    for row_number in (2, second_header_row):
        for cell in comparison[row_number]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        comparison.row_dimensions[row_number].height = 46
    for row_number in (distribution_total_row, analysis_total_row):
        for cell in comparison[row_number]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = total_fill
    for row_number in range(3, distribution_total_row + 1):
        for column in ("C", "D", "F", "H"):
            comparison[f"{column}{row_number}"].number_format = "0.0%"
        for column in ("E", "G", "I"):
            comparison[f"{column}{row_number}"].number_format = '0.0" pp"'
            value = comparison[f"{column}{row_number}"].value
            if isinstance(value, (int, float)) and value < 0:
                comparison[f"{column}{row_number}"].font = Font(color="FF0000", bold=True)
    for row_number in range(second_header_row + 1, analysis_total_row + 1):
        for column in ("B", "C", "G", "I", "K"):
            comparison[f"{column}{row_number}"].number_format = '#,##0;[Red]-#,##0;-'
        for column in ("D", "E", "H", "J", "L"):
            comparison[f"{column}{row_number}"].number_format = '0.0"%"'
            value = comparison[f"{column}{row_number}"].value
            if isinstance(value, (int, float)):
                comparison[f"{column}{row_number}"].font = Font(color="00B050" if value >= 0 else "FF0000", bold=True)
        comparison[f"F{row_number}"].number_format = '0.0" pp"'
        signal = comparison[f"M{row_number}"].value
        if signal == "Echilibrat":
            comparison[f"M{row_number}"].fill = signal_balanced_fill
            comparison[f"M{row_number}"].font = Font(color="006100", bold=True)
        elif signal == "Peste sezonier":
            comparison[f"M{row_number}"].fill = signal_seasonal_fill
            comparison[f"M{row_number}"].font = Font(color="9C6500", bold=True)
        elif signal == "Peste AI":
            comparison[f"M{row_number}"].fill = signal_ai_fill
            comparison[f"M{row_number}"].font = Font(color="9C0006", bold=True)
    comparison.freeze_panes = "A3"
    comparison.column_dimensions["A"].width = 23
    comparison.column_dimensions["B"].width = 15
    for column in ("C", "D", "E", "F", "G", "H", "I", "J", "K", "L"):
        comparison.column_dimensions[column].width = 18
    comparison.column_dimensions["M"].width = 17
    for row_number in (distribution_total_row, analysis_total_row):
        for cell in comparison[row_number]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = total_fill
    for worksheet in (summary, parameters):
        for cell in worksheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in worksheet.columns:
            letter = get_column_letter(column[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 34)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    stamp = business_today().strftime("%Y%m%d")
    return output, f"targete_{scenario['target_month']}_scenariu_{scenario_id}_{stamp}.xlsx"
