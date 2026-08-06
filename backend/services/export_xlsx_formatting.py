"""Shared XLSX formatting primitives, isolated from export package imports."""
from typing import Any

from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from services.spreadsheet_safety import append_openpyxl_row


def number_format(column_type: str) -> str | None:
    return {"currency": "#,##0.00", "percent": "0.00", "integer": "0"}.get(column_type)


def safe_filename(value: str) -> str:
    safe = "".join(
        character if character.isalnum() or character in ("-", "_", ".") else "_"
        for character in value
    ).strip("._")
    if not safe:
        safe = "export_retail"
    if not safe.lower().endswith(".xlsx"):
        safe += ".xlsx"
    return safe[:140]


def days_filename_suffix(selected_days: list[int] | None) -> str:
    if not selected_days:
        return ""
    value = (
        "-".join(str(day) for day in selected_days)
        if len(selected_days) <= 10
        else f"{len(selected_days)}selectate"
    )
    return f"_zile_{value}"


def configure_day_axis(chart: LineChart) -> None:
    chart.x_axis.title = "Zi"
    chart.x_axis.delete = False
    chart.x_axis.axPos = "b"
    chart.x_axis.tickLblPos = "nextTo"
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1
    chart.x_axis.majorTickMark = "out"
    chart.x_axis.noMultiLvlLbl = True


def write_table_sheet(
    worksheet: Any,
    columns: list[dict[str, str]],
    rows: list[dict[str, Any]],
    *,
    header_fill: str,
) -> None:
    append_openpyxl_row(worksheet, [column["label"] for column in columns])
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="1f2937")
        cell.fill = PatternFill("solid", fgColor=header_fill)
    for row in rows:
        append_openpyxl_row(worksheet, [row.get(column["key"]) for column in columns])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = max(
            10,
            min(30, len(column["label"]) + 3),
        )
        cell_format = number_format(column["type"])
        if cell_format:
            for cell in worksheet[letter][1:]:
                cell.number_format = cell_format


def add_daily_comparison_chart(
    worksheet: Any,
    *,
    months: list[str],
    metric_label: str,
    max_row: int,
    first_data_col: int,
) -> None:
    if not months or max_row < 2:
        return
    chart = LineChart()
    chart.title = f"Comparatie zilnica - {metric_label}"
    chart.y_axis.title = metric_label
    configure_day_axis(chart)
    chart.visible_cells_only = True
    chart.style = 13
    data = Reference(
        worksheet,
        min_col=first_data_col,
        max_col=first_data_col + len(months) - 1,
        min_row=1,
        max_row=max_row,
    )
    categories = Reference(
        worksheet,
        min_col=first_data_col - 1,
        min_row=2,
        max_row=max_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 20
    worksheet.add_chart(
        chart,
        f"{get_column_letter(first_data_col + len(months) + 3)}2",
    )
