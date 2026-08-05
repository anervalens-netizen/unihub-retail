from __future__ import annotations

import hashlib
import json
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any, TypedDict

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from business_clock import business_today
from services.spreadsheet_safety import append_openpyxl_row

from repositories.target_calculator import (
    TargetCalculatorRepository,
    TargetScenarioAlgorithmMismatch,
    TargetScenarioFinalizedError,
    TargetScenarioVersionConflict,
)
from services.forecast import get_forecast_factor
from services.fiscal_rules import (
    STANDARD_VAT_RULESET_ID,
    net_to_gross,
    standard_vat_rule,
    standard_vat_ruleset_hash,
)
from services.target_rule_registry import (
    TargetRuleSet,
    TargetRuleSetValidationError,
    profitability_assumptions as rule_set_profitability_assumptions,
    store_salary_parameters,
    target_rule_set_from_snapshot,
    validate_store_exception_scope,
    validate_target_rule_set,
)

MONEY = Decimal("0.01")
DEFAULT_MIN_FLOOR = Decimal("35000")
DEFAULT_PREVIOUS_MONTH_FLOOR_PCT = Decimal("0.90")
DEFAULT_PREVIOUS_MONTH_CAP_PCT = Decimal("1.70")
DEFAULT_SEASONALITY_YEARS = 3
MAX_SEASONALITY_YEARS = 3
DEFAULT_TREND_WEIGHT = Decimal("0.10")
DEFAULT_TREND_ADJUSTMENT_MIN = Decimal("0.95")
DEFAULT_TREND_ADJUSTMENT_MAX = Decimal("1.10")
DEFAULT_SEASONALITY_MIN = Decimal("0.70")
DEFAULT_SEASONALITY_MAX = Decimal("1.70")
MIN_SEASONALITY_BASE = Decimal("10000")
CALCULATION_METHOD = "seasonal_blended_multiyear_v2_ruleset"
SALARY_PNL_FACTOR = Decimal("1.6955")
MEAL_VOUCHERS_PER_AGENT = Decimal("480")
SALES_COMMISSION_RATE = Decimal("0.03")
SALARY_ASSUMED_ATTAINMENT = Decimal("0.90")
DEFAULT_STORE_AGENT_COUNT = 2
SUN_PLAZA_AGENT_COUNT = 3
BASE_SALARY_DEFAULT = Decimal("2400")
BASE_SALARY_HIGH = Decimal("2600")
BASE_SALARY_HIGH_SITE_CODES = {
    "AFICOTRO",
    "AUCHMIL2",
    "AUCHMILI",
    "AUCHTRIC",
    "CCTCIT",
    "CJIULMALL",
    "CJPPOL",
    "CLUJCFPOL",
    "CORALEX",
    "COTROCENI",
    "CRFFEER",
    "CTAUCH",
    "CTCITYPRK",
    "CTCORA",
    "CTCRFTOM",
    "CTVIVO",
    "MC-MEGAMALL",
    "MCRFBAL",
    "MEGAMALL",
    "PRKLK",
    "PROM",
    "PROMEN",
    "SUNPLZ",
    "TMACUH",
    "TMSHOPCITY",
    "UNIRII",
}
PROFITABILITY_REQUIRED_CATEGORIES = {"v11", "c11", "c4", "c5", "c6"}
ROMANIAN_MONTH_NAMES = (
    "",
    "ianuarie",
    "februarie",
    "martie",
    "aprilie",
    "mai",
    "iunie",
    "iulie",
    "august",
    "septembrie",
    "octombrie",
    "noiembrie",
    "decembrie",
)

STRONG_SEASONALITY_WEIGHTS = {
    "store": Decimal("0.50"),
    "zone": Decimal("0.30"),
    "network": Decimal("0.20"),
}
WEAK_SEASONALITY_WEIGHTS = {
    "store": Decimal("0.30"),
    "zone": Decimal("0.40"),
    "network": Decimal("0.30"),
}
NEW_STORE_SEASONALITY_WEIGHTS = {
    "store": Decimal("0"),
    "zone": Decimal("0.60"),
    "network": Decimal("0.40"),
}


class SourceMetric(TypedDict):
    target: Decimal
    actual_realized: Decimal
    realized: Decimal
    forecast_factor: Decimal
    is_forecast: bool


class PeriodTotals(TypedDict):
    target: Decimal
    realized: Decimal


class SeasonalityPair(TypedDict):
    year_offset: int
    base_month: str
    target_month: str


def money(value: Decimal | int | str | float) -> Decimal:
    return Decimal(str(value)).quantize(MONEY, rounding=ROUND_HALF_UP)


def percent_change(new_value: float, base_value: float) -> float | None:
    if base_value <= 0:
        return None
    return round((new_value - base_value) * 100 / base_value, 2)


def realized_for_calculation(actual_realized: Decimal, forecast_factor: Decimal) -> Decimal:
    return money(actual_realized * forecast_factor)


def shift_month(month: str, offset: int) -> str:
    try:
        year, month_number = (int(value) for value in month.split("-"))
        if month_number < 1 or month_number > 12:
            raise ValueError
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Luna trebuie sa fie in format YYYY-MM") from exc
    index = year * 12 + month_number - 1 + offset
    return f"{index // 12:04d}-{index % 12 + 1:02d}"


def month_label_ro(month: str) -> str:
    try:
        month_number = int(month.split("-")[1])
        return ROMANIAN_MONTH_NAMES[month_number]
    except (IndexError, ValueError):
        return month


def source_month_configuration(target_month: str) -> list[dict[str, str]]:
    pairs = seasonality_pair_configuration(target_month, DEFAULT_SEASONALITY_YEARS)
    return build_source_month_configuration(target_month, pairs)


def seasonality_pair_configuration(target_month: str, years: int) -> list[SeasonalityPair]:
    years = max(1, min(int(years), MAX_SEASONALITY_YEARS))
    return [
        {
            "year_offset": year_offset,
            "base_month": shift_month(target_month, -1 - 12 * year_offset),
            "target_month": shift_month(target_month, -12 * year_offset),
        }
        for year_offset in range(1, years + 1)
    ]


def build_source_month_configuration(target_month: str, pairs: list[SeasonalityPair]) -> list[dict[str, str]]:
    source_months: list[dict[str, str]] = []
    for pair in sorted(pairs, key=lambda item: item["base_month"]):
        source_months.extend([
            {
                "month": pair["base_month"],
                "label": f"Baza sezoniera Y-{pair['year_offset']}",
                "role": f"seasonality_base_y{pair['year_offset']}",
            },
            {
                "month": pair["target_month"],
                "label": f"Luna target Y-{pair['year_offset']}",
                "role": f"seasonality_target_y{pair['year_offset']}",
            },
        ])
    source_months.append({
        "month": shift_month(target_month, -1),
        "label": "Forecast luna curenta",
        "role": "floor_reference",
    })
    return source_months


def unique_months(months: list[str]) -> list[str]:
    return list(dict.fromkeys(months))


def clamp_decimal(value: Decimal, minimum: Decimal, maximum: Decimal) -> Decimal:
    return min(max(value, minimum), maximum)


def seasonal_year_weights(count: int) -> list[Decimal]:
    if count <= 1:
        return [Decimal("1")]
    if count == 2:
        return [Decimal("0.70"), Decimal("0.30")]
    return [Decimal("0.50"), Decimal("0.30"), Decimal("0.20")]


def weighted_ratio(
    pairs: list[SeasonalityPair],
    value_by_month: dict[str, Decimal],
    *,
    minimum_base: Decimal = Decimal("0"),
) -> tuple[Decimal | None, list[dict[str, Any]]]:
    usable: list[tuple[SeasonalityPair, Decimal]] = []
    details: list[dict[str, Any]] = []
    for pair in pairs:
        base_value = money(value_by_month.get(pair["base_month"], Decimal("0")))
        target_value = money(value_by_month.get(pair["target_month"], Decimal("0")))
        ratio = target_value / base_value if base_value > minimum_base and target_value > 0 else None
        details.append({
            "year_offset": pair["year_offset"],
            "base_month": pair["base_month"],
            "target_month": pair["target_month"],
            "base_value": float(base_value),
            "target_value": float(target_value),
            "ratio": float(ratio.quantize(Decimal("0.0001"))) if ratio is not None else None,
        })
        if ratio is not None:
            usable.append((pair, ratio))

    if not usable:
        return None, details

    weights = seasonal_year_weights(len(usable))
    factor = sum((ratio * weights[index] for index, (_, ratio) in enumerate(usable)), Decimal("0"))
    return factor, details


def weighted_available(components: dict[str, tuple[Decimal | None, Decimal]]) -> Decimal | None:
    total_weight = sum(weight for value, weight in components.values() if value is not None and weight > 0)
    if total_weight <= 0:
        return None
    return sum(
        (
            value * weight / total_weight
            for value, weight in components.values()
            if value is not None and weight > 0
        ),
        Decimal("0"),
    )


class TargetBudgetInfeasibleError(ValueError):
    def __init__(self, requested_total: Decimal, floor_total: Decimal, cap_total: Decimal | None = None):
        self.requested_total = money(requested_total)
        self.floor_total = money(floor_total)
        self.cap_total = money(cap_total) if cap_total is not None else None
        if self.requested_total < self.floor_total:
            detail = "Bugetul este sub suma floor-urilor"
        else:
            detail = "Bugetul depaseste suma cap-urilor"
        super().__init__(detail)


def _mark_bound(row: dict[str, Any], *, floor: bool = False, cap: bool = False) -> None:
    if floor:
        row["is_floor_limited"] = True
        row["allocation_reason"] = "floor"
        if "FLOOR_APPLIED" not in row["flags"]:
            row["flags"].append("FLOOR_APPLIED")
    if cap:
        row["is_cap_limited"] = True
        row["allocation_reason"] = "cap"
        if "CAP_APPLIED" not in row["flags"]:
            row["flags"].append("CAP_APPLIED")


def _normalize_bounds(rows: list[dict[str, Any]], include_caps: bool) -> tuple[Decimal, Decimal | None]:
    floor_total = Decimal("0")
    cap_total = Decimal("0") if include_caps else None
    for row in rows:
        row["floor_target"] = money(row["floor_target"])
        row.setdefault("is_floor_limited", False)
        row.setdefault("is_cap_limited", False)
        row.setdefault("allocation_reason", "proportional")
        row.setdefault("flags", [])
        floor_total += row["floor_target"]
        if include_caps:
            row["cap_target"] = money(row["cap_target"])
            if row["cap_target"] < row["floor_target"]:
                raise ValueError("Cap-ul unei locatii nu poate fi sub floor.")
            cap_total = (cap_total or Decimal("0")) + row["cap_target"]
    return floor_total, cap_total


def _apply_rounding_difference(
    rows: list[dict[str, Any]],
    requested_total: Decimal,
    *,
    include_caps: bool,
) -> None:
    """Distribute the final cent residual deterministically over all available capacity."""
    rounded_total = sum((row["proposed_target"] for row in rows), Decimal("0"))
    difference = money(requested_total - rounded_total)
    if not difference:
        return

    increase = difference > 0
    remaining = abs(difference)
    while remaining > 0:
        candidates: list[tuple[Decimal, str, int, dict[str, Any]]] = []
        for index, row in enumerate(rows):
            capacity = (
                row["cap_target"] - row["proposed_target"]
                if increase and include_caps
                else row["proposed_target"] - row["floor_target"]
                if not increase
                else remaining
            )
            capacity = money(capacity)
            if capacity > 0:
                candidates.append((capacity, str(row.get("site_code", "")), index, row))
        if not candidates:
            floor_total, cap_total = _normalize_bounds(rows, include_caps)
            raise TargetBudgetInfeasibleError(requested_total, floor_total, cap_total)

        progressed = False
        for capacity, _site_code, _index, row in sorted(
            candidates,
            key=lambda candidate: (-candidate[0], candidate[1], candidate[2]),
        ):
            # Both operands are strictly positive by the loop and candidate guards.
            step = min(capacity, remaining)
            row["proposed_target"] = money(
                row["proposed_target"] + step if increase else row["proposed_target"] - step
            )
            remaining = money(remaining - step)
            progressed = True
            if include_caps and row["proposed_target"] == row["cap_target"]:
                _mark_bound(row, cap=True)
            if row["proposed_target"] == row["floor_target"]:
                _mark_bound(row, floor=True)
            if not remaining:
                return
        if not progressed:  # pragma: no cover - defensive guard for malformed bounds
            floor_total, cap_total = _normalize_bounds(rows, include_caps)
            raise TargetBudgetInfeasibleError(requested_total, floor_total, cap_total)


def allocate_with_floors(
    rows: list[dict[str, Any]],
    requested_total: Decimal,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Allocate an exactly feasible budget proportionally while honoring floors."""
    if not rows:
        return rows, []
    requested_total = money(requested_total)
    floor_total, _ = _normalize_bounds(rows, include_caps=False)
    if floor_total > requested_total:
        raise TargetBudgetInfeasibleError(requested_total, floor_total)

    remaining = set(range(len(rows)))
    assigned: dict[int, Decimal] = {}
    remaining_budget = requested_total
    while remaining:
        weight_total = sum((rows[index]["calculated_weight"] for index in remaining), Decimal("0"))
        allocations = {
            index: (
                remaining_budget * rows[index]["calculated_weight"] / weight_total
                if weight_total > 0
                else remaining_budget / Decimal(len(remaining))
            )
            for index in remaining
        }
        below_floor = {index for index in remaining if allocations[index] < rows[index]["floor_target"]}
        if not below_floor:
            assigned.update(allocations)
            break
        for index in below_floor:
            assigned[index] = rows[index]["floor_target"]
            remaining_budget -= rows[index]["floor_target"]
            _mark_bound(rows[index], floor=True)
        remaining -= below_floor

    for index, row in enumerate(rows):
        row["proposed_target"] = money(assigned[index])
    _apply_rounding_difference(rows, requested_total, include_caps=False)
    if sum((row["proposed_target"] for row in rows), Decimal("0")) != requested_total:
        raise TargetBudgetInfeasibleError(requested_total, floor_total)
    return rows, []


def allocate_with_bounds(
    rows: list[dict[str, Any]],
    requested_total: Decimal,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Allocate only a budget that can exactly satisfy every floor and cap."""
    if not rows:
        return rows, []
    requested_total = money(requested_total)
    floor_total, cap_total = _normalize_bounds(rows, include_caps=True)
    assert cap_total is not None
    if requested_total < floor_total or requested_total > cap_total:
        raise TargetBudgetInfeasibleError(requested_total, floor_total, cap_total)

    remaining = set(range(len(rows)))
    assigned: dict[int, Decimal] = {}
    remaining_budget = requested_total
    while remaining:
        weight_total = sum((rows[index]["calculated_weight"] for index in remaining), Decimal("0"))
        allocations = {
            index: (
                remaining_budget * rows[index]["calculated_weight"] / weight_total
                if weight_total > 0
                else remaining_budget / Decimal(len(remaining))
            )
            for index in remaining
        }
        fixed = False
        for index in sorted(remaining):
            row = rows[index]
            if allocations[index] < row["floor_target"]:
                assigned[index] = row["floor_target"]
                remaining_budget -= row["floor_target"]
                remaining.remove(index)
                _mark_bound(row, floor=True)
                fixed = True
            elif allocations[index] > row["cap_target"]:
                assigned[index] = row["cap_target"]
                remaining_budget -= row["cap_target"]
                remaining.remove(index)
                _mark_bound(row, cap=True)
                fixed = True
        if not fixed:
            assigned.update(allocations)
            break

    for index, row in enumerate(rows):
        row["proposed_target"] = money(assigned[index])
    _apply_rounding_difference(rows, requested_total, include_caps=True)
    for row in rows:
        if row["proposed_target"] == row["floor_target"]:
            _mark_bound(row, floor=True)
        if row["proposed_target"] == row["cap_target"]:
            _mark_bound(row, cap=True)
    final_total = sum((row["proposed_target"] for row in rows), Decimal("0"))
    if final_total != requested_total or any(
        row["proposed_target"] < row["floor_target"] or row["proposed_target"] > row["cap_target"]
        for row in rows
    ):
        raise TargetBudgetInfeasibleError(requested_total, floor_total, cap_total)
    return rows, []

class TargetCalculatorService:
    def __init__(self, repo: TargetCalculatorRepository):
        self.repo = repo

    async def get_context(self) -> dict[str, Any]:
        latest_month = await self.repo.get_latest_sales_month()
        if not latest_month:
            raise HTTPException(status_code=404, detail="Nu exista date de vanzari pentru calculator.")
        suggested_month = shift_month(latest_month, 1)
        target_total = await self.repo.get_target_total(suggested_month)
        if target_total == 0:
            target_total = await self.repo.get_target_total(latest_month)
        cohort = await self.repo.get_active_cohort(latest_month, suggested_month)
        return {
            "latest_sales_month": latest_month,
            "suggested_target_month": suggested_month,
            "suggested_cohort_month": latest_month,
            "suggested_total_target": float(target_total),
            "default_min_floor": float(DEFAULT_MIN_FLOOR),
            "default_previous_month_floor_pct": float(DEFAULT_PREVIOUS_MONTH_FLOOR_PCT),
            "default_previous_month_cap_pct": float(DEFAULT_PREVIOUS_MONTH_CAP_PCT),
            "default_seasonality_years": DEFAULT_SEASONALITY_YEARS,
            "active_store_count": len(cohort),
            "regionals": sorted({row["regional"] for row in cohort}),
        }

    async def calculate(self, payload: dict[str, Any]) -> dict[str, Any]:
        target_month = payload["target_month"]
        seasonality_years = int(payload.get("seasonality_years") or DEFAULT_SEASONALITY_YEARS)
        seasonality_years = max(1, min(seasonality_years, MAX_SEASONALITY_YEARS))
        source_pairs = seasonality_pair_configuration(target_month, seasonality_years)
        source_months = build_source_month_configuration(target_month, source_pairs)
        latest_before_target = await self.repo.get_latest_sales_month(before_month=target_month)
        cohort_month = payload.get("cohort_month") or latest_before_target
        if not cohort_month:
            raise HTTPException(
                status_code=400,
                detail="Nu exista o luna cu vanzari anterioara lunii pentru care se calculeaza targetul.",
            )
        if cohort_month >= target_month:
            raise HTTPException(
                status_code=400,
                detail="Cohorta activa trebuie sa provina dintr-o luna anterioara lunii de target.",
            )

        cohort = await self.repo.get_active_cohort(cohort_month, target_month)
        if not cohort:
            raise HTTPException(status_code=400, detail="Luna de cohorta nu are magazine active.")

        total_target = money(payload["total_target"])
        min_floor = money(payload.get("min_floor", DEFAULT_MIN_FLOOR))
        floor_pct = Decimal(str(payload.get("previous_month_floor_pct", DEFAULT_PREVIOUS_MONTH_FLOOR_PCT)))
        cap_pct = Decimal(str(payload.get("previous_month_cap_pct", DEFAULT_PREVIOUS_MONTH_CAP_PCT)))
        trend_weight = Decimal(str(payload.get("trend_weight", DEFAULT_TREND_WEIGHT)))
        seasonality_min = Decimal(str(payload.get("seasonality_min", DEFAULT_SEASONALITY_MIN)))
        seasonality_max = Decimal(str(payload.get("seasonality_max", DEFAULT_SEASONALITY_MAX)))
        trend_min = Decimal(str(payload.get("trend_adjustment_min", DEFAULT_TREND_ADJUSTMENT_MIN)))
        trend_max = Decimal(str(payload.get("trend_adjustment_max", DEFAULT_TREND_ADJUSTMENT_MAX)))
        if (
            total_target <= 0
            or min_floor < 0
            or floor_pct < 0
            or cap_pct <= 0
            or cap_pct < floor_pct
            or trend_weight < 0
            or seasonality_min <= 0
            or seasonality_max < seasonality_min
            or trend_min <= 0
            or trend_max < trend_min
        ):
            raise HTTPException(status_code=400, detail="Parametrii de calcul nu sunt valizi.")

        rule_record = await self.repo.get_effective_target_rule_set(target_month)
        if not rule_record:
            raise HTTPException(
                status_code=409,
                detail="Nu exista un rule-set Target efectiv pentru luna ceruta; nu s-a creat nicio propunere.",
            )
        try:
            target_rule_set = validate_target_rule_set(dict(rule_record), target_month)
        except TargetRuleSetValidationError as exc:
            raise HTTPException(
                status_code=409,
                detail=f"Rule-set-ul Target este invalid; nu s-a creat nicio propunere. {exc}",
            ) from None
        exception_codes = sorted(target_rule_set.rules["store_exceptions"])
        if exception_codes:
            master_rows = await self.repo.get_target_rule_exception_master(exception_codes)
            try:
                validate_store_exception_scope(
                    target_rule_set,
                    cohort=[dict(row) for row in cohort],
                    master_rows=[dict(row) for row in master_rows],
                )
            except TargetRuleSetValidationError as exc:
                raise HTTPException(
                    status_code=409,
                    detail=f"Rule-set-ul Target nu se reconciliaza cu master/cohort; nu s-a creat nicio propunere. {exc}",
                ) from None

        months = unique_months([item["month"] for item in source_months])
        site_codes = [row["site_code"] for row in cohort]
        metrics = await self.repo.get_source_metrics(site_codes, months)
        async with self.repo.pool.acquire() as conn:
            forecast_factors = {
                month: Decimal(str(await get_forecast_factor(conn, month)))
                for month in months
            }
        calculation_input_sha256 = self._canonical_input_hash({
            "target_month": target_month,
            "cohort_month": cohort_month,
            "source_months": source_months,
            "cohort": [dict(row) for row in cohort],
            "source_metrics": [dict(row) for row in metrics],
            "forecast_factors": {month: str(forecast_factors[month]) for month in sorted(forecast_factors)},
        })
        metric_map: dict[tuple[str, str], SourceMetric] = {
            (row["site_code"], row["import_month"]): {
                "target": Decimal(row["target"] or 0),
                "actual_realized": money(Decimal(row["realized"] or 0)),
                "realized": realized_for_calculation(
                    Decimal(row["realized"] or 0),
                    forecast_factors[row["import_month"]],
                ),
                "forecast_factor": forecast_factors[row["import_month"]],
                "is_forecast": forecast_factors[row["import_month"]] > Decimal("1"),
            }
            for row in metrics
        }
        for site_code in site_codes:
            for month in months:
                metric_map.setdefault((site_code, month), {
                    "target": Decimal("0"),
                    "actual_realized": Decimal("0"),
                    "realized": Decimal("0"),
                    "forecast_factor": forecast_factors[month],
                    "is_forecast": forecast_factors[month] > Decimal("1"),
                })
        totals: dict[str, PeriodTotals] = {
            month: {
                "target": sum((metric_map[(site_code, month)]["target"] for site_code in site_codes), Decimal("0")),
                "realized": sum((metric_map[(site_code, month)]["realized"] for site_code in site_codes), Decimal("0")),
            }
            for month in months
        }
        regionals = sorted({row["regional"] for row in cohort})
        site_regional = {row["site_code"]: row["regional"] for row in cohort}
        regional_month_values: dict[tuple[str, str], Decimal] = {
            (regional, month): sum(
                (
                    metric_map[(site_code, month)]["realized"]
                    for site_code in site_codes
                    if site_regional[site_code] == regional
                ),
                Decimal("0"),
            )
            for regional in regionals
            for month in months
        }

        warnings: list[str] = []
        for item in source_months:
            total = totals[item["month"]]
            if total["target"] == 0 and total["realized"] == 0:
                warnings.append(f"Nu exista date pentru perioada de referinta {item['month']}.")
            if forecast_factors[item["month"]] > Decimal("1"):
                warnings.append(
                    f"Perioada {item['month']} este partiala; vanzarile folosite in calcul sunt forecastate "
                    f"cu factor {forecast_factors[item['month']]:.4f}x pe baza importului disponibil."
                )
        if seasonality_years > 1:
            warnings.append(
                f"Formula foloseste sezonalitate multi-year pe pana la {seasonality_years} ani; anii fara date suficiente sunt sariti automat."
            )

        calculated_rows: list[dict[str, Any]] = []
        current_month = shift_month(target_month, -1)
        network_values = {month: totals[month]["realized"] for month in months}
        network_factor, network_years = weighted_ratio(source_pairs, network_values)
        if network_factor is None:
            network_factor = Decimal("1")
        for cohort_row in cohort:
            site_code = cohort_row["site_code"]
            regional = cohort_row["regional"]
            history: list[dict[str, Any]] = []
            for item in source_months:
                metric = metric_map[(site_code, item["month"])]
                total = totals[item["month"]]
                shares: list[Decimal] = []
                if total["target"] > 0:
                    shares.append(metric["target"] / total["target"])
                if total["realized"] > 0:
                    shares.append(metric["realized"] / total["realized"])
                period_weight = sum(shares, Decimal("0")) / Decimal(len(shares)) if shares else Decimal("0")
                attainment = (
                    metric["realized"] / metric["target"] * Decimal("100")
                    if metric["target"] > 0 else None
                )
                history.append({
                    "month": item["month"],
                    "label": item["label"],
                    "role": item["role"],
                    "target": float(money(metric["target"])),
                    "realized": float(money(metric["realized"])),
                    "actual_realized": float(money(metric["actual_realized"])),
                    "is_forecast": metric["is_forecast"],
                    "forecast_factor": float(metric["forecast_factor"]),
                    "attainment_pct": float(attainment.quantize(MONEY)) if attainment is not None else None,
                    "weight": float(period_weight),
                })

            store_values = {month: metric_map[(site_code, month)]["realized"] for month in months}
            regional_values = {month: regional_month_values[(regional, month)] for month in months}
            store_factor, store_years = weighted_ratio(
                source_pairs,
                store_values,
                minimum_base=MIN_SEASONALITY_BASE,
            )
            zone_factor, zone_years = weighted_ratio(source_pairs, regional_values)
            last_year_store_factor, _ = weighted_ratio(source_pairs[:1], store_values, minimum_base=MIN_SEASONALITY_BASE)
            multiyear_store_factor, _ = weighted_ratio(source_pairs, store_values, minimum_base=MIN_SEASONALITY_BASE)

            weights = STRONG_SEASONALITY_WEIGHTS
            flags: list[str] = []
            usable_store_years = sum(1 for item in store_years if item["ratio"] is not None)
            store_ratios = [
                Decimal(str(item["ratio"]))
                for item in store_years
                if item["ratio"] is not None
            ]
            if store_factor is None:
                weights = NEW_STORE_SEASONALITY_WEIGHTS
                flags.extend(["NEW_STORE", "LOW_HISTORY"])
            elif usable_store_years <= 1 and seasonality_years > 1:
                weights = WEAK_SEASONALITY_WEIGHTS
                flags.append("LOW_HISTORY")
            elif (
                seasonality_years > 1
                and any(item["year_offset"] == 1 and item["ratio"] is None for item in store_years)
            ):
                weights = WEAK_SEASONALITY_WEIGHTS
                flags.append("LOW_RECENT_HISTORY")
            if store_ratios and (min(store_ratios) < Decimal("0.50") or max(store_ratios) > Decimal("2.00")):
                weights = WEAK_SEASONALITY_WEIGHTS
                flags.append("EXTREME_SEASONALITY")

            blended = weighted_available({
                "store": (store_factor, weights["store"]),
                "zone": (zone_factor, weights["zone"]),
                "network": (network_factor, weights["network"]),
            })
            raw_blended = blended if blended is not None else Decimal("1")
            seasonality_factor = clamp_decimal(raw_blended, seasonality_min, seasonality_max)
            if seasonality_factor != raw_blended:
                flags.append("SEASONALITY_CAPPED")

            current_forecast = metric_map[(site_code, current_month)]["realized"]
            trend_base = metric_map[(site_code, source_pairs[0]["base_month"])]["realized"]
            if trend_base > MIN_SEASONALITY_BASE:
                trend_ratio = current_forecast / trend_base
                trend_adjustment_raw = Decimal("1") + ((trend_ratio - Decimal("1")) * trend_weight)
            else:
                trend_ratio = None
                trend_adjustment_raw = Decimal("1")
            trend_adjustment = clamp_decimal(trend_adjustment_raw, trend_min, trend_max)
            if trend_adjustment != trend_adjustment_raw:
                flags.append("TREND_ADJUSTMENT_CAPPED")

            raw_estimate = money(current_forecast * seasonality_factor * trend_adjustment)
            floor_target = max(min_floor, money(current_forecast * floor_pct))
            cap_target = max(floor_target, money(current_forecast * cap_pct))
            calculation_details = {
                "method": CALCULATION_METHOD,
                "seasonality_years": seasonality_years,
                "current_month": current_month,
                "current_forecast": float(money(current_forecast)),
                "seasonality": {
                    "store_factor": float(store_factor.quantize(Decimal("0.0001"))) if store_factor is not None else None,
                    "zone_factor": float(zone_factor.quantize(Decimal("0.0001"))) if zone_factor is not None else None,
                    "network_factor": float(network_factor.quantize(Decimal("0.0001"))),
                    "blended_factor": float(raw_blended.quantize(Decimal("0.0001"))),
                    "used_factor": float(seasonality_factor.quantize(Decimal("0.0001"))),
                    "last_year_store_factor": (
                        float(last_year_store_factor.quantize(Decimal("0.0001")))
                        if last_year_store_factor is not None else None
                    ),
                    "multiyear_store_factor": (
                        float(multiyear_store_factor.quantize(Decimal("0.0001")))
                        if multiyear_store_factor is not None else None
                    ),
                    "weights": {key: float(value) for key, value in weights.items()},
                    "store_years": store_years,
                    "zone_years": zone_years,
                    "network_years": network_years,
                    "min": float(seasonality_min),
                    "max": float(seasonality_max),
                },
                "trend": {
                    "base_month": source_pairs[0]["base_month"],
                    "ratio": float(trend_ratio.quantize(Decimal("0.0001"))) if trend_ratio is not None else None,
                    "weight": float(trend_weight),
                    "raw_adjustment": float(trend_adjustment_raw.quantize(Decimal("0.0001"))),
                    "used_adjustment": float(trend_adjustment.quantize(Decimal("0.0001"))),
                    "min": float(trend_min),
                    "max": float(trend_max),
                },
                "raw_estimate": float(raw_estimate),
                "floor_target": float(floor_target),
                "cap_target": float(cap_target),
                "flags": [],
                "allocation_reason": "proportional",
            }
            calculated_rows.append({
                "site_code": site_code,
                "locatie": cohort_row["locatie"],
                "firma": cohort_row["firma"],
                "regional": regional,
                "asm": cohort_row["asm"],
                "calculated_weight": raw_estimate,
                "floor_target": floor_target,
                "cap_target": cap_target,
                "proposed_target": Decimal("0"),
                "is_floor_limited": False,
                "is_cap_limited": False,
                "allocation_reason": "proportional",
                "flags": flags,
                "history": history,
                "calculation_details": calculation_details,
            })

        if sum((row["calculated_weight"] for row in calculated_rows), Decimal("0")) == 0:
            equal_weight = Decimal("1") / Decimal(len(calculated_rows))
            for row in calculated_rows:
                row["calculated_weight"] = equal_weight
                row["flags"].append("LOW_HISTORY")
            warnings.append("Datele istorice nu contin estimari sezoniere utilizabile; targetul a fost distribuit uniform.")
        else:
            raw_total = sum((row["calculated_weight"] for row in calculated_rows), Decimal("0"))
            for row in calculated_rows:
                row["calculated_weight"] = row["calculated_weight"] / raw_total

        floor_total = sum((row["floor_target"] for row in calculated_rows), Decimal("0"))
        cap_total = sum((row["cap_target"] for row in calculated_rows), Decimal("0"))
        if total_target < floor_total:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Targetul total {total_target:,.0f} RON este sub suma floor-urilor calculate "
                    f"{floor_total:,.0f} RON. Ajusteaza bugetul sau floor-ul operational; propunerea nu a fost salvata."
                ).replace(",", "."),
            )
        if cap_total < total_target:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Targetul total {total_target:,.0f} RON depaseste cap-ul maxim calculat "
                    f"{cap_total:,.0f} RON. Verifica valoarea bugetului sau mareste cap-ul operational."
                ).replace(",", "."),
            )

        try:
            calculated_rows, allocation_warnings = allocate_with_bounds(calculated_rows, total_target)
        except TargetBudgetInfeasibleError as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Bugetul Target este infezabil; propunerea nu a fost salvata. {exc}",
            ) from None
        warnings.extend(allocation_warnings)
        for row in calculated_rows:
            flags = list(dict.fromkeys(row["flags"]))
            row["calculation_details"]["flags"] = flags
            row["calculation_details"]["allocation_reason"] = row["allocation_reason"]
            row["calculation_details"]["is_floor_limited"] = row["is_floor_limited"]
            row["calculation_details"]["is_cap_limited"] = row["is_cap_limited"]

        profitability_inputs = await self.repo.get_profitability_inputs(
            site_codes=site_codes,
            target_month=target_month,
        )
        forecast_coverage, _forecast_values = self._forecast_coverage(
            calculated_rows,
            profitability_inputs,
        )
        profitability_inputs["forecast_coverage"] = forecast_coverage
        if forecast_coverage["mode"] != "uniform":
            raise self._forecast_coverage_error(forecast_coverage)
        profitability_input_sha256 = self._canonical_input_hash(
            self._profitability_input_payload(profitability_inputs)
        )
        profitability_summary = self._populate_profitability(
            {
                "target_month": target_month,
                "rule_set_snapshot": target_rule_set.snapshot(),
                "calculation_params": {
                    "profitability": rule_set_profitability_assumptions(target_rule_set),
                },
            },
            calculated_rows,
            profitability_inputs,
        )
        profitability_summary["input_sha256"] = profitability_input_sha256
        for row in calculated_rows:
            row["profitability_snapshot"] = row.pop("profitability")
        try:
            scenario_id = await self.repo.save_draft_scenario(
                {
                    "target_month": target_month,
                    "cohort_month": cohort_month,
                    "total_target": total_target,
                    "min_floor": min_floor,
                    "previous_month_floor_pct": floor_pct,
                    "calculation_method": CALCULATION_METHOD,
                    "source_months": source_months,
                    "warnings": warnings,
                    "rule_set_id": target_rule_set.rule_set_id,
                    "rule_set_hash": target_rule_set.rules_hash,
                    "rule_set_snapshot": target_rule_set.snapshot(),
                    "calculation_input_sha256": calculation_input_sha256,
                    "profitability_input_sha256": profitability_input_sha256,
                    "calculation_params": {
                        "seasonality_years": seasonality_years,
                        "seasonality_min": float(seasonality_min),
                        "seasonality_max": float(seasonality_max),
                        "trend_weight": float(trend_weight),
                        "trend_adjustment_min": float(trend_min),
                        "trend_adjustment_max": float(trend_max),
                        "previous_month_cap_pct": float(cap_pct),
                        "minimum_seasonality_base": float(MIN_SEASONALITY_BASE),
                        "strong_weights": {key: float(value) for key, value in STRONG_SEASONALITY_WEIGHTS.items()},
                        "weak_weights": {key: float(value) for key, value in WEAK_SEASONALITY_WEIGHTS.items()},
                        "new_store_weights": {key: float(value) for key, value in NEW_STORE_SEASONALITY_WEIGHTS.items()},
                        "profitability": rule_set_profitability_assumptions(target_rule_set),
                        "profitability_summary": profitability_summary,
                    },
                },
                calculated_rows,
                payload.get("expected_revision"),
            )
        except TargetScenarioFinalizedError:
            raise HTTPException(
                status_code=409,
                detail="Targetul acestei luni a fost deja finalizat si nu mai poate fi recalculat.",
            ) from None
        except TargetScenarioAlgorithmMismatch:
            raise HTTPException(
                status_code=409,
                detail="Scenariul draft apartine altei versiuni de algoritm; nu poate fi rescris.",
            ) from None
        except TargetScenarioVersionConflict:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Scenariul a fost modificat de alt utilizator. "
                    "Reincarca datele inainte de recalculare."
                ),
            ) from None
        return await self.get_scenario_detail(scenario_id)

    async def list_scenarios(self) -> list[dict[str, Any]]:
        rows = await self.repo.list_scenarios()
        serialized = [self._serialize_header(dict(row)) for row in rows]
        for row in serialized:
            if row.get("rule_set_snapshot") is None:
                for key in (
                    "rule_set_id",
                    "rule_set_hash",
                    "rule_set_snapshot",
                    "calculation_input_sha256",
                    "profitability_input_sha256",
                ):
                    row.pop(key, None)
        return serialized

    async def get_scenario_detail(self, scenario_id: int) -> dict[str, Any]:
        scenario = await self.repo.get_scenario(scenario_id)
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenariul de target nu exista.")
        rows = await self.repo.get_scenario_rows(scenario_id)
        header = self._serialize_header(dict(scenario))
        serialized_rows = [self._serialize_row(dict(row)) for row in rows]
        legacy_unversioned = header.get("rule_set_snapshot") is None
        profitability_summary = await self._attach_profitability(header, serialized_rows)
        if legacy_unversioned:
            for key in (
                "rule_set_id",
                "rule_set_hash",
                "rule_set_snapshot",
                "calculation_input_sha256",
                "profitability_input_sha256",
            ):
                header.pop(key, None)
            for row in serialized_rows:
                for key in (
                    "cap_target",
                    "is_cap_limited",
                    "manager_override_target",
                    "manager_override_reason",
                    "manager_override_actor",
                    "manager_override_at",
                    "manager_override_revision",
                    "profitability_snapshot",
                ):
                    row.pop(key, None)
        for row in serialized_rows:
            row.pop("manager_override_actor", None)
        proposed_total = sum(row["proposed_target"] for row in serialized_rows)
        final_total = sum((row["final_target"] or 0) for row in serialized_rows)
        pending_final_count = sum(1 for row in serialized_rows if row["final_target"] is None)
        detail = {
            **header,
            "store_count": len(serialized_rows),
            "proposed_total": proposed_total,
            "final_total": final_total,
            "remaining_difference": header["total_target"] - final_total,
            "pending_final_count": pending_final_count,
            "floor_limited_count": sum(1 for row in serialized_rows if row["is_floor_limited"]),
            "manual_adjustments_count": sum(
                1 for row in serialized_rows
                if row["final_target"] is not None and abs(row["final_target"] - row["proposed_target"]) > 0.01
            ),
            "rows": serialized_rows,
            "regional_summary": self._regional_summary(serialized_rows),
            "source_summary": self._source_summary(serialized_rows),
            "profitability_summary": profitability_summary,
        }
        if not legacy_unversioned:
            detail["cap_limited_count"] = sum(1 for row in serialized_rows if row.get("is_cap_limited"))
            detail["manager_overrides_count"] = sum(
                1 for row in serialized_rows if row.get("manager_override_target") is not None
            )
        return detail

    @staticmethod
    def _profitability_assumptions(target_month: str) -> dict[str, Any]:
        vat_rule = standard_vat_rule(target_month)
        return {
            "vat_ruleset_id": STANDARD_VAT_RULESET_ID,
            "vat_ruleset_hash": standard_vat_ruleset_hash(),
            "vat_rule_id": vat_rule.rule_id,
            "vat_effective_from": vat_rule.effective_from.isoformat(),
            "vat_multiplier": float(vat_rule.multiplier),
            "vat_rate": float(vat_rule.rate),
            "salary_pnl_factor": float(SALARY_PNL_FACTOR),
            "meal_vouchers_per_agent": float(MEAL_VOUCHERS_PER_AGENT),
            "sales_commission_rate": float(SALES_COMMISSION_RATE),
            "salary_assumed_attainment": float(SALARY_ASSUMED_ATTAINMENT),
            "default_store_agent_count": DEFAULT_STORE_AGENT_COUNT,
            "sun_plaza_agent_count": SUN_PLAZA_AGENT_COUNT,
            "base_salary_default": float(BASE_SALARY_DEFAULT),
            "base_salary_high": float(BASE_SALARY_HIGH),
        }

    @staticmethod
    def _legacy_profitability_assumptions() -> dict[str, Any]:
        assumptions = TargetCalculatorService._profitability_assumptions("2025-08")
        assumptions.update({
            "vat_ruleset_id": "legacy-unversioned",
            "vat_ruleset_hash": None,
            "vat_rule_id": "legacy-unversioned",
            "vat_effective_from": None,
        })
        return assumptions

    @staticmethod
    def _saved_profitability_assumptions(scenario: dict[str, Any]) -> dict[str, Any]:
        """Normalize historical snapshots without persisting invented metadata."""
        legacy = TargetCalculatorService._legacy_profitability_assumptions()
        raw = (scenario.get("calculation_params") or {}).get("profitability")
        if not isinstance(raw, dict):
            return legacy
        assumptions = {**legacy, **raw}
        if "vat_multiplier" not in raw:
            saved_rate = raw.get("vat_rate")
            if saved_rate is not None:
                assumptions["vat_multiplier"] = float(
                    Decimal("1") + Decimal(str(saved_rate))
                )
            else:
                assumptions["vat_multiplier"] = legacy["vat_multiplier"]
        if not raw.get("vat_rule_id"):
            assumptions["vat_rule_id"] = "legacy-unversioned"
        if not raw.get("vat_ruleset_id"):
            assumptions["vat_ruleset_id"] = "legacy-unversioned"
        return assumptions

    @staticmethod
    def _saved_target_rule_set(scenario: dict[str, Any]) -> TargetRuleSet | None:
        snapshot = scenario.get("rule_set_snapshot")
        if snapshot is None:
            return None
        if isinstance(snapshot, str):
            try:
                snapshot = json.loads(snapshot)
            except json.JSONDecodeError as exc:
                raise HTTPException(status_code=409, detail="Snapshotul Target nu este JSON valid.") from exc
        try:
            rule_set = target_rule_set_from_snapshot(snapshot, scenario["target_month"])
        except TargetRuleSetValidationError as exc:
            raise HTTPException(
                status_code=409,
                detail=f"Snapshotul rule-set-ului Target este invalid: {exc}",
            ) from None
        if rule_set is None:
            raise HTTPException(status_code=409, detail="Snapshotul rule-set-ului Target este incomplet.")
        return rule_set

    @staticmethod
    def _canonical_input_hash(payload: Any) -> str:
        canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
        return hashlib.sha256(canonical.encode("utf-8")).hexdigest()

    @staticmethod
    def _profitability_input_payload(inputs: dict[str, Any]) -> dict[str, Any]:
        forecast_run = inputs.get("forecast_run")
        return {
            "pnl_months": list(inputs.get("pnl_months") or []),
            "pnl_rows": [dict(record) for record in inputs.get("pnl_rows") or []],
            "forecast_run": dict(forecast_run) if forecast_run else None,
            "forecast_rows": [dict(record) for record in inputs.get("forecast_rows") or []],
            "forecast_coverage": inputs.get("forecast_coverage"),
        }

    @staticmethod
    def _forecast_coverage(
        rows: list[dict[str, Any]],
        inputs: dict[str, Any],
    ) -> tuple[dict[str, Any], dict[str, Decimal]]:
        """Derive the v2 forecast contract from explicit per-store source presence."""
        expected_site_codes = sorted({str(row["site_code"]) for row in rows})
        records_by_site = {
            str(record["site_code"]): record
            for record in inputs.get("forecast_rows") or []
        }
        forecast_values: dict[str, Decimal] = {}
        covered_site_codes: list[str] = []
        missing_site_codes: list[str] = []
        cutoff_values: list[str] = []
        for site_code in expected_site_codes:
            record = records_by_site.get(site_code)
            forecast_present = bool(record and record.get("forecast_present"))
            realized_present = bool(record and record.get("realized_present"))
            forecast_sales = record.get("forecast_sales") if record else None
            cutoff_date = record.get("cutoff_date") if record else None
            if not forecast_present or not realized_present or forecast_sales is None or cutoff_date is None:
                missing_site_codes.append(site_code)
                continue
            forecast_values[site_code] = money(Decimal(forecast_sales))
            covered_site_codes.append(site_code)
            cutoff_values.append(str(cutoff_date))

        distinct_cutoffs = sorted(set(cutoff_values))
        uniform = (
            inputs.get("forecast_run") is not None
            and len(covered_site_codes) == len(expected_site_codes)
            and len(distinct_cutoffs) == 1
        )
        coverage = {
            "mode": "uniform" if uniform else "nonuniform",
            "cutoff": distinct_cutoffs[0] if uniform else None,
            "cutoff_min": min(cutoff_values) if cutoff_values else None,
            "cutoff_max": max(cutoff_values) if cutoff_values else None,
            "expected_store_count": len(expected_site_codes),
            "covered_store_count": len(covered_site_codes),
            "missing_site_codes": missing_site_codes,
        }
        return coverage, forecast_values

    @staticmethod
    def _forecast_coverage_error(coverage: dict[str, Any]) -> HTTPException:
        return HTTPException(
            status_code=409,
            detail={
                "message": "Forecastul curent nu are coverage uniform complet; propunerea nu a fost salvată.",
                "forecast_coverage": coverage,
            },
        )

    def _populate_profitability(
        self,
        scenario: dict[str, Any],
        rows: list[dict[str, Any]],
        inputs: dict[str, Any],
    ) -> dict[str, Any]:
        weight_total = sum((Decimal(str(row["calculated_weight"])) for row in rows), Decimal("0"))
        for row in rows:
            weight = Decimal(str(row["calculated_weight"]))
            row["normalized_weight"] = float(weight / weight_total) if weight_total > 0 else 0.0

        pnl_months = list(inputs.get("pnl_months") or [])
        pnl_values: dict[tuple[str, str], Decimal] = {
            (record["site_code"], record["category_code"]): Decimal(record["amount"] or 0)
            for record in inputs.get("pnl_rows") or []
        }
        forecast_coverage_contract, forecast_values = self._forecast_coverage(rows, inputs)
        forecast_run_record = inputs.get("forecast_run")
        forecast_run = dict(forecast_run_record) if forecast_run_record else None
        saved_target_rule_set = self._saved_target_rule_set(scenario)
        saved_profitability = (
            rule_set_profitability_assumptions(saved_target_rule_set)
            if saved_target_rule_set is not None
            else self._saved_profitability_assumptions(scenario)
        )
        vat_multiplier = Decimal(str(saved_profitability["vat_multiplier"]))
        if saved_target_rule_set is not None:
            salary_rules = saved_target_rule_set.rules["salary"]
            salary_pnl_factor = Decimal(str(salary_rules["pnl_factor"]))
            meal_vouchers = Decimal(str(salary_rules["meal_vouchers_per_agent"]))
            commission_rate = Decimal(str(salary_rules["sales_commission_rate"]))
            assumed_attainment = Decimal(str(salary_rules["assumed_attainment"]))
        else:
            salary_pnl_factor = SALARY_PNL_FACTOR
            meal_vouchers = MEAL_VOUCHERS_PER_AGENT
            commission_rate = SALES_COMMISSION_RATE
            assumed_attainment = SALARY_ASSUMED_ATTAINMENT

        salary_total = Decimal("0")
        opex_total = Decimal("0")
        break_even_total = Decimal("0")
        forecast_total = Decimal("0")
        forecast_below_break_even_count = 0
        target_below_break_even_count = 0
        complete_pnl_count = 0
        for row in rows:
            site_code = row["site_code"]
            if saved_target_rule_set is not None:
                agents, base_salary = store_salary_parameters(saved_target_rule_set, site_code)
            else:
                agents = SUN_PLAZA_AGENT_COUNT if site_code == "SUNPLZ" else DEFAULT_STORE_AGENT_COUNT
                base_salary = BASE_SALARY_HIGH if site_code in BASE_SALARY_HIGH_SITE_CODES else BASE_SALARY_DEFAULT
            calculated_target = money(row["proposed_target"])
            salary_source = (
                Decimal(agents) * (base_salary + meal_vouchers)
                + calculated_target * assumed_attainment * commission_rate
            )
            salary_cost = money(salary_source * salary_pnl_factor)
            salary_total += salary_cost

            categories = {
                category: pnl_values.get((site_code, category))
                for category in PROFITABILITY_REQUIRED_CATEGORIES
            }
            pnl_complete = len(pnl_months) == 3 and all(value is not None for value in categories.values())
            accessory_margin: Decimal | None = None
            opex: Decimal | None = None
            break_even: Decimal | None = None
            if pnl_complete:
                accessory_revenue = categories["v11"] or Decimal("0")
                accessory_cogs = categories["c11"] or Decimal("0")
                if accessory_revenue > 0:
                    accessory_margin = (accessory_revenue - accessory_cogs) / accessory_revenue
                if accessory_margin is not None and accessory_margin > 0:
                    opex = money(
                        (
                            (categories["c4"] or Decimal("0"))
                            + (categories["c5"] or Decimal("0"))
                            + (categories["c6"] or Decimal("0"))
                        )
                        / Decimal(len(pnl_months))
                    )
                    net_break_even = (salary_cost + opex) / accessory_margin
                    break_even = net_to_gross(net_break_even, scenario["target_month"])
                    if vat_multiplier != standard_vat_rule(scenario["target_month"]).multiplier:
                        break_even = money(net_break_even * vat_multiplier)
                    complete_pnl_count += 1
                    opex_total += opex
                    break_even_total += break_even

            forecast = forecast_values.get(site_code)
            if forecast is not None:
                forecast_total += forecast

            anomaly_flags: list[str] = []
            if not pnl_complete or break_even is None:
                anomaly_flags.append("PNL_INCOMPLETE")
            if forecast is None:
                anomaly_flags.append("FORECAST_MISSING")
            if break_even is not None and calculated_target < break_even:
                anomaly_flags.append("TARGET_BELOW_BREAK_EVEN")
                target_below_break_even_count += 1
            if break_even is not None and forecast is not None and forecast < break_even:
                anomaly_flags.append("FORECAST_BELOW_BREAK_EVEN")
                forecast_below_break_even_count += 1
            elif forecast is not None and forecast < calculated_target:
                anomaly_flags.append("FORECAST_BELOW_TARGET")

            row["profitability"] = {
                "agent_count": agents,
                "base_salary_per_agent": float(base_salary),
                "salary_cost_at_90_pct": float(salary_cost),
                "operating_costs": float(opex) if opex is not None else None,
                "accessory_margin_pct": (
                    float(accessory_margin * Decimal("100")) if accessory_margin is not None else None
                ),
                "break_even_gross_sales": float(break_even) if break_even is not None else None,
                "forecast_sales": float(forecast) if forecast is not None else None,
                "anomaly_flags": anomaly_flags,
            }

        forecast_coverage = int(forecast_coverage_contract["covered_store_count"])
        forecast_complete = forecast_coverage_contract["mode"] == "uniform"
        source_status = "ready" if complete_pnl_count == len(rows) and forecast_complete else "partial"
        return {
            "status": source_status,
            "pnl_months": pnl_months,
            "pnl_store_count": complete_pnl_count,
            "forecast_store_count": forecast_coverage,
            "forecast_run": {
                "id": int(forecast_run["id"]),
                "model_name": forecast_run["model_name"],
                "model_mode": forecast_run["model_mode"],
                "variant": forecast_run["variant"],
                "generated_at": forecast_run["generated_at"],
                "source_month": forecast_run["source_month"],
            } if forecast_run else None,
            "forecast_coverage": forecast_coverage_contract,
            "assumptions": saved_profitability,
            "salary_total": float(money(salary_total)),
            "operating_costs_total": float(money(opex_total)) if complete_pnl_count else None,
            "break_even_total": float(money(break_even_total)) if complete_pnl_count else None,
            "forecast_total": float(money(forecast_total)) if forecast_complete else None,
            "forecast_below_break_even_count": forecast_below_break_even_count,
            "target_below_break_even_count": target_below_break_even_count,
        }

    def _frozen_profitability(self, scenario: dict[str, Any], rows: list[dict[str, Any]]) -> dict[str, Any]:
        rule_set = self._saved_target_rule_set(scenario)
        if (
            rule_set is None
            or scenario.get("rule_set_id") != rule_set.rule_set_id
            or scenario.get("rule_set_hash") != rule_set.rules_hash
        ):
            raise HTTPException(status_code=409, detail="Snapshotul rule-set-ului Target nu corespunde antetului salvat.")
        calculation_params = scenario.get("calculation_params") or {}
        summary = calculation_params.get("profitability_summary")
        expected_hash = scenario.get("profitability_input_sha256")
        if not isinstance(summary, dict) or not isinstance(expected_hash, str) or len(expected_hash) != 64:
            raise HTTPException(status_code=409, detail="Snapshotul de profitabilitate Target este incomplet.")
        if summary.get("input_sha256") != expected_hash:
            raise HTTPException(status_code=409, detail="Hashul snapshotului de profitabilitate Target nu corespunde.")
        for row in rows:
            snapshot = row.get("profitability_snapshot")
            if isinstance(snapshot, str):
                try:
                    snapshot = json.loads(snapshot)
                except json.JSONDecodeError as exc:
                    raise HTTPException(status_code=409, detail="Snapshotul per magazin Target nu este JSON valid.") from exc
            if not isinstance(snapshot, dict):
                raise HTTPException(status_code=409, detail="Snapshotul per magazin Target lipseste.")
            row["profitability"] = snapshot
        return summary

    async def _attach_profitability(
        self,
        scenario: dict[str, Any],
        rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        if scenario.get("rule_set_snapshot") is not None:
            return self._frozen_profitability(scenario, rows)
        inputs = await self.repo.get_profitability_inputs(
            site_codes=[row["site_code"] for row in rows],
            target_month=scenario["target_month"],
        )
        return self._populate_profitability(scenario, rows, inputs)

    async def save_final_targets(
        self,
        scenario_id: int,
        rows: list[dict[str, Any]],
        expected_revision: int,
        *,
        actor: str | None = None,
    ) -> dict[str, Any]:
        if len({row["site_code"] for row in rows}) != len(rows):
            raise HTTPException(status_code=400, detail="Aceeasi locatie apare de mai multe ori in salvare.")
        if actor is not None:
            current = await self.get_scenario_detail(scenario_id)
            proposed_by_site = {row["site_code"]: row["proposed_target"] for row in current["rows"]}
            for row in rows:
                proposed = proposed_by_site.get(row["site_code"])
                final_target = row.get("final_target")
                if proposed is not None and final_target is not None and money(final_target) != money(proposed):
                    reason = str(row.get("override_reason") or row.get("note") or "").strip()
                    if not reason:
                        raise HTTPException(
                            status_code=400,
                            detail="Override-ul managerial necesita un motiv explicit.",
                        )
        try:
            updated = await self.repo.update_final_targets(
                scenario_id,
                rows,
                expected_revision,
                actor=actor,
            )
        except TargetScenarioVersionConflict:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Scenariul a fost modificat de alt utilizator. "
                    "Reincarca datele inainte de salvare."
                ),
            ) from None
        if updated != len(rows):
            scenario = await self.repo.get_scenario(scenario_id)
            if not scenario:
                raise HTTPException(status_code=404, detail="Scenariul de target nu exista.")
            if scenario["status"] != "draft":
                raise HTTPException(status_code=409, detail="Un scenariu finalizat nu mai poate fi editat.")
            raise HTTPException(status_code=400, detail="Una sau mai multe locatii nu apartin scenariului.")
        return await self.get_scenario_detail(scenario_id)

    async def finalize(self, scenario_id: int, expected_revision: int) -> dict[str, Any]:
        scenario = await self.get_scenario_detail(scenario_id)
        if scenario["calculation_method"] != CALCULATION_METHOD:
            raise HTTPException(
                status_code=409,
                detail="Scenariul a fost calculat cu o formula veche. Genereaza o propunere noua inainte de finalizare.",
            )
        if scenario["pending_final_count"] > 0:
            raise HTTPException(
                status_code=400,
                detail="Toate locatiile trebuie sa aiba target final completat inainte de finalizare.",
            )
        if money(scenario["final_total"]) != money(scenario["total_target"]):
            raise HTTPException(
                status_code=400,
                detail="Totalul targetelor finale trebuie sa fie egal cu bugetul scenariului inainte de finalizare.",
            )
        try:
            finalized = await self.repo.finalize_scenario(
                scenario_id,
                expected_revision,
            )
        except TargetScenarioVersionConflict:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Scenariul a fost modificat de alt utilizator. "
                    "Reincarca datele inainte de finalizare."
                ),
            ) from None
        if not finalized:
            raise HTTPException(status_code=409, detail="Scenariul nu poate fi finalizat.")
        return await self.get_scenario_detail(scenario_id)

    def _manager_allocation_analysis(self, scenario: dict[str, Any]) -> list[dict[str, Any]]:
        target_month = scenario["target_month"]
        previous_year_base_month = shift_month(target_month, -13)
        previous_year_target_month = shift_month(target_month, -12)
        previous_month = shift_month(target_month, -1)
        rows_by_manager: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in scenario["rows"]:
            rows_by_manager[row["regional"]].append(row)

        def period_value(row: dict[str, Any], month: str) -> float:
            period = next((item for item in row["history"] if item["month"] == month), None)
            return float((period or {}).get("realized") or 0)

        def build(manager: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
            target = sum(float(row["proposed_target"]) for row in rows)
            previous = sum(period_value(row, previous_month) for row in rows)
            previous_year_base = sum(period_value(row, previous_year_base_month) for row in rows)
            previous_year_target = sum(period_value(row, previous_year_target_month) for row in rows)
            forecast_values = [(row.get("profitability") or {}).get("forecast_sales") for row in rows]
            forecast = (
                sum(float(value) for value in forecast_values if value is not None)
                if all(value is not None for value in forecast_values)
                else None
            )
            seasonality_pct = percent_change(previous_year_target, previous_year_base)
            seasonal_target = (
                previous * (1 + seasonality_pct / 100)
                if seasonality_pct is not None
                else None
            )
            target_vs_previous_pct = percent_change(target, previous)
            target_vs_seasonal_pct = (
                percent_change(target, seasonal_target)
                if seasonal_target is not None
                else None
            )
            target_vs_forecast_pct = (
                percent_change(target, forecast)
                if forecast is not None
                else None
            )
            if target_vs_forecast_pct is not None and target_vs_forecast_pct >= 5:
                signal = "Peste AI"
            elif target_vs_seasonal_pct is not None and round(target_vs_seasonal_pct, 1) >= 3:
                signal = "Peste sezonier"
            else:
                signal = "Echilibrat"
            return {
                "manager": manager,
                "store_count": len(rows),
                "target": target,
                "previous": previous,
                "previous_year_base": previous_year_base,
                "previous_year_target": previous_year_target,
                "forecast": forecast,
                "target_vs_previous_pct": target_vs_previous_pct,
                "seasonality_pct": seasonality_pct,
                "seasonality_deviation_pp": (
                    target_vs_previous_pct - seasonality_pct
                    if target_vs_previous_pct is not None and seasonality_pct is not None
                    else None
                ),
                "seasonal_target": seasonal_target,
                "target_vs_seasonal_pct": target_vs_seasonal_pct,
                "target_vs_previous_year_pct": percent_change(target, previous_year_target),
                "target_vs_forecast_pct": target_vs_forecast_pct,
                "signal": signal,
            }

        managers = [
            build(manager, manager_rows)
            for manager, manager_rows in rows_by_manager.items()
        ]
        managers.sort(key=lambda item: (-item["target"], item["manager"]))
        network = build("TOTAL REȚEA", list(scenario["rows"]))
        network["signal"] = "Rețea"
        for item in [*managers, network]:
            item["target_share"] = (
                item["target"] / network["target"] if network["target"] > 0 else 0
            )
            item["previous_share"] = (
                item["previous"] / network["previous"] if network["previous"] > 0 else 0
            )
            item["previous_year_share"] = (
                item["previous_year_target"] / network["previous_year_target"]
                if network["previous_year_target"] > 0
                else 0
            )
            item["forecast_share"] = (
                item["forecast"] / network["forecast"]
                if item["forecast"] is not None and network["forecast"]
                else None
            )
            item["target_vs_previous_share_pp"] = (
                (item["target_share"] - item["previous_share"]) * 100
            )
            item["target_vs_previous_year_share_pp"] = (
                (item["target_share"] - item["previous_year_share"]) * 100
            )
            item["target_vs_forecast_share_pp"] = (
                (item["target_share"] - item["forecast_share"]) * 100
                if item["forecast_share"] is not None
                else None
            )
        return [*managers, network]

    async def export_excel(self, scenario_id: int) -> tuple[BytesIO, str]:
        scenario = await self.get_scenario_detail(scenario_id)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Target + profitabilitate"
        target_month = scenario["target_month"]
        comparison_months = [
            shift_month(target_month, -13),
            shift_month(target_month, -12),
            shift_month(target_month, -1),
        ]
        headers = ["Firma", "Manager", "Nume locație", "Cod locație"]
        for month in comparison_months:
            headers.extend([
                f"Target {month}",
                f"Realizat {month}",
                f"% {month}",
            ])
        headers.extend([
            "Pondere calcul",
            f"Calcul target {month_label_ro(target_month)}",
            "Propunere manager",
            "Cheltuieli salariale la 90% - P&L estimat",
            "Cheltuieli operaționale estimate",
            "Break-even vânzări brute",
            f"Forecast {month_label_ro(target_month)}",
        ])
        append_openpyxl_row(sheet, ["SUBTOTAL", *([""] * (len(headers) - 1))])
        append_openpyxl_row(sheet, headers)
        sorted_rows = sorted(
            scenario["rows"],
            key=lambda row: (-row["proposed_target"], row["locatie"], row["site_code"]),
        )
        for row in sorted_rows:
            history_by_month = {period["month"]: period for period in row["history"]}
            values: list[Any] = [
                row["firma"], row["regional"], row["locatie"], row["site_code"],
            ]
            for month in comparison_months:
                history = history_by_month.get(month) or {}
                attainment = history.get("attainment_pct")
                values.extend([
                    history.get("target", 0),
                    history.get("realized", 0),
                    None if attainment is None else attainment / 100,
                ])
            profitability = row.get("profitability") or {}
            values.extend([
                row.get("normalized_weight", row["calculated_weight"]),
                row["proposed_target"],
                row["final_target"],
                profitability.get("salary_cost_at_90_pct"),
                profitability.get("operating_costs"),
                profitability.get("break_even_gross_sales"),
                profitability.get("forecast_sales"),
            ])
            append_openpyxl_row(sheet, values)
        last_row = sheet.max_row
        sheet.freeze_panes = "E3"
        sheet.auto_filter.ref = f"A2:T{last_row}"

        total_columns = ("E", "F", "H", "I", "K", "L", "N", "O", "P", "Q", "R", "S", "T")
        for column in total_columns:
            sheet[f"{column}1"] = f"=SUBTOTAL(109,{column}3:{column}{last_row})"
        for percentage_column, target_column, realized_column in (
            ("G", "E", "F"),
            ("J", "H", "I"),
            ("M", "K", "L"),
        ):
            sheet[f"{percentage_column}1"] = (
                f'=IF({target_column}1=0,0,{realized_column}1/{target_column}1)'
            )

        comparison = workbook.create_sheet("Comparație manageri")
        manager_analysis = self._manager_allocation_analysis(scenario)
        append_openpyxl_row(comparison, ["1. Distribuția targetului", *([""] * 8)])
        comparison.merge_cells("A1:I1")
        append_openpyxl_row(comparison, [
            "Manager", "Nr. locații", "Pondere target calculat",
            f"Pondere realizat {comparison_months[2]}",
            f"Δ vs {comparison_months[2]} (pp)",
            f"Pondere realizat {comparison_months[1]}",
            f"Δ vs {comparison_months[1]} (pp)",
            f"Pondere forecast {target_month}",
            "Δ vs forecast (pp)",
        ])
        for item in manager_analysis:
            append_openpyxl_row(comparison, [
                item["manager"], item["store_count"], item["target_share"],
                item["previous_share"], item["target_vs_previous_share_pp"],
                item["previous_year_share"], item["target_vs_previous_year_share_pp"],
                item["forecast_share"], item["target_vs_forecast_share_pp"],
            ])
        distribution_total_row = comparison.max_row
        append_openpyxl_row(comparison, [])
        second_title_row = distribution_total_row + 2
        append_openpyxl_row(
            comparison,
            ["2. Target vs lună precedentă, an precedent, sezonalitate și forecast AI", *([""] * 12)],
        )
        comparison.merge_cells(start_row=second_title_row, start_column=1, end_row=second_title_row, end_column=13)
        second_header_row = second_title_row + 1
        append_openpyxl_row(comparison, [
            "Manager", f"Target {target_month}", f"Realizat {comparison_months[2]}",
            f"Target vs {comparison_months[2]}", "Sezonalitate istorică",
            "Abatere sezonalitate (pp)", "Target sezonier estimat",
            "Gap vs target sezonier", f"Realizat {comparison_months[1]}",
            f"Target vs {comparison_months[1]}", f"Forecast AI {target_month}",
            "Target vs AI", "Semnal alocare",
        ])
        for item in manager_analysis:
            append_openpyxl_row(comparison, [
                item["manager"], item["target"], item["previous"],
                item["target_vs_previous_pct"], item["seasonality_pct"],
                item["seasonality_deviation_pp"], item["seasonal_target"],
                item["target_vs_seasonal_pct"], item["previous_year_target"],
                item["target_vs_previous_year_pct"], item["forecast"],
                item["target_vs_forecast_pct"], item["signal"],
            ])
        analysis_total_row = comparison.max_row

        summary = workbook.create_sheet("Rezumat calcul")
        append_openpyxl_row(summary, [
            "Regional", "Magazine", "Floor", "Target propus", "Target final", "Diferenta",
            "Luna curenta", "Forecast luna curenta", "% crestere propus vs luna curenta",
            "Baza anul trecut", "Target anul trecut", "Realizat baza anul trecut",
            "Realizat target anul trecut", "% crestere anul trecut",
        ])
        for row in scenario["regional_summary"]:
            append_openpyxl_row(summary, [
                row["regional"], row["store_count"], row["floor_total"],
                row["proposed_total"], row["final_total"], row["final_total"] - row["proposed_total"],
                row.get("current_month"), row.get("current_forecast_total"), row.get("proposed_growth_vs_current_pct"),
                row.get("last_year_base_month"), row.get("last_year_target_month"),
                row.get("last_year_base_total"), row.get("last_year_target_total"),
                row.get("last_year_growth_pct"),
            ])

        parameters = workbook.create_sheet("Parametri")
        append_openpyxl_row(parameters, ["Parametru", "Valoare"])
        append_openpyxl_row(parameters, ["Scenariu", scenario["id"]])
        append_openpyxl_row(parameters, ["Status", scenario["status"]])
        append_openpyxl_row(parameters, ["Luna target", scenario["target_month"]])
        append_openpyxl_row(parameters, ["Luna cohorta magazine active", scenario["cohort_month"]])
        append_openpyxl_row(parameters, ["Target total", scenario["total_target"]])
        append_openpyxl_row(parameters, ["Prag minim absolut", scenario["min_floor"]])
        append_openpyxl_row(parameters, ["Floor fata de luna precedenta", scenario["previous_month_floor_pct"]])
        append_openpyxl_row(parameters, ["Metoda", scenario["calculation_method"]])
        for key, value in (scenario.get("calculation_params") or {}).items():
            append_openpyxl_row(parameters, [f"Parametru {key}", json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value])
        for item in scenario["source_months"]:
            append_openpyxl_row(parameters, [item["label"], item["month"]])
        for item in scenario["source_summary"]:
            if item["is_forecast"]:
                append_openpyxl_row(parameters, [
                    f"Forecast {item['month']}",
                    f"{item['forecast_factor']:.4f}x; importat {item['actual_realized']:.2f}; folosit {item['realized']:.2f}",
                ])
        for warning in scenario["warnings"]:
            append_openpyxl_row(parameters, ["Atentionare", warning])
        profitability_summary = scenario.get("profitability_summary") or {}
        append_openpyxl_row(parameters, ["Status surse profitabilitate", profitability_summary.get("status")])
        append_openpyxl_row(parameters, ["Luni P&L reale", ", ".join(profitability_summary.get("pnl_months") or [])])
        forecast_run = profitability_summary.get("forecast_run") or {}
        append_openpyxl_row(parameters, ["Forecast run", forecast_run.get("id")])
        append_openpyxl_row(parameters, ["Forecast model", forecast_run.get("model_name")])
        append_openpyxl_row(parameters, ["Forecast variant", forecast_run.get("variant")])

        navy_fill = PatternFill("solid", fgColor="17365D")
        subtotal_fill = PatternFill("solid", fgColor="D9E2F3")
        percentage_fill = PatternFill("solid", fgColor="F3F4F6")
        manager_fill = PatternFill("solid", fgColor="FFF8D9")
        break_even_fill = PatternFill("solid", fgColor="FFF7ED")
        forecast_fill = PatternFill("solid", fgColor="EFF8F1")
        red_fill = PatternFill("solid", fgColor="F4CCCC")
        green_fill = PatternFill("solid", fgColor="E2F0D9")
        red_font = Font(color="9C0006", bold=True)
        amber_font = Font(color="C65911", bold=True)
        green_font = Font(color="00B050", bold=True)

        for cell in sheet[1]:
            cell.font = Font(color="111827", bold=True)
            cell.fill = subtotal_fill
        for cell in sheet[2]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[2].height = 52
        for row_number in range(1, last_row + 1):
            for column in ("G", "J", "M", "N"):
                sheet[f"{column}{row_number}"].number_format = "0.0%"
        for row_number in range(3, last_row + 1):
            for column in ("G", "J", "M", "N"):
                sheet[f"{column}{row_number}"].fill = percentage_fill
            sheet[f"P{row_number}"].fill = manager_fill
            sheet[f"S{row_number}"].fill = break_even_fill
            sheet[f"T{row_number}"].fill = forecast_fill
            sheet[f"O{row_number}"].font = Font(color="111827", bold=True)
            sheet[f"S{row_number}"].font = Font(color="8A4B16", bold=True)
            sheet[f"T{row_number}"].font = Font(color="27633B", bold=True)
        for row_number in range(1, last_row + 1):
            for column in ("E", "F", "H", "I", "K", "L", "O", "P", "Q", "R", "S", "T"):
                sheet[f"{column}{row_number}"].number_format = '#,##0;[Red]-#,##0;-'

        for column in ("G", "J", "M"):
            data_range = f"{column}3:{column}{last_row}"
            sheet.conditional_formatting.add(
                data_range,
                CellIsRule(operator="lessThan", formula=["0.9"], font=red_font),
            )
            sheet.conditional_formatting.add(
                data_range,
                CellIsRule(operator="between", formula=["0.9", "0.999999999"], font=amber_font),
            )
            sheet.conditional_formatting.add(
                data_range,
                CellIsRule(operator="greaterThanOrEqual", formula=["1"], font=green_font),
            )
        sheet.conditional_formatting.add(
            f"O3:O{last_row}",
            FormulaRule(formula=["AND(ISNUMBER($S3),$O3<$S3)"], fill=red_fill, font=red_font),
        )
        sheet.conditional_formatting.add(
            f"P3:P{last_row}",
            FormulaRule(formula=["AND(ISNUMBER($P3),ISNUMBER($S3),$P3<$S3)"], fill=red_fill, font=red_font),
        )
        sheet.conditional_formatting.add(
            f"T3:T{last_row}",
            FormulaRule(formula=["AND(ISNUMBER($T3),ISNUMBER($S3),$T3<$S3)"], fill=red_fill, font=red_font),
        )
        sheet.conditional_formatting.add(
            f"T3:T{last_row}",
            FormulaRule(formula=["AND(ISNUMBER($T3),ISNUMBER($S3),$T3>=$S3)"], fill=green_fill, font=green_font),
        )

        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 29
        sheet.column_dimensions["D"].width = 15
        for column in ("E", "F", "H", "I", "K", "L", "O", "P", "Q", "R", "S", "T"):
            sheet.column_dimensions[column].width = 15
        for column in ("G", "J", "M", "N"):
            sheet.column_dimensions[column].width = 11
        sheet.column_dimensions["Q"].width = 20
        sheet.column_dimensions["R"].width = 19
        sheet.column_dimensions["S"].width = 18

        section_fill = PatternFill("solid", fgColor="17365D")
        header_fill = PatternFill("solid", fgColor="5B9BD5")
        total_fill = PatternFill("solid", fgColor="1F2937")
        signal_balanced_fill = PatternFill("solid", fgColor="C6EFCE")
        signal_seasonal_fill = PatternFill("solid", fgColor="FFF2CC")
        signal_ai_fill = PatternFill("solid", fgColor="F4CCCC")
        for row_number in (1, second_title_row):
            for cell in comparison[row_number]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = section_fill
        for row_number in (2, second_header_row):
            for cell in comparison[row_number]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
            comparison.row_dimensions[row_number].height = 46
        for row_number in (distribution_total_row, analysis_total_row):
            for cell in comparison[row_number]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = total_fill
        for row_number in range(3, distribution_total_row + 1):
            for column in ("C", "D", "F", "H"):
                comparison[f"{column}{row_number}"].number_format = "0.0%"
            for column in ("E", "G", "I"):
                comparison[f"{column}{row_number}"].number_format = '0.0" pp"'
                value = comparison[f"{column}{row_number}"].value
                if isinstance(value, (int, float)) and value < 0:
                    comparison[f"{column}{row_number}"].font = Font(color="FF0000", bold=True)
        for row_number in range(second_header_row + 1, analysis_total_row + 1):
            for column in ("B", "C", "G", "I", "K"):
                comparison[f"{column}{row_number}"].number_format = '#,##0;[Red]-#,##0;-'
            for column in ("D", "E", "H", "J", "L"):
                comparison[f"{column}{row_number}"].number_format = '0.0"%"'
                value = comparison[f"{column}{row_number}"].value
                if isinstance(value, (int, float)):
                    comparison[f"{column}{row_number}"].font = Font(
                        color="00B050" if value >= 0 else "FF0000",
                        bold=True,
                    )
            comparison[f"F{row_number}"].number_format = '0.0" pp"'
            signal = comparison[f"M{row_number}"].value
            if signal == "Echilibrat":
                comparison[f"M{row_number}"].fill = signal_balanced_fill
                comparison[f"M{row_number}"].font = Font(color="006100", bold=True)
            elif signal == "Peste sezonier":
                comparison[f"M{row_number}"].fill = signal_seasonal_fill
                comparison[f"M{row_number}"].font = Font(color="9C6500", bold=True)
            elif signal == "Peste AI":
                comparison[f"M{row_number}"].fill = signal_ai_fill
                comparison[f"M{row_number}"].font = Font(color="9C0006", bold=True)
        comparison.freeze_panes = "A3"
        comparison.column_dimensions["A"].width = 23
        comparison.column_dimensions["B"].width = 15
        for column in ("C", "D", "E", "F", "G", "H", "I", "J", "K", "L"):
            comparison.column_dimensions[column].width = 18
        comparison.column_dimensions["M"].width = 17
        for row_number in (distribution_total_row, analysis_total_row):
            for cell in comparison[row_number]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = total_fill

        for worksheet in (summary, parameters):
            for cell in worksheet[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = navy_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for column in worksheet.columns:
                letter = get_column_letter(column[0].column)
                max_length = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 34)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        stamp = business_today().strftime("%Y%m%d")
        return output, f"targete_{scenario['target_month']}_scenariu_{scenario_id}_{stamp}.xlsx"

    async def get_store_detail(self, scenario_id: int, site_code: str) -> dict[str, Any]:
        data = await self.repo.get_store_detail(scenario_id, site_code)
        if not data:
            raise HTTPException(status_code=404, detail="Locatia nu exista in documentul de target.")

        scenario = dict(data["scenario"])
        history = [self._serialize_store_history(dict(row)) for row in data["history"]]
        agents = [self._serialize_store_agent(dict(row)) for row in data["agents"]]
        latest = next((row for row in reversed(history) if row["total_sales"] > 0 or row["target_value"] > 0), None)
        sales_values = [row["total_sales"] for row in history]
        best_month = max(history, key=lambda row: row["total_sales"]) if history else None
        avg_sales = sum(sales_values) / len(sales_values) if sales_values else 0

        return {
            "site_code": scenario["site_code"],
            "locatie": scenario["locatie"],
            "firma": scenario["firma"],
            "regional": scenario["regional"],
            "asm": scenario["asm"],
            "target_month": scenario["target_month"],
            "cohort_month": scenario["cohort_month"],
            "proposed_target": float(scenario["proposed_target"] or 0),
            "final_target": float(scenario["final_target"]) if scenario["final_target"] is not None else None,
            "history": history,
            "latest": latest,
            "best_month": best_month,
            "avg_sales_16m": round(avg_sales, 2),
            "agents": agents,
        }

    def _serialize_header(self, row: dict[str, Any]) -> dict[str, Any]:
        for key in ("total_target", "min_floor", "previous_month_floor_pct", "proposed_total", "final_total"):
            if key in row:
                row[key] = float(row[key] or 0)
        for key in ("source_months", "warnings", "calculation_params", "rule_set_snapshot"):
            if key in row and isinstance(row[key], str):
                row[key] = json.loads(row[key])
        row.setdefault("source_months", [])
        row.setdefault("warnings", [])
        row.setdefault("calculation_params", {})
        row.setdefault("rule_set_snapshot", None)
        if "store_count" in row:
            row["store_count"] = int(row["store_count"])
        row["pending_final_count"] = int(row.get("pending_final_count") or 0)
        return row

    def _serialize_row(self, row: dict[str, Any]) -> dict[str, Any]:
        for key in ("calculated_weight", "floor_target", "cap_target", "proposed_target"):
            if key in row:
                row[key] = float(row[key] or 0)
        for key in ("final_target", "manager_override_target"):
            row[key] = float(row[key]) if row.get(key) is not None else None
        if isinstance(row.get("profitability_snapshot"), str):
            row["profitability_snapshot"] = json.loads(row["profitability_snapshot"])
        if isinstance(row.get("history"), str):
            row["history"] = json.loads(row["history"])
        if isinstance(row.get("calculation_details"), str):
            row["calculation_details"] = json.loads(row["calculation_details"])
        row.setdefault("calculation_details", {})
        return row

    def _regional_summary(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        summary: dict[str, dict[str, Any]] = defaultdict(
            lambda: {
                "store_count": 0,
                "floor_total": 0.0,
                "proposed_total": 0.0,
                "final_total": 0.0,
                "current_month": None,
                "current_forecast_total": 0.0,
                "last_year_base_month": None,
                "last_year_target_month": None,
                "last_year_base_total": 0.0,
                "last_year_target_total": 0.0,
            }
        )
        for row in rows:
            data = summary[row["regional"]]
            data["store_count"] += 1
            data["floor_total"] += row["floor_target"]
            data["proposed_total"] += row["proposed_target"]
            data["final_total"] += row["final_target"] or 0
            details = row.get("calculation_details") or {}
            if isinstance(details, str):
                details = json.loads(details)
            history = row.get("history") or []
            if isinstance(history, str):
                history = json.loads(history)

            current_month = details.get("current_month")
            current_forecast = details.get("current_forecast")
            if current_forecast is None:
                current_period = next((item for item in history if item.get("role") == "floor_reference"), None)
                current_month = current_month or (current_period or {}).get("month")
                current_forecast = (current_period or {}).get("realized")
            if current_month:
                data["current_month"] = current_month
            data["current_forecast_total"] += float(current_forecast or 0)

            seasonality = details.get("seasonality") or {}
            last_year = next(
                (item for item in seasonality.get("store_years") or [] if item.get("year_offset") == 1),
                None,
            )
            if last_year is None:
                base_period = next((item for item in history if item.get("role") == "seasonality_base_y1"), None)
                target_period = next((item for item in history if item.get("role") == "seasonality_target_y1"), None)
                last_year = {
                    "base_month": (base_period or {}).get("month"),
                    "target_month": (target_period or {}).get("month"),
                    "base_value": (base_period or {}).get("realized"),
                    "target_value": (target_period or {}).get("realized"),
                }
            if last_year.get("base_month"):
                data["last_year_base_month"] = last_year["base_month"]
            if last_year.get("target_month"):
                data["last_year_target_month"] = last_year["target_month"]
            data["last_year_base_total"] += float(last_year.get("base_value") or 0)
            data["last_year_target_total"] += float(last_year.get("target_value") or 0)
        return [
            {
                "regional": regional,
                **values,
                "proposed_growth_vs_current_pct": percent_change(
                    values["proposed_total"],
                    values["current_forecast_total"],
                ),
                "final_growth_vs_current_pct": percent_change(
                    values["final_total"],
                    values["current_forecast_total"],
                ),
                "last_year_growth_pct": percent_change(
                    values["last_year_target_total"],
                    values["last_year_base_total"],
                ),
            }
            for regional, values in sorted(summary.items())
        ]

    def _source_summary(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        summary: dict[str, dict[str, Any]] = defaultdict(
            lambda: {
                "target": 0.0,
                "realized": 0.0,
                "actual_realized": 0.0,
                "is_forecast": False,
                "forecast_factor": 1.0,
                "label": "",
            }
        )
        for row in rows:
            for period in row["history"]:
                values = summary[period["month"]]
                values["label"] = period["label"]
                values["target"] += period["target"]
                values["realized"] += period["realized"]
                values["actual_realized"] += period.get("actual_realized", period["realized"])
                if period.get("is_forecast", False):
                    values["is_forecast"] = True
                    values["forecast_factor"] = period.get("forecast_factor", 1.0)
        return [
            {
                "month": month,
                **values,
                "attainment_pct": values["realized"] / values["target"] * 100 if values["target"] else None,
            }
            for month, values in summary.items()
        ]

    def _serialize_store_history(self, row: dict[str, Any]) -> dict[str, Any]:
        total_sales = float(row["total_sales"] or 0)
        target = float(row["target_value"] or 0)
        total_quantity = int(row["total_quantity"] or 0)
        receipt_count = int(row["receipt_count"] or 0)
        receipt_2plus = int(row["receipt_2plus_count"] or 0)
        focus_quantity = int(row["focus_quantity"] or 0)
        return {
            "month": row["import_month"],
            "total_sales": total_sales,
            "target_value": target,
            "target_pct": total_sales / target * 100 if target else None,
            "total_quantity": total_quantity,
            "receipt_count": receipt_count,
            "cartele_qty": int(row["cartele_qty"] or 0),
            "avg_receipt": total_sales / receipt_count if receipt_count else None,
            "bon2acc_pct": receipt_2plus / receipt_count * 100 if receipt_count else None,
            "focus_pct": focus_quantity / total_quantity * 100 if total_quantity else None,
            "active_agents": int(row["active_agents"] or 0),
            "working_days": int(row["working_days"] or 0),
        }

    def _serialize_store_agent(self, row: dict[str, Any]) -> dict[str, Any]:
        total_sales = float(row["total_sales"] or 0)
        total_quantity = int(row["total_quantity"] or 0)
        receipt_count = int(row["receipt_count"] or 0)
        receipt_2plus = int(row["receipt_2plus_count"] or 0)
        focus_quantity = int(row["focus_quantity"] or 0)
        return {
            "agent": row["agent"],
            "total_sales": total_sales,
            "sales_share_pct": float(row["sales_share_pct"] or 0),
            "total_quantity": total_quantity,
            "receipt_count": receipt_count,
            "avg_receipt": total_sales / receipt_count if receipt_count else None,
            "bon2acc_pct": receipt_2plus / receipt_count * 100 if receipt_count else None,
            "focus_pct": focus_quantity / total_quantity * 100 if total_quantity else None,
            "active_months_16": int(row["active_months_16"] or 0),
            "sales_16m": float(row["sales_16m"] or 0),
        }
