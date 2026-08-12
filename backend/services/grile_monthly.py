"""Operatii lunare native pentru grile in UniHub Retail.

Inlocuieste proxy-ul catre aplicatia veche `grile-salarii`: Retail citeste
registrul din DB (`grile_sheets` + `stores`), genereaza Excelul final, arhiva
XLSX/ZIP si ruleaza resetul controlat direct cu Google APIs.
"""

from __future__ import annotations

import asyncio
import json
import os
import shutil
import threading
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Awaitable, Callable, Coroutine, Literal
from uuid import uuid4

import asyncpg
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from repositories.grile_monthly_operations import (
    MonthlyExecutionLease,
    ResetItemInput,
    attach_job as persist_monthly_operation_job,
    claim_reset_item as persist_reset_item_claim,
    ensure_reset_items as persist_reset_items,
    fail as persist_monthly_operation_failure,
    fail_queued as persist_queued_monthly_operation_failure,
    get_execution_lease as fetch_monthly_execution_lease,
    mark_cancelled_uncertain as persist_cancelled_uncertain,
    finish as persist_monthly_operation_result,
    finish_reset_success as persist_reset_success,
    finish_reset_item as persist_reset_item_result,
    get_latest_manifest as fetch_latest_monthly_manifest,
    get_manifest as fetch_monthly_manifest,
    get_operation_manifest as fetch_operation_manifest,
    get_previous_completed_reset_item as fetch_previous_completed_reset_item,
    heartbeat as persist_monthly_operation_heartbeat,
    approve_manifest as persist_monthly_manifest_approval,
    persist_manifest_result,
    record_reset_item_backup as persist_reset_item_backup,
    record_reset_item_rollback as persist_reset_item_rollback,
    prepare_reset_clear as persist_reset_clear_intent,
    confirm_reset_clear as persist_reset_clear_confirmation,
    prepare_reset_rollback as persist_reset_rollback_intent,
    confirm_reset_rollback as persist_reset_rollback_confirmation,
    claim_reconciliation_candidates,
    list_reset_items_for_reconciliation,
    mark_item_recovery_required,
    mark_item_safe_retry,
    mark_reconciliation_result,
    operation_to_dict as persisted_operation_to_dict,
    reserve as persist_monthly_operation_reservation,
    start as persist_monthly_operation_start,
)
from services.grile_monthly_google import (
    GoogleAdapterClosed,
    GoogleSyncAdapter,
    call_with_backoff,
)
from services.spreadsheet_safety import TrustedFormula, append_openpyxl_row
from services.grile_monthly_state import (
    GrileMonthlyRetryBlockedError,
    MonthlyOperationReservation,
    MonthlyOperationStartResult,
    safe_persisted_result,
)
from services.grile_constants import (
    GOOGLE_API_RETRY_ATTEMPTS,
    GOOGLE_API_RETRY_BASE_DELAY_SECONDS,
)
from services.grile_monthly_integrity import (
    MonthlyIntegrityError,
    base_manifest,
    canonical_snapshot,
    decimal_text,
    file_sha256,
    finalize_manifest,
    manifest_sha256,
    parse_required_decimal,
    relative_artifact,
    resolve_artifact_path,
    secure_directory, secure_file,
    secure_write_json,
    snapshot_sha256,
    utc_now,
    validate_verified_manifest,
    verify_artifacts,
)

GOOGLE_OPERATION_DEADLINE_SECONDS = 120.0
MONTHLY_OPERATION_HEARTBEAT_SECONDS = 60.0

RO_MONTHS = [
    "", "Ianuarie", "Februarie", "Martie", "Aprilie", "Mai", "Iunie",
    "Iulie", "August", "Septembrie", "Octombrie", "Noiembrie", "Decembrie",
]

VALID_OPS = {"finalize", "archive", "reset"}
MANIFEST_ATTEMPT_STATUSES = (
    "building",
    "failed",
    "verified",
    "approved",
    "consumed",
    "rolled_back",
    "uncertain",
)
VALID_DOWNLOADS = {"final": "final", "archive": "archive"}
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUTS_DIR = Path(os.getenv("GRILE_OUTPUTS_DIR", BASE_DIR / "outputs" / "grile"))
FINAL_EXPORT_NAME_PREFIX = "Tabel Salarii -"
ARCHIVE_DIR_NAME = "archive"
RESET_RANGES = [
    "Grila!D8",
    "Grila!D22",
    "Grila!P5:P36",
    "Grila!Q5:S36",
    "Grila!U5:U36",
    "Grila!V5:X36",
    "Grila!B32:F46",
    "Grila!F12:F14",
    "Grila!F26:F28",
    "Pontaj!C8:AG31",
]
RESET_RANGES_V3 = [
    "Grila!D8",
    "Grila!D22",
    "Grila!D36",
    "Grila!P5:P50",
    "Grila!Q5:S50",
    "Grila!U5:U50",
    "Grila!V5:X50",
    "Grila!Z5:Z50",
    "Grila!AA5:AC50",
    "Grila!B46:F60",
    "Grila!F12:F14",
    "Grila!F26:F28",
    "Grila!F40:F42",
    "Pontaj!C8:AG31",
]
GRILA_CELLS = {
    1: {
        "agent": "D2",
        "base_salary": "D3",
        "sales_commission_cells": ["G8", "G9", "G12", "G13", "G14"],
        "bonuri": "D4",
        "extra_hours_pay": "G10",
        "extra_location_commission": "G11",
        "worked_hours": "Pontaj!AH8",
    },
    2: {
        "agent": "D16",
        "base_salary": "D17",
        "sales_commission_cells": ["G22", "G23", "G26", "G27", "G28"],
        "bonuri": "D18",
        "extra_hours_pay": "G24",
        "extra_location_commission": "G25",
        "worked_hours": "Pontaj!AH11",
    },
}
GRILA_CELLS_V3 = {
    **GRILA_CELLS,
    3: {
        "agent": "D30",
        "base_salary": "D31",
        "sales_commission_cells": ["G36", "G37", "G40", "G41", "G42"],
        "bonuri": "D32",
        "extra_hours_pay": "G38",
        "extra_location_commission": "G39",
        "worked_hours": "Pontaj!AH14",
    },
}
HEADERS = [
    "Nr",
    "Manager",
    "Magazin",
    "Agent",
    "Salariu baza",
    "Comision vanzare",
    "Flip",
    "Comision vanzare zile suplimentare",
    "Incentive lunar",
    "Plata ore suplimentare",
    "Total salariu",
    "Salariu Cash",
    "Bonuri",
    "Data angajarii",
    "Data plecarii",
    "Nr. Ore lucrate",
    "Zile CO luna in curs",
]
AUDIT_HEADERS = [
    "Company",
    "Store",
    "Slot",
    "Agent",
    "Sheet ID",
    "Comision vanzare",
    "Comision supl",
    "Plata ore supl",
    "Bonuri",
    "Ore lucrate",
    "Source",
    "Status",
    "Error",
]
_TRANSIENT = {429, 500, 502, 503, 504}


@dataclass(frozen=True)
class StoreEntry:
    company: str
    store: str
    sheet_id: str
    site_code: str
    manager: str
    is_closed: bool = False
    template_version: str = "v2"


def cells_for_entry(entry: StoreEntry) -> dict[int, dict[str, Any]]:
    return GRILA_CELLS_V3 if entry.template_version == "v3" else GRILA_CELLS


def reset_ranges_for_entry(entry: StoreEntry) -> list[str]:
    return list(RESET_RANGES_V3 if entry.template_version == "v3" else RESET_RANGES)


@dataclass
class ExtractedAgentRow:
    company: str
    store: str
    slot: int
    agent: Any
    base_salary: Any
    sales_commission: Any
    extra_location_commission: Any
    extra_hours_pay: Any
    bonuri: Any
    worked_hours: Any
    status: str
    error: str
    sheet_id: str
    site_code: str = ""
    error_code: str = ""


@dataclass(frozen=True)
class MonthlyExecution:
    path: Path
    manifest: dict[str, Any]
    rollback: Callable[[], Awaitable[dict[str, Any]]] | None = None


class MonthlyManifestError(MonthlyIntegrityError):
    def __init__(self, code: str, message: str, manifest: dict[str, Any]):
        super().__init__(code, message)
        self.manifest = manifest


def ro_month_label(ym: str) -> str:
    """`2026-05` -> `Mai 2026`."""
    year, month = ym.split("-")
    return f"{RO_MONTHS[int(month)]} {year}"


def next_ym(ym: str) -> str:
    year, month = (int(x) for x in ym.split("-"))
    month += 1
    if month > 12:
        month = 1
        year += 1
    return f"{year:04d}-{month:02d}"


def _sa_file() -> Path:
    return Path(
        os.getenv(
            "GRILE_GOOGLE_SA_FILE",
            BASE_DIR / "config" / "google" / "service-account.json",
        )
    )


def get_credentials() -> Any:
    from google.oauth2.service_account import Credentials

    path = _sa_file()
    if not path.exists():
        raise FileNotFoundError(
            f"Service account Google lipsa: {path}. Pune fisierul sau seteaza GRILE_GOOGLE_SA_FILE."
        )
    return Credentials.from_service_account_file(str(path), scopes=SCOPES)


def build_google_services() -> tuple[Any, Any]:
    creds = get_credentials()
    sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)
    return sheets, drive


def safe_filename(value: str) -> str:
    import re

    cleaned = value.replace("/", " - ").replace("\\", " - ")
    cleaned = re.sub(r'[<>:"|?*]', "_", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned.rstrip(". ") or "untitled"


def month_slug(month: str) -> str:
    import re

    slug = re.sub(r"[^0-9A-Za-zĂÂÎȘȚăâîșț]+", "-", month.strip())
    return slug.strip("-")


def build_final_export_path(outputs_dir: Path, month: str) -> Path:
    return outputs_dir / f"{FINAL_EXPORT_NAME_PREFIX} {month}.xlsx"


def build_archive_dir(outputs_dir: Path, month: str) -> Path:
    return outputs_dir / ARCHIVE_DIR_NAME / month


def build_archive_manifest_path(outputs_dir: Path, month: str) -> Path:
    return build_archive_dir(outputs_dir, month) / f"archive-manifest-{month_slug(month)}.json"


def build_archive_zip_path(outputs_dir: Path, month: str) -> Path:
    return build_archive_dir(outputs_dir, month) / f"Grile - {month}.zip"


def build_reset_report_path(outputs_dir: Path, next_month: str) -> Path:
    return outputs_dir / f"reset-report-{month_slug(next_month)}.json"


def build_reset_dry_run_report_path(outputs_dir: Path, next_month: str) -> Path:
    return outputs_dir / f"reset-dry-run-{month_slug(next_month)}.json"


def build_reset_backup_dir(outputs_dir: Path, closing_month: str, operation_id: int) -> Path:
    return outputs_dir / "reset-backups" / month_slug(closing_month) / str(operation_id)


def build_store_export_path(outputs_dir: Path, month: str, entry: StoreEntry) -> Path:
    return build_archive_dir(outputs_dir, month) / safe_filename(entry.company) / f"{safe_filename(entry.store)}.xlsx"


def build_manager_zip_path(outputs_dir: Path, month: str, manager: str) -> Path:
    return build_archive_dir(outputs_dir, month) / "ASM" / f"Grile - {month} - {safe_filename(manager)}.zip"


def resolve_output_path(month: str, only: str | None, output_dir: Path) -> Path:
    output_path = build_final_export_path(output_dir, month)
    if only:
        output_path = output_path.with_name(
            f"{output_path.stem} - TEST {safe_filename(only)}{output_path.suffix}"
        )
    return output_path


def _is_transient(exc: Exception) -> bool:
    status = getattr(getattr(exc, "resp", None), "status", None)
    return status in _TRANSIENT


def _google_error_code(exc: Exception) -> str:
    status = getattr(getattr(exc, "resp", None), "status", None)
    if status == 429:
        return "google_rate_limited"
    if status in {500, 502, 503, 504}:
        return "google_unavailable"
    if isinstance(exc, (TimeoutError, asyncio.TimeoutError)):
        return "google_timeout"
    return "google_request_failed"


def retry_api(fn, *, label: str, attempts: int = 4, base_delay: float = 1.0):
    last: Exception | None = None
    for attempt in range(attempts):
        try:
            return fn()
        except Exception as exc:  # noqa: BLE001
            last = exc
            if isinstance(exc, MonthlyIntegrityError):
                raise
            if attempt < attempts - 1 and _is_transient(exc):
                # Compatibility-only synchronous callers.  Worker operations
                # use call_with_backoff(), which yields with asyncio.sleep.
                threading.Event().wait(base_delay * (2**attempt))
                continue
            raise MonthlyIntegrityError(_google_error_code(exc), f"{label} failed") from exc
    assert last is not None
    raise last


async def _google_request(
    adapter: GoogleSyncAdapter,
    operation: str,
    request: dict[str, Any],
    *,
    label: str,
    destructive: bool = False,
) -> Any:
    try:
        deadline = asyncio.get_running_loop().time() + GOOGLE_OPERATION_DEADLINE_SECONDS
        return await call_with_backoff(
            adapter,
            operation,
            request,
            label=label,
            attempts=GOOGLE_API_RETRY_ATTEMPTS,
            base_delay=GOOGLE_API_RETRY_BASE_DELAY_SECONDS,
            destructive=destructive,
            deadline=deadline,
        )
    except MonthlyIntegrityError:
        raise
    except Exception as exc:  # noqa: BLE001 - provider boundary is classified here
        raise MonthlyIntegrityError(_google_error_code(exc), f"{label} failed") from exc


def _company_from_values(registry_key: str | None, fallback: str | None) -> str:
    raw = (registry_key or "").split("/", 1)[0].strip() or (fallback or "").strip()
    normalized = raw.casefold()
    if normalized == "mobicell":
        return "Mobicell"
    if normalized == "mobiup":
        return "Mobiup"
    raise MonthlyIntegrityError(
        "unknown_company",
        "Grile registry company is missing or unsupported",
    )


def _store_from_values(registry_key: str | None, fallback: str | None) -> str:
    if registry_key and "/" in registry_key:
        return registry_key.split("/", 1)[1].strip()
    return (fallback or "").strip()


async def load_entries(
    pool: asyncpg.Pool,
    only: str | None = None,
    *,
    month: str | None = None,
) -> list[StoreEntry]:
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                gs.site_code,
                gs.sheet_id,
                gs.registry_key,
                s.locatie,
                s.firma,
                s.asm,
                gs.template_version
            FROM grile_sheets gs
            JOIN stores s ON s.site_code = gs.site_code
            WHERE gs.is_active = true
              AND s.is_active = true
              AND ($1::TEXT IS NULL OR gs.active_from_month IS NULL OR gs.active_from_month <= $1)
            ORDER BY COALESCE(gs.registry_key, s.firma || '/' || s.locatie)
            """,
            month,
        )

    entries = [
        StoreEntry(
            company=_company_from_values(r["registry_key"], r["firma"]),
            store=_store_from_values(r["registry_key"], r["locatie"]),
            sheet_id=r["sheet_id"],
            site_code=r["site_code"],
            manager=(r["asm"] or "Neatribuit").strip() or "Neatribuit",
            is_closed=str(r["locatie"] or "").strip().upper().startswith("INCHIS "),
            template_version=r.get("template_version") or "v2",
        )
        for r in rows
    ]
    if only:
        needle = only.casefold()
        entries = [
            e for e in entries
            if needle in f"{e.company}/{e.store}/{e.site_code}/{e.manager}".casefold()
        ]
    if not entries:
        raise RuntimeError("No active grile matched the requested filter.")
    return entries


def _operation_to_dict(row: asyncpg.Record | None) -> dict[str, Any] | None:
    return persisted_operation_to_dict(row)


def _safe_operation_result(operation: dict[str, Any] | None) -> dict[str, Any] | None:
    """Return an independent copy of a persisted JSON result for replays."""
    return safe_persisted_result(operation)


async def reserve_monthly_operation(
    pool: asyncpg.Pool,
    *,
    op: str,
    month: str,
    only: str | None,
    dry_run: bool,
    requested_by_sub: str,
    approved_manifest_id: int | None = None,
) -> MonthlyOperationReservation:
    if op not in VALID_OPS:
        raise ValueError(f"Operatie necunoscuta: {op}")
    return await persist_monthly_operation_reservation(
        pool,
        op=op,
        month=month,
        only=only,
        dry_run=dry_run,
        requested_by_sub=requested_by_sub,
        approved_manifest_id=approved_manifest_id,
    )


async def attach_monthly_operation_job(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    job_id: str,
) -> bool:
    return await persist_monthly_operation_job(
        pool,
        operation_id=operation_id,
        job_id=job_id,
    )


async def start_monthly_operation(
    pool: asyncpg.Pool,
    operation_id: int,
    *,
    execution_owner: str | None = None,
) -> MonthlyOperationStartResult:
    return await persist_monthly_operation_start(
        pool,
        operation_id,
        execution_owner=execution_owner,
    )


async def heartbeat_monthly_operation(
    pool: asyncpg.Pool,
    operation_id: int,
    *,
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    return await persist_monthly_operation_heartbeat(
        pool,
        operation_id,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
    )


async def _run_with_monthly_lease(
    pool: asyncpg.Pool,
    operation_id: int,
    *,
    execution_owner: str,
    execution_epoch: int,
    operation: Coroutine[Any, Any, MonthlyExecution],
    heartbeat_interval: float = MONTHLY_OPERATION_HEARTBEAT_SECONDS,
) -> MonthlyExecution:
    """Keep the DB lease alive and abort work immediately when its fence is lost."""

    async def monitor() -> None:
        while True:
            await asyncio.sleep(heartbeat_interval)
            alive = await heartbeat_monthly_operation(
                pool,
                operation_id,
                execution_owner=execution_owner,
                execution_epoch=execution_epoch,
            )
            if not alive:
                raise MonthlyIntegrityError(
                    "operation_lease_lost",
                    "Monthly operation lease was lost",
                )

    operation_task: asyncio.Task[MonthlyExecution] = asyncio.create_task(operation)
    heartbeat_task = asyncio.create_task(monitor())
    try:
        done, _ = await asyncio.wait(
            {operation_task, heartbeat_task},
            return_when=asyncio.FIRST_COMPLETED,
        )
        if operation_task in done:
            return await operation_task
        operation_task.cancel()
        await asyncio.gather(operation_task, return_exceptions=True)
        await heartbeat_task
        raise AssertionError("Lease monitor exited without a result")
    finally:
        heartbeat_task.cancel()
        await asyncio.gather(heartbeat_task, return_exceptions=True)


async def finish_monthly_operation(
    pool: asyncpg.Pool,
    operation_id: int,
    *,
    result: dict[str, Any],
    error_message: str | None = None,
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    return await persist_monthly_operation_result(
        pool,
        operation_id,
        result=result,
        error_message=error_message,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
    )


async def fail_monthly_operation(
    pool: asyncpg.Pool,
    operation_id: int,
    *,
    error_message: str,
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    return await persist_monthly_operation_failure(
        pool,
        operation_id,
        error_message=error_message,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
    )


async def fail_queued_monthly_operation(
    pool: asyncpg.Pool,
    operation_id: int,
    *,
    error_message: str,
) -> bool:
    return await persist_queued_monthly_operation_failure(
        pool,
        operation_id,
        error_message=error_message,
    )


async def get_monthly_execution_lease(
    pool: asyncpg.Pool,
    operation_id: int,
    *,
    execution_owner: str,
) -> MonthlyExecutionLease | None:
    return await fetch_monthly_execution_lease(
        pool,
        operation_id,
        execution_owner=execution_owner,
    )


async def mark_monthly_operation_cancelled_uncertain(
    pool: asyncpg.Pool,
    operation_id: int,
    *,
    error_message: str,
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    return await persist_cancelled_uncertain(
        pool,
        operation_id,
        error_message=error_message,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
    )


async def get_monthly_manifest(
    pool: asyncpg.Pool,
    manifest_id: int,
) -> dict[str, Any] | None:
    return await fetch_monthly_manifest(pool, manifest_id)


async def get_latest_monthly_manifest(
    pool: asyncpg.Pool,
    *,
    month: str,
) -> dict[str, Any] | None:
    return await fetch_latest_monthly_manifest(
        pool,
        closing_month=month,
        operation="archive",
        statuses=MANIFEST_ATTEMPT_STATUSES,
    )


async def ensure_reset_items(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    closing_month_key: str,
    next_month_key: str,
    entries: list[StoreEntry],
    execution_owner: str,
    execution_epoch: int,
) -> None:
    await persist_reset_items(
        pool,
        operation_id=operation_id,
        closing_month=closing_month_key,
        next_month=next_month_key,
        entries=[
            ResetItemInput(
                site_code=entry.site_code,
                sheet_id=entry.sheet_id,
                company=entry.company,
                store=entry.store,
                ranges=tuple(reset_ranges_for_entry(entry)),
            )
            for entry in entries
        ],
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
    )


async def get_previous_completed_reset_item(
    pool: asyncpg.Pool,
    *,
    closing_month_key: str,
    site_code: str,
) -> asyncpg.Record | None:
    return await fetch_previous_completed_reset_item(
        pool,
        closing_month=closing_month_key,
        site_code=site_code,
    )


async def mark_reset_item_running(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    site_code: str,
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    return await persist_reset_item_claim(
        pool,
        operation_id=operation_id,
        site_code=site_code,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
    )


async def finish_reset_item(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    site_code: str,
    status: Literal["completed", "error", "skipped"],
    execution_owner: str,
    execution_epoch: int,
    error_message: str | None = None,
) -> bool:
    return await persist_reset_item_result(
        pool,
        operation_id=operation_id,
        site_code=site_code,
        status=status,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
        error_message=error_message,
    )


async def record_reset_item_backup(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    site_code: str,
    backup_path: str,
    backup_sha256: str,
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    return await persist_reset_item_backup(
        pool,
        operation_id=operation_id,
        site_code=site_code,
        backup_path=backup_path,
        backup_sha256=backup_sha256,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
    )


async def record_reset_item_rollback(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    site_code: str,
    restored: bool,
    execution_owner: str,
    execution_epoch: int,
    error_message: str | None = None,
) -> bool:
    return await persist_reset_item_rollback(
        pool,
        operation_id=operation_id,
        site_code=site_code,
        restored=restored,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
        error_message=error_message,
    )


def validate_archive_manifest(manifest: dict[str, Any], expected_count: int) -> tuple[bool, list[str]]:
    errors: list[str] = []
    if manifest.get("registry_count") != expected_count:
        errors.append(f"registry_count mismatch: {manifest.get('registry_count')} != {expected_count}")
    if manifest.get("exported_count") != expected_count:
        errors.append(f"exported_count mismatch: {manifest.get('exported_count')} != {expected_count}")
    if manifest.get("error_count") != 0:
        errors.append(f"archive has {manifest.get('error_count')} export errors")

    stores = manifest.get("stores")
    if not isinstance(stores, list) or len(stores) != expected_count:
        count = len(stores) if isinstance(stores, list) else "invalid"
        errors.append(f"stores count mismatch: {count} != {expected_count}")
    else:
        for store in stores:
            company = store.get("company", "?")
            name = store.get("store", "?")
            if store.get("status") != "OK":
                errors.append(f"{company}/{name} status is {store.get('status')}")
            xlsx_path = Path(str(store.get("xlsx_path", "")))
            if not xlsx_path.exists() or xlsx_path.stat().st_size == 0:
                errors.append(f"missing or empty export: {xlsx_path}")

    zip_path = Path(str(manifest.get("zip_path", "")))
    if not zip_path.exists() or zip_path.stat().st_size == 0:
        errors.append(f"missing or empty archive zip: {zip_path}")
    return not errors, errors


def scalar(values: list[list[Any]]) -> Any:
    if not values or not values[0]:
        return ""
    return values[0][0]


def to_number(value: Any, *, field: str = "value") -> float:
    return float(parse_required_decimal(value, field=field))


def sum_scalars(value_ranges: list[dict[str, Any]], *, field: str = "value") -> float:
    return float(
        sum(
            (
                Decimal("0")
                if scalar(vr.get("values", [])) in (None, "")
                else parse_required_decimal(scalar(vr.get("values", [])), field=field)
                for vr in value_ranges
            ),
            start=Decimal("0"),
        )
    )


def _closed_empty_slot(entry: StoreEntry, agent: Any, slot_values: list[Any]) -> bool:
    """Accept template defaults only for an explicitly closed store with no work."""
    if not entry.is_closed or agent not in (None, ""):
        return False
    # Base salary and meal vouchers are template defaults even without an agent.
    # Any worked hours, commission or extra payment still requires an agent name.
    work_values = [*slot_values[2:9], slot_values[10]]
    return all(value in (None, "", 0, 0.0, False) for value in work_values)


def _error_row(
    entry: StoreEntry,
    *,
    slot: int,
    code: str,
) -> ExtractedAgentRow:
    return ExtractedAgentRow(
        site_code=entry.site_code,
        company=entry.company,
        store=entry.store,
        slot=slot,
        agent="",
        base_salary="",
        sales_commission="",
        extra_location_commission="",
        extra_hours_pay="",
        bonuri="",
        worked_hours="",
        status="ERROR",
        error_code=code,
        error=code,
        sheet_id=entry.sheet_id,
    )


def extract_store_rows(
    sheets_svc: Any,
    entry: StoreEntry,
    *,
    value_ranges: list[dict[str, Any]] | None = None,
) -> list[ExtractedAgentRow]:
    supplied_value_ranges = value_ranges

    def read_values() -> list[dict[str, Any]]:
        if supplied_value_ranges is not None:
            return supplied_value_ranges
        ranges = []
        for cells in cells_for_entry(entry).values():
            ranges += [
                f"Grila!{cells['agent']}",
                f"Grila!{cells['base_salary']}",
                *[f"Grila!{cell}" for cell in cells["sales_commission_cells"]],
                f"Grila!{cells['extra_location_commission']}",
                f"Grila!{cells['extra_hours_pay']}",
                f"Grila!{cells['bonuri']}",
                cells["worked_hours"],
            ]
        response = sheets_svc.spreadsheets().values().batchGet(
            spreadsheetId=entry.sheet_id,
            ranges=ranges,
            valueRenderOption="UNFORMATTED_VALUE",
        ).execute()
        value_ranges = response.get("valueRanges") if isinstance(response, dict) else None
        if not isinstance(value_ranges, list) or len(value_ranges) != len(ranges):
            raise MonthlyIntegrityError(
                "google_response_incomplete",
                "Google sheet response is incomplete",
            )
        return value_ranges

    try:
        value_ranges = retry_api(
            read_values,
            label="Google sheet read",
            attempts=GOOGLE_API_RETRY_ATTEMPTS,
            base_delay=GOOGLE_API_RETRY_BASE_DELAY_SECONDS,
        )
        rows: list[ExtractedAgentRow] = []
        idx = 0
        for slot in cells_for_entry(entry):
            slot_ranges = value_ranges[idx : idx + 11]
            idx += 11
            slot_values = [scalar(item.get("values", [])) for item in slot_ranges]
            agent = slot_values[0]
            # Template formulas can leave numeric zeroes in every salary cell
            # of an unused slot. Treat that slot as empty only when its agent
            # cell is blank and every remaining value is blank/zero. A blank
            # agent with meaningful data remains a fail-closed error.
            if agent in (None, "") and all(
                value in (None, "", 0, 0.0, False)
                for value in slot_values[1:]
            ):
                continue
            if _closed_empty_slot(entry, agent, slot_values):
                continue
            if (
                not isinstance(agent, str)
                or not agent.strip()
                or any(ord(char) < 32 or ord(char) == 127 for char in agent)
            ):
                rows.append(_error_row(entry, slot=slot, code="missing_or_invalid_agent"))
                continue
            try:
                base_salary = to_number(slot_values[1], field="base_salary")
                sales_commission = sum_scalars(slot_ranges[2:7], field="sales_commission")
                commission = to_number(slot_values[7], field="extra_location_commission")
                extra_hours = to_number(slot_values[8], field="extra_hours_pay")
                bonuri = to_number(slot_values[9], field="meal_vouchers")
                worked_hours = to_number(slot_values[10], field="worked_hours")
            except MonthlyIntegrityError as exc:
                rows.append(_error_row(entry, slot=slot, code=exc.code))
                continue
            rows.append(
                ExtractedAgentRow(
                    site_code=entry.site_code,
                    company=entry.company,
                    store=entry.store,
                    slot=slot,
                    agent=agent.strip(),
                    base_salary=base_salary,
                    sales_commission=sales_commission,
                    extra_location_commission=commission,
                    extra_hours_pay=extra_hours,
                    bonuri=bonuri,
                    worked_hours=worked_hours,
                    status="OK",
                    error_code="",
                    error="",
                    sheet_id=entry.sheet_id,
                )
            )
        if not rows and entry.is_closed:
            return []
        if not rows:
            return [_error_row(entry, slot=0, code="store_has_no_agent")]
        seen_agents: set[str] = set()
        deduplicated: list[ExtractedAgentRow] = []
        for row in rows:
            normalized = str(row.agent).strip().casefold()
            if row.status == "OK" and normalized in seen_agents:
                deduplicated.append(_error_row(entry, slot=row.slot, code="duplicate_agent"))
                continue
            if row.status == "OK":
                seen_agents.add(normalized)
            deduplicated.append(row)
        return deduplicated
    except Exception as exc:  # noqa: BLE001
        code = exc.code if isinstance(exc, MonthlyIntegrityError) else _google_error_code(exc)
        return [_error_row(entry, slot=0, code=code)]


def make_output_row(row: ExtractedAgentRow, nr: int, metadata: dict[str, Any]) -> list[Any]:
    excel_row = nr + 1
    return [
        nr,
        metadata.get("Manager", ""),
        row.store,
        row.agent,
        row.base_salary,
        row.sales_commission,
        metadata.get("Flip", ""),
        row.extra_location_commission,
        metadata.get("Incentive lunar", ""),
        row.extra_hours_pay,
        TrustedFormula(f"=SUM(E{excel_row}:J{excel_row},M{excel_row})"),
        TrustedFormula(f"=K{excel_row}-M{excel_row}"),
        row.bonuri,
        metadata.get("Data angajarii", ""),
        metadata.get("Data plecarii", ""),
        row.worked_hours,
        metadata.get("Zile CO luna in curs", ""),
    ]


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 8,
        "B": 22,
        "C": 30,
        "D": 30,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 26,
        "I": 16,
        "J": 18,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 16,
        "O": 16,
        "P": 16,
        "Q": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_workbook(
    rows: list[ExtractedAgentRow],
    output_path: Path,
    metadata_by_company_store: dict[tuple[str, str], dict[str, Any]],
) -> None:
    wb = Workbook()
    wb.active.title = "Mobiup"
    wb.create_sheet("Mobicell")
    ws_audit = wb.create_sheet("Audit")

    for ws in (wb["Mobiup"], wb["Mobicell"]):
        append_openpyxl_row(ws, HEADERS)

    counters = {"Mobiup": 1, "Mobicell": 1}
    for row in rows:
        if row.status != "OK":
            continue
        ws = wb[row.company]
        metadata = metadata_by_company_store.get((row.company, row.store), {})
        append_openpyxl_row(ws, make_output_row(row, counters[row.company], metadata))
        counters[row.company] += 1

    append_openpyxl_row(ws_audit, AUDIT_HEADERS)
    for row in rows:
        append_openpyxl_row(
            ws_audit,
            [
                row.company,
                row.store,
                row.slot,
                row.agent,
                row.sheet_id,
                row.sales_commission,
                row.extra_location_commission,
                row.extra_hours_pay,
                row.bonuri,
                row.worked_hours,
                f"https://docs.google.com/spreadsheets/d/{row.sheet_id}",
                row.status,
                row.error,
            ]
        )

    for ws in wb.worksheets:
        style_sheet(ws)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            col = get_column_letter(col_idx)
            if ws.column_dimensions[col].width is None:
                ws.column_dimensions[col].width = 14

    secure_directory(output_path.parent)
    wb.save(output_path)


def _validate_finalization_coverage(
    entries: list[StoreEntry],
    rows: list[ExtractedAgentRow],
) -> tuple[int, int, int, int, list[str]]:
    errors: list[str] = []
    entries_by_site: dict[str, StoreEntry] = {}
    sheet_ids: set[str] = set()
    for entry in entries:
        if entry.site_code in entries_by_site or entry.sheet_id in sheet_ids:
            errors.append("duplicate_registry_entry")
        entries_by_site[entry.site_code] = entry
        sheet_ids.add(entry.sheet_id)

    rows_by_site: dict[str, list[ExtractedAgentRow]] = {}
    for row in rows:
        expected = entries_by_site.get(row.site_code)
        if expected is None:
            errors.append("unexpected_store")
            continue
        if (
            row.sheet_id != expected.sheet_id
            or row.company != expected.company
            or row.store != expected.store
        ):
            errors.append("contradictory_store_metadata")
        rows_by_site.setdefault(row.site_code, []).append(row)

    expected_agents = 0
    processed_agents = 0
    processed_stores = 0
    for site_code in entries_by_site:
        store_rows = rows_by_site.get(site_code, [])
        if not store_rows and entries_by_site[site_code].is_closed:
            processed_stores += 1
            continue
        valid_slots = set(cells_for_entry(entries_by_site[site_code]))
        slot_rows = [row for row in store_rows if row.slot in valid_slots]
        expected_agents += len(slot_rows)
        valid_rows = [row for row in slot_rows if row.status == "OK"]
        processed_agents += len(valid_rows)
        store_errors = [row.error_code or "store_read_failed" for row in store_rows if row.status != "OK"]
        if not store_rows:
            store_errors.append("store_not_processed")
        if not slot_rows:
            store_errors.append("store_has_no_agent")
        if len({str(row.agent).strip().casefold() for row in valid_rows}) != len(valid_rows):
            store_errors.append("duplicate_agent")
        if store_errors:
            errors.extend(store_errors)
        else:
            processed_stores += 1

    return (
        len(entries_by_site),
        processed_stores,
        expected_agents,
        processed_agents,
        sorted(set(errors)),
    )


def _control_totals(rows: list[ExtractedAgentRow]) -> dict[str, str]:
    fields = (
        "base_salary",
        "sales_commission",
        "extra_location_commission",
        "extra_hours_pay",
        "bonuri",
        "worked_hours",
    )
    totals: dict[str, str] = {}
    valid_rows = [row for row in rows if row.status == "OK"]
    for field in fields:
        total = sum(
            (Decimal(str(getattr(row, field))) for row in valid_rows),
            start=Decimal("0"),
        )
        totals[field] = decimal_text(total)
    totals["salary_components"] = decimal_text(
        sum(
            (
                Decimal(str(row.base_salary))
                + Decimal(str(row.sales_commission))
                + Decimal(str(row.extra_location_commission))
                + Decimal(str(row.extra_hours_pay))
                + Decimal(str(row.bonuri))
                for row in valid_rows
            ),
            start=Decimal("0"),
        )
    )
    return totals


def _source_registry(entries: list[StoreEntry]) -> list[dict[str, str]]:
    return sorted(
        (
            {
                "site_code": entry.site_code,
                "sheet_id": entry.sheet_id,
                "template_version": entry.template_version,
            }
            for entry in entries
        ),
        key=lambda item: (item["site_code"], item["sheet_id"]),
    )


def _with_source_registry(
    manifest: dict[str, Any],
    entries: list[StoreEntry],
) -> dict[str, Any]:
    enriched = dict(manifest)
    enriched["source_registry"] = _source_registry(entries)
    return finalize_manifest(enriched)


def _validate_final_workbook(path: Path, *, expected_agents: int) -> None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if set(workbook.sheetnames) != {"Mobiup", "Mobicell", "Audit"}:
                raise MonthlyIntegrityError("workbook_structure_invalid", "Workbook structure is invalid")
            agent_rows = sum(
                max(workbook[company].max_row - 1, 0)
                for company in ("Mobiup", "Mobicell")
            )
            if agent_rows != expected_agents:
                raise MonthlyIntegrityError("workbook_coverage_incomplete", "Workbook coverage is incomplete")
            audit = workbook["Audit"]
            statuses = [row[11].value for row in audit.iter_rows(min_row=2) if len(row) >= 12]
            if len(statuses) != expected_agents or any(status != "OK" for status in statuses):
                raise MonthlyIntegrityError("workbook_audit_invalid", "Workbook audit is invalid")
        finally:
            workbook.close()
    except MonthlyIntegrityError:
        raise
    except Exception as exc:
        raise MonthlyIntegrityError("workbook_invalid", "Workbook cannot be verified") from exc


def _staging_dir(operation: str, operation_id: int | None) -> Path:
    suffix = str(operation_id) if operation_id is not None else "direct"
    path = OUTPUTS_DIR / ".staging" / f"{operation}-{suffix}"
    if path.exists():
        shutil.rmtree(path)
    secure_directory(path)
    return path


def _promote_file(staged: Path, destination: Path) -> None:
    secure_directory(destination.parent)
    revision: Path | None = None
    if destination.exists():
        revision_dir = OUTPUTS_DIR / ".revisions"
        secure_directory(revision_dir)
        revision = revision_dir / f"{destination.name}.{file_sha256(destination)[:16]}"
        if revision.exists():
            destination.unlink()
        else:
            os.replace(destination, revision)
    try:
        os.replace(staged, destination)
        secure_file(destination)
    except Exception:
        if revision is not None and revision.exists() and not destination.exists():
            os.replace(revision, destination)
        raise


async def _finalize_month_execution(
    pool: asyncpg.Pool,
    month: str,
    *,
    month_key: str,
    requested_by_sub: str,
    operation_id: int | None,
    only: str | None = None,
    delay: float = 1.1,
    google_adapter: GoogleSyncAdapter | None = None,
) -> MonthlyExecution:
    entries = await load_entries(pool, only=only, month=month_key)
    metadata = {(e.company, e.store): {"Manager": e.manager} for e in entries}
    sheets_svc = None
    if google_adapter is None:
        sheets_svc, _ = build_google_services()
    all_rows: list[ExtractedAgentRow] = []
    for idx, entry in enumerate(entries, start=1):
        if google_adapter is None:
            assert sheets_svc is not None
            all_rows.extend(extract_store_rows(sheets_svc, entry))
        else:
            ranges: list[str] = []
            for cells in cells_for_entry(entry).values():
                ranges += [
                    f"Grila!{cells['agent']}",
                    f"Grila!{cells['base_salary']}",
                    *[f"Grila!{cell}" for cell in cells["sales_commission_cells"]],
                    f"Grila!{cells['extra_location_commission']}",
                    f"Grila!{cells['extra_hours_pay']}",
                    f"Grila!{cells['bonuri']}",
                    cells["worked_hours"],
                ]
            response = await _google_request(
                google_adapter,
                "read_values",
                {
                    "spreadsheet_id": entry.sheet_id,
                    "ranges": ranges,
                    "value_render_option": "UNFORMATTED_VALUE",
                },
                label="Google sheet read",
            )
            value_ranges = response.get("valueRanges") if isinstance(response, dict) else None
            if not isinstance(value_ranges, list) or len(value_ranges) != len(ranges):
                raise MonthlyIntegrityError(
                    "google_response_incomplete",
                    "Google sheet response is incomplete",
                )
            all_rows.extend(extract_store_rows(None, entry, value_ranges=value_ranges))
        if delay > 0 and idx < len(entries):
            await asyncio.sleep(delay)

    expected_stores, processed_stores, expected_agents, processed_agents, errors = (
        _validate_finalization_coverage(entries, all_rows)
    )
    totals = _control_totals(all_rows)
    if errors or processed_stores != expected_stores or processed_agents != expected_agents:
        failed = base_manifest(
            month=month_key,
            operation="finalize",
            requested_by_sub=requested_by_sub,
            expected_stores=expected_stores,
            expected_agents=expected_agents,
            processed_stores=processed_stores,
            processed_agents=processed_agents,
            control_totals=totals,
            artifacts=[],
            errors=errors or ["coverage_incomplete"],
            status="failed",
        )
        raise MonthlyManifestError("finalization_incomplete", "Finalization coverage is incomplete", failed)

    stage_dir = _staging_dir("finalize", operation_id)
    staged_path = stage_dir / "candidate.xlsx"
    output_path = resolve_output_path(month, only, OUTPUTS_DIR)
    try:
        build_workbook(all_rows, staged_path, metadata)
        secure_file(staged_path)
        _validate_final_workbook(staged_path, expected_agents=expected_agents)
        _promote_file(staged_path, output_path)
    except Exception as exc:
        code = exc.code if isinstance(exc, MonthlyIntegrityError) else "workbook_promotion_failed"
        failed = base_manifest(
            month=month_key,
            operation="finalize",
            requested_by_sub=requested_by_sub,
            expected_stores=expected_stores,
            expected_agents=expected_agents,
            processed_stores=processed_stores,
            processed_agents=processed_agents,
            control_totals=totals,
            artifacts=[],
            errors=[code],
            status="failed",
        )
        raise MonthlyManifestError(code, "Final workbook could not be verified", failed) from exc
    finally:
        shutil.rmtree(stage_dir, ignore_errors=True)

    artifact = relative_artifact(output_path, root=OUTPUTS_DIR, kind="final_workbook")
    manifest = _with_source_registry(
        base_manifest(
            month=month_key,
            operation="finalize",
            requested_by_sub=requested_by_sub,
            expected_stores=expected_stores,
            expected_agents=expected_agents,
            processed_stores=processed_stores,
            processed_agents=processed_agents,
            control_totals=totals,
            artifacts=[artifact],
        ),
        entries,
    )
    validate_verified_manifest(manifest, operation="finalize")
    verify_artifacts(manifest, root=OUTPUTS_DIR)
    return MonthlyExecution(path=output_path, manifest=manifest)


async def finalize_month(
    pool: asyncpg.Pool,
    month: str,
    only: str | None = None,
    delay: float = 1.1,
    *,
    month_key: str | None = None,
    requested_by_sub: str = "direct-execution",
    operation_id: int | None = None,
    google_adapter: GoogleSyncAdapter | None = None,
) -> Path:
    execution = await _finalize_month_execution(
        pool,
        month,
        month_key=month_key or month,
        requested_by_sub=requested_by_sub,
        operation_id=operation_id,
        only=only,
        delay=delay,
        google_adapter=google_adapter,
    )
    return execution.path


def export_sheet_xlsx(drive_service: Any, entry: StoreEntry, output_path: Path) -> dict[str, Any]:
    result = {
        "company": entry.company,
        "store": entry.store,
        "site_code": entry.site_code,
        "manager": entry.manager,
        "sheet_id": entry.sheet_id,
        "template_version": entry.template_version,
        "status": "OK",
        "xlsx_path": str(output_path),
        "bytes": 0,
        "error": "",
    }
    secure_directory(output_path.parent)
    request = drive_service.files().export_media(fileId=entry.sheet_id, mimeType=XLSX_MIME)
    try:
        with output_path.open("wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        secure_file(output_path)
        result["bytes"] = output_path.stat().st_size
        if result["bytes"] == 0:
            raise MonthlyIntegrityError("empty_source_backup", "Exported source backup is empty")
        return result
    except Exception:
        output_path.unlink(missing_ok=True)
        raise


def write_exported_xlsx(entry: StoreEntry, output_path: Path, content: bytes) -> dict[str, Any]:
    if not isinstance(content, bytes) or not content:
        raise MonthlyIntegrityError("empty_source_backup", "Exported source backup is empty")
    secure_directory(output_path.parent)
    try:
        output_path.write_bytes(content)
        secure_file(output_path)
        if output_path.stat().st_size == 0:
            raise MonthlyIntegrityError("empty_source_backup", "Exported source backup is empty")
    except Exception:
        output_path.unlink(missing_ok=True)
        raise
    return {
        "company": entry.company,
        "store": entry.store,
        "site_code": entry.site_code,
        "manager": entry.manager,
        "sheet_id": entry.sheet_id,
        "template_version": entry.template_version,
        "status": "OK",
        "xlsx_path": str(output_path),
        "bytes": output_path.stat().st_size,
        "error": "",
    }


def create_archive_zip(zip_path: Path, exported_files: list[Path], archive_dir: Path) -> None:
    secure_directory(zip_path.parent)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in exported_files:
            zf.write(path, path.relative_to(archive_dir).as_posix())


def create_manager_zips(output_dir: Path, month: str, results: list[dict[str, Any]]) -> dict[str, Path]:
    archive_dir = build_archive_dir(output_dir, month)
    files_by_manager: dict[str, list[Path]] = {}
    for item in results:
        if item.get("status") != "OK":
            continue
        files_by_manager.setdefault(item.get("manager") or "Neatribuit", []).append(Path(item["xlsx_path"]))

    zip_paths: dict[str, Path] = {}
    for manager, files in sorted(files_by_manager.items()):
        zip_path = build_manager_zip_path(output_dir, month, manager)
        create_archive_zip(zip_path, files, archive_dir)
        zip_paths[manager] = zip_path
    return zip_paths


def summarize_archive_results(
    month: str,
    registry_count: int,
    results: list[dict[str, Any]],
    zip_path: Path,
    manager_zip_paths: dict[str, Path] | None = None,
) -> dict[str, Any]:
    return {
        "month": month,
        "created_at": utc_now(),
        "registry_count": registry_count,
        "exported_count": sum(1 for item in results if item.get("status") == "OK"),
        "error_count": sum(1 for item in results if item.get("status") != "OK"),
        "zip_path": str(zip_path),
        "manager_zip_paths": {manager: str(path) for manager, path in sorted((manager_zip_paths or {}).items())},
        "stores": results,
    }


def _validate_archive_zip(zip_path: Path, *, expected_files: int) -> None:
    try:
        with zipfile.ZipFile(zip_path) as archive:
            members = [item for item in archive.infolist() if not item.is_dir()]
            if len(members) != expected_files or len({item.filename for item in members}) != expected_files:
                raise MonthlyIntegrityError("archive_coverage_incomplete", "Archive coverage is incomplete")
            if archive.testzip() is not None:
                raise MonthlyIntegrityError("archive_corrupt", "Archive is corrupt")
    except MonthlyIntegrityError:
        raise
    except Exception as exc:
        raise MonthlyIntegrityError("archive_invalid", "Archive cannot be verified") from exc


def _validate_source_workbook(path: Path) -> None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if not {"Grila", "Pontaj"}.issubset(workbook.sheetnames):
                raise MonthlyIntegrityError(
                    "source_workbook_partial",
                    "Source workbook is missing required sheets",
                )
        finally:
            workbook.close()
    except MonthlyIntegrityError:
        raise
    except Exception as exc:
        raise MonthlyIntegrityError("source_workbook_invalid", "Source workbook is invalid") from exc


def _future_artifact(
    staged_path: Path,
    *,
    staged_archive_dir: Path,
    official_archive_dir: Path,
    kind: str,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    relative_inside = staged_path.resolve().relative_to(staged_archive_dir.resolve())
    future_path = official_archive_dir / relative_inside
    artifact = {
        "kind": kind,
        "path": future_path.resolve().relative_to(OUTPUTS_DIR.resolve()).as_posix(),
        "bytes": staged_path.stat().st_size,
        "sha256": file_sha256(staged_path),
    }
    if extra:
        artifact.update(extra)
    return artifact


def _promote_directory(
    staged: Path,
    destination: Path,
    *,
    verify: Callable[[], None] | None = None,
) -> None:
    secure_directory(destination.parent)
    revision: Path | None = None
    promoted = False
    if destination.exists():
        revision_dir = OUTPUTS_DIR / ".revisions"
        secure_directory(revision_dir)
        revision = revision_dir / f"archive-{safe_filename(destination.name)}-{time.time_ns()}"
        os.replace(destination, revision)
    try:
        os.replace(staged, destination)
        promoted = True
        if verify is not None:
            verify()
    except Exception as exc:
        rollback_error: Exception | None = None
        if promoted and destination.exists():
            try:
                os.replace(destination, staged)
            except Exception:  # noqa: BLE001 - remove unverified output
                try:
                    shutil.rmtree(destination)
                except Exception as remove_error:  # noqa: BLE001 - surfaced below
                    rollback_error = remove_error
        if revision is not None and revision.exists() and not destination.exists():
            try:
                os.replace(revision, destination)
            except Exception as restore_error:  # noqa: BLE001 - surfaced below
                rollback_error = restore_error
        if rollback_error is not None:
            raise MonthlyIntegrityError(
                "archive_promotion_rollback_failed",
                "Archive promotion rollback failed",
            ) from exc
        raise


async def _archive_month_execution(
    pool: asyncpg.Pool,
    month: str,
    *,
    month_key: str,
    requested_by_sub: str,
    operation_id: int | None,
    only: str | None = None,
    delay: float = 0.5,
    google_adapter: GoogleSyncAdapter | None = None,
) -> MonthlyExecution:
    if only:
        failed = base_manifest(
            month=month_key,
            operation="archive",
            requested_by_sub=requested_by_sub,
            expected_stores=0,
            expected_agents=0,
            processed_stores=0,
            processed_agents=0,
            control_totals={},
            artifacts=[],
            errors=["partial_archive_forbidden"],
            status="failed",
        )
        raise MonthlyManifestError("partial_archive_forbidden", "Partial archive is not allowed", failed)

    final_record = await fetch_latest_monthly_manifest(
        pool,
        closing_month=month_key,
        operation="finalize",
        statuses=MANIFEST_ATTEMPT_STATUSES,
    )
    final_manifest = final_record.get("manifest") if final_record else None
    if (
        final_record is None
        or final_record.get("status") != "verified"
        or not isinstance(final_manifest, dict)
    ):
        failed = base_manifest(
            month=month_key,
            operation="archive",
            requested_by_sub=requested_by_sub,
            expected_stores=0,
            expected_agents=0,
            processed_stores=0,
            processed_agents=0,
            control_totals={},
            artifacts=[],
            errors=["verified_finalization_missing"],
            status="failed",
        )
        raise MonthlyManifestError("verified_finalization_missing", "Verified finalization is required", failed)
    validate_verified_manifest(final_manifest, operation="finalize")
    verify_artifacts(final_manifest, root=OUTPUTS_DIR)

    entries = await load_entries(pool, month=month_key)
    expected = final_manifest["expected"]
    finalized_registry = final_manifest.get("source_registry")
    current_registry = _source_registry(entries)
    if (
        len(entries) != expected["stores"]
        or len({entry.site_code for entry in entries}) != len(entries)
        or len({entry.sheet_id for entry in entries}) != len(entries)
        or finalized_registry != current_registry
    ):
        failed = base_manifest(
            month=month_key,
            operation="archive",
            requested_by_sub=requested_by_sub,
            expected_stores=int(expected["stores"]),
            expected_agents=int(expected["agents"]),
            processed_stores=0,
            processed_agents=0,
            control_totals=final_manifest.get("control_totals", {}),
            artifacts=[],
            errors=["registry_changed_or_duplicate_after_finalization"],
            status="failed",
        )
        raise MonthlyManifestError(
            "registry_changed_or_duplicate_after_finalization",
            "Registry changed after finalization",
            failed,
        )

    drive_service = None
    if google_adapter is None:
        _, drive_service = build_google_services()
    stage_root = _staging_dir("archive", operation_id)
    staged_archive_dir = build_archive_dir(stage_root, month)
    official_archive_dir = build_archive_dir(OUTPUTS_DIR, month)
    results: list[dict[str, Any]] = []
    exported_files: list[Path] = []
    errors: list[str] = []
    try:
        for idx, entry in enumerate(entries, start=1):
            output_path = build_store_export_path(stage_root, month, entry)
            try:
                if google_adapter is None:
                    assert drive_service is not None
                    result = retry_api(
                        lambda entry=entry, output_path=output_path: export_sheet_xlsx(
                            drive_service,
                            entry,
                            output_path,
                        ),
                        label="Google source export",
                        attempts=GOOGLE_API_RETRY_ATTEMPTS,
                        base_delay=GOOGLE_API_RETRY_BASE_DELAY_SECONDS,
                    )
                else:
                    content = await _google_request(
                        google_adapter,
                        "export_xlsx",
                        {"spreadsheet_id": entry.sheet_id, "mime_type": XLSX_MIME},
                        label="Google source export",
                    )
                    result = write_exported_xlsx(entry, output_path, content)
                _validate_source_workbook(Path(result["xlsx_path"]))
            except MonthlyIntegrityError as exc:
                result = {
                    "company": entry.company,
                    "store": entry.store,
                    "site_code": entry.site_code,
                    "manager": entry.manager,
                    "sheet_id": entry.sheet_id,
                    "template_version": entry.template_version,
                    "status": "ERROR",
                    "xlsx_path": "",
                    "bytes": 0,
                    "error": exc.code,
                }
                errors.append(exc.code)
            results.append(result)
            if result["status"] == "OK":
                exported_files.append(Path(result["xlsx_path"]))
            if delay > 0 and idx < len(entries):
                await asyncio.sleep(delay)

        if errors or len(exported_files) != len(entries):
            failed = base_manifest(
                month=month_key,
                operation="archive",
                requested_by_sub=requested_by_sub,
                expected_stores=len(entries),
                expected_agents=int(expected["agents"]),
                processed_stores=len(exported_files),
                processed_agents=int(expected["agents"]) if not errors else 0,
                control_totals=final_manifest.get("control_totals", {}),
                artifacts=[],
                errors=errors or ["archive_coverage_incomplete"],
                status="failed",
            )
            raise MonthlyManifestError("archive_incomplete", "Archive is incomplete", failed)

        zip_path = build_archive_zip_path(stage_root, month)
        create_archive_zip(zip_path, exported_files, staged_archive_dir)
        secure_file(zip_path)
        _validate_archive_zip(zip_path, expected_files=len(entries))
        manager_zip_paths = create_manager_zips(stage_root, month, results)
        for path in manager_zip_paths.values():
            secure_file(path)

        source_backups = [
            _future_artifact(
                Path(result["xlsx_path"]),
                staged_archive_dir=staged_archive_dir,
                official_archive_dir=official_archive_dir,
                kind="source_workbook",
                extra={
                    "site_code": result["site_code"],
                    "sheet_id": result["sheet_id"],
                    "template_version": result.get("template_version", "v2"),
                },
            )
            for result in results
        ]
        archive_artifacts = [
            _future_artifact(
                zip_path,
                staged_archive_dir=staged_archive_dir,
                official_archive_dir=official_archive_dir,
                kind="archive_zip",
            ),
            *source_backups,
            *[
                _future_artifact(
                    path,
                    staged_archive_dir=staged_archive_dir,
                    official_archive_dir=official_archive_dir,
                    kind="manager_archive_zip",
                )
                for path in manager_zip_paths.values()
            ],
            *[dict(item) for item in final_manifest["artifacts"]],
        ]
        manifest = base_manifest(
            month=month_key,
            operation="archive",
            requested_by_sub=requested_by_sub,
            expected_stores=len(entries),
            expected_agents=int(expected["agents"]),
            processed_stores=len(entries),
            processed_agents=int(expected["agents"]),
            control_totals=final_manifest.get("control_totals", {}),
            artifacts=archive_artifacts,
            source_backups=source_backups,
        )
        validate_verified_manifest(manifest, operation="archive")
        manifest_path = build_archive_manifest_path(stage_root, month)
        secure_write_json(manifest_path, manifest)
        official_manifest_path = build_archive_manifest_path(OUTPUTS_DIR, month)
        _promote_directory(
            staged_archive_dir,
            official_archive_dir,
            verify=lambda: verify_artifacts(manifest, root=OUTPUTS_DIR),
        )
        return MonthlyExecution(path=official_manifest_path, manifest=manifest)
    finally:
        shutil.rmtree(stage_root, ignore_errors=True)


async def archive_month(
    pool: asyncpg.Pool,
    month: str,
    only: str | None = None,
    delay: float = 0.5,
    *,
    month_key: str | None = None,
    requested_by_sub: str = "direct-execution",
    operation_id: int | None = None,
    google_adapter: GoogleSyncAdapter | None = None,
) -> Path:
    execution = await _archive_month_execution(
        pool,
        month,
        month_key=month_key or month,
        requested_by_sub=requested_by_sub,
        operation_id=operation_id,
        only=only,
        delay=delay,
        google_adapter=google_adapter,
    )
    return execution.path


def public_manifest_payload(record: dict[str, Any]) -> dict[str, Any]:
    raw_manifest = record.get("manifest")
    manifest: dict[str, Any] = raw_manifest if isinstance(raw_manifest, dict) else {}
    return {
        "id": record.get("id"),
        "operation_id": record.get("operation_id"),
        "month": record.get("closing_month"),
        "operation": record.get("operation"),
        "status": record.get("status"),
        "expected": manifest.get("expected", {}),
        "processed": manifest.get("processed", {}),
        "error_count": record.get("error_count", 0),
        "manifest_sha256": record.get("manifest_sha256"),
        "approved": bool(record.get("approved_by_sub")),
        "created_at": _public_timestamp(record.get("created_at")),
        "verified_at": _public_timestamp(record.get("verified_at")),
        "approved_at": _public_timestamp(record.get("approved_at")),
        "consumed_at": _public_timestamp(record.get("consumed_at")),
    }


def _public_timestamp(value: Any) -> Any:
    return value.isoformat() if isinstance(value, datetime) else value


async def approve_monthly_manifest(
    pool: asyncpg.Pool,
    *,
    manifest_id: int,
    approved_by_sub: str,
) -> dict[str, Any]:
    record = await fetch_monthly_manifest(pool, manifest_id)
    if record is None:
        raise FileNotFoundError("Manifestul nu exista.")
    manifest = record.get("manifest")
    if record.get("operation") != "archive" or record.get("status") != "verified":
        raise MonthlyIntegrityError("manifest_not_approvable", "Manifest is not approvable")
    if not isinstance(manifest, dict):
        raise MonthlyIntegrityError("manifest_invalid", "Manifest is invalid")
    validate_verified_manifest(manifest, operation="archive")
    verify_artifacts(manifest, root=OUTPUTS_DIR)
    current_sha = manifest.get("manifest_sha256")
    if not isinstance(current_sha, str):
        raise MonthlyIntegrityError("manifest_hash_invalid", "Manifest hash is invalid")
    approved_manifest = dict(manifest)
    approved_manifest["status"] = "approved"
    approved_manifest["approved_by_sub"] = approved_by_sub
    approved_manifest["approved_at"] = utc_now()
    approved_manifest = finalize_manifest(approved_manifest)
    approved = await persist_monthly_manifest_approval(
        pool,
        manifest_id=manifest_id,
        expected_sha256=current_sha,
        approved_by_sub=approved_by_sub,
        approved_manifest=approved_manifest,
    )
    if approved is None:
        current = await fetch_monthly_manifest(pool, manifest_id)
        if current is not None and current.get("status") in {"approved", "consumed"}:
            return public_manifest_payload(current)
        raise MonthlyIntegrityError("manifest_approval_race", "Manifest approval changed concurrently")
    return public_manifest_payload(approved)


def _read_reset_snapshot(sheets_svc: Any, entry: StoreEntry) -> dict[str, Any]:
    reset_ranges = reset_ranges_for_entry(entry)

    def read() -> dict[str, Any]:
        response = sheets_svc.spreadsheets().values().batchGet(
            spreadsheetId=entry.sheet_id,
            ranges=reset_ranges,
            valueRenderOption="FORMULA",
            dateTimeRenderOption="SERIAL_NUMBER",
        ).execute()
        value_ranges = response.get("valueRanges") if isinstance(response, dict) else None
        if not isinstance(value_ranges, list) or len(value_ranges) != len(reset_ranges):
            raise MonthlyIntegrityError(
                "backup_response_incomplete",
                "Google backup response is incomplete",
            )
        return canonical_snapshot(value_ranges)

    return retry_api(
        read,
        label="Google reset backup",
        attempts=GOOGLE_API_RETRY_ATTEMPTS,
        base_delay=GOOGLE_API_RETRY_BASE_DELAY_SECONDS,
    )


async def _read_reset_snapshot_async(
    google_adapter: GoogleSyncAdapter,
    entry: StoreEntry,
) -> dict[str, Any]:
    response = await _google_request(
        google_adapter,
        "read_values",
        {
            "spreadsheet_id": entry.sheet_id,
            "ranges": reset_ranges_for_entry(entry),
            "value_render_option": "FORMULA",
            "date_time_render_option": "SERIAL_NUMBER",
        },
        label="Google reset readback",
    )
    value_ranges = response.get("valueRanges") if isinstance(response, dict) else None
    ranges = reset_ranges_for_entry(entry)
    if not isinstance(value_ranges, list) or len(value_ranges) != len(ranges):
        raise MonthlyIntegrityError("backup_response_incomplete", "Google backup response is incomplete")
    return canonical_snapshot(value_ranges)


def _restore_reset_snapshot(
    sheets_svc: Any,
    entry: StoreEntry,
    snapshot: dict[str, Any],
) -> None:
    value_ranges = snapshot.get("value_ranges")
    if not isinstance(value_ranges, list):
        raise MonthlyIntegrityError("backup_invalid", "Reset backup is invalid")
    data = [
        {
            "range": item["range"],
            "majorDimension": item.get("majorDimension", "ROWS"),
            "values": item.get("values", []),
        }
        for item in value_ranges
        if isinstance(item, dict) and item.get("values")
    ]

    def restore() -> dict[str, Any]:
        return sheets_svc.spreadsheets().values().batchUpdate(
            spreadsheetId=entry.sheet_id,
            body={"valueInputOption": "USER_ENTERED", "data": data},
        ).execute()

    if data:
        retry_api(
            restore,
            label="Google reset rollback",
            attempts=GOOGLE_API_RETRY_ATTEMPTS,
            base_delay=GOOGLE_API_RETRY_BASE_DELAY_SECONDS,
        )
    restored = _read_reset_snapshot(sheets_svc, entry)
    if snapshot_sha256(restored) != snapshot_sha256(snapshot):
        raise MonthlyIntegrityError("rollback_verification_failed", "Reset rollback verification failed")


async def _restore_reset_snapshot_async(
    google_adapter: GoogleSyncAdapter,
    entry: StoreEntry,
    snapshot: dict[str, Any],
) -> None:
    value_ranges = snapshot.get("value_ranges")
    if not isinstance(value_ranges, list):
        raise MonthlyIntegrityError("backup_invalid", "Reset backup is invalid")
    data = [
        {
            "range": item["range"],
            "majorDimension": item.get("majorDimension", "ROWS"),
            "values": item.get("values", []),
        }
        for item in value_ranges
        if isinstance(item, dict) and item.get("values")
    ]
    if data:
        await _google_request(
            google_adapter,
            "restore",
            {"spreadsheet_id": entry.sheet_id, "data": data},
            label="Google reset rollback",
            destructive=True,
        )
    restored = await _read_reset_snapshot_async(google_adapter, entry)
    if snapshot_sha256(restored) != snapshot_sha256(snapshot):
        raise MonthlyIntegrityError("rollback_verification_failed", "Reset rollback verification failed")


def _verify_reset_cleared(sheets_svc: Any, entry: StoreEntry) -> None:
    snapshot = _read_reset_snapshot(sheets_svc, entry)
    value_ranges = snapshot.get("value_ranges", [])
    if any(item.get("values") for item in value_ranges if isinstance(item, dict)):
        raise MonthlyIntegrityError("reset_verification_failed", "Reset verification failed")


async def _verify_reset_cleared_async(
    google_adapter: GoogleSyncAdapter,
    entry: StoreEntry,
) -> None:
    snapshot = await _read_reset_snapshot_async(google_adapter, entry)
    value_ranges = snapshot.get("value_ranges", [])
    if any(item.get("values") for item in value_ranges if isinstance(item, dict)):
        raise MonthlyIntegrityError("reset_verification_failed", "Reset verification failed")


def reset_store(sheets_svc: Any | None, entry: StoreEntry, *, dry_run: bool) -> dict[str, Any]:
    reset_ranges = reset_ranges_for_entry(entry)
    result = {
        "company": entry.company,
        "store": entry.store,
        "site_code": entry.site_code,
        "sheet_id": entry.sheet_id,
        "status": "DRY_RUN" if dry_run else "OK",
        "error": "",
        "ranges": reset_ranges,
    }
    if dry_run:
        return result

    assert sheets_svc is not None
    try:
        def clear() -> dict[str, Any]:
            return sheets_svc.spreadsheets().values().batchClear(
                spreadsheetId=entry.sheet_id,
                body={"ranges": reset_ranges},
            ).execute()

        retry_api(
            clear,
            label="Google reset",
            attempts=GOOGLE_API_RETRY_ATTEMPTS,
            base_delay=GOOGLE_API_RETRY_BASE_DELAY_SECONDS,
        )
        return result
    except Exception as exc:  # noqa: BLE001
        result["status"] = "ERROR"
        result["error"] = exc.code if isinstance(exc, MonthlyIntegrityError) else _google_error_code(exc)
        return result


async def _rollback_reset_entries(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    entries: list[StoreEntry],
    sheets_svc: Any,
    snapshots: dict[str, dict[str, Any]],
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    rollback_failed = False
    for entry in reversed(entries):
        try:
            _restore_reset_snapshot(sheets_svc, entry, snapshots[entry.site_code])
            restored = True
        except Exception:  # noqa: BLE001 - checkpoint below records the uncertainty
            restored = False
            rollback_failed = True
        try:
            rollback_recorded = await record_reset_item_rollback(
                pool,
                operation_id=operation_id,
                site_code=entry.site_code,
                restored=restored,
                execution_owner=execution_owner,
                execution_epoch=execution_epoch,
                error_message="reset_rolled_back" if restored else "reset_rollback_failed",
            )
        except Exception:  # noqa: BLE001 - Google restoration still took precedence
            rollback_recorded = False
        if not rollback_recorded:
            rollback_failed = True
    return not rollback_failed


async def _rollback_reset_entries_cancel_safe(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    entries: list[StoreEntry],
    sheets_svc: Any,
    snapshots: dict[str, dict[str, Any]],
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    """Finish rollback after task cancellation; any second interruption is uncertain."""
    current = asyncio.current_task()
    if current is not None:
        while current.cancelling():
            current.uncancel()
    rollback_task = asyncio.create_task(
        _rollback_reset_entries(
            pool,
            operation_id=operation_id,
            entries=entries,
            sheets_svc=sheets_svc,
            snapshots=snapshots,
            execution_owner=execution_owner,
            execution_epoch=execution_epoch,
        )
    )
    try:
        return await asyncio.shield(rollback_task)
    except BaseException:  # cancellation or provider/DB failure leaves an uncertain checkpoint
        return False


async def _rollback_reset_entries_adapter(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    entries: list[StoreEntry],
    google_adapter: GoogleSyncAdapter,
    snapshots: dict[str, dict[str, Any]],
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    rollback_failed = False
    for entry in reversed(entries):
        intent = await persist_reset_rollback_intent(
            pool,
            operation_id=operation_id,
            site_code=entry.site_code,
            execution_owner=execution_owner,
            execution_epoch=execution_epoch,
        )
        if intent is None:
            rollback_failed = True
            continue
        restored = False
        try:
            await _restore_reset_snapshot_async(
                google_adapter,
                entry,
                snapshots[entry.site_code],
            )
            restored = True
        except BaseException:  # noqa: BLE001 - the fenced checkpoint records uncertainty
            rollback_failed = True
        confirmed = await persist_reset_rollback_confirmation(
            pool,
            operation_id=operation_id,
            site_code=entry.site_code,
            execution_owner=execution_owner,
            execution_epoch=execution_epoch,
            fence_epoch=int(intent["fence_epoch"]),
            restored=restored,
            error_message="reset_rolled_back" if restored else "reset_rollback_failed",
        )
        if not confirmed:
            rollback_failed = True
    return not rollback_failed


async def _rollback_reset_entries_adapter_cancel_safe(
    pool: asyncpg.Pool,
    *,
    operation_id: int,
    entries: list[StoreEntry],
    google_adapter: GoogleSyncAdapter,
    snapshots: dict[str, dict[str, Any]],
    execution_owner: str,
    execution_epoch: int,
) -> bool:
    current = asyncio.current_task()
    if current is not None:
        while current.cancelling():
            current.uncancel()
    task = asyncio.create_task(
        _rollback_reset_entries_adapter(
            pool,
            operation_id=operation_id,
            entries=entries,
            google_adapter=google_adapter,
            snapshots=snapshots,
            execution_owner=execution_owner,
            execution_epoch=execution_epoch,
        )
    )
    try:
        return await asyncio.shield(task)
    except BaseException:
        return False


@dataclass(slots=True)
class ResetRunContext:
    pool: asyncpg.Pool
    closing_month: str
    next_month: str
    closing_month_key: str
    next_month_key: str
    requested_by_sub: str
    operation_id: int | None
    approved_manifest_id: int | None
    only: str | None
    dry_run: bool
    google_adapter: GoogleSyncAdapter | None
    execution_owner: str | None
    execution_epoch: int | None


def _reset_manifest(
    context: ResetRunContext,
    *,
    expected: dict[str, Any] | None = None,
    archive_manifest: dict[str, Any] | None = None,
    processed_stores: int = 0,
    source_backups: list[dict[str, Any]] | None = None,
    errors: list[str],
    status: str = "failed",
    artifacts: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    expected = expected or {"stores": 0, "agents": 0}
    return base_manifest(
        month=context.closing_month_key,
        operation="reset",
        requested_by_sub=context.requested_by_sub,
        expected_stores=int(expected["stores"]),
        expected_agents=int(expected["agents"]),
        processed_stores=processed_stores,
        processed_agents=(
            int(expected["agents"]) if status == "verified" else 0
        ),
        control_totals=(archive_manifest or {}).get("control_totals", {}),
        artifacts=artifacts or [],
        source_backups=source_backups or [],
        errors=errors,
        status=status,
    )


def _validate_reset_request(context: ResetRunContext) -> None:
    if not context.dry_run and (
        context.operation_id is None or context.approved_manifest_id is None
    ):
        manifest = _reset_manifest(
            context, errors=["approved_manifest_required"]
        )
        raise MonthlyManifestError(
            "approved_manifest_required", "Approved manifest is required", manifest
        )
    if context.only and not context.dry_run:
        manifest = _reset_manifest(
            context, errors=["partial_live_reset_forbidden"]
        )
        raise MonthlyManifestError(
            "partial_live_reset_forbidden",
            "Partial live reset is forbidden",
            manifest,
        )


async def _load_reset_archive(context: ResetRunContext) -> dict[str, Any]:
    latest = await fetch_latest_monthly_manifest(
        context.pool,
        closing_month=context.closing_month_key,
        operation="archive",
        statuses=MANIFEST_ATTEMPT_STATUSES,
    )
    if context.approved_manifest_id is not None:
        prerequisite = await fetch_monthly_manifest(
            context.pool, context.approved_manifest_id
        )
        if (
            latest is None
            or prerequisite is None
            or latest.get("id") != prerequisite.get("id")
        ):
            prerequisite = None
    else:
        prerequisite = latest
    archive = prerequisite.get("manifest") if prerequisite else None
    allowed = {"approved"} if not context.dry_run else {"verified", "approved"}
    if (
        prerequisite is None
        or prerequisite.get("status") not in allowed
        or not isinstance(archive, dict)
    ):
        manifest = _reset_manifest(
            context, errors=["verified_archive_required"]
        )
        raise MonthlyManifestError(
            "verified_archive_required", "Verified archive is required", manifest
        )
    validate_verified_manifest(archive, operation="archive")
    verify_artifacts(archive, root=OUTPUTS_DIR)
    if not context.dry_run and (
        context.execution_owner is None or context.execution_epoch is None
    ):
        raise MonthlyIntegrityError(
            "operation_lease_missing", "Reset operation lease is missing"
        )
    return archive


async def _load_reset_entries(
    context: ResetRunContext, archive: dict[str, Any]
) -> tuple[list[StoreEntry], dict[str, Any]]:
    entries = await load_entries(
        context.pool, only=context.only, month=context.closing_month_key
    )
    expected = archive["expected"]
    source_backups = archive.get("source_backups")
    archived_by_site = (
        {
            item.get("site_code"): item
            for item in source_backups
            if isinstance(item, dict)
            and isinstance(item.get("site_code"), str)
        }
        if isinstance(source_backups, list)
        else {}
    )
    current_sites = {entry.site_code for entry in entries}
    valid = (
        len(entries) == int(expected["stores"])
        and set(archived_by_site) == current_sites
        and len(current_sites) == len(entries)
        and isinstance(source_backups, list)
        and len(source_backups) == len(entries)
        and all(
            archived_by_site[entry.site_code].get("sheet_id") == entry.sheet_id
            and archived_by_site[entry.site_code].get("template_version", "v2")
            == entry.template_version
            for entry in entries
        )
    )
    if not valid:
        manifest = _reset_manifest(
            context,
            expected=expected,
            archive_manifest=archive,
            errors=["registry_or_archive_coverage_changed"],
        )
        raise MonthlyManifestError(
            "registry_or_archive_coverage_changed",
            "Registry or archive coverage changed before reset",
            manifest,
        )
    return entries, expected


async def _prepare_reset_execution(
    context: ResetRunContext, entries: list[StoreEntry]
) -> Any | None:
    sheets_service = None
    if context.google_adapter is None:
        sheets_service, _ = build_google_services()
    if context.operation_id is not None and not context.dry_run:
        assert context.execution_owner is not None
        assert context.execution_epoch is not None
        await ensure_reset_items(
            context.pool,
            operation_id=context.operation_id,
            closing_month_key=context.closing_month_key,
            next_month_key=context.next_month_key,
            entries=entries,
            execution_owner=context.execution_owner,
            execution_epoch=context.execution_epoch,
        )
    return sheets_service


async def _read_reset_entry_snapshot(
    context: ResetRunContext, sheets_service: Any, entry: StoreEntry
) -> dict[str, Any]:
    if context.google_adapter is None:
        return _read_reset_snapshot(sheets_service, entry)
    return await _read_reset_snapshot_async(context.google_adapter, entry)


async def _persist_reset_backup(
    context: ResetRunContext,
    *,
    entry: StoreEntry,
    snapshot: dict[str, Any],
    backup_dir: Path,
) -> dict[str, Any]:
    assert context.operation_id is not None
    assert context.execution_owner is not None
    assert context.execution_epoch is not None
    token = manifest_sha256({"site_code": entry.site_code})[:20]
    backup_path = backup_dir / f"source-{token}.json"
    secure_write_json(
        backup_path,
        {
            "schema_version": 1,
            "operation_id": context.operation_id,
            "closing_month": context.closing_month_key,
            "site_code": entry.site_code,
            "sheet_id": entry.sheet_id,
            "template_version": entry.template_version,
            "snapshot": snapshot,
            "snapshot_sha256": snapshot_sha256(snapshot),
            "created_at": utc_now(),
        },
    )
    artifact = relative_artifact(
        backup_path, root=OUTPUTS_DIR, kind="reset_source_snapshot"
    )
    artifact.update(
        site_code=entry.site_code,
        sheet_id=entry.sheet_id,
        template_version=entry.template_version,
    )
    recorded = await record_reset_item_backup(
        context.pool,
        operation_id=context.operation_id,
        site_code=entry.site_code,
        backup_path=artifact["path"],
        backup_sha256=artifact["sha256"],
        execution_owner=context.execution_owner,
        execution_epoch=context.execution_epoch,
    )
    if not recorded:
        raise MonthlyIntegrityError(
            "backup_checkpoint_failed", "Backup checkpoint failed"
        )
    return artifact


async def _capture_reset_backups(
    context: ResetRunContext,
    *,
    entries: list[StoreEntry],
    sheets_service: Any,
    expected: dict[str, Any],
    archive: dict[str, Any],
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    snapshots: dict[str, dict[str, Any]] = {}
    artifacts: list[dict[str, Any]] = []
    backup_dir = (
        build_reset_backup_dir(
            OUTPUTS_DIR, context.closing_month, context.operation_id
        )
        if context.operation_id is not None and not context.dry_run
        else None
    )
    try:
        for entry in entries:
            snapshot = await _read_reset_entry_snapshot(
                context, sheets_service, entry
            )
            snapshots[entry.site_code] = snapshot
            if backup_dir is not None:
                artifacts.append(
                    await _persist_reset_backup(
                        context,
                        entry=entry,
                        snapshot=snapshot,
                        backup_dir=backup_dir,
                    )
                )
        if artifacts:
            verify_artifacts({"artifacts": artifacts}, root=OUTPUTS_DIR)
        return snapshots, artifacts
    except BaseException as exc:
        code = (
            exc.code
            if isinstance(exc, MonthlyIntegrityError)
            else "reset_preflight_failed"
        )
        manifest = _reset_manifest(
            context,
            expected=expected,
            archive_manifest=archive,
            processed_stores=len(snapshots),
            source_backups=artifacts,
            errors=[code],
        )
        raise MonthlyManifestError(
            code, "Reset preflight failed", manifest
        ) from exc


def _reset_report_artifact(
    context: ResetRunContext,
    *,
    expected: dict[str, Any],
    processed_stores: int,
    dry_run: bool,
) -> tuple[Path, dict[str, Any]]:
    report = {
        "schema_version": 1,
        "operation": "reset",
        "month": context.closing_month_key,
        "next_month": context.next_month_key,
        "dry_run": dry_run,
        "expected_store_count": int(expected["stores"]),
        "processed_store_count": processed_stores,
        "error_count": 0,
        "created_at": utc_now(),
    }
    if not dry_run:
        report["approved_manifest_id"] = context.approved_manifest_id
    report_path = (
        build_reset_dry_run_report_path(OUTPUTS_DIR, context.next_month)
        if dry_run
        else build_reset_report_path(OUTPUTS_DIR, context.next_month)
    )
    stage_dir = _staging_dir(
        "reset-dry-run" if dry_run else "reset", context.operation_id
    )
    staged_report = stage_dir / "report.json"
    try:
        secure_write_json(staged_report, report)
        _promote_file(staged_report, report_path)
    finally:
        shutil.rmtree(stage_dir, ignore_errors=True)
    artifact = relative_artifact(
        report_path,
        root=OUTPUTS_DIR,
        kind="reset_dry_run_report" if dry_run else "reset_report",
    )
    return report_path, artifact


def _build_reset_dry_run(
    context: ResetRunContext,
    *,
    expected: dict[str, Any],
    snapshots: dict[str, dict[str, Any]],
    archive: dict[str, Any],
) -> MonthlyExecution:
    report_path, artifact = _reset_report_artifact(
        context,
        expected=expected,
        processed_stores=len(snapshots),
        dry_run=True,
    )
    manifest = _reset_manifest(
        context,
        expected=expected,
        archive_manifest=archive,
        processed_stores=len(snapshots),
        source_backups=archive.get("source_backups", []),
        artifacts=[artifact],
        errors=[],
        status="verified",
    )
    return MonthlyExecution(path=report_path, manifest=manifest)


async def _clear_reset_entry(
    context: ResetRunContext,
    *,
    entry: StoreEntry,
    sheets_service: Any,
) -> None:
    assert context.operation_id is not None
    assert context.execution_owner is not None
    assert context.execution_epoch is not None
    await heartbeat_monthly_operation(
        context.pool,
        context.operation_id,
        execution_owner=context.execution_owner,
        execution_epoch=context.execution_epoch,
    )
    claimed = await mark_reset_item_running(
        context.pool,
        operation_id=context.operation_id,
        site_code=entry.site_code,
        execution_owner=context.execution_owner,
        execution_epoch=context.execution_epoch,
    )
    if not claimed:
        raise MonthlyIntegrityError(
            "reset_checkpoint_claim_failed", "Reset checkpoint claim failed"
        )
    if context.google_adapter is None:
        result = reset_store(sheets_service, entry, dry_run=False)
        if result["status"] != "OK":
            raise MonthlyIntegrityError(result["error"], "Google reset failed")
        _verify_reset_cleared(sheets_service, entry)
        persisted = await finish_reset_item(
            context.pool,
            operation_id=context.operation_id,
            site_code=entry.site_code,
            status="completed",
            execution_owner=context.execution_owner,
            execution_epoch=context.execution_epoch,
        )
    else:
        intent = await persist_reset_clear_intent(
            context.pool,
            operation_id=context.operation_id,
            site_code=entry.site_code,
            execution_owner=context.execution_owner,
            execution_epoch=context.execution_epoch,
        )
        if intent is None:
            raise MonthlyIntegrityError(
                "operation_lease_lost", "Reset clear intent was fenced"
            )
        await _google_request(
            context.google_adapter,
            "clear",
            {
                "spreadsheet_id": entry.sheet_id,
                "ranges": reset_ranges_for_entry(entry),
            },
            label="Google reset",
            destructive=True,
        )
        await _verify_reset_cleared_async(context.google_adapter, entry)
        persisted = await persist_reset_clear_confirmation(
            context.pool,
            operation_id=context.operation_id,
            site_code=entry.site_code,
            execution_owner=context.execution_owner,
            execution_epoch=context.execution_epoch,
            fence_epoch=int(intent["fence_epoch"]),
        )
    if not persisted:
        raise MonthlyIntegrityError(
            "reset_checkpoint_finish_failed", "Reset checkpoint finish failed"
        )


async def _rollback_touched_entries(
    context: ResetRunContext,
    *,
    entries: list[StoreEntry],
    sheets_service: Any,
    snapshots: dict[str, dict[str, Any]],
) -> bool:
    assert context.operation_id is not None
    if context.execution_owner is None or context.execution_epoch is None:
        return False
    if context.google_adapter is None:
        return await _rollback_reset_entries_cancel_safe(
            context.pool,
            operation_id=context.operation_id,
            entries=entries,
            sheets_svc=sheets_service,
            snapshots=snapshots,
            execution_owner=context.execution_owner,
            execution_epoch=context.execution_epoch,
        )
    return await _rollback_reset_entries_adapter_cancel_safe(
        context.pool,
        operation_id=context.operation_id,
        entries=entries,
        google_adapter=context.google_adapter,
        snapshots=snapshots,
        execution_owner=context.execution_owner,
        execution_epoch=context.execution_epoch,
    )


async def _execute_reset_effects(
    context: ResetRunContext,
    *,
    entries: list[StoreEntry],
    sheets_service: Any,
    snapshots: dict[str, dict[str, Any]],
    backups: list[dict[str, Any]],
    expected: dict[str, Any],
    archive: dict[str, Any],
) -> list[StoreEntry]:
    touched: list[StoreEntry] = []
    try:
        for entry in entries:
            touched.append(entry)
            await _clear_reset_entry(
                context, entry=entry, sheets_service=sheets_service
            )
        return touched
    except BaseException as exc:
        code = (
            exc.code
            if isinstance(exc, MonthlyIntegrityError)
            else (
                "reset_cancelled"
                if isinstance(exc, asyncio.CancelledError)
                else "reset_failed"
            )
        )
        rollback_ok = await _rollback_touched_entries(
            context,
            entries=touched,
            sheets_service=sheets_service,
            snapshots=snapshots,
        )
        status = "rolled_back" if rollback_ok else "uncertain"
        manifest = _reset_manifest(
            context,
            expected=expected,
            archive_manifest=archive,
            source_backups=backups,
            errors=[code, "rollback_verified" if rollback_ok else "rollback_failed"],
            status=status,
        )
        raise MonthlyManifestError(
            status, "Reset failed and rollback was evaluated", manifest
        ) from exc


async def _reset_rollback_manifest(
    context: ResetRunContext,
    *,
    touched: list[StoreEntry],
    sheets_service: Any,
    snapshots: dict[str, dict[str, Any]],
    backups: list[dict[str, Any]],
    expected: dict[str, Any],
    archive: dict[str, Any],
) -> dict[str, Any]:
    rollback_ok = await _rollback_touched_entries(
        context,
        entries=touched,
        sheets_service=sheets_service,
        snapshots=snapshots,
    )
    return _reset_manifest(
        context,
        expected=expected,
        archive_manifest=archive,
        source_backups=backups,
        errors=[
            "reset_commit_failed",
            "rollback_verified" if rollback_ok else "rollback_failed",
        ],
        status="rolled_back" if rollback_ok else "uncertain",
    )


async def _build_live_reset_execution(
    context: ResetRunContext,
    *,
    entries: list[StoreEntry],
    touched: list[StoreEntry],
    sheets_service: Any,
    snapshots: dict[str, dict[str, Any]],
    backups: list[dict[str, Any]],
    expected: dict[str, Any],
    archive: dict[str, Any],
) -> MonthlyExecution:
    async def rollback_after_commit_failure() -> dict[str, Any]:
        return await _reset_rollback_manifest(
            context,
            touched=touched,
            sheets_service=sheets_service,
            snapshots=snapshots,
            backups=backups,
            expected=expected,
            archive=archive,
        )

    try:
        report_path, artifact = _reset_report_artifact(
            context,
            expected=expected,
            processed_stores=len(entries),
            dry_run=False,
        )
        manifest = _reset_manifest(
            context,
            expected=expected,
            archive_manifest=archive,
            processed_stores=len(entries),
            source_backups=backups,
            artifacts=[artifact],
            errors=[],
            status="verified",
        )
        validate_verified_manifest(manifest, operation="reset")
        verify_artifacts(manifest, root=OUTPUTS_DIR)
        return MonthlyExecution(
            path=report_path,
            manifest=manifest,
            rollback=rollback_after_commit_failure,
        )
    except BaseException as exc:
        failed = await rollback_after_commit_failure()
        raise MonthlyManifestError(
            str(failed["status"]),
            "Reset output failed and rollback was evaluated",
            failed,
        ) from exc


async def _reset_month_execution(
    pool: asyncpg.Pool,
    closing_month: str,
    next_month: str,
    *,
    closing_month_key: str,
    next_month_key: str,
    requested_by_sub: str,
    operation_id: int | None,
    approved_manifest_id: int | None,
    only: str | None = None,
    dry_run: bool = True,
    google_adapter: GoogleSyncAdapter | None = None,
    execution_owner: str | None = None,
    execution_epoch: int | None = None,
) -> MonthlyExecution:
    context = ResetRunContext(
        pool=pool,
        closing_month=closing_month,
        next_month=next_month,
        closing_month_key=closing_month_key,
        next_month_key=next_month_key,
        requested_by_sub=requested_by_sub,
        operation_id=operation_id,
        approved_manifest_id=approved_manifest_id,
        only=only,
        dry_run=dry_run,
        google_adapter=google_adapter,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
    )
    _validate_reset_request(context)
    archive = await _load_reset_archive(context)
    entries, expected = await _load_reset_entries(context, archive)
    sheets_service = await _prepare_reset_execution(context, entries)
    snapshots, backups = await _capture_reset_backups(
        context,
        entries=entries,
        sheets_service=sheets_service,
        expected=expected,
        archive=archive,
    )
    if dry_run:
        return _build_reset_dry_run(
            context, expected=expected, snapshots=snapshots, archive=archive
        )
    touched = await _execute_reset_effects(
        context,
        entries=entries,
        sheets_service=sheets_service,
        snapshots=snapshots,
        backups=backups,
        expected=expected,
        archive=archive,
    )
    return await _build_live_reset_execution(
        context,
        entries=entries,
        touched=touched,
        sheets_service=sheets_service,
        snapshots=snapshots,
        backups=backups,
        expected=expected,
        archive=archive,
    )


async def reset_month(
    pool: asyncpg.Pool,
    closing_month: str,
    next_month: str,
    only: str | None = None,
    dry_run: bool = True,
    operation_id: int | None = None,
    closing_month_key: str | None = None,
    next_month_key: str | None = None,
    requested_by_sub: str = "direct-execution",
    approved_manifest_id: int | None = None,
    google_adapter: GoogleSyncAdapter | None = None,
    execution_owner: str | None = None,
    execution_epoch: int | None = None,
) -> Path:
    execution = await _reset_month_execution(
        pool,
        closing_month,
        next_month,
        closing_month_key=closing_month_key or closing_month,
        next_month_key=next_month_key or next_month,
        requested_by_sub=requested_by_sub,
        operation_id=operation_id,
        approved_manifest_id=approved_manifest_id,
        only=only,
        dry_run=dry_run,
        google_adapter=google_adapter,
        execution_owner=execution_owner,
        execution_epoch=execution_epoch,
    )
    return execution.path


def _reconciliation_entry(item: dict[str, Any]) -> StoreEntry:
    return StoreEntry(
        str(item.get("company") or ""),
        str(item.get("store") or ""),
        str(item.get("sheet_id") or ""),
        str(item.get("site_code") or ""),
        "Neatribuit",
    )


def _read_reset_backup(item: dict[str, Any]) -> dict[str, Any]:
    raw_path = Path(str(item.get("backup_path") or ""))
    path = raw_path if raw_path.is_absolute() else OUTPUTS_DIR / raw_path
    if not path.exists() or file_sha256(path) != item.get("backup_sha256"):
        raise MonthlyIntegrityError("backup_invalid", "Reset backup cannot be verified")
    payload = json.loads(path.read_text(encoding="utf-8"))
    snapshot = payload.get("snapshot") if isinstance(payload, dict) else None
    if not isinstance(snapshot, dict):
        raise MonthlyIntegrityError("backup_invalid", "Reset backup snapshot is invalid")
    if payload.get("snapshot_sha256") != snapshot_sha256(snapshot):
        raise MonthlyIntegrityError("backup_invalid", "Reset backup hash is invalid")
    return snapshot


def _snapshot_is_cleared(snapshot: dict[str, Any]) -> bool:
    return not any(
        item.get("values")
        for item in snapshot.get("value_ranges", [])
        if isinstance(item, dict)
    )


async def reconcile_monthly_operations(
    pool: asyncpg.Pool,
    google_adapter: GoogleSyncAdapter,
) -> int:
    """Recover stale monthly work without replaying a reset automatically."""
    owner = f"reconciler-{uuid4().hex}"
    candidates = await claim_reconciliation_candidates(
        pool,
        execution_owner=owner,
    )
    reconciled = 0
    for operation in candidates:
        operation_id = int(operation["id"])
        epoch = int(operation.get("execution_epoch", 0))
        items = await list_reset_items_for_reconciliation(pool, operation_id)
        classifications: list[Literal["safe_retry", "rolled_back", "recovery_required"]] = []
        for item in items:
            site_code = str(item["site_code"])
            phase = str(item.get("checkpoint_phase") or "legacy_unknown")
            if phase == "legacy_unknown":
                await mark_item_recovery_required(
                    pool,
                    operation_id=operation_id,
                    site_code=site_code,
                    execution_owner=owner,
                    execution_epoch=epoch,
                    error_message="legacy_unknown_recovery_required",
                )
                classifications.append("recovery_required")
                continue

            try:
                snapshot = _read_reset_backup(item)
                current = await _read_reset_snapshot_async(
                    google_adapter,
                    _reconciliation_entry(item),
                )
                if snapshot_sha256(current) == snapshot_sha256(snapshot):
                    if phase == "snapshot_persisted":
                        safe = await mark_item_safe_retry(
                            pool,
                            operation_id=operation_id,
                            site_code=site_code,
                            execution_owner=owner,
                            execution_epoch=epoch,
                        )
                        classifications.append("safe_retry" if safe else "recovery_required")
                    else:
                        intent = await persist_reset_rollback_intent(
                            pool,
                            operation_id=operation_id,
                            site_code=site_code,
                            execution_owner=owner,
                            execution_epoch=epoch,
                        )
                        restored = bool(intent) and await persist_reset_rollback_confirmation(
                            pool,
                            operation_id=operation_id,
                            site_code=site_code,
                            execution_owner=owner,
                            execution_epoch=epoch,
                            fence_epoch=int(intent["fence_epoch"]) if intent else -1,
                            restored=True,
                            error_message="rollback_verified_readback",
                        )
                        classifications.append("rolled_back" if restored else "recovery_required")
                elif _snapshot_is_cleared(current):
                    intent = await persist_reset_rollback_intent(
                        pool,
                        operation_id=operation_id,
                        site_code=site_code,
                        execution_owner=owner,
                        execution_epoch=epoch,
                    )
                    if intent is None:
                        classifications.append("recovery_required")
                        continue
                    restored = False
                    try:
                        await _restore_reset_snapshot_async(
                            google_adapter,
                            _reconciliation_entry(item),
                            snapshot,
                        )
                        restored = True
                    except BaseException:  # noqa: BLE001 - recovery is fail-closed
                        restored = False
                    confirmed = await persist_reset_rollback_confirmation(
                        pool,
                        operation_id=operation_id,
                        site_code=site_code,
                        execution_owner=owner,
                        execution_epoch=epoch,
                        fence_epoch=int(intent["fence_epoch"]),
                        restored=restored,
                        error_message="rollback_verified" if restored else "recovery_required",
                    )
                    classifications.append("rolled_back" if restored and confirmed else "recovery_required")
                else:
                    classifications.append("recovery_required")
            except asyncio.CancelledError:
                raise
            except BaseException:  # noqa: BLE001 - read/hash failures block retry
                classifications.append("recovery_required")

            if classifications[-1] == "recovery_required":
                await mark_item_recovery_required(
                    pool,
                    operation_id=operation_id,
                    site_code=site_code,
                    execution_owner=owner,
                    execution_epoch=epoch,
                    error_message="recovery_required",
                )

        classification: Literal["safe_retry", "rolled_back", "recovery_required"]
        if "recovery_required" in classifications:
            classification = "recovery_required"
        elif "rolled_back" in classifications:
            classification = "rolled_back"
        else:
            classification = "safe_retry"
        await mark_reconciliation_result(
            pool,
            operation_id=operation_id,
            execution_owner=owner,
            execution_epoch=epoch,
            classification=classification,
            error_message=classification,
            alert=classification == "recovery_required",
        )
        reconciled += 1
    return reconciled


async def _monthly_operation_pool() -> Any:
    from db.connection import get_pool

    return await get_pool()


def _monthly_run_ports() -> Any:
    from services.grile_monthly_orchestration import MonthlyRunPorts

    return MonthlyRunPorts(
        valid_ops=frozenset(VALID_OPS),
        owner_hex=lambda: uuid4().hex,
        get_pool=_monthly_operation_pool,
        start_operation=start_monthly_operation,
        heartbeat_operation=heartbeat_monthly_operation,
        finish_operation=finish_monthly_operation,
        run_with_lease=_run_with_monthly_lease,
        persist_manifest=persist_manifest_result,
        persist_reset_success=persist_reset_success,
        fetch_manifest=fetch_monthly_manifest,
        finalize_execution=_finalize_month_execution,
        archive_execution=_archive_month_execution,
        reset_execution=_reset_month_execution,
        base_manifest=base_manifest,
        public_manifest_payload=public_manifest_payload,
        finalize_manifest=finalize_manifest,
        ro_month_label=ro_month_label,
        next_month=next_ym,
        utc_now=utc_now,
        manifest_error_type=MonthlyManifestError,
        integrity_error_type=MonthlyIntegrityError,
    )


async def run_monthly_op(
    *,
    op: str | None = None,
    month: str | None = None,
    only: str | None = None,
    dry_run: bool = True,
    operation_id: int | None = None,
    google_adapter: GoogleSyncAdapter | None = None,
    execution_owner_hint: str | None = None,
) -> dict[str, Any]:
    from services.grile_monthly_orchestration import orchestrate_monthly_operation

    return await orchestrate_monthly_operation(
        _monthly_run_ports(),
        op=op,
        month=month,
        only=only,
        dry_run=dry_run,
        operation_id=operation_id,
        google_adapter=google_adapter,
        execution_owner_hint=execution_owner_hint,
    )


async def fetch_download(kind: str, month: str) -> tuple[bytes, str, str]:
    if kind not in VALID_DOWNLOADS:
        raise ValueError(f"Tip download necunoscut: {kind}")
    month_label = ro_month_label(month)
    if kind == "final":
        path = build_final_export_path(OUTPUTS_DIR, month_label)
        filename = f"Tabel Salarii - {month_label}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        path = build_archive_zip_path(OUTPUTS_DIR, month_label)
        filename = f"Arhiva Grile - {month_label}.zip"
        media_type = "application/zip"

    if not path.exists() or path.stat().st_size == 0:
        raise FileNotFoundError(f"Fisierul {kind} pentru {month_label} nu exista inca.")
    return await asyncio.to_thread(path.read_bytes), filename, media_type
