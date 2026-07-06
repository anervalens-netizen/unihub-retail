from __future__ import annotations

from pathlib import Path
from typing import Sequence

import asyncpg
from openpyxl import load_workbook

from retail_filters import retail_exclusion_clauses
from services.phone_models import extract_phone_model_keys, phone_model_metadata


_REPO_ROOT = Path(__file__).resolve().parents[2]
_PREMIUM_CAMERA_FILE = _REPO_ROOT / "data" / "folii premium camera.xlsx"


_BRAND_GROUP_SQL = """
CASE
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%apple%'
         OR LOWER(COALESCE(st.item_name, '')) LIKE '%iphone%'
    THEN 'Apple'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%samsung%' THEN 'Samsung'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%honor%' THEN 'Honor'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%xiaomi%' THEN 'Xiaomi'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%huawei%' THEN 'Huawei'
    WHEN LOWER(COALESCE(st.item_name, '')) LIKE '%motorola%' THEN 'Motorola'
    ELSE 'Altele'
END
""".strip()

_MONTH_INDEX_SQL = (
    "(split_part({alias}, '-', 1)::INT * 12 + split_part({alias}, '-', 2)::INT)"
)

_RETAIL_TRANSACTION_EXCLUSIONS_SQL = "\n          AND ".join(
    retail_exclusion_clauses(site_alias="st", store_alias="s")
)

_PREMIUM_GLASS_INSERT_SQL = """
WITH target_models(model_key, model_label, model_regex, exclude_regex) AS (
    VALUES
        ('iphone_15', 'iPhone 15', 'IPHONE 15', 'IPHONE 15 PRO|IPHONE 15 PLUS -'),
        ('iphone_15_pro', 'iPhone 15 Pro', 'IPHONE 15 PRO|15 PRO/', 'IPHONE 15 PRO MAX|15 PRO MAX'),
        ('iphone_15_pro_max', 'iPhone 15 Pro Max', 'IPHONE 15 PRO MAX|15 PRO MAX', NULL),
        ('iphone_16', 'iPhone 16', 'IPHONE 16|/16([[:space:]]|-|$)', 'IPHONE 16 PRO|16 PRO|IPHONE 15 PLUS/16 PLUS|IPHONE 16 PLUS -'),
        ('iphone_16_pro', 'iPhone 16 Pro', 'IPHONE 16 PRO|16 PRO/', 'IPHONE 16 PRO MAX|16 PRO MAX'),
        ('iphone_16_pro_max', 'iPhone 16 Pro Max', 'IPHONE 16 PRO MAX|16 PRO MAX', NULL),
        ('iphone_17', 'iPhone 17', 'IPHONE 17|PRO/17([[:space:]]|-|$)', 'IPHONE 17 PRO|IPHONE 17 AIR'),
        ('iphone_17_pro', 'iPhone 17 Pro', 'IPHONE 17 PRO|17 PRO/', 'IPHONE 17 PRO MAX|17 PRO MAX'),
        ('iphone_17_pro_max', 'iPhone 17 Pro Max', 'IPHONE 17 PRO MAX|17 PRO MAX', NULL),
        ('samsung_s26', 'Samsung S26', 'SAMSUNG GALAXY S26 5G|S26 5G', 'S26 PLUS|S26 ULTRA'),
        ('samsung_s26_ultra', 'Samsung S26 Ultra', 'SAMSUNG GALAXY S26 ULTRA|S26 ULTRA', NULL)
),
camera_products AS (
    SELECT *
    FROM UNNEST(
        $1::TEXT[],
        $2::TEXT[],
        $3::BOOLEAN[],
        $4::TEXT[],
        $5::TEXT[]
    ) AS t(item_code, item_name, is_premium_glass, model_key, model_label)
),
source_products AS (
    SELECT DISTINCT
        st.item_code,
        st.item_name,
        UPPER(COALESCE(st.item_name, '')) AS item_name_upper
    FROM sales_transactions st
    WHERE LOWER(TRIM(COALESCE(st.category, ''))) = 'folii sticla'
      AND st.item_code IS NOT NULL
      AND TRIM(st.item_code) != ''
),
screen_matches AS (
    SELECT DISTINCT ON (sp.item_code, tm.model_key)
        sp.item_code,
        sp.item_name,
        (sp.item_name_upper ~ '(SAPPHIRE|CERAMIC|CORNING)') AS is_premium_glass,
        tm.model_key,
        tm.model_label
    FROM source_products sp
    JOIN target_models tm
        ON sp.item_name_upper ~ tm.model_regex
       AND (tm.exclude_regex IS NULL OR sp.item_name_upper !~ tm.exclude_regex)
    WHERE NOT (sp.item_code = ANY($1::TEXT[]))
    ORDER BY sp.item_code, tm.model_key, sp.item_name
),
camera_matches AS (
    SELECT DISTINCT
        item_code,
        item_name,
        is_premium_glass,
        model_key,
        model_label
    FROM camera_products
)
INSERT INTO premium_glass_item_models (
    item_code,
    item_name,
    is_premium_glass,
    model_key,
    model_label
)
SELECT DISTINCT ON (sp.item_code, sp.model_key)
    sp.item_code,
    sp.item_name,
    sp.is_premium_glass,
    sp.model_key,
    sp.model_label
FROM (
    SELECT * FROM camera_matches
    UNION ALL
    SELECT * FROM screen_matches
) sp
ORDER BY sp.item_code, sp.model_key, sp.item_name;

""".strip()


def _load_premium_camera_rows(
    path: Path = _PREMIUM_CAMERA_FILE,
) -> tuple[list[str], list[str], list[bool], list[str], list[str]]:
    if not path.exists():
        return [], [], [], [], []

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Folii Camera"] if "Folii Camera" in workbook.sheetnames else workbook.active
    headers = [str(value or "").strip().lower() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {name: index for index, name in enumerate(headers)}
    code_idx = header_index.get("cod")
    premium_idx = header_index.get("premium")
    name_idx = header_index.get("itemname")
    if code_idx is None or premium_idx is None or name_idx is None:
        return [], [], [], [], []

    item_codes: list[str] = []
    item_names: list[str] = []
    premium_flags: list[bool] = []
    model_keys: list[str] = []
    model_labels: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = str(row[code_idx] or "").strip()
        item_name = str(row[name_idx] or "").strip()
        premium_value = str(row[premium_idx] or "").strip().lower()
        if not code or not item_name:
            continue
        is_premium = premium_value in {"da", "yes", "true", "1"}
        for model in sorted(extract_phone_model_keys(item_name)):
            metadata = phone_model_metadata(model)
            if metadata is None:
                continue
            model_key, model_label = metadata
            item_codes.append(code)
            item_names.append(item_name)
            premium_flags.append(is_premium)
            model_keys.append(model_key)
            model_labels.append(model_label)
    return item_codes, item_names, premium_flags, model_keys, model_labels


async def refresh_premium_glass_indicators(conn: asyncpg.Connection) -> None:
    await conn.execute("TRUNCATE premium_glass_item_models")
    await conn.execute(_PREMIUM_GLASS_INSERT_SQL, *_load_premium_camera_rows())
    await conn.execute("ANALYZE premium_glass_item_models")


async def rebuild_reporting_month(
    conn: asyncpg.Connection,
    import_month: str,
) -> None:
    await conn.execute("DROP TABLE IF EXISTS tmp_reporting_receipts")
    await conn.execute(
        "DELETE FROM reporting_category_month WHERE import_month = $1",
        import_month,
    )
    await conn.execute(
        "DELETE FROM reporting_focus_item_month WHERE import_month = $1",
        import_month,
    )
    await conn.execute(
        "DELETE FROM reporting_agent_month WHERE import_month = $1",
        import_month,
    )
    await conn.execute(
        "DELETE FROM reporting_item_day WHERE import_month = $1",
        import_month,
    )
    await conn.execute(
        "DELETE FROM reporting_item_month WHERE import_month = $1",
        import_month,
    )
    await conn.execute(
        "DELETE FROM reporting_agent_day WHERE import_month = $1",
        import_month,
    )

    await conn.execute(
        """
        CREATE TEMP TABLE tmp_reporting_receipts (
            import_month TEXT NOT NULL,
            sale_date DATE NOT NULL,
            site_code TEXT NOT NULL,
            agent TEXT NOT NULL,
            bon_nr TEXT NOT NULL,
            qty_positive INTEGER NOT NULL
        )
        """
    )
    await conn.execute(
        f"""
        INSERT INTO tmp_reporting_receipts (
            import_month,
            sale_date,
            site_code,
            agent,
            bon_nr,
            qty_positive
        )
        SELECT
            st.import_month,
            st.sale_date,
            st.site_code,
            st.agent,
            st.bon_nr,
            COALESCE(
                SUM(CASE WHEN st.quantity > 0 THEN st.quantity ELSE 0 END),
                0
            )::INT AS qty_positive
        FROM sales_transactions st
        JOIN stores s ON s.site_code = st.site_code
        WHERE st.import_month = $1
          AND {_RETAIL_TRANSACTION_EXCLUSIONS_SQL}
        GROUP BY st.import_month, st.sale_date, st.site_code, st.agent, st.bon_nr
        """,
        import_month,
    )
    await conn.execute(
        """
        CREATE INDEX idx_tmp_reporting_receipts_keys
            ON tmp_reporting_receipts (import_month, sale_date, site_code, agent, bon_nr)
        """
    )

    await conn.execute(
        f"""
        INSERT INTO reporting_agent_day (
            import_month,
            sale_date,
            site_code,
            locatie,
            firma,
            regional,
            asm,
            agent,
            total_sales,
            total_quantity,
            focus_quantity,
            receipt_count,
            receipt_2plus_count,
            receipt_1_count,
            receipt_2_count,
            receipt_3_count,
            receipt_4plus_count
        )
        SELECT
            st.import_month,
            st.sale_date,
            st.site_code,
            s.locatie,
            s.firma,
            s.regional,
            s.asm,
            st.agent,
            COALESCE(SUM(st.total_value), 0)::NUMERIC(12, 2) AS total_sales,
            COALESCE(
                SUM(CASE WHEN NOT st.is_return AND st.quantity > 0 THEN st.quantity ELSE 0 END),
                0
            )::INT AS total_quantity,
            COALESCE(
                SUM(CASE WHEN fp.item_code IS NOT NULL AND st.quantity > 0 THEN st.quantity ELSE 0 END),
                0
            )::INT AS focus_quantity,
            COUNT(DISTINCT st.bon_nr)::INT AS receipt_count,
            COUNT(DISTINCT st.bon_nr) FILTER (WHERE tr.qty_positive >= 2)::INT AS receipt_2plus_count,
            COUNT(DISTINCT st.bon_nr) FILTER (WHERE tr.qty_positive = 1)::INT AS receipt_1_count,
            COUNT(DISTINCT st.bon_nr) FILTER (WHERE tr.qty_positive = 2)::INT AS receipt_2_count,
            COUNT(DISTINCT st.bon_nr) FILTER (WHERE tr.qty_positive = 3)::INT AS receipt_3_count,
            COUNT(DISTINCT st.bon_nr) FILTER (WHERE tr.qty_positive >= 4)::INT AS receipt_4plus_count
        FROM sales_transactions st
        JOIN stores s ON s.site_code = st.site_code
        LEFT JOIN focus_products fp ON fp.item_code = st.item_code
        LEFT JOIN tmp_reporting_receipts tr
            ON tr.import_month = st.import_month
            AND tr.sale_date = st.sale_date
            AND tr.site_code = st.site_code
            AND tr.agent = st.agent
            AND tr.bon_nr = st.bon_nr
        WHERE st.import_month = $1
          AND {_RETAIL_TRANSACTION_EXCLUSIONS_SQL}
        GROUP BY
            st.import_month,
            st.sale_date,
            st.site_code,
            s.locatie,
            s.firma,
            s.regional,
            s.asm,
            st.agent
        """,
        import_month,
    )

    await conn.execute(
        f"""
        INSERT INTO reporting_agent_month (
            import_month,
            site_code,
            locatie,
            firma,
            regional,
            asm,
            agent,
            total_sales,
            total_quantity,
            focus_quantity,
            receipt_count,
            receipt_2plus_count,
            receipt_1_count,
            receipt_2_count,
            receipt_3_count,
            receipt_4plus_count,
            working_days
        )
        SELECT
            import_month,
            site_code,
            locatie,
            firma,
            regional,
            asm,
            agent,
            COALESCE(SUM(total_sales), 0)::NUMERIC(12, 2) AS total_sales,
            COALESCE(SUM(total_quantity), 0)::INT AS total_quantity,
            COALESCE(SUM(focus_quantity), 0)::INT AS focus_quantity,
            COALESCE(SUM(receipt_count), 0)::INT AS receipt_count,
            COALESCE(SUM(receipt_2plus_count), 0)::INT AS receipt_2plus_count,
            COALESCE(SUM(receipt_1_count), 0)::INT AS receipt_1_count,
            COALESCE(SUM(receipt_2_count), 0)::INT AS receipt_2_count,
            COALESCE(SUM(receipt_3_count), 0)::INT AS receipt_3_count,
            COALESCE(SUM(receipt_4plus_count), 0)::INT AS receipt_4plus_count,
            COUNT(*)::INT AS working_days
        FROM reporting_agent_day
        WHERE import_month = $1
        GROUP BY import_month, site_code, locatie, firma, regional, asm, agent
        """,
        import_month,
    )

    await conn.execute(
        f"""
        INSERT INTO reporting_item_day (
            import_month,
            sale_date,
            site_code,
            locatie,
            firma,
            regional,
            asm,
            agent,
            item_code,
            item_name,
            total_sales,
            net_quantity,
            positive_quantity,
            return_quantity,
            receipt_count
        )
        SELECT
            st.import_month,
            st.sale_date,
            st.site_code,
            s.locatie,
            s.firma,
            s.regional,
            s.asm,
            st.agent,
            st.item_code,
            COALESCE(MAX(st.item_name), st.item_code) AS item_name,
            COALESCE(SUM(st.total_value), 0)::NUMERIC(12, 2) AS total_sales,
            COALESCE(SUM(st.quantity), 0)::INT AS net_quantity,
            COALESCE(SUM(CASE WHEN st.quantity > 0 THEN st.quantity ELSE 0 END), 0)::INT AS positive_quantity,
            COALESCE(SUM(CASE WHEN st.quantity < 0 THEN st.quantity ELSE 0 END), 0)::INT AS return_quantity,
            COUNT(DISTINCT st.bon_nr)::INT AS receipt_count
        FROM sales_transactions st
        JOIN stores s ON s.site_code = st.site_code
        WHERE st.import_month = $1
          AND {_RETAIL_TRANSACTION_EXCLUSIONS_SQL}
        GROUP BY
            st.import_month,
            st.sale_date,
            st.site_code,
            s.locatie,
            s.firma,
            s.regional,
            s.asm,
            st.agent,
            st.item_code
        """,
        import_month,
    )

    await conn.execute(
        f"""
        INSERT INTO reporting_focus_item_month (
            import_month,
            site_code,
            locatie,
            firma,
            regional,
            asm,
            agent,
            item_code,
            item_name,
            focus_subcategory,
            total_sales,
            total_quantity
        )
        SELECT
            st.import_month,
            st.site_code,
            s.locatie,
            s.firma,
            s.regional,
            s.asm,
            st.agent,
            st.item_code,
            COALESCE(MAX(fp.item_name), MAX(st.item_name), st.item_code) AS item_name,
            COALESCE(
                NULLIF(TRIM(MAX(st.subcategory)), ''),
                NULLIF(TRIM(MAX(st.category)), ''),
                'Necategorizat'
            ) AS focus_subcategory,
            COALESCE(SUM(st.total_value), 0)::NUMERIC(12, 2) AS total_sales,
            COALESCE(
                SUM(CASE WHEN NOT st.is_return AND st.quantity > 0 THEN st.quantity ELSE 0 END),
                0
            )::INT AS total_quantity
        FROM sales_transactions st
        JOIN stores s ON s.site_code = st.site_code
        JOIN focus_products fp ON fp.item_code = st.item_code
        WHERE st.import_month = $1
          AND {_RETAIL_TRANSACTION_EXCLUSIONS_SQL}
        GROUP BY
            st.import_month,
            st.site_code,
            s.locatie,
            s.firma,
            s.regional,
            s.asm,
            st.agent,
            st.item_code
        """,
        import_month,
    )

    await conn.execute(
        """
        INSERT INTO reporting_item_month (
            import_month,
            site_code,
            locatie,
            firma,
            regional,
            asm,
            agent,
            item_code,
            item_name,
            total_sales,
            net_quantity,
            positive_quantity,
            return_quantity,
            receipt_count
        )
        SELECT
            import_month,
            site_code,
            locatie,
            firma,
            regional,
            asm,
            agent,
            item_code,
            MAX(item_name) AS item_name,
            COALESCE(SUM(total_sales), 0)::NUMERIC(12, 2) AS total_sales,
            COALESCE(SUM(net_quantity), 0)::INT AS net_quantity,
            COALESCE(SUM(positive_quantity), 0)::INT AS positive_quantity,
            COALESCE(SUM(return_quantity), 0)::INT AS return_quantity,
            COALESCE(SUM(receipt_count), 0)::INT AS receipt_count
        FROM reporting_item_day
        WHERE import_month = $1
        GROUP BY
            import_month,
            site_code,
            locatie,
            firma,
            regional,
            asm,
            agent,
            item_code
        """,
        import_month,
    )

    await conn.execute(
        f"""
        INSERT INTO reporting_category_month (
            import_month,
            site_code,
            locatie,
            firma,
            regional,
            asm,
            agent,
            category,
            subcategory,
            brand_group,
            total_sales,
            total_quantity
        )
        SELECT
            st.import_month,
            st.site_code,
            s.locatie,
            s.firma,
            s.regional,
            s.asm,
            st.agent,
            COALESCE(NULLIF(TRIM(st.category), ''), 'Necategorizat') AS category,
            COALESCE(
                NULLIF(TRIM(st.subcategory), ''),
                NULLIF(TRIM(st.category), ''),
                'Necategorizat'
            ) AS subcategory,
            {_BRAND_GROUP_SQL} AS brand_group,
            COALESCE(SUM(st.total_value), 0)::NUMERIC(12, 2) AS total_sales,
            COALESCE(
                SUM(CASE WHEN NOT st.is_return AND st.quantity > 0 THEN st.quantity ELSE 0 END),
                0
            )::INT AS total_quantity
        FROM sales_transactions st
        JOIN stores s ON s.site_code = st.site_code
        WHERE st.import_month = $1
          AND {_RETAIL_TRANSACTION_EXCLUSIONS_SQL}
        GROUP BY
            st.import_month,
            st.site_code,
            s.locatie,
            s.firma,
            s.regional,
            s.asm,
            st.agent,
            COALESCE(NULLIF(TRIM(st.category), ''), 'Necategorizat'),
            COALESCE(
                NULLIF(TRIM(st.subcategory), ''),
                NULLIF(TRIM(st.category), ''),
                'Necategorizat'
            ),
            {_BRAND_GROUP_SQL}
        """,
        import_month,
    )

    await conn.execute("ANALYZE reporting_agent_day")
    await conn.execute("ANALYZE reporting_agent_month")
    await conn.execute("ANALYZE reporting_item_day")
    await conn.execute("ANALYZE reporting_item_month")
    await conn.execute("ANALYZE reporting_focus_item_month")
    await conn.execute("ANALYZE reporting_category_month")
    await refresh_premium_glass_indicators(conn)
    await conn.execute("DROP TABLE IF EXISTS tmp_reporting_receipts")


async def rebuild_agent_lifecycle_reporting(conn: asyncpg.Connection) -> None:
    month_index = _MONTH_INDEX_SQL.format(alias="import_month")
    prev_month_index = _MONTH_INDEX_SQL.format(alias="prev_active_month")
    latest_month_index = _MONTH_INDEX_SQL.format(alias="lm.latest_month")
    max_month_index = _MONTH_INDEX_SQL.format(alias="MAX(lc.import_month)")

    await conn.execute("DELETE FROM reporting_agent_lifecycle_month")
    await conn.execute("DELETE FROM reporting_agent_profile")

    await conn.execute(
        f"""
        INSERT INTO reporting_agent_lifecycle_month (
            import_month,
            agent,
            total_sales,
            total_quantity,
            receipt_count,
            working_days,
            active_store_count,
            active_firma_count,
            active_regional_count,
            active_asm_count,
            first_seen_month,
            prev_active_month,
            gap_since_prev_active_months,
            is_new,
            is_reactivated,
            is_active
        )
        WITH base AS (
            SELECT
                import_month,
                agent,
                COALESCE(SUM(total_sales), 0)::NUMERIC(12, 2) AS total_sales,
                COALESCE(SUM(total_quantity), 0)::INT AS total_quantity,
                COALESCE(SUM(receipt_count), 0)::INT AS receipt_count,
                COALESCE(SUM(working_days), 0)::INT AS working_days,
                COUNT(DISTINCT site_code)::INT AS active_store_count,
                COUNT(DISTINCT firma)::INT AS active_firma_count,
                COUNT(DISTINCT regional)::INT AS active_regional_count,
                COUNT(DISTINCT asm)::INT AS active_asm_count
            FROM reporting_agent_month
            GROUP BY import_month, agent
        ),
        sequenced AS (
            SELECT
                base.*,
                MIN(import_month) OVER (PARTITION BY agent) AS first_seen_month,
                LAG(import_month) OVER (PARTITION BY agent ORDER BY import_month) AS prev_active_month
            FROM base
        )
        SELECT
            import_month,
            agent,
            total_sales,
            total_quantity,
            receipt_count,
            working_days,
            active_store_count,
            active_firma_count,
            active_regional_count,
            active_asm_count,
            first_seen_month,
            prev_active_month,
            CASE
                WHEN prev_active_month IS NULL THEN 0
                ELSE GREATEST({month_index} - ({prev_month_index}) - 1, 0)
            END::INT AS gap_since_prev_active_months,
            (prev_active_month IS NULL) AS is_new,
            (
                prev_active_month IS NOT NULL
                AND GREATEST({month_index} - ({prev_month_index}) - 1, 0) >= 2
            ) AS is_reactivated,
            true AS is_active
        FROM sequenced
        """
    )

    await conn.execute(
        f"""
        INSERT INTO reporting_agent_profile (
            agent,
            first_seen_month,
            last_seen_month,
            active_months_count,
            distinct_store_count,
            distinct_firma_count,
            distinct_regional_count,
            distinct_asm_count,
            months_since_last_seen,
            reactivation_count,
            longest_active_streak,
            career_total_sales,
            career_total_quantity,
            avg_monthly_sales,
            best_month,
            best_month_sales,
            current_status
        )
        WITH lifecycle AS (
            SELECT
                agent,
                import_month,
                total_sales,
                total_quantity,
                first_seen_month,
                is_reactivated,
                ROW_NUMBER() OVER (PARTITION BY agent ORDER BY import_month) AS row_num,
                { _MONTH_INDEX_SQL.format(alias='import_month') } AS month_index
            FROM reporting_agent_lifecycle_month
        ),
        streaks AS (
            SELECT
                agent,
                COUNT(*)::INT AS streak_length
            FROM (
                SELECT
                    agent,
                    month_index - row_num AS streak_group
                FROM lifecycle
            ) grouped
            GROUP BY agent, streak_group
        ),
        best_months AS (
            SELECT DISTINCT ON (agent)
                agent,
                import_month AS best_month,
                total_sales AS best_month_sales
            FROM reporting_agent_lifecycle_month
            ORDER BY agent, total_sales DESC, import_month DESC
        ),
        store_scope AS (
            SELECT
                agent,
                COUNT(DISTINCT site_code)::INT AS distinct_store_count,
                COUNT(DISTINCT firma)::INT AS distinct_firma_count,
                COUNT(DISTINCT regional)::INT AS distinct_regional_count,
                COUNT(DISTINCT asm)::INT AS distinct_asm_count
            FROM reporting_agent_month
            GROUP BY agent
        ),
        latest_month AS (
            SELECT MAX(import_month) AS latest_month
            FROM reporting_agent_lifecycle_month
        )
        SELECT
            lc.agent,
            MIN(lc.first_seen_month) AS first_seen_month,
            MAX(lc.import_month) AS last_seen_month,
            COUNT(*)::INT AS active_months_count,
            COALESCE(MAX(ss.distinct_store_count), 0)::INT AS distinct_store_count,
            COALESCE(MAX(ss.distinct_firma_count), 0)::INT AS distinct_firma_count,
            COALESCE(MAX(ss.distinct_regional_count), 0)::INT AS distinct_regional_count,
            COALESCE(MAX(ss.distinct_asm_count), 0)::INT AS distinct_asm_count,
            GREATEST(
                ({latest_month_index}) - ({max_month_index}),
                0
            )::INT AS months_since_last_seen,
            COUNT(*) FILTER (WHERE lc.is_reactivated)::INT AS reactivation_count,
            COALESCE(MAX(st.streak_length), 0)::INT AS longest_active_streak,
            COALESCE(SUM(lc.total_sales), 0)::NUMERIC(12, 2) AS career_total_sales,
            COALESCE(SUM(lc.total_quantity), 0)::INT AS career_total_quantity,
            CASE
                WHEN COUNT(*) > 0 THEN ROUND(COALESCE(SUM(lc.total_sales), 0) / COUNT(*), 2)
                ELSE 0
            END::NUMERIC(12, 2) AS avg_monthly_sales,
            MAX(bm.best_month) AS best_month,
            COALESCE(MAX(bm.best_month_sales), 0)::NUMERIC(12, 2) AS best_month_sales,
            CASE
                WHEN GREATEST(
                    ({latest_month_index}) - ({max_month_index}),
                    0
                ) >= 2
                THEN 'churned'
                WHEN GREATEST(
                    ({latest_month_index}) - ({max_month_index}),
                    0
                ) = 1
                THEN 'inactive_recent'
                ELSE 'active'
            END AS current_status
        FROM lifecycle lc
        CROSS JOIN latest_month lm
        LEFT JOIN streaks st
            ON st.agent = lc.agent
        LEFT JOIN best_months bm
            ON bm.agent = lc.agent
        LEFT JOIN store_scope ss
            ON ss.agent = lc.agent
        GROUP BY lc.agent, lm.latest_month
        """
    )

    await conn.execute("ANALYZE reporting_agent_lifecycle_month")
    await conn.execute("ANALYZE reporting_agent_profile")


async def list_completed_import_months(conn: asyncpg.Connection) -> list[str]:
    rows = await conn.fetch(
        """
        SELECT DISTINCT import_month
        FROM import_snapshots
        WHERE status = 'completed'
        ORDER BY import_month ASC
        """
    )
    return [str(row["import_month"]) for row in rows]


async def rebuild_reporting_all(
    conn: asyncpg.Connection,
    months: Sequence[str] | None = None,
) -> list[str]:
    target_months = list(months) if months is not None else await list_completed_import_months(conn)
    for import_month in target_months:
        async with conn.transaction():
            await rebuild_reporting_month(conn, import_month)
    async with conn.transaction():
        await rebuild_agent_lifecycle_reporting(conn)
    return target_months
