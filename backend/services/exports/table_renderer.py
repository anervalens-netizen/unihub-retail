"""Normal-workbook table layout helpers (no repository/SQL imports)."""
from __future__ import annotations

import resource
import time
from hashlib import sha256
from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import SpooledTemporaryFile
from typing import Any, TYPE_CHECKING
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from business_clock import business_now
from services.spreadsheet_safety import append_openpyxl_row
from services.spreadsheet_safety import set_openpyxl_cell

from .artifact import XlsxArtifact
from .catalog import COMPARISON_LEVELS, DAILY_EVOLUTION_METRICS, DATASETS
from .calculations import pct
from .metrics import EXPORT_BUILD_SECONDS, EXPORT_CELLS, EXPORT_OUTPUT_BYTES, EXPORT_PEAK_RSS_BYTES, EXPORT_REJECTED_TOTAL
from .validation import (
    EXPORT_MAX_OUTPUT_BYTES,
    EXPORT_MAX_PEAK_RSS_BYTES,
    EXPORT_MAX_PROCESS_RSS_BYTES,
    ExportValidationError,
)
from services.export_xlsx_formatting import (
    add_daily_comparison_chart,
    configure_day_axis,
    number_format,
    safe_filename,
    write_table_sheet,
)

XLSX_SPOOL_MAX_MEMORY_BYTES = 8 * 1024 * 1024


def _current_rss_bytes() -> int:
    """Sample current Linux RSS; fall back to the process lifetime peak."""
    try:
        resident_pages = int(Path("/proc/self/statm").read_text(encoding="ascii").split()[1])
        return resident_pages * resource.getpagesize()
    except (OSError, ValueError, IndexError):
        return resource.getrusage(resource.RUSAGE_SELF).ru_maxrss * 1024

class XlsxRenderers:
    """OpenPyXL-only renderer mixin; never loads repository data."""

    if TYPE_CHECKING:
        def _days_filename_suffix(self, selected_days: list[int] | None) -> str: ...
        def _compute_metrics(self, row: Any) -> dict[str, Any]: ...
        def _json_value(self, value: Any) -> Any: ...

    def _render_simple_table_xlsx(
        self,
        request: dict[str, Any],
        result: dict[str, Any],
        selected_days: list[int] | None,
        _daily_rows: list[Any] | None,
    ) -> XlsxArtifact:
        """Stream a plain table without retaining an openpyxl cell graph."""
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Raport")
        columns = result["columns"]
        rows = result["rows"]
        header = []
        for column in columns:
            cell = WriteOnlyCell(ws)
            set_openpyxl_cell(cell, column["label"])
            cell.font = Font(bold=True, color="1f2937")
            cell.fill = PatternFill("solid", fgColor="EEF2FF")
            header.append(cell)
        ws.append(header)
        for row in rows:
            output = []
            for column in columns:
                cell = WriteOnlyCell(ws)
                set_openpyxl_cell(cell, row.get(column["key"]))
                number_format = self._excel_number_format(column["type"])
                if number_format:
                    cell.number_format = number_format
                output.append(cell)
            ws.append(output)

        cfg = wb.create_sheet("Configuratie")
        cfg.append(["Optiune", "Valoare"])
        cfg.append(["Dataset", DATASETS[request["dataset"]]["label"]])
        cfg.append(["Luni", ", ".join(request["months"])])
        cfg.append(["Zile", ", ".join(str(day) for day in selected_days) if selected_days else "Toata luna"])
        cfg.append(["Generat", business_now().strftime("%Y-%m-%d %H:%M")])
        cfg.append(["Randuri", len(rows)])
        return self._spool_workbook(
            wb,
            self._safe_filename(str(request.get("filename") or "export_retail.xlsx")),
            cells=(len(rows) + 1) * len(columns) + 6 * 2,
        )

    def _render_table_xlsx(
        self,
        request: dict[str, Any],
        result: dict[str, Any],
        selected_days: list[int] | None,
        daily_rows: list[Any] | None,
    ) -> XlsxArtifact:
        wb = Workbook()
        ws = wb.active
        ws.title = "Raport"

        columns = result["columns"]
        rows = result["rows"]
        append_openpyxl_row(ws, [column["label"] for column in columns])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1f2937")
            cell.fill = PatternFill("solid", fgColor="EEF2FF")

        for row in rows:
            append_openpyxl_row(ws, [row.get(column["key"]) for column in columns])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for idx, column in enumerate(columns, start=1):
            letter = get_column_letter(idx)
            width = max(12, min(28, len(column["label"]) + 3))
            ws.column_dimensions[letter].width = width
            number_format = self._excel_number_format(column["type"])
            if number_format:
                for cell in ws[letter][1:]:
                    cell.number_format = number_format

        cfg = wb.create_sheet("Configuratie")
        append_openpyxl_row(cfg, ["Optiune", "Valoare"])
        append_openpyxl_row(cfg, ["Dataset", DATASETS[request["dataset"]]["label"]])
        append_openpyxl_row(cfg, ["Luni", ", ".join(request["months"])])
        append_openpyxl_row(cfg, ["Zile", ", ".join(str(day) for day in selected_days) if selected_days else "Toata luna"])
        append_openpyxl_row(cfg, ["Generat", business_now().strftime("%Y-%m-%d %H:%M")])
        append_openpyxl_row(cfg, ["Randuri", len(rows)])
        for cell in cfg[1]:
            cell.font = Font(bold=True)
        cfg.column_dimensions["A"].width = 24
        cfg.column_dimensions["B"].width = 64

        if request.get("daily_metrics") and daily_rows is not None:
            self._add_daily_evolution_sheet(
                wb,
                months=sorted(request["months"]),
                metrics=[metric for metric in request["daily_metrics"] if metric in DAILY_EVOLUTION_METRICS],
                records=daily_rows,
            )

        filename = request.get("filename") or (
            f"export_retail_{request['dataset']}_{'_'.join(request['months'])}"
            f"{self._days_filename_suffix(selected_days)}.xlsx"
        )
        return self._spool_workbook(
            wb,
            self._safe_filename(str(filename)),
            cells=(len(rows) + 1) * len(columns),
        )

    def _render_daily_comparison_xlsx(
        self,
        request: dict[str, Any],
        months: list[str],
        metrics: list[str],
        levels: list[str],
        include_closed_stores: bool,
        selected_days: list[int] | None,
        tables: list[tuple[str, dict[str, Any]]],
    ) -> XlsxArtifact:
        wb = Workbook()
        first_sheet = True
        total_rows = 0

        for level, table in tables:
            sheet_name = COMPARISON_LEVELS[level]["sheet"]
            ws = wb.active if first_sheet else wb.create_sheet(sheet_name)
            ws.title = sheet_name
            first_sheet = False
            self._write_table_sheet(ws, table["columns"], table["rows"], header_fill="DCFCE7")
            total_rows += len(table["rows"])
            self._add_daily_comparison_chart(
                ws,
                months=months,
                metric=metrics[0],
                max_row=len(table["rows"]) + 1,
                first_data_col=len(COMPARISON_LEVELS[level]["dimensions"]) + 2,
            )

        cfg = wb.create_sheet("Configuratie")
        append_openpyxl_row(cfg, ["Optiune", "Valoare"])
        append_openpyxl_row(cfg, ["Tip export", "Evolutie zilnica comparativa"])
        append_openpyxl_row(cfg, ["Luni", ", ".join(months)])
        append_openpyxl_row(cfg, ["Zile", ", ".join(str(day) for day in selected_days) if selected_days else "Toata luna"])
        append_openpyxl_row(cfg, ["Metrici zilnice", ", ".join(DAILY_EVOLUTION_METRICS[item].label for item in metrics)])
        append_openpyxl_row(cfg, ["Niveluri", ", ".join(str(COMPARISON_LEVELS[item]["label"]) for item in levels)])
        append_openpyxl_row(cfg, ["Include magazine inchise", "Da" if include_closed_stores else "Nu"])
        append_openpyxl_row(cfg, ["Generat", business_now().strftime("%Y-%m-%d %H:%M")])
        append_openpyxl_row(cfg, ["Randuri", total_rows])
        for cell in cfg[1]:
            cell.font = Font(bold=True)
        cfg.column_dimensions["A"].width = 28
        cfg.column_dimensions["B"].width = 72

        filename = request.get("filename") or (
            f"export_retail_evolutie_zilnica_{'_'.join(months)}"
            f"{self._days_filename_suffix(selected_days)}.xlsx"
        )
        return self._spool_workbook(
            wb,
            self._safe_filename(str(filename)),
            cells=sum(len(table[1]["rows"]) * len(table[1]["columns"]) for table in tables),
        )

    @staticmethod
    def _spool_workbook(
        wb: Workbook,
        filename: str,
        *,
        cells: int = 0,
        row_count: int | None = None,
    ) -> XlsxArtifact:
        stream = SpooledTemporaryFile(max_size=XLSX_SPOOL_MAX_MEMORY_BYTES, mode="w+b")
        started_at = time.perf_counter()
        baseline_peak_rss = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss * 1024
        baseline_current_rss = _current_rss_bytes()
        try:
            if (
                EXPORT_MAX_PROCESS_RSS_BYTES == 0
                or max(baseline_peak_rss, baseline_current_rss)
                > EXPORT_MAX_PROCESS_RSS_BYTES
            ):
                EXPORT_REJECTED_TOTAL.labels("peak_rss_bytes").inc()
                raise ExportValidationError("Procesul web depaseste limita absoluta de memorie RSS.")
            wb.save(stream)
            size = stream.tell()
            if size > EXPORT_MAX_OUTPUT_BYTES:
                EXPORT_REJECTED_TOTAL.labels("output_bytes").inc()
                raise ExportValidationError("Fisierul XLSX depaseste limita de dimensiune de output.")
            stream.seek(0)
            digest = sha256()
            for chunk in iter(lambda: stream.read(256 * 1024), b""):
                digest.update(chunk)
            stream.seek(0)
            peak_rss = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss * 1024
            current_rss = _current_rss_bytes()
            build_seconds = time.perf_counter() - started_at
            EXPORT_BUILD_SECONDS.observe(build_seconds)
            EXPORT_OUTPUT_BYTES.set(size)
            EXPORT_CELLS.set(cells)
            EXPORT_PEAK_RSS_BYTES.set(peak_rss)
            # Current RSS attributes retained memory to this build, while the
            # lifetime peak enforces the absolute web-process safety ceiling.
            current_growth = max(0, current_rss - baseline_current_rss)
            peak_growth = max(0, peak_rss - baseline_peak_rss)
            if (
                EXPORT_MAX_PEAK_RSS_BYTES == 0
                or max(peak_rss, current_rss) > EXPORT_MAX_PROCESS_RSS_BYTES
                or max(current_growth, peak_growth) > EXPORT_MAX_PEAK_RSS_BYTES
            ):
                EXPORT_REJECTED_TOTAL.labels("peak_rss_bytes").inc()
                raise ExportValidationError("Exportul depaseste limita de memorie RSS.")
            return XlsxArtifact(
                stream=stream,
                filename=filename,
                size=size,
                sha256=digest.hexdigest(),
                peak_rss_bytes=peak_rss,
                build_seconds=build_seconds,
                cell_count=cells,
                row_count=row_count,
            )
        except Exception:
            stream.close()
            raise
        finally:
            wb.close()

    def _safe_filename(self, value: str) -> str:
        return safe_filename(value)

    def _excel_number_format(self, column_type: str) -> str | None:
        return number_format(column_type)

    def _write_table_sheet(
        self,
        ws: Any,
        columns: list[dict[str, str]],
        rows: list[dict[str, Any]],
        *,
        header_fill: str,
    ) -> None:
        write_table_sheet(ws, columns, rows, header_fill=header_fill)

    def _add_daily_comparison_chart(
        self,
        ws: Any,
        *,
        months: list[str],
        metric: str,
        max_row: int,
        first_data_col: int,
    ) -> None:
        add_daily_comparison_chart(
            ws,
            months=months,
            metric_label=DAILY_EVOLUTION_METRICS[metric].label,
            max_row=max_row,
            first_data_col=first_data_col,
        )

    def _add_daily_evolution_sheet(
        self,
        wb: Workbook,
        *,
        months: list[str],
        metrics: list[str],
        records: list[Any],
    ) -> None:
        if not metrics:
            return
        ws = wb.create_sheet("Evolutie zilnica")
        values: dict[tuple[int, str], dict[str, Any]] = {}
        max_day = 31
        for record in records:
            day = int(record["day_of_month"] or 0)
            if day <= 0:
                continue
            max_day = max(max_day, day)
            values[(day, str(record["import_month"]))] = self._compute_metrics(record)

        headers = ["Zi"]
        for metric in metrics:
            definition = DAILY_EVOLUTION_METRICS[metric]
            for month in months:
                headers.append(f"{month} {definition.label}")
            if len(months) == 2:
                if definition.type == "percent":
                    headers.append(f"Delta pp {definition.label}")
                else:
                    headers.append(f"Delta {definition.label}")
                    headers.append(f"Delta % {definition.label}")
        append_openpyxl_row(ws, headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1f2937")
            cell.fill = PatternFill("solid", fgColor="DCFCE7")

        for day in range(1, max_day + 1):
            row: list[Any] = [day]
            for metric in metrics:
                month_values: list[Any] = []
                for month in months:
                    value = values.get((day, month), {}).get(metric)
                    month_values.append(value)
                    row.append(self._json_value(value))
                if len(months) == 2:
                    left = month_values[0]
                    right = month_values[1]
                    if left is not None and right is not None:
                        delta = Decimal(str(right)) - Decimal(str(left))
                        row.append(self._json_value(delta))
                        if DAILY_EVOLUTION_METRICS[metric].type != "percent":
                            row.append(self._json_value(pct(delta, Decimal(str(left))) if Decimal(str(left)) != 0 else None))
                    else:
                        row.append(None)
                        if DAILY_EVOLUTION_METRICS[metric].type != "percent":
                            row.append(None)
            append_openpyxl_row(ws, row)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions["A"].width = 8
        for idx, header in enumerate(headers[1:], start=2):
            letter = get_column_letter(idx)
            ws.column_dimensions[letter].width = max(14, min(26, len(header) + 2))
        if months:
            chart = LineChart()
            chart.title = f"Evolutie zilnica - {DAILY_EVOLUTION_METRICS[metrics[0]].label}"
            chart.y_axis.title = DAILY_EVOLUTION_METRICS[metrics[0]].label
            self._configure_day_axis(chart)
            data = Reference(ws, min_col=2, max_col=1 + len(months), min_row=1, max_row=max_day + 1)
            categories = Reference(ws, min_col=1, min_row=2, max_row=max_day + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 18
            ws.add_chart(chart, f"{get_column_letter(len(headers) + 2)}2")

    def _configure_day_axis(self, chart: LineChart) -> None:
        configure_day_axis(chart)
