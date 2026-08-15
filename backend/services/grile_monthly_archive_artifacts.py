"""Filesystem and workbook primitives for monthly Grile archives."""

from __future__ import annotations

import os
import shutil
import time
import zipfile
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from services.grile_monthly_integrity import MonthlyIntegrityError, file_sha256
from services.grile_monthly_types import StoreEntry


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def export_sheet_xlsx(
    drive_service: Any,
    entry: StoreEntry,
    output_path: Path,
    *,
    downloader_type: Any,
    secure_directory: Callable[[Path], None],
    secure_file: Callable[[Path], None],
) -> dict[str, Any]:
    result = _export_result(entry, output_path)
    secure_directory(output_path.parent)
    request = drive_service.files().export_media(fileId=entry.sheet_id, mimeType=XLSX_MIME)
    try:
        with output_path.open("wb") as handle:
            downloader = downloader_type(handle, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        secure_file(output_path)
        result["bytes"] = output_path.stat().st_size
        _require_nonempty_export(result["bytes"])
        return result
    except Exception:
        output_path.unlink(missing_ok=True)
        raise


def write_exported_xlsx(
    entry: StoreEntry,
    output_path: Path,
    content: bytes,
    *,
    secure_directory: Callable[[Path], None],
    secure_file: Callable[[Path], None],
) -> dict[str, Any]:
    if not isinstance(content, bytes) or not content:
        _require_nonempty_export(0)
    secure_directory(output_path.parent)
    try:
        output_path.write_bytes(content)
        secure_file(output_path)
        size = output_path.stat().st_size
        _require_nonempty_export(size)
    except Exception:
        output_path.unlink(missing_ok=True)
        raise
    result = _export_result(entry, output_path)
    result["bytes"] = size
    return result


def _require_nonempty_export(size: int) -> None:
    if size == 0:
        raise MonthlyIntegrityError(
            "empty_source_backup",
            "Exported source backup is empty",
        )


def _export_result(entry: StoreEntry, output_path: Path) -> dict[str, Any]:
    return {
        "company": entry.company,
        "store": entry.store,
        "site_code": entry.site_code,
        "manager": entry.manager,
        "sheet_id": entry.sheet_id,
        "template_version": entry.template_version,
        "status": "OK",
        "xlsx_path": str(output_path),
        "bytes": 0,
        "error": "",
    }


def create_archive_zip(
    zip_path: Path,
    exported_files: list[Path],
    archive_dir: Path,
    *,
    secure_directory: Callable[[Path], None],
) -> None:
    secure_directory(zip_path.parent)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in exported_files:
            archive.write(path, path.relative_to(archive_dir).as_posix())


def create_manager_zips(
    output_dir: Path,
    month: str,
    results: list[dict[str, Any]],
    *,
    build_archive_dir: Callable[[Path, str], Path],
    build_manager_zip_path: Callable[[Path, str, str], Path],
    create_zip: Callable[[Path, list[Path], Path], None],
) -> dict[str, Path]:
    archive_dir = build_archive_dir(output_dir, month)
    files_by_manager: dict[str, list[Path]] = {}
    for item in results:
        if item.get("status") == "OK":
            files_by_manager.setdefault(
                item.get("manager") or "Neatribuit",
                [],
            ).append(Path(item["xlsx_path"]))
    zip_paths: dict[str, Path] = {}
    for manager, files in sorted(files_by_manager.items()):
        zip_path = build_manager_zip_path(output_dir, month, manager)
        create_zip(zip_path, files, archive_dir)
        zip_paths[manager] = zip_path
    return zip_paths


def summarize_archive_results(
    month: str,
    registry_count: int,
    results: list[dict[str, Any]],
    zip_path: Path,
    manager_zip_paths: dict[str, Path] | None,
    *,
    now: Callable[[], str],
) -> dict[str, Any]:
    return {
        "month": month,
        "created_at": now(),
        "registry_count": registry_count,
        "exported_count": sum(item.get("status") == "OK" for item in results),
        "error_count": sum(item.get("status") != "OK" for item in results),
        "zip_path": str(zip_path),
        "manager_zip_paths": {
            manager: str(path) for manager, path in sorted((manager_zip_paths or {}).items())
        },
        "stores": results,
    }


def validate_archive_zip(zip_path: Path, *, expected_files: int) -> None:
    try:
        with zipfile.ZipFile(zip_path) as archive:
            members = [item for item in archive.infolist() if not item.is_dir()]
            unique = {item.filename for item in members}
            if len(members) != expected_files or len(unique) != expected_files:
                raise MonthlyIntegrityError(
                    "archive_coverage_incomplete",
                    "Archive coverage is incomplete",
                )
            if archive.testzip() is not None:
                raise MonthlyIntegrityError("archive_corrupt", "Archive is corrupt")
    except MonthlyIntegrityError:
        raise
    except Exception as exc:
        raise MonthlyIntegrityError("archive_invalid", "Archive cannot be verified") from exc


def validate_source_workbook(path: Path) -> None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if not {"Grila", "Pontaj"}.issubset(workbook.sheetnames):
                raise MonthlyIntegrityError(
                    "source_workbook_partial",
                    "Source workbook is missing required sheets",
                )
        finally:
            workbook.close()
    except MonthlyIntegrityError:
        raise
    except Exception as exc:
        raise MonthlyIntegrityError("source_workbook_invalid", "Source workbook is invalid") from exc


def future_artifact(
    staged_path: Path,
    *,
    outputs_dir: Path,
    staged_archive_dir: Path,
    official_archive_dir: Path,
    kind: str,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    relative_inside = staged_path.resolve().relative_to(staged_archive_dir.resolve())
    future_path = official_archive_dir / relative_inside
    artifact = {
        "kind": kind,
        "path": future_path.resolve().relative_to(outputs_dir.resolve()).as_posix(),
        "bytes": staged_path.stat().st_size,
        "sha256": file_sha256(staged_path),
    }
    if extra:
        artifact.update(extra)
    return artifact


def promote_directory(
    staged: Path,
    destination: Path,
    *,
    outputs_dir: Path,
    safe_filename: Callable[[str], str],
    secure_directory: Callable[[Path], None],
    verify: Callable[[], None] | None = None,
) -> None:
    secure_directory(destination.parent)
    revision = _preserve_directory_revision(
        outputs_dir,
        destination,
        safe_filename,
        secure_directory,
    )
    promoted = False
    try:
        os.replace(staged, destination)
        promoted = True
        if verify is not None:
            verify()
    except Exception as exc:
        rollback_error = _rollback_directory_promotion(
            staged,
            destination,
            revision,
            promoted,
        )
        if rollback_error is not None:
            raise MonthlyIntegrityError(
                "archive_promotion_rollback_failed",
                "Archive promotion rollback failed",
            ) from exc
        raise


def _preserve_directory_revision(
    outputs_dir: Path,
    destination: Path,
    safe_filename: Callable[[str], str],
    secure_directory: Callable[[Path], None],
) -> Path | None:
    if not destination.exists():
        return None
    revision_dir = outputs_dir / ".revisions"
    secure_directory(revision_dir)
    revision = revision_dir / f"archive-{safe_filename(destination.name)}-{time.time_ns()}"
    os.replace(destination, revision)
    return revision


def _rollback_directory_promotion(
    staged: Path,
    destination: Path,
    revision: Path | None,
    promoted: bool,
) -> Exception | None:
    rollback_error: Exception | None = None
    if promoted and destination.exists():
        try:
            os.replace(destination, staged)
        except Exception:
            try:
                shutil.rmtree(destination)
            except Exception as exc:  # noqa: BLE001 - returned to caller
                rollback_error = exc
    if revision is not None and revision.exists() and not destination.exists():
        try:
            os.replace(revision, destination)
        except Exception as exc:  # noqa: BLE001 - returned to caller
            rollback_error = exc
    return rollback_error
