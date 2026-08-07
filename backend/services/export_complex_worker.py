"""Process-isolated XLSX renderer for the complex daily comparison export."""

from __future__ import annotations

import os
import resource
import tempfile
import time
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from services.spreadsheet_safety import append_openpyxl_row
from services.export_xlsx_formatting import (
    add_daily_comparison_chart,
    days_filename_suffix,
    safe_filename,
    write_table_sheet,
)


def _peak_rss_bytes() -> int:
    """Return Linux RSS in bytes (the supported production platform)."""
    return resource.getrusage(resource.RUSAGE_SELF).ru_maxrss * 1024


def _enforce_memory_limit(limit_bytes: int) -> None:
    """Fence a chart writer before it can allocate an unbounded cell graph.

    The web process only sends charted exports to this spawned process.  Keep
    the inherited hard limit intact: an operator may have configured a lower
    cgroup/systemd ceiling, which must always win.
    """
    if limit_bytes <= 0:
        raise ValueError("complex export memory limit must be positive")
    soft, hard = resource.getrlimit(resource.RLIMIT_AS)
    effective = limit_bytes if hard == resource.RLIM_INFINITY else min(limit_bytes, hard)
    if soft != resource.RLIM_INFINITY:
        effective = min(effective, soft)
    resource.setrlimit(resource.RLIMIT_AS, (effective, hard))


def _assert_memory_budget(limit_bytes: int) -> int:
    peak_rss = _peak_rss_bytes()
    if peak_rss > limit_bytes:
        raise MemoryError("complex export exceeded RSS budget")
    return peak_rss


def _private_output_path(output_path: str | None) -> Path:
    if output_path is None:
        descriptor, path_value = tempfile.mkstemp(
            prefix="unihub-export-",
            suffix=".xlsx",
        )
        os.close(descriptor)
        return Path(path_value)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    descriptor = os.open(path, os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o600)
    os.close(descriptor)
    return path


def _save_hashed_workbook(
    workbook: Workbook,
    *,
    max_output_bytes: int,
    max_peak_rss_bytes: int,
    output_path: str | None = None,
) -> dict[str, Any]:
    """Save to a private temporary artifact and attest its exact bytes."""
    _assert_memory_budget(max_peak_rss_bytes)
    path = _private_output_path(output_path)
    try:
        workbook.save(path)
        size = path.stat().st_size
        if size > max_output_bytes:
            raise ValueError("complex export exceeded output budget")
        peak_rss = _assert_memory_budget(max_peak_rss_bytes)
        digest = sha256()
        with path.open("rb") as artifact:
            for chunk in iter(lambda: artifact.read(256 * 1024), b""):
                digest.update(chunk)
        return {
            "path": str(path),
            "size": size,
            "sha256": digest.hexdigest(),
            "peak_rss": peak_rss,
        }
    except Exception:
        path.unlink(missing_ok=True)
        raise


def render_daily_comparison_xlsx(payload: dict[str, Any]) -> dict[str, Any]:
    """Render one complex workbook and return only a temporary path and budgets."""
    started_at = time.perf_counter()
    request = payload["request"]
    months: list[str] = payload["months"]
    metrics: list[str] = payload["metrics"]
    levels: list[str] = payload["levels"]
    level_config: dict[str, dict[str, Any]] = payload["level_config"]
    metric_labels: dict[str, str] = payload["metric_labels"]
    selected_days: list[int] | None = payload["selected_days"]
    tables: list[tuple[str, dict[str, Any]]] = payload["tables"]
    include_closed_stores = bool(payload["include_closed_stores"])

    max_output_bytes = int(payload["max_output_bytes"])
    max_peak_rss_bytes = int(payload["max_peak_rss_bytes"])
    _enforce_memory_limit(max_peak_rss_bytes)
    workbook = Workbook()
    first_sheet = True
    total_rows = 0
    try:
        for level, table in tables:
            sheet_name = level_config[level]["sheet"]
            ws = workbook.active if first_sheet else workbook.create_sheet(sheet_name)
            ws.title = sheet_name
            first_sheet = False
            write_table_sheet(
                ws,
                table["columns"],
                table["rows"],
                header_fill="DCFCE7",
            )
            total_rows += len(table["rows"])
            add_daily_comparison_chart(
                ws,
                months=months,
                metric_label=metric_labels[metrics[0]],
                max_row=len(table["rows"]) + 1,
                first_data_col=len(level_config[level]["dimensions"]) + 2,
            )

        config = workbook.create_sheet("Configuratie")
        append_openpyxl_row(config, ["Optiune", "Valoare"])
        append_openpyxl_row(config, ["Tip export", "Evolutie zilnica comparativa"])
        append_openpyxl_row(config, ["Luni", ", ".join(months)])
        append_openpyxl_row(
            config,
            ["Zile", ", ".join(str(day) for day in selected_days) if selected_days else "Toata luna"],
        )
        append_openpyxl_row(config, ["Metrici zilnice", ", ".join(metric_labels[item] for item in metrics)])
        append_openpyxl_row(config, ["Niveluri", ", ".join(str(level_config[item]["label"]) for item in levels)])
        append_openpyxl_row(config, ["Include magazine inchise", "Da" if include_closed_stores else "Nu"])
        append_openpyxl_row(config, ["Generat", time.strftime("%Y-%m-%d %H:%M")])
        append_openpyxl_row(config, ["Randuri", total_rows])
        for cell in config[1]:
            cell.font = Font(bold=True)
        config.column_dimensions["A"].width = 28
        config.column_dimensions["B"].width = 72

        filename = request.get("filename") or (
            f"export_retail_evolutie_zilnica_{'_'.join(months)}"
            f"{days_filename_suffix(selected_days)}.xlsx"
        )
        artifact = _save_hashed_workbook(
            workbook,
            max_output_bytes=max_output_bytes,
            max_peak_rss_bytes=max_peak_rss_bytes,
            output_path=payload.get("output_path"),
        )
        return {
            **artifact,
            "filename": safe_filename(str(filename)),
            "build_seconds": time.perf_counter() - started_at,
        }
    finally:
        workbook.close()


def render_daily_metrics_xlsx(payload: dict[str, Any]) -> dict[str, Any]:
    """Build the charted daily-metrics variant in the same fenced process.

    The compatibility renderer deliberately remains the single source for the
    workbook layout.  This worker owns only isolation, output attestation and
    lifecycle, so no database work or web-request state crosses the boundary.
    """
    started_at = time.perf_counter()
    max_output_bytes = int(payload["max_output_bytes"])
    max_peak_rss_bytes = int(payload["max_peak_rss_bytes"])
    _enforce_memory_limit(max_peak_rss_bytes)

    # Import at execution time: this module is imported by exports.py and the
    # spawned child must not form an import cycle while module initialization is
    # still in progress.
    from services.exports import ExportsService

    service = ExportsService(None)  # type: ignore[arg-type]
    artifact = None
    try:
        artifact = service._render_table_xlsx(
            payload["request"],
            payload["result"],
            payload["selected_days"],
            payload["daily_rows"],
        )
        path = _private_output_path(
            str(payload["output_path"]) if payload.get("output_path") else None
        )
        try:
            with path.open("wb") as destination:
                for chunk in artifact.iter_chunks():
                    destination.write(chunk)
            size = path.stat().st_size
            if size > max_output_bytes:
                raise ValueError("complex export exceeded output budget")
            peak_rss = _assert_memory_budget(max_peak_rss_bytes)
            digest = sha256()
            with path.open("rb") as output:
                for chunk in iter(lambda: output.read(256 * 1024), b""):
                    digest.update(chunk)
            return {
                "path": str(path),
                "filename": safe_filename(str(payload["filename"])),
                "size": size,
                "sha256": digest.hexdigest(),
                "peak_rss": peak_rss,
                "build_seconds": time.perf_counter() - started_at,
            }
        except Exception:
            path.unlink(missing_ok=True)
            raise
    finally:
        if artifact is not None:
            artifact.close()
