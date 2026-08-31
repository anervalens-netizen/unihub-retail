from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import math
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from services.spreadsheet_safety import (
    TrustedFormula,
    SpreadsheetUploadError,
    SpreadsheetUploadLimits,
    SpreadsheetParserMeasurement,
    ERP_RECONCILIATION_SPREADSHEET_LIMITS,
    HISTORY_SPREADSHEET_LIMITS,
    PROMO_ACTUALS_SPREADSHEET_LIMITS,
    SALES_SPREADSHEET_LIMITS,
    TARGETS_SPREADSHEET_LIMITS,
    append_openpyxl_row,
    csv_cell_value,
    google_sheets_value,
    sanitize_dataframe_text,
    sanitize_spreadsheet_text,
    spreadsheet_cell_value,
    validate_spreadsheet_upload,
)
from services import legacy_xls


def workbook_bytes(rows: int = 1) -> bytes:
    workbook = Workbook()
    for index in range(rows):
        workbook.active.cell(row=index + 1, column=1, value=index)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_xlsx_structural_preflight_accepts_bounded_workbook() -> None:
    content = workbook_bytes(2)
    stats = validate_spreadsheet_upload(content, ".xlsx")

    assert stats.compressed_bytes == len(content)
    assert stats.uncompressed_bytes is not None
    assert stats.uncompressed_bytes > len(content)
    assert stats.cells == 2


def test_spreadsheet_preflight_rejects_signature_and_cell_budget() -> None:
    with pytest.raises(SpreadsheetUploadError, match="semnătură"):
        validate_spreadsheet_upload(b"not-a-workbook", ".xlsx")
    with pytest.raises(SpreadsheetUploadError, match="celule"):
        validate_spreadsheet_upload(
            workbook_bytes(2),
            ".xlsx",
            limits=SpreadsheetUploadLimits(max_cells=1),
        )


def test_spreadsheet_preflight_rejects_expansion_budget() -> None:
    with pytest.raises(SpreadsheetUploadError, match="decomprimat"):
        validate_spreadsheet_upload(
            workbook_bytes(),
            ".xlsx",
            limits=SpreadsheetUploadLimits(max_uncompressed_bytes=1),
        )


def test_xls_preflight_reports_unavailable_structure_honestly() -> None:
    content = bytes.fromhex("d0cf11e0a1b11ae1") + b"legacy"

    stats = validate_spreadsheet_upload(content, ".xls")

    assert stats.source_bytes == len(content)
    assert stats.compressed_bytes is None
    assert stats.uncompressed_bytes is None
    assert stats.cells is None
    assert stats.format == "xls"


def test_numeric_legacy_sheet_selection_is_delegated_to_bounded_child(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, object] = {}

    def bounded_parse(source: bytes, *, sheets: object, limits: object) -> object:
        captured["source"] = source
        captured["sheets"] = sheets
        captured["limits"] = limits
        return legacy_xls.LegacyXlsWorkbook(
            (legacy_xls.LegacyXlsSheet("first", (("header",), ("value",))),)
        )

    monkeypatch.setattr(legacy_xls, "parse_legacy_xls", bounded_parse)
    frame = legacy_xls.read_legacy_xls_frame(b"untrusted-xls", sheet_name=0)

    assert captured["sheets"] == [0]
    assert frame.to_dict(orient="records") == [{"header": "value"}]


def test_import_policies_are_explicit_and_parser_measurement_is_finite() -> None:
    assert len(
        {
            SALES_SPREADSHEET_LIMITS.max_cells,
            PROMO_ACTUALS_SPREADSHEET_LIMITS.max_cells,
            ERP_RECONCILIATION_SPREADSHEET_LIMITS.max_cells,
            TARGETS_SPREADSHEET_LIMITS.max_cells,
            HISTORY_SPREADSHEET_LIMITS.max_cells,
        }
    ) == 5
    assert SALES_SPREADSHEET_LIMITS.max_member_bytes == 128 * 1024 * 1024
    assert (
        SALES_SPREADSHEET_LIMITS.max_member_bytes
        <= SALES_SPREADSHEET_LIMITS.max_uncompressed_bytes
    )
    content = workbook_bytes(2)
    measurement = SpreadsheetParserMeasurement("test_parser")
    with measurement:
        measurement.set_preflight(validate_spreadsheet_upload(content, ".xlsx"))
        measurement.set_rows(2)

    resources = measurement.as_dict()
    for key in ("source_bytes", "compressed_bytes", "expanded_bytes", "cells", "rows", "parse_seconds", "peak_rss_bytes"):
        value = resources[key]
        assert value is not None
        assert math.isfinite(float(value))


@pytest.mark.parametrize(
    "value",
    [
        "=1+1",
        "+SUM(1,1)",
        "-1+2",
        "@SUM(1,1)",
        "\t=1",
        "\r=1",
        "\n=1",
        "-123",
    ],
)
def test_untrusted_formula_prefixes_are_explicit_text(value: str) -> None:
    wb = Workbook()
    ws = wb.active
    append_openpyxl_row(ws, [value])
    cell = ws["A1"]
    assert cell.data_type == "s"
    assert cell.value == "'" + value


def test_native_values_and_trusted_formula_round_trip() -> None:
    wb = Workbook()
    ws = wb.active
    append_openpyxl_row(
        ws,
        [
            "Buna ziua",
            "",
            None,
            -123,
            -1.5,
            Decimal("-2.3"),
            True,
            date(2026, 1, 2),
            datetime(2026, 1, 2, 3, 4),
            TrustedFormula("=SUM(A1:A1)"),
        ],
    )
    stream = BytesIO()
    wb.save(stream)
    reopened = load_workbook(BytesIO(stream.getvalue()), data_only=False)
    assert reopened.active["J1"].data_type == "f"
    assert reopened.active["J1"].value == "=SUM(A1:A1)"
    assert reopened.active["D1"].value == -123
    with ZipFile(BytesIO(stream.getvalue())) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml").decode()
    assert "<f>" in xml


def test_untrusted_hyperlink_is_not_formula_in_xlsx_xml() -> None:
    payload = '=HYPERLINK("https://example.invalid","click")'
    wb = Workbook()
    append_openpyxl_row(wb.active, [payload])
    stream = BytesIO()
    wb.save(stream)
    with ZipFile(BytesIO(stream.getvalue())) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml").decode()
    assert "<f>HYPERLINK" not in xml
    assert (
        load_workbook(BytesIO(stream.getvalue()), data_only=False)
        .active["A1"]
        .data_type
        == "s"
    )


def test_invalid_trusted_formula_is_rejected() -> None:
    with pytest.raises(ValueError, match="must start"):
        TrustedFormula("SUM(A1:A2)")


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("Buna ziua", "Buna ziua"),
        ("", ""),
        (None, None),
        ("Ștefan", "Ștefan"),
        (-4, -4),
        (-1.5, -1.5),
        (Decimal("-2.5"), Decimal("-2.5")),
        (True, True),
        (date(2026, 1, 2), date(2026, 1, 2)),
        (datetime(2026, 1, 2, 3, 4), datetime(2026, 1, 2, 3, 4)),
    ],
)
def test_central_value_api_preserves_safe_native_values(
    value: object,
    expected: object,
) -> None:
    assert spreadsheet_cell_value(value) == expected


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@", "\t", "\r", "\n"])
def test_csv_and_google_values_neutralize_all_formula_prefixes(prefix: str) -> None:
    value = prefix + "1+1"
    assert csv_cell_value(value) == "'" + value
    assert google_sheets_value(value) == "'" + value


def test_append_rows_do_not_overwrite_leading_or_fully_empty_rows() -> None:
    wb = Workbook()
    ws = wb.active
    append_openpyxl_row(ws, [None, "primul"])
    append_openpyxl_row(ws, ["al doilea", "valoare"])
    append_openpyxl_row(ws, [None, None])
    append_openpyxl_row(ws, ["dupa-rand-gol"])

    assert ws["B1"].value == "primul"
    assert ws["A2"].value == "al doilea"
    assert ws["B2"].value == "valoare"
    assert ws["A3"].value is None
    assert ws["A4"].value == "dupa-rand-gol"


def test_non_finite_decimal_and_broken_string_conversion_are_safe() -> None:
    class BrokenString:
        def __str__(self) -> str:
            raise RuntimeError("no string")

    assert spreadsheet_cell_value(Decimal("NaN")) == "NaN"
    assert spreadsheet_cell_value(Decimal("Infinity")) == "Infinity"
    assert spreadsheet_cell_value(BrokenString()) == "<BrokenString>"
    assert sanitize_spreadsheet_text("=1+1") == "'=1+1"


def test_dataframe_boundary_only_neutralizes_text_columns() -> None:
    import pandas as pd

    safe = sanitize_dataframe_text(
        pd.DataFrame({"text": ["=1+1"], "amount": [-12]})
    )
    assert safe.loc[0, "text"] == "'=1+1"
    assert safe.loc[0, "amount"] == -12


def test_dataframe_boundary_preserves_missing_and_extension_text() -> None:
    import pandas as pd

    frame = pd.DataFrame(
        {
            "string_text": pd.Series(["=1+1", pd.NA], dtype="string"),
            "category_text": pd.Series(["+SUM(1,1)", None], dtype="category"),
            "object_text": ["@SUM(1,1)", None],
            "amount": [-12, -13],
        }
    )

    safe = sanitize_dataframe_text(frame)

    assert safe.loc[0, "string_text"] == "'=1+1"
    assert safe.loc[1, "string_text"] is None
    assert safe.loc[0, "category_text"] == "'+SUM(1,1)"
    assert safe.loc[1, "category_text"] is None
    assert safe.loc[0, "object_text"] == "'@SUM(1,1)"
    assert safe.loc[1, "object_text"] is None
    assert list(safe["amount"]) == [-12, -13]
