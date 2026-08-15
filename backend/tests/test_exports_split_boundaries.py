from __future__ import annotations

import hashlib
import asyncio
import resource
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any, cast
from unittest.mock import MagicMock, mock_open

import pytest
from openpyxl import load_workbook

import services.export_complex_worker as child_renderer
import services.exports.service as service_module
from repositories.export_daily_comparison_query import (
    build_daily_comparison_rows_query,
)
from services.export_xlsx_formatting import days_filename_suffix
from services.exports import ExportValidationError, ExportsService


def metric_row(**overrides: Any) -> dict[str, Any]:
    value: dict[str, Any] = {
        "import_month": "2026-05",
        "day_of_month": 1,
        "total_sales": 100,
        "total_quantity": 10,
        "total_receipts": 5,
        "receipt_2plus_count": 1,
        "focus_quantity": 2,
        "target": 200,
        "working_days": 1,
        "store_count": 1,
        "agent_count": 1,
        "incentive_sales": 0,
        "incentive_quantity": 0,
        "incentive_bonus": 0,
        "promo_sales": 0,
        "promo_quantity": 0,
    }
    value.update(overrides)
    return value


def daily_metrics_request() -> dict[str, Any]:
    return {
        "export_mode": "table",
        "dataset": "stores",
        "months": ["2026-05", "2026-06"],
        "dimensions": ["site_code"],
        "metrics": ["total_sales"],
        "monthly_metrics": [],
        "daily_metrics": ["total_sales", "proc_bon2acc"],
        "selected_days": [1],
    }


def test_daily_comparison_query_boundary_preserves_scope_and_campaign_order() -> None:
    query, params = build_daily_comparison_rows_query(
        level="stores",
        months=["2026-05", "2026-06"],
        filters={
            "firma": ["Mobiup"],
            "regional": ["RM 1"],
            "agent": ["Agent 1"],
        },
        include_closed_stores=False,
        campaign_codes_by_month={"2026-06": ["P1", "P2"]},
        selected_days=[1, 9],
        include_campaign_metrics=True,
        limit=25,
    )

    assert params == [
        ["2026-05", "2026-06"],
        ["Mobiup"],
        ["RM 1"],
        ["Agent 1"],
        [1, 9],
        ["2026-06", "2026-06"],
        ["P1", "P2"],
        True,
        25,
    ]
    assert "agg.site_code AS site_code" in query
    assert "campaign.site_code IS NOT DISTINCT FROM base.site_code" in query
    assert "s.is_active = TRUE" in query
    assert "UNNEST($6::TEXT[], $7::TEXT[])" in query
    assert "AND $8::BOOLEAN" in query
    assert query.rstrip().endswith("LIMIT $9")


def test_complex_request_boundary_accepts_only_supported_durable_shapes() -> None:
    service = ExportsService(cast(Any, None))
    daily_request = daily_metrics_request()

    assert ExportsService.is_complex_request(daily_request)
    assert service.validate_complex_request(daily_request) == "daily_metrics"
    assert service.validate_complex_request(
        {
            "export_mode": "daily_comparison",
            "months": ["2026-05", "2026-06"],
            "daily_metrics": ["total_sales"],
            "comparison_levels": ["general"],
            "selected_days": [1],
        }
    ) == "daily_comparison"
    assert not ExportsService.is_complex_request({"export_mode": "table"})

    invalid_requests = [
        {"export_mode": "table", "dataset": "stores", "months": ["2026-05"]},
        {
            **daily_request,
            "dataset": "incentive_products",
        },
        {
            **daily_request,
            "months": [],
        },
        {
            **daily_request,
            "daily_metrics": ["total_sales"] * 8,
        },
        {
            **daily_request,
            "monthly_metrics": ["total_sales"] * 20,
        },
    ]
    for request in invalid_requests:
        with pytest.raises(ExportValidationError):
            service.validate_complex_request(request)


def test_in_process_daily_renderer_preserves_layout_deltas_and_formats() -> None:
    service = ExportsService(cast(Any, None))
    request = daily_metrics_request()
    result = {
        "columns": [
            {"key": "site_code", "label": "Magazin", "type": "text", "group": "Identificare"},
            {"key": "total_sales", "label": "Vânzări", "type": "currency", "group": "Metrici"},
        ],
        "rows": [{"site_code": "S1", "total_sales": 250}],
    }
    daily_rows = [
        metric_row(import_month="2026-05", total_sales=100, receipt_2plus_count=1),
        metric_row(import_month="2026-06", total_sales=150, receipt_2plus_count=3),
    ]

    artifact = service._render_table_xlsx(request, result, [1], daily_rows)
    try:
        workbook = load_workbook(BytesIO(b"".join(artifact.iter_chunks())))
    finally:
        artifact.close()

    assert workbook.sheetnames == ["Raport", "Configuratie", "Evolutie zilnica"]
    assert workbook["Raport"]["B2"].number_format == "#,##0.00"
    evolution = workbook["Evolutie zilnica"]
    assert evolution["D2"].value == 50
    assert evolution["E2"].value == 50
    assert len(evolution._charts) == 1
    workbook.close()


@pytest.mark.asyncio
async def test_complex_artifact_adoption_rejects_worker_and_attestation_failures(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    service = ExportsService(cast(Any, None))

    async def rejected(*_args: Any, **_kwargs: Any) -> dict[str, Any]:
        raise service_module.ExportRendererProcessError(
            "renderer_failed", "worker rejected payload"
        )

    monkeypatch.setattr(service_module, "run_export_renderer_process", rejected)
    with pytest.raises(ExportValidationError, match="siguranta"):
        await service._run_complex_renderer(
            "daily_metrics",
            {"filename": "report.xlsx"},
            cells=1,
        )

    missing_directory = tmp_path / "unihub-export-operation-missing"
    missing_directory.mkdir()
    missing = missing_directory / "missing.xlsx"

    async def missing_result(*_args: Any, **_kwargs: Any) -> dict[str, Any]:
        return {
            "path": str(missing),
            "size": 1,
            "sha256": "0" * 64,
            "peak_rss": 1,
            "filename": "report.xlsx",
            "operation_directory": str(missing_directory),
        }

    monkeypatch.setattr(service_module, "run_export_renderer_process", missing_result)
    with pytest.raises(ExportValidationError, match="lipseste"):
        await service._run_complex_renderer(
            "daily_metrics",
            {"filename": "report.xlsx"},
            cells=1,
        )

    tampered_directory = tmp_path / "unihub-export-operation-tampered"
    tampered_directory.mkdir()
    tampered = tampered_directory / "tampered.xlsx"
    tampered.write_bytes(b"not-the-attested-content")

    async def tampered_result(*_args: Any, **_kwargs: Any) -> dict[str, Any]:
        return {
            "path": str(tampered),
            "size": tampered.stat().st_size,
            "sha256": "0" * 64,
            "peak_rss": 1,
            "filename": "report.xlsx",
            "operation_directory": str(tampered_directory),
        }

    monkeypatch.setattr(service_module, "run_export_renderer_process", tampered_result)
    with pytest.raises(ExportValidationError, match="integritate"):
        await service._run_complex_renderer(
            "daily_metrics",
            {"filename": "report.xlsx"},
            cells=1,
        )
    assert not tampered.exists()


@pytest.mark.asyncio
async def test_complex_renderer_rejects_unknown_renderer_name() -> None:
    service = ExportsService(cast(Any, None))

    with pytest.raises(ExportValidationError, match="siguranta"):
        await service._run_complex_renderer(
            cast(Any, "nightly_metrics"),
            {"filename": "report.xlsx"},
            cells=1,
        )


@pytest.mark.asyncio
async def test_complex_artifact_adoption_rejects_path_outside_operation_directory(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    service = ExportsService(cast(Any, None))

    operation_directory = tmp_path / "unihub-export-operation-outside"
    operation_directory.mkdir()
    outside_file = tmp_path / "outside.xlsx"
    outside_file.write_bytes(b"payload")

    async def outside_result(*_args: Any, **_kwargs: Any) -> dict[str, Any]:
        return {
            "path": str(outside_file),
            "size": outside_file.stat().st_size,
            "sha256": "0" * 64,
            "peak_rss": 1,
            "filename": "report.xlsx",
            "operation_directory": str(operation_directory),
        }

    monkeypatch.setattr(service_module, "run_export_renderer_process", outside_result)
    with pytest.raises(ExportValidationError, match="invalida"):
        await service._run_complex_renderer(
            "daily_metrics",
            {"filename": "report.xlsx"},
            cells=1,
        )


@pytest.mark.asyncio
async def test_cancelled_artifact_adoption_propagates_to_killable_process_boundary(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = ExportsService(cast(Any, None))
    started = asyncio.Event()
    cancelled = asyncio.Event()

    async def renderer_process(*_args: Any, **_kwargs: Any) -> dict[str, Any]:
        started.set()
        try:
            await asyncio.Future()
        except asyncio.CancelledError:
            cancelled.set()
            raise
        raise AssertionError("renderer process returned unexpectedly")

    monkeypatch.setattr(service_module, "run_export_renderer_process", renderer_process)
    task = asyncio.create_task(
        service._run_complex_renderer(
            "daily_metrics",
            {"filename": "report.xlsx"},
            cells=1,
        )
    )
    await started.wait()
    task.cancel()
    with pytest.raises(asyncio.CancelledError):
        await task
    assert cancelled.is_set()


def test_child_renderers_emit_hashed_artifacts_without_process_coverage_gaps(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(child_renderer, "_enforce_memory_limit", lambda _limit: None)
    monkeypatch.setattr(child_renderer, "_assert_memory_budget", lambda _limit: 4096)
    comparison_payload = {
        "request": {"filename": "comparison.xlsx"},
        "months": ["2026-05", "2026-06"],
        "metrics": ["total_sales"],
        "levels": ["general"],
        "level_config": {
            "general": {"label": "General", "sheet": "General", "dimensions": []}
        },
        "metric_labels": {"total_sales": "Vânzări"},
        "selected_days": [1],
        "tables": [
            (
                "general",
                {
                    "columns": [
                        {"key": "day_of_month", "label": "Zi", "type": "integer"},
                        {"key": "2026-05:total_sales", "label": "Mai", "type": "currency"},
                        {"key": "2026-06:total_sales", "label": "Iunie", "type": "currency"},
                    ],
                    "rows": [
                        {
                            "day_of_month": 1,
                            "2026-05:total_sales": 100,
                            "2026-06:total_sales": 150,
                        }
                    ],
                },
            )
        ],
        "include_closed_stores": False,
        "max_output_bytes": 64 * 1024 * 1024,
        "max_peak_rss_bytes": 512 * 1024 * 1024,
    }
    comparison = child_renderer.render_daily_comparison_xlsx(comparison_payload)
    comparison_path = Path(comparison["path"])
    try:
        content = comparison_path.read_bytes()
        assert hashlib.sha256(content).hexdigest() == comparison["sha256"]
        workbook = load_workbook(BytesIO(content))
        assert workbook.sheetnames == ["General", "Configuratie"]
        assert len(workbook["General"]._charts) == 1
        workbook.close()
    finally:
        comparison_path.unlink(missing_ok=True)

    request = daily_metrics_request()
    metrics_payload = {
        "request": request,
        "result": {
            "columns": [
                {"key": "site_code", "label": "Magazin", "type": "text"},
                {"key": "total_sales", "label": "Vânzări", "type": "currency"},
            ],
            "rows": [{"site_code": "S1", "total_sales": 250}],
        },
        "selected_days": [1],
        "daily_rows": [
            metric_row(import_month="2026-05", total_sales=100),
            metric_row(import_month="2026-06", total_sales=150),
        ],
        "filename": "daily.xlsx",
        "max_output_bytes": 64 * 1024 * 1024,
        "max_peak_rss_bytes": 512 * 1024 * 1024,
    }
    metrics = child_renderer.render_daily_metrics_xlsx(metrics_payload)
    metrics_path = Path(metrics["path"])
    try:
        content = metrics_path.read_bytes()
        assert hashlib.sha256(content).hexdigest() == metrics["sha256"]
        workbook = load_workbook(BytesIO(content))
        assert "Evolutie zilnica" in workbook.sheetnames
        workbook.close()
    finally:
        metrics_path.unlink(missing_ok=True)

    metrics_payload["max_output_bytes"] = 1
    with pytest.raises(ValueError, match="output budget"):
        child_renderer.render_daily_metrics_xlsx(metrics_payload)


def test_child_memory_and_output_fences_fail_closed(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    with pytest.raises(ValueError, match="positive"):
        child_renderer._enforce_memory_limit(0)

    real_peak_rss = child_renderer._peak_rss_bytes
    monkeypatch.setattr(child_renderer, "_peak_rss_bytes", lambda: 101)
    with pytest.raises(MemoryError, match="RSS"):
        child_renderer._assert_memory_budget(100)
    monkeypatch.setattr(child_renderer, "_peak_rss_bytes", real_peak_rss)

    monkeypatch.setattr(
        child_renderer.resource,
        "getrusage",
        lambda *_args: SimpleNamespace(ru_maxrss=1),
    )
    assert child_renderer._peak_rss_bytes() == 1024
    assert child_renderer._assert_memory_budget(2048) == 1024

    with monkeypatch.context() as statm_context:
        statm_context.setattr(Path, "open", mock_open(read_data="7 3\n"))
        statm_context.setattr(child_renderer.resource, "getpagesize", lambda: 4096)
        assert child_renderer._virtual_memory_bytes() == 7 * 4096

    set_limit = MagicMock()
    monkeypatch.setattr(child_renderer.resource, "setrlimit", set_limit)
    monkeypatch.setattr(child_renderer, "_peak_rss_bytes", lambda: 1024)
    monkeypatch.setattr(child_renderer, "_virtual_memory_bytes", lambda: 8192)
    monkeypatch.setattr(
        child_renderer.resource,
        "getrlimit",
        lambda *_args: (resource.RLIM_INFINITY, resource.RLIM_INFINITY),
    )
    child_renderer._enforce_memory_limit(4096)
    set_limit.assert_called_with(resource.RLIMIT_AS, (11264, resource.RLIM_INFINITY))

    monkeypatch.setattr(child_renderer.resource, "getrlimit", lambda *_args: (10000, 12000))
    child_renderer._enforce_memory_limit(4096)
    set_limit.assert_called_with(resource.RLIMIT_AS, (10000, 12000))

    monkeypatch.setattr(child_renderer, "_peak_rss_bytes", lambda: 4096)
    with pytest.raises(MemoryError, match="before rendering"):
        child_renderer._enforce_memory_limit(4096)

    monkeypatch.setattr(child_renderer, "_assert_memory_budget", lambda _limit: 1)
    from openpyxl import Workbook

    workbook = Workbook()
    with pytest.raises(ValueError, match="output budget"):
        child_renderer._save_hashed_workbook(
            workbook,
            max_output_bytes=1,
            max_peak_rss_bytes=100,
        )
    workbook.close()

    explicit_output = tmp_path / "private" / "artifact.xlsx"
    assert child_renderer._private_output_path(str(explicit_output)) == explicit_output
    assert explicit_output.stat().st_mode & 0o777 == 0o600
    with pytest.raises(FileExistsError):
        child_renderer._private_output_path(str(explicit_output))

    assert days_filename_suffix(None) == ""
    assert days_filename_suffix([1, 2]) == "_zile_1-2"
    assert days_filename_suffix(list(range(1, 12))) == "_zile_11selectate"
