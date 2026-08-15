from __future__ import annotations

import hashlib
import json
import os
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from services.dashboard_specials import (
    _generated_config_path,
    _promotion_products_cache,
    validate_special_cards_config,
)
from services.promo_copurchase import load_promo_actual_units
import services.promo_generation_migration as migration_module
from services.promo_generation_migration import (
    PromoGenerationMigrationError,
    migrate_legacy_promo_generation,
)


def _workbook(site: str, code: str, quantity: int, *, sheet_name: str = "AccesoriPromoLunar") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["SiteCode", "Cod", "Promo Luna Curenta", "PromoValoare Luna Curenta"])
    sheet.append([site, code, quantity, "20.00"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _rule_master_workbook() -> bytes:
    workbook = Workbook()
    trigger = workbook.active
    trigger.title = "Produse trigger"
    trigger.append(["Cod", "Denumire"])
    trigger.append(["TRIGGER-1", "Produs trigger"])
    discounted = workbook.create_sheet("Produse discount")
    discounted.append(["Cod", "Denumire"])
    discounted.append(["DISCOUNT-1", "Produs discount"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _add_external_rule_master_to_v1(data_dir: Path, pointer_path: Path, tmp_path: Path) -> Path:
    pointer = json.loads(pointer_path.read_text())
    config_path = data_dir / "promo_generations" / pointer["config_file"]
    config = json.loads(config_path.read_text())
    external_rule_master = tmp_path / "private-rules" / "rule-master.xlsx"
    external_rule_master.parent.mkdir()
    external_rule_master.write_bytes(_rule_master_workbook())
    config["promotions"].append(
        {
            "key": "external-rule",
            "rule_type": "trigger_discounted",
            "source_file": str(external_rule_master),
            "trigger_sheet": "Produse trigger",
            "discounted_sheet": "Produse discount",
            "start_date": "2026-10-01",
            "end_date": "2026-10-31",
        }
    )
    config_bytes = (json.dumps(config, sort_keys=True, separators=(",", ":")) + "\n").encode()
    config_path.write_bytes(config_bytes)
    pointer["config_sha256"] = hashlib.sha256(config_bytes).hexdigest()
    pointer["material_sha256"] = validate_special_cards_config(config)[1]
    pointer_path.write_bytes(
        (json.dumps(pointer, sort_keys=True, separators=(",", ":")) + "\n").encode()
    )
    return external_rule_master


def _downgrade_pointer_to_pre_rule_manifest_v2(pointer_path: Path) -> None:
    pointer = json.loads(pointer_path.read_text())
    pointer.pop("rule_sources", None)
    pointer.pop("rule_sources_sha256", None)
    pointer_path.write_bytes(
        (json.dumps(pointer, sort_keys=True, separators=(",", ":")) + "\n").encode()
    )


def _legacy_generation(tmp_path: Path, *, tamper_source: bool = False) -> tuple[Path, Path]:
    data_dir = tmp_path / "data"
    root = data_dir / "promo_generations"
    generation_id = "1" * 32
    legacy_dir = root / generation_id
    legacy_dir.mkdir(parents=True)
    sources = [
        ("2026-06-30", _workbook("S1", "I1", 2, sheet_name="Sheet1")),
        ("2026-07-31", _workbook("S2", "I2", 3)),
        ("2026-08-05", _workbook("S3", "I3", 4)),
    ]
    manifest: list[dict[str, str]] = []
    promotions: list[dict[str, object]] = []
    for index, (cutoff, content) in enumerate(sources, start=1):
        # Deliberately external to the legacy generation: migration must copy
        # the approved inputs into its immutable v2 directory.
        source = tmp_path / "legacy-inputs" / f"promo-{index}.xlsx"
        source.parent.mkdir(exist_ok=True)
        source.write_bytes(content)
        source.chmod(0o644)
        digest = hashlib.sha256(content).hexdigest()
        manifest.append({"file": str(source), "sha256": digest})
        month = cutoff[:7]
        promotions.append(
            {
                "key": f"promo-{index}",
                "start_date": f"{month}-01",
                "end_date": cutoff,
                "item_codes": [f"I{index}"],
                "actuals_source_file": str(source),
                "actuals_sheet": "Sheet1" if index == 1 else "AccesoriPromoLunar",
                "actuals_cutoff_date": cutoff,
            }
        )
    promotions.append(
        {
            "key": "rule-only",
            "start_date": "2026-09-01",
            "end_date": "2026-09-30",
            "item_codes": ["I9"],
        }
    )
    config = {"promotions": promotions, "incentives": []}
    config_bytes = (json.dumps(config, sort_keys=True, separators=(",", ":")) + "\n").encode()
    (legacy_dir / "hub_specials.json").write_bytes(config_bytes)
    pointer = {
        "version": 1,
        "generation_id": generation_id,
        "config_file": f"{generation_id}/hub_specials.json",
        "config_sha256": hashlib.sha256(config_bytes).hexdigest(),
        "actuals_sha256": hashlib.sha256(b"legacy").hexdigest(),
        "actuals": manifest,
        "material_sha256": validate_special_cards_config(config)[1],
        "previous_generation_id": None,
    }
    pointer_path = root / "current.json"
    pointer_path.write_bytes((json.dumps(pointer, sort_keys=True, separators=(",", ":")) + "\n").encode())
    if tamper_source:
        Path(manifest[0]["file"]).write_bytes(b"changed")
    return data_dir, pointer_path


def _assert_shared_generation_artifacts(
    config_path: Path,
    *,
    pointer_before: bytes,
    pointer: dict[str, object],
) -> None:
    recovery = config_path.parent / "previous-current-v1.json"
    assert recovery.read_bytes() == pointer_before
    assert hashlib.sha256(recovery.read_bytes()).hexdigest() == pointer["previous_pointer_sha256"]
    assert os.stat(recovery).st_mode & 0o777 == 0o660
    config = json.loads(config_path.read_text())
    for promotion in config["promotions"]:
        if not promotion.get("actuals_source_file"):
            assert promotion["key"] == "rule-only"
            continue
        source = Path(promotion["actuals_source_file"])
        material = Path(promotion["actuals_material_file"])
        assert source.is_file() and material.is_file()
        assert os.stat(source).st_mode & 0o777 == 0o660
        assert os.stat(material).st_mode & 0o777 == 0o660
        units, error = load_promo_actual_units(promotion, item_codes=promotion["item_codes"])
        assert error is None
        assert units
    assert os.stat(config_path.parent).st_mode & 0o777 == 0o770


def test_migrates_all_legacy_sources_once_and_is_idempotent(tmp_path: Path) -> None:
    data_dir, pointer_path = _legacy_generation(tmp_path)
    pointer_before = pointer_path.read_bytes()

    dry_run = migrate_legacy_promo_generation(data_dir=data_dir)

    assert dry_run.status == "dry_run"
    assert dry_run.source_count == 3
    assert dry_run.promotion_count == 4
    assert pointer_path.read_bytes() == pointer_before
    assert not list((data_dir / "promo_generations").glob(".staging-*"))

    applied = migrate_legacy_promo_generation(data_dir=data_dir, apply=True)

    assert applied.status == "migrated"
    pointer = json.loads(pointer_path.read_text())
    assert pointer["version"] == 2
    assert pointer["previous_generation_id"] == "1" * 32
    assert len(pointer["actuals"]) == 3
    assert len(pointer["actuals_materials"]) == 3
    config_path = _generated_config_path(data_dir)
    assert config_path is not None
    _assert_shared_generation_artifacts(
        config_path,
        pointer_before=pointer_before,
        pointer=pointer,
    )

    retry = migrate_legacy_promo_generation(data_dir=data_dir, apply=True)
    assert retry.status == "already_v2"
    assert pointer_path.read_bytes() == json.dumps(pointer, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode() + b"\n"


def test_migration_rejects_tampered_source_without_changing_pointer(tmp_path: Path) -> None:
    data_dir, pointer_path = _legacy_generation(tmp_path, tamper_source=True)
    pointer_before = pointer_path.read_bytes()

    with pytest.raises(PromoGenerationMigrationError, match="hashului aprobat"):
        migrate_legacy_promo_generation(data_dir=data_dir, apply=True)

    assert pointer_path.read_bytes() == pointer_before
    assert len(list((data_dir / "promo_generations").iterdir())) == 2


def test_migration_rejects_unparseable_approved_source_without_changing_pointer(tmp_path: Path) -> None:
    data_dir, pointer_path = _legacy_generation(tmp_path)
    pointer = json.loads(pointer_path.read_text())
    source = Path(pointer["actuals"][0]["file"])
    invalid = b"not-an-xlsx"
    source.write_bytes(invalid)
    pointer["actuals"][0]["sha256"] = hashlib.sha256(invalid).hexdigest()
    pointer_path.write_bytes((json.dumps(pointer, sort_keys=True, separators=(",", ":")) + "\n").encode())
    pointer_before = pointer_path.read_bytes()

    with pytest.raises(PromoGenerationMigrationError, match="limitele structurale"):
        migrate_legacy_promo_generation(data_dir=data_dir, apply=True)

    assert pointer_path.read_bytes() == pointer_before


def test_migration_rejects_one_source_with_conflicting_cutoffs(tmp_path: Path) -> None:
    data_dir, pointer_path = _legacy_generation(tmp_path)
    pointer = json.loads(pointer_path.read_text())
    config_path = data_dir / "promo_generations" / pointer["config_file"]
    config = json.loads(config_path.read_text())
    duplicate = dict(config["promotions"][0])
    duplicate["key"] = "conflicting-cutoff"
    duplicate["item_codes"] = ["I4"]
    duplicate["actuals_cutoff_date"] = "2026-06-29"
    duplicate["end_date"] = "2026-06-30"
    config["promotions"].append(duplicate)
    config_bytes = (json.dumps(config, sort_keys=True, separators=(",", ":")) + "\n").encode()
    config_path.write_bytes(config_bytes)
    pointer["config_sha256"] = hashlib.sha256(config_bytes).hexdigest()
    pointer["material_sha256"] = validate_special_cards_config(config)[1]
    pointer_path.write_bytes((json.dumps(pointer, sort_keys=True, separators=(",", ":")) + "\n").encode())
    pointer_before = pointer_path.read_bytes()

    with pytest.raises(PromoGenerationMigrationError, match="Metadatele"):
        migrate_legacy_promo_generation(data_dir=data_dir)

    assert pointer_path.read_bytes() == pointer_before


def test_migration_rechecks_pointer_before_atomic_switch(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    data_dir, pointer_path = _legacy_generation(tmp_path)
    original_write = migration_module._write_durable_private_file

    def write_and_move_pointer(path: Path, content: bytes) -> None:
        original_write(path, content)
        if path.name == "hub_specials.json":
            pointer_path.write_bytes(b'{"generation_id":"changed"}\n')

    monkeypatch.setattr(migration_module, "_write_durable_private_file", write_and_move_pointer)

    with pytest.raises(PromoGenerationMigrationError, match="s-a schimbat"):
        migrate_legacy_promo_generation(data_dir=data_dir, apply=True)

    assert pointer_path.read_bytes() == b'{"generation_id":"changed"}\n'


def test_migration_v1_owns_external_rule_master_in_first_atomic_switch(
    tmp_path: Path,
) -> None:
    data_dir, pointer_path = _legacy_generation(tmp_path)
    external_rule_master = _add_external_rule_master_to_v1(data_dir, pointer_path, tmp_path)
    legacy_pointer = pointer_path.read_bytes()

    migrated = migrate_legacy_promo_generation(data_dir=data_dir, apply=True)

    assert migrated.status == "migrated"
    assert pointer_path.read_bytes() != legacy_pointer
    pointer = json.loads(pointer_path.read_text())
    assert pointer["previous_generation_id"] == "1" * 32
    assert len(pointer["rule_sources"]) == 1
    config_path = _generated_config_path(data_dir)
    assert config_path is not None
    config = json.loads(config_path.read_text())
    rule = next(item for item in config["promotions"] if item["key"] == "external-rule")
    owned_source = Path(rule["source_file"])
    assert owned_source.parent.resolve() == config_path.parent.resolve()
    assert owned_source.is_file()
    assert not owned_source.is_symlink()
    assert rule["source_sha256"] == hashlib.sha256(owned_source.read_bytes()).hexdigest()

    external_rule_master.unlink()
    _promotion_products_cache.clear()
    validate_special_cards_config(config)
    pointer_after_migration = pointer_path.read_bytes()
    retry = migrate_legacy_promo_generation(data_dir=data_dir, apply=True)
    assert retry.status == "already_v2"
    assert pointer_path.read_bytes() == pointer_after_migration


def test_migration_v1_rule_ownership_fault_preserves_active_pointer(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    data_dir, pointer_path = _legacy_generation(tmp_path)
    _add_external_rule_master_to_v1(data_dir, pointer_path, tmp_path)
    pointer_before = pointer_path.read_bytes()
    original_write = migration_module._write_durable_private_file

    def fail_owned_rule(path: Path, content: bytes) -> None:
        if path.name.startswith("promo_rule-"):
            raise OSError("owned rule publication fault")
        original_write(path, content)

    monkeypatch.setattr(
        migration_module,
        "_write_durable_private_file",
        fail_owned_rule,
    )

    with pytest.raises(OSError, match="owned rule publication fault"):
        migrate_legacy_promo_generation(data_dir=data_dir, apply=True)

    assert pointer_path.read_bytes() == pointer_before
    assert not list((data_dir / "promo_generations").glob(".staging-*"))


def test_migration_upgrades_active_unowned_v2_atomically_and_idempotently(
    tmp_path: Path,
) -> None:
    data_dir, pointer_path = _legacy_generation(tmp_path)
    migrate_legacy_promo_generation(data_dir=data_dir, apply=True)
    external_rule_master = _add_external_rule_master_to_v1(
        data_dir, pointer_path, tmp_path
    )
    _downgrade_pointer_to_pre_rule_manifest_v2(pointer_path)
    unowned_pointer = pointer_path.read_bytes()
    unowned_generation_id = json.loads(unowned_pointer)["generation_id"]

    upgraded = migrate_legacy_promo_generation(data_dir=data_dir, apply=True)

    assert upgraded.status == "migrated"
    assert pointer_path.read_bytes() != unowned_pointer
    pointer = json.loads(pointer_path.read_text())
    assert pointer["previous_generation_id"] == unowned_generation_id
    assert len(pointer["rule_sources"]) == 1
    config_path = _generated_config_path(data_dir)
    assert config_path is not None
    config = json.loads(config_path.read_text())
    rule = next(item for item in config["promotions"] if item["key"] == "external-rule")
    owned_source = Path(rule["source_file"])
    assert owned_source.parent.resolve() == config_path.parent.resolve()
    assert owned_source.is_file() and not owned_source.is_symlink()
    assert rule["source_sha256"] == hashlib.sha256(owned_source.read_bytes()).hexdigest()
    external_rule_master.unlink()
    _promotion_products_cache.clear()
    validate_special_cards_config(config)
    pointer_after_upgrade = pointer_path.read_bytes()
    retry = migrate_legacy_promo_generation(data_dir=data_dir, apply=True)
    assert retry.status == "already_v2"
    assert pointer_path.read_bytes() == pointer_after_upgrade


def test_migration_unowned_v2_upgrade_fault_preserves_active_pointer(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    data_dir, pointer_path = _legacy_generation(tmp_path)
    migrate_legacy_promo_generation(data_dir=data_dir, apply=True)
    _add_external_rule_master_to_v1(data_dir, pointer_path, tmp_path)
    _downgrade_pointer_to_pre_rule_manifest_v2(pointer_path)
    pointer_before = pointer_path.read_bytes()
    original_write = migration_module._write_durable_private_file

    def fail_successor_config(path: Path, content: bytes) -> None:
        if path.name == "hub_specials.json" and path.parent.name.startswith(
            ".staging-"
        ):
            raise OSError("successor publication fault")
        original_write(path, content)

    monkeypatch.setattr(
        migration_module,
        "_write_durable_private_file",
        fail_successor_config,
    )

    with pytest.raises(OSError, match="successor publication fault"):
        migrate_legacy_promo_generation(data_dir=data_dir, apply=True)

    assert pointer_path.read_bytes() == pointer_before
    assert not list((data_dir / "promo_generations").glob(".staging-*"))
