from __future__ import annotations

from io import BytesIO
from pathlib import Path
from unittest.mock import AsyncMock
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from services.exports import ExportsService
from services.grile_monthly import ExtractedAgentRow, build_workbook
from services.target_calculator import TargetCalculatorService


def _assert_no_untrusted_formula_xml(content: bytes, allowed: set[str] = set()) -> None:
    with ZipFile(BytesIO(content)) as archive:
        formulas: list[str] = []
        for name in archive.namelist():
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                xml = archive.read(name).decode()
                formulas.extend(part.split("</f>", 1)[0] for part in xml.split("<f>")[1:])
    assert set(formulas) <= allowed


class _ExportsRepo:
    async def fetch_daily_comparison_rows(self, **_kwargs: object) -> list[dict[str, object]]:
        return []


def _block_raw_append(monkeypatch: pytest.MonkeyPatch) -> None:
    def blocked_append(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("raw Worksheet.append is forbidden in runtime writers")

    monkeypatch.setattr(Worksheet, "append", blocked_append)


@pytest.mark.asyncio
async def test_exports_standard_and_daily_writers_use_boundary_and_neutralize_xml(monkeypatch: pytest.MonkeyPatch) -> None:
    service = ExportsService(_ExportsRepo())  # type: ignore[arg-type]
    monkeypatch.setattr(service, "build_report", AsyncMock(return_value={
        "columns": [
            {"key": "agent", "label": "Agent", "type": "text"},
            {"key": "locatie", "label": "Locatie", "type": "text"},
            {"key": "asm", "label": "ASM", "type": "text"},
            {"key": "sales", "label": "Vanzari", "type": "currency"},
        ],
        "rows": [{"agent": '=HYPERLINK("https://example.invalid","x")', "locatie": "+SUM(1,1)", "asm": "@SUM(1,1)", "sales": 12}],
    }))
    _block_raw_append(monkeypatch)

    standard, _ = await service.build_xlsx({"dataset": "stores", "months": ["2026-06"]})
    standard_book = load_workbook(BytesIO(standard), data_only=False)
    row = standard_book["Raport"][2]
    assert [cell.data_type for cell in row[:3]] == ["s", "s", "s"]
    assert row[0].value.startswith("'") and row[1].value.startswith("'") and row[2].value.startswith("'")
    assert row[3].value == 12
    _assert_no_untrusted_formula_xml(standard)

    monkeypatch.setattr(service, "_daily_comparison_params", lambda _request: (["=1+1"], ["total_sales"], ["general"], {}, False, []))
    daily_artifact = await service._build_daily_comparison_xlsx({})
    try:
        daily = b"".join(daily_artifact.iter_chunks())
    finally:
        daily_artifact.close()
    daily_book = load_workbook(BytesIO(daily), data_only=False)
    month_cell = daily_book["Configuratie"]["B3"]
    assert month_cell.value == "'=1+1" and month_cell.data_type == "s"
    _assert_no_untrusted_formula_xml(daily)


@pytest.mark.asyncio
async def test_target_calculator_writer_uses_boundary_and_neutralizes_all_sheets(monkeypatch: pytest.MonkeyPatch) -> None:
    service = TargetCalculatorService(object())  # type: ignore[arg-type]
    detail = {
        "id": 7, "status": "=1+1", "target_month": "2026-06", "cohort_month": "2026-05", "total_target": 100,
        "min_floor": 1, "previous_month_floor_pct": 0.9, "calculation_method": "=method", "calculation_params": {},
        "source_months": [{"month": "2026-05", "label": "=luna"}],
        "rows": [{"firma": "=1+1", "regional": "+SUM(1,1)", "asm": "@SUM(1,1)", "locatie": "-1+2", "site_code": "\t=1+1", "floor_target": 1, "calculated_weight": 1, "normalized_weight": 1, "proposed_target": 2, "final_target": 2, "note": '=HYPERLINK("https://example.invalid","x")', "history": [{"month": "2026-05", "target": 1, "realized": 1, "actual_realized": 1, "attainment_pct": 100, "is_forecast": False}], "calculation_details": {"seasonality": {}, "trend": {}, "flags": []}, "profitability": {"salary_cost_at_90_pct": 1, "operating_costs": 1, "break_even_gross_sales": 1, "forecast_sales": 2, "anomaly_flags": []}}],
        "regional_summary": [{"regional": "\n=1+1", "store_count": 1, "floor_total": 1, "proposed_total": 2, "final_total": 2}],
        "source_summary": [], "warnings": ["\n=1+1"], "profitability_summary": {},
    }
    service.get_scenario_detail = AsyncMock(return_value=detail)  # type: ignore[method-assign]
    _block_raw_append(monkeypatch)
    output, _ = await service.export_excel(7)
    content = output.getvalue()
    book = load_workbook(BytesIO(content), data_only=False)
    assert book["Target + profitabilitate"]["A3"].data_type == "s"
    assert book["Target + profitabilitate"]["A3"].value == "'=1+1"
    assert book["Parametri"]["B3"].value == "'=1+1"
    trusted_formulas = {
        *(f"SUBTOTAL(109,{column}3:{column}3)" for column in ("E", "F", "H", "I", "K", "L", "N", "O", "P", "Q", "R", "S", "T")),
        "IF(E1=0,0,F1/E1)",
        "IF(H1=0,0,I1/H1)",
        "IF(K1=0,0,L1/K1)",
    }
    _assert_no_untrusted_formula_xml(content, trusted_formulas)


def test_grile_writer_uses_boundary_neutralizes_external_values_and_keeps_only_trusted_formulas(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _block_raw_append(monkeypatch)
    output = tmp_path / "grile.xlsx"
    build_workbook([ExtractedAgentRow("Mobiup", "+Store", 1, "=Agent", 1, 2, 3, 4, 5, 6, "OK", "@error", "sheet-id")], output, {("Mobiup", "+Store"): {"Manager": "-Manager"}})
    content = output.read_bytes()
    book = load_workbook(BytesIO(content), data_only=False)
    sheet = book["Mobiup"]
    assert sheet["B2"].value == "'-Manager"
    assert sheet["C2"].value == "'+Store"
    assert sheet["D2"].value == "'=Agent"
    assert sheet["K2"].value == "=SUM(E2:J2,M2)" and sheet["K2"].data_type == "f"
    assert sheet["L2"].value == "=K2-M2" and sheet["L2"].data_type == "f"
    _assert_no_untrusted_formula_xml(content, {"SUM(E2:J2,M2)", "K2-M2"})
