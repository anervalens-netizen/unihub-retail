from __future__ import annotations
import pytest
from unittest.mock import AsyncMock, MagicMock, patch


@pytest.fixture(scope="module")
def anyio_backend():
    return "asyncio"


@pytest.mark.anyio
async def test_filters_options_endpoint():
    from services.filter_options import FilterOptionsService
    from repositories.filters import FiltersRepository
    from db.connection import get_pool

    pool = await get_pool()
    repo = FiltersRepository(pool)
    svc = FilterOptionsService(repo)
    result = await svc.get_options("2026-03")
    assert hasattr(result, "firme")
    assert hasattr(result, "regionali")


@pytest.mark.anyio
async def test_filters_months_endpoint():
    from services.filter_options import FilterOptionsService
    from repositories.filters import FiltersRepository
    from db.connection import get_pool

    pool = await get_pool()
    repo = FiltersRepository(pool)
    svc = FilterOptionsService(repo)
    months = await svc.get_available_months()
    assert isinstance(months, list)
    if months:
        assert all(isinstance(m, str) for m in months)


@pytest.mark.anyio
async def test_filters_options_cache_clear():
    from services.filter_options import FilterOptionsService
    FilterOptionsService.clear_cache()
    # After clearing cache, a new request should refetch
    from repositories.filters import FiltersRepository
    from db.connection import get_pool
    pool = await get_pool()
    repo = FiltersRepository(pool)
    svc = FilterOptionsService(repo)
    result = await svc.get_options("2026-03")
    assert hasattr(result, "firme")


@pytest.mark.anyio
async def test_stores_list_service():
    from services.stores import StoresService
    from repositories.stores import StoresRepository
    from db.connection import get_pool

    pool = await get_pool()
    repo = StoresRepository(pool)
    svc = StoresService(repo, pool)
    stores = await svc.get_active_stores()
    assert isinstance(stores, list)


@pytest.mark.anyio
async def test_stores_save_targets():
    from services.stores import StoresService
    from repositories.stores import StoresRepository
    from db.connection import get_pool

    pool = await get_pool()
    repo = StoresRepository(pool)
    svc = StoresService(repo, pool)
    # Save empty targets - should not crash
    inserted = await svc.save_targets([])
    assert inserted == 0


@pytest.mark.anyio
async def test_imports_history_service():
    from services.imports import ImportsService
    from repositories.imports import ImportsRepository
    from db.connection import get_pool

    pool = await get_pool()
    repo = ImportsRepository(pool)
    svc = ImportsService(repo, pool)
    history = await svc.get_import_history()
    assert isinstance(history, list)


@pytest.mark.anyio
async def test_imports_job_status_not_found(monkeypatch):
    from services.imports import ImportsService
    from repositories.imports import ImportsRepository
    from db.connection import get_pool
    import services.imports as imports_service
    from services.jobs import JobResult, JobStatus

    async def fake_get_job_status(job_id: str) -> JobResult:
        return JobResult(job_id=job_id, status=JobStatus.NOT_FOUND)

    monkeypatch.setattr(imports_service, "get_job_status", fake_get_job_status)

    pool = await get_pool()
    repo = ImportsRepository(pool)
    svc = ImportsService(repo, pool)
    result = await svc.get_import_job_status("nonexistent-job-id")
    assert result.job_id == "nonexistent-job-id"
    assert result.status in ("not_found", "queued")


def test_filters_normalize():
    from services.filters import normalize_filter

    assert normalize_filter(None) is None
    assert normalize_filter("") is None
    assert normalize_filter("Toate") is None
    assert normalize_filter("Toti") is None
    assert normalize_filter("  Test  ") == "Test"
    assert normalize_filter("Firma 1") == "Firma 1"


def test_filters_base_filter_values_no_filters():
    from services.filters import base_filter_values

    result = base_filter_values("2026-03", None, None, None, None, None)
    assert isinstance(result, tuple)
    assert len(result) == 2


def test_filters_base_filter_values_with_filters():
    from services.filters import base_filter_values

    params, where = base_filter_values("2026-03", "Firma 1", "Regional 1", "ASM 1", "ST001", "Agent 1")
    assert isinstance(params, list)
    assert isinstance(where, dict)


@pytest.mark.anyio
async def test_forecast_factor():
    from services.forecast import get_forecast_factor
    from db.connection import get_pool

    pool = await get_pool()
    async with pool.acquire() as conn:
        result = await get_forecast_factor(conn, "2026-03")
        assert result > 0


def test_product_lists_load_codes_empty():
    from services.product_lists import load_product_code_rows
    import tempfile, os

    # Create a minimal Excel file with correct column name
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value="Cod produs")
        ws.cell(row=2, column=1, value="PROD001")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
            wb.save(path)

        codes = load_product_code_rows(path)
        assert isinstance(codes, list)
        assert len(codes) == 1
        assert codes[0]["item_code"] == "PROD001"
    finally:
        if os.path.exists(path):
            os.unlink(path)
