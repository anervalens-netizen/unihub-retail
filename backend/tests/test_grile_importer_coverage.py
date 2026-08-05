from __future__ import annotations

import asyncio
import json
from datetime import date
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock, MagicMock

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

import services.grile_monthly as grile
import services.importer as importer


def grile_row(*, status: str = "OK") -> grile.ExtractedAgentRow:
    return grile.ExtractedAgentRow(
        company="Mobiup",
        store="Park Lake",
        slot=1,
        agent="Agent Test" if status == "OK" else "",
        base_salary=2600 if status == "OK" else "",
        sales_commission=300 if status == "OK" else "",
        extra_location_commission=25 if status == "OK" else "",
        extra_hours_pay=150 if status == "OK" else "",
        bonuri=480 if status == "OK" else "",
        worked_hours=176 if status == "OK" else "",
        status=status,
        error="" if status == "OK" else "invalid",
        sheet_id="sheet-1",
        site_code="SITE01",
        error_code="" if status == "OK" else "invalid",
    )


def sales_row(**overrides: Any) -> dict[str, Any]:
    row: dict[str, Any] = {
        "Data": "01.07.2099",
        "SiteCode": "SITE01",
        "ItemCode": "ITEM01",
        "ItemName": "Produs",
        "Cantitate": 2,
        "Brand": "Brand",
        "Pret": 10.125,
        "Valoare": 20.255,
        "Locatie": "Magazin",
        "Firma": "Mobiup",
        "ASM": "Manager",
        "Regional": "Regional",
        "Nr": "BON1",
        "Categorie": "Accesorii",
        "SubCategorie": "Test",
        "Agent": "Agent",
    }
    row.update(overrides)
    return row


def sales_workbook(rows: list[dict[str, Any]]) -> bytes:
    output = BytesIO()
    pd.DataFrame(rows, columns=importer.SALES_COLUMNS).to_excel(output, index=False)
    return output.getvalue()


def sales_frame(**overrides: Any) -> pd.DataFrame:
    row = sales_row(**overrides)
    if "Data" not in overrides:
        row["Data"] = date(2099, 7, 1)
    return pd.DataFrame([row], columns=importer.SALES_COLUMNS)


def test_grile_defensive_manifest_and_result_copy() -> None:
    ok, errors = grile.validate_archive_manifest(
        {
            "registry_count": 1,
            "exported_count": 1,
            "error_count": 0,
            "stores": {},
            "zip_path": "/definitely-missing/archive.zip",
        },
        expected_count=1,
    )
    assert not ok
    assert any(error == "stores count mismatch: invalid != 1" for error in errors)

    source = {"result": {"status": "completed", "nested": {"value": 1}}}
    copied = grile._safe_operation_result(source)
    assert copied == source["result"]
    assert copied is not source["result"]


def test_grile_workbook_applies_default_column_width(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def clear_widths(worksheet: Any) -> None:
        for column in "ABCDEFGHIJKLMNOPQ":
            worksheet.column_dimensions[column].__dict__["width"] = None

    monkeypatch.setattr(grile, "style_sheet", clear_widths)
    output = tmp_path / "default-width.xlsx"

    grile.build_workbook([grile_row()], output, {})

    workbook = load_workbook(output)
    assert workbook["Mobiup"].column_dimensions["A"].width == 14


@pytest.mark.asyncio
async def test_approve_manifest_rejects_missing_manifest_hash(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        grile,
        "fetch_monthly_manifest",
        AsyncMock(
            return_value={
                "operation": "archive",
                "status": "verified",
                "manifest": {"status": "verified"},
            }
        ),
    )
    monkeypatch.setattr(grile, "validate_verified_manifest", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(grile, "verify_artifacts", lambda *_args, **_kwargs: None)

    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        await grile.approve_monthly_manifest(
            MagicMock(), manifest_id=4, approved_by_sub="qa"
        )

    assert exc_info.value.code == "manifest_hash_invalid"


@pytest.mark.asyncio
async def test_reset_rollback_reports_checkpoint_record_failure(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    entry = grile.StoreEntry("Mobiup", "Park Lake", "sheet-1", "SITE01", "ASM", False, "v2")
    monkeypatch.setattr(grile, "_restore_reset_snapshot", lambda *_args: None)
    record = AsyncMock(side_effect=RuntimeError("checkpoint unavailable"))
    monkeypatch.setattr(grile, "record_reset_item_rollback", record)

    restored = await grile._rollback_reset_entries(
        object(),
        operation_id=7,
        entries=[entry],
        sheets_svc=object(),
        snapshots={"SITE01": {}},
        execution_owner="worker-a",
        execution_epoch=1,
    )

    assert restored is False
    record.assert_awaited_once()


@pytest.mark.asyncio
async def test_cancel_safe_rollback_clears_existing_cancellation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class FakeCurrent:
        def __init__(self) -> None:
            self.uncancel_calls = 0

        def cancelling(self) -> int:
            return 1 if self.uncancel_calls == 0 else 0

        def uncancel(self) -> None:
            self.uncancel_calls += 1

    fake_current = FakeCurrent()
    monkeypatch.setattr(grile.asyncio, "current_task", lambda: fake_current)
    rollback = AsyncMock(return_value=True)
    monkeypatch.setattr(grile, "_rollback_reset_entries", rollback)

    result = await grile._rollback_reset_entries_cancel_safe(
        object(),
        operation_id=1,
        entries=[],
        sheets_svc=object(),
        snapshots={},
        execution_owner="worker-a",
        execution_epoch=1,
    )

    assert result is True
    assert fake_current.uncancel_calls == 1
    rollback.assert_awaited_once()


@pytest.mark.asyncio
async def test_cancel_safe_rollback_returns_false_on_second_interruption(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    started = asyncio.Event()
    release = asyncio.Event()

    async def slow_rollback(*_args: Any, **_kwargs: Any) -> bool:
        started.set()
        await release.wait()
        return True

    monkeypatch.setattr(grile, "_rollback_reset_entries", slow_rollback)
    task = asyncio.create_task(
        grile._rollback_reset_entries_cancel_safe(
            object(),
            operation_id=1,
                entries=[],
                sheets_svc=object(),
                snapshots={},
                execution_owner="worker-a",
                execution_epoch=1,
        )
    )
    await started.wait()
    task.cancel()
    assert await task is False
    release.set()
    await asyncio.sleep(0)


@pytest.mark.asyncio
async def test_run_monthly_op_persists_uncertain_manifest_on_cancellation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    pool = object()
    monkeypatch.setattr("db.connection.get_pool", AsyncMock(return_value=pool))
    monkeypatch.setattr(
        grile,
        "start_monthly_operation",
        AsyncMock(
            return_value=grile.MonthlyOperationStartResult(
                "started",
                12,
                operation={
                    "op": "reset",
                    "closing_month": "2026-06",
                    "only_filter": None,
                    "dry_run": False,
                    "requested_by_sub": "qa-subject",
                    "approved_manifest_id": None,
                },
            )
        ),
    )
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())

    async def cancel_execution(*_args: Any, **_kwargs: Any) -> None:
        raise asyncio.CancelledError

    monkeypatch.setattr(grile, "_reset_month_execution", cancel_execution)
    persisted = AsyncMock(
        return_value={
            "id": 12,
            "operation_id": 12,
            "closing_month": "2026-06",
            "operation": "reset",
            "status": "uncertain",
            "manifest": {},
            "error_count": 2,
        }
    )
    monkeypatch.setattr(grile, "persist_manifest_result", persisted)
    finished = AsyncMock(return_value=True)
    monkeypatch.setattr(grile, "finish_monthly_operation", finished)

    result = await grile.run_monthly_op(operation_id=12)

    assert result["status"] == "failed"
    assert result["operation_id"] == 12
    persisted.assert_awaited_once()
    assert persisted.await_args is not None
    assert persisted.await_args.kwargs["error_code"] == "monthly_operation_cancelled"
    assert persisted.await_args.kwargs["manifest"]["status"] == "uncertain"
    finished.assert_awaited_once()


@pytest.mark.asyncio
async def test_run_monthly_op_requires_month_before_pool_access() -> None:
    with pytest.raises(ValueError, match="month is required"):
        await grile.run_monthly_op(op="finalize", month=None)


def test_importer_path_and_input_validation_branches(tmp_path: Path) -> None:
    source = tmp_path / "sales.xlsx"
    source.write_bytes(sales_workbook([sales_row()]))
    loaded = importer.load_sales_dataframe(source)
    assert loaded.loc[0, "SiteCode"] == "SITE01"

    with pytest.raises(ValueError, match="Data"):
        importer.load_sales_dataframe(sales_workbook([sales_row(Data="not-a-date")]))

    with pytest.raises(ValueError, match="coloane obligatorii"):
        importer.validate_sales_dataframe(pd.DataFrame({"Data": [date(2099, 7, 1)]}))

    with pytest.raises(ValueError, match="Data"):
        importer.validate_sales_dataframe(sales_frame(Data=None))
    with pytest.raises(ValueError, match="Cantitate"):
        importer.validate_sales_dataframe(sales_frame(Cantitate="invalid"))
    with pytest.raises(ValueError, match="Pret"):
        importer.validate_sales_dataframe(sales_frame(Pret="invalid"))


@pytest.mark.asyncio
async def test_importer_coverage_helper_and_path_digest(tmp_path: Path) -> None:
    conn = MagicMock()
    conn.execute = AsyncMock()
    await importer.record_coverage_report(conn, 17, {"missing": ["SITE02"]})
    assert conn.execute.await_count == 1
    assert json.loads(conn.execute.await_args_list[0].args[2]) == {"missing": ["SITE02"]}

    with pytest.raises(ValueError, match="at least 60"):
        await importer.reserve_snapshot(conn, "2099-07", "sales.xlsx", 1, lease_seconds=59)

    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source bytes")
    frame = sales_frame()
    monkeypatch = pytest.MonkeyPatch()
    try:
        monkeypatch.setattr(importer, "load_sales_dataframe", MagicMock(return_value=frame))
        run_import = AsyncMock(return_value="imported")
        monkeypatch.setattr(importer, "import_sales_dataframe", run_import)
        result = await importer.import_sales_file(conn, source, "source.xlsx")
    finally:
        monkeypatch.undo()

    assert result == "imported"
    run_import.assert_awaited_once()
    assert run_import.await_args is not None
    assert run_import.await_args.kwargs["source_sha256"] == sha256(b"source bytes").hexdigest()


def test_load_targets_skips_unversioned_invalid_and_blank_columns(tmp_path: Path) -> None:
    source = tmp_path / "targets.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["", None, 2099, None])
    sheet.append(["SiteCode", "TG L01", "TG invalid", "TG L02"])
    sheet.append(["SITE01", 100, 999, 200])
    sheet.append(["   ", 300, 999, 400])
    workbook.save(source)

    assert importer.load_targets_dataframe(source) == [
        {
            "site_code": "SITE01",
            "import_month": "2099-02",
            "target_value": importer.Decimal("200.00"),
        }
    ]
