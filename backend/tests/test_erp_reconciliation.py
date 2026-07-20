from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook

from schemas.campaigns import PromoIncentiveSummary
from services.erp_reconciliation import (
    AGENT_REQUIRED_COLUMNS,
    COMMON_METRIC_COLUMNS,
    STORE_AGENT_DERIVED_METRIC_COLUMNS,
    STORE_REQUIRED_COLUMNS,
    ErpReportValidationError,
    parse_erp_report,
    reconcile_erp_report,
)


def workbook_bytes(
    *,
    agent_bon2: int = 2,
    days_elapsed: int = 16,
    include_store_derived_metrics: bool = False,
    store_focus_quantity: int = 3,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    stores = workbook.create_sheet("Locatii")
    store_columns = (
        *STORE_REQUIRED_COLUMNS,
        *(
            STORE_AGENT_DERIVED_METRIC_COLUMNS
            if include_store_derived_metrics
            else ()
        ),
    )
    stores.append([None] * len(store_columns))
    stores.append(list(store_columns))
    store_values = {
        "Firma": "MobiUp",
        "CodLocatie": "S1",
        "Locatie": "MAGAZIN TEST",
        "AccValTarget": 1000,
        "AccValRealizat": 500,
        "AccQttyRealizat": 10,
        "NrBonuri": 4,
        "NrBon2Acc": agent_bon2,
        "AccFocusQtty": store_focus_quantity,
        "Audio": 1,
        "Battery": 1,
        "Suporti": 1,
        "FoliiQtty": 4,
        "Folii Sticla": 3,
        "Folii TPU": 1,
        "Still&Protectie": 2,
        "Incarcare&Transfer": 1,
        "ZileLuna": 31,
        "ZileTrecute": days_elapsed,
        "ZileRamase": 31 - days_elapsed,
    }
    stores.append([store_values[column] for column in store_columns])
    stores.append(
        [
            "MobiUp" if column == "Firma" else
            "TR1" if column == "CodLocatie" else
            "TR Test" if column == "Locatie" else
            0
            for column in store_columns
        ]
    )

    agents = workbook.create_sheet("Agenti")
    agents.append([None] * len(AGENT_REQUIRED_COLUMNS))
    agents.append(list(AGENT_REQUIRED_COLUMNS))
    agent_values = {
        "Firma": "MobiUp",
        "CodLocatie": "S1",
        "Locatie": "MAGAZIN TEST",
        "Agent": "AGENT1",
        **{column: store_values[column] for column in COMMON_METRIC_COLUMNS},
        "AccFocusQtty": 3,
    }
    agents.append([agent_values[column] for column in AGENT_REQUIRED_COLUMNS])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def reference(*, bon2: int = 2) -> dict:
    return {
        "snapshot": {"id": 1},
        "retail_cutoff_date": date(2026, 7, 16),
        "stores": [
            {
                "site_code": "S1",
                "locatie": "MAGAZIN TEST",
                "total_sales": Decimal("500.00"),
                "total_quantity": 10,
                "focus_quantity": 3,
                "receipt_count": 3,
                "receipt_2plus_count": bon2,
                "target_value": Decimal("1000.00"),
                "agent_count": 1,
            }
        ],
        "agents": [
            {
                "site_code": "S1",
                "agent": "AGENT1",
                "locatie": "MAGAZIN TEST",
                "total_sales": Decimal("500.00"),
                "total_quantity": 10,
                "focus_quantity": 3,
                "receipt_count": 3,
                "receipt_2plus_count": bon2,
            }
        ],
        "receipt_rows": [
            {
                "site_code": "S1",
                "agent": "AGENT1",
                "all_receipts": 4,
                "positive_receipts": 3,
                "return_only_receipts": 1,
            }
        ],
        "focus_rows": [
            {"focus_subcategory": "Casti intraauriculare", "quantity": 1},
            {"focus_subcategory": "Baterie Externa", "quantity": 1},
            {"focus_subcategory": "Suport auto", "quantity": 1},
        ],
        "category_rows": [
            {"category": "Folii Sticla", "subcategory": "Folii sticla 2.5D", "quantity": 3},
            {"category": "Folii TPU", "subcategory": "Folie TPU", "quantity": 1},
            {"category": "Stil si Protectie", "subcategory": "Capac protectie", "quantity": 2},
            {"category": "Incarcare si Transfer", "subcategory": "Cablu de date", "quantity": 1},
        ],
    }


def test_parse_erp_report_uses_detail_sheets_and_excludes_tr() -> None:
    parsed = parse_erp_report(workbook_bytes(), "2026-07")

    assert parsed.cutoff_date == date(2026, 7, 16)
    assert list(parsed.stores) == [("S1",)]
    assert list(parsed.agents) == [("S1", "AGENT1")]
    assert parsed.stores[("S1",)]["AccValRealizat"] == Decimal("500")
    assert parsed.stores[("S1",)]["AccFocusQtty"] == Decimal("3")
    assert parsed.stores[("S1",)]["FoliiQtty"] == Decimal("4")


def test_parse_erp_report_rejects_month_day_mismatch() -> None:
    with pytest.raises(ErpReportValidationError, match="luna 2026-06 are 30"):
        parse_erp_report(workbook_bytes(), "2026-06")


def test_parse_erp_report_validates_optional_store_metrics_when_present() -> None:
    with pytest.raises(
        ErpReportValidationError,
        match="Foile Locatii si Agenti nu au acelasi total pentru AccFocusQtty",
    ):
        parse_erp_report(
            workbook_bytes(
                include_store_derived_metrics=True,
                store_focus_quantity=4,
            ),
            "2026-07",
        )


def test_reconciliation_explains_returns_without_failing() -> None:
    parsed = parse_erp_report(workbook_bytes(), "2026-07")
    result = reconcile_erp_report(
        parsed,
        reference(),
        PromoIncentiveSummary(incentive_sold_qty=8, incentive_qty=7),
        import_month="2026-07",
        filename="erp.xlsx",
        file_digest="abc123",
    )

    assert result.status == "ok"
    assert result.issue_count == 0
    receipts = next(metric for metric in result.metrics if metric.key == "receipts")
    assert receipts.status == "explained"
    assert receipts.difference == 1
    assert result.app_only_metrics[1].value == 8


def test_reconciliation_accepts_retail_snapshot_beyond_report_cutoff() -> None:
    parsed = parse_erp_report(workbook_bytes(days_elapsed=16), "2026-07")
    later_reference = reference()
    later_reference["retail_cutoff_date"] = date(2026, 7, 17)

    result = reconcile_erp_report(
        parsed,
        later_reference,
        None,
        import_month="2026-07",
        filename="erp.xlsx",
        file_digest="abc123",
    )

    assert result.report_cutoff_date == date(2026, 7, 16)
    assert result.retail_cutoff_date == date(2026, 7, 17)
    assert result.cutoff_matches is True
    assert result.status == "ok"


def test_reconciliation_reports_bon2_difference_at_agent_level() -> None:
    parsed = parse_erp_report(workbook_bytes(agent_bon2=3), "2026-07")
    result = reconcile_erp_report(
        parsed,
        reference(bon2=2),
        None,
        import_month="2026-07",
        filename="erp.xlsx",
        file_digest="abc123",
    )

    assert result.status == "differences"
    assert result.issue_count == 1
    assert result.issues[0].scope == "agent"
    assert result.issues[0].metric == "Bonuri 2+"
    assert result.issues[0].difference == 1
