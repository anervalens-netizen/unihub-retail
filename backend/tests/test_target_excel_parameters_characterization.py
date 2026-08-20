from __future__ import annotations

from io import BytesIO
from typing import Any

import pytest
from openpyxl import load_workbook

from services.target_calculator.export import build_target_excel


def _scenario() -> dict[str, Any]:
    return {
        "id": 9,
        "status": "draft",
        "target_month": "2026-06",
        "cohort_month": "2026-05",
        "total_target": 100.0,
        "min_floor": 10.0,
        "previous_month_floor_pct": 0.9,
        "calculation_method": "characterization",
        "calculation_params": {"strong_weights": {"store": 0.5}},
        "source_months": [
            {"label": "Baza", "month": "2025-05"},
        ],
        "source_summary": [
            {
                "month": "2026-05",
                "is_forecast": True,
                "forecast_factor": 1.2,
                "actual_realized": 100.0,
                "realized": 120.0,
            },
        ],
        "warnings": ["Atentionare test"],
        "profitability_summary": {
            "status": "ready",
            "pnl_months": ["2026-02", "2026-03"],
            "forecast_run": {
                "id": 44,
                "model_name": "TimesFM + XGRegressor",
                "variant": "august_exact",
            },
        },
        "regional_summary": [],
        "rows": [
            {
                "firma": "Mobiup",
                "regional": "Regional A",
                "locatie": "Magazin 1",
                "site_code": "SITE01",
                "calculated_weight": 1.0,
                "normalized_weight": 1.0,
                "proposed_target": 100.0,
                "final_target": 100.0,
                "history": [
                    {
                        "month": "2025-05",
                        "target": 100.0,
                        "realized": 100.0,
                        "attainment_pct": 100.0,
                    },
                    {
                        "month": "2025-06",
                        "target": 100.0,
                        "realized": 100.0,
                        "attainment_pct": 100.0,
                    },
                    {
                        "month": "2026-05",
                        "target": 100.0,
                        "realized": 100.0,
                        "attainment_pct": 100.0,
                    },
                ],
                "profitability": {
                    "salary_cost_at_90_pct": 10.0,
                    "operating_costs": 20.0,
                    "break_even_gross_sales": 30.0,
                    "forecast_sales": 110.0,
                },
            },
        ],
    }


@pytest.mark.asyncio
async def test_target_excel_parameters_sheet_contract_is_exact() -> None:
    scenario = _scenario()

    async def load_scenario(scenario_id: int) -> dict[str, Any]:
        assert scenario_id == 9
        return scenario

    output, _filename = await build_target_excel(9, load_scenario)
    workbook = load_workbook(BytesIO(output.getvalue()))
    parameters = workbook["Parametri"]

    assert [list(row) for row in parameters.iter_rows(values_only=True)] == [
        ["Parametru", "Valoare"],
        ["Scenariu", 9],
        ["Status", "draft"],
        ["Luna target", "2026-06"],
        ["Luna cohorta magazine active", "2026-05"],
        ["Target total", 100],
        ["Prag minim absolut", 10],
        ["Floor fata de luna precedenta", 0.9],
        ["Metoda", "characterization"],
        ["Parametru strong_weights", '{"store": 0.5}'],
        ["Baza", "2025-05"],
        ["Forecast 2026-05", "1.2000x; importat 100.00; folosit 120.00"],
        ["Atentionare", "Atentionare test"],
        ["Status surse profitabilitate", "ready"],
        ["Luni P&L reale", "2026-02, 2026-03"],
        ["Forecast run", 44],
        ["Forecast model", "TimesFM + XGRegressor"],
        ["Forecast variant", "august_exact"],
    ]
