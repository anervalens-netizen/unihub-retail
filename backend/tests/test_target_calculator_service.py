from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

import services.target_calculator as target_module
from repositories.target_calculator import (
    TargetScenarioFinalizedError,
    TargetScenarioVersionConflict,
)
from services.target_calculator import (
    CALCULATION_METHOD,
    TargetCalculatorService,
    allocate_with_floors,
    shift_month,
)


def make_service() -> tuple[TargetCalculatorService, MagicMock]:
    repo = MagicMock()
    repo.get_latest_sales_month = AsyncMock()
    repo.get_target_total = AsyncMock()
    repo.get_active_cohort = AsyncMock()
    repo.get_source_metrics = AsyncMock()
    repo.save_draft_scenario = AsyncMock()
    repo.list_scenarios = AsyncMock()
    repo.get_scenario = AsyncMock()
    repo.get_scenario_rows = AsyncMock()
    repo.get_profitability_inputs = AsyncMock(return_value={
        "pnl_months": [],
        "pnl_rows": [],
        "forecast_run": None,
        "forecast_rows": [],
    })
    repo.update_final_targets = AsyncMock()
    repo.finalize_scenario = AsyncMock()
    repo.get_store_detail = AsyncMock()
    connection = MagicMock()
    repo.pool.acquire.return_value.__aenter__ = AsyncMock(return_value=connection)
    repo.pool.acquire.return_value.__aexit__ = AsyncMock(return_value=False)
    return TargetCalculatorService(repo), repo


def scenario_header(**overrides: object) -> dict:
    row: dict[str, object] = {
        "id": 9,
        "target_month": "2026-06",
        "cohort_month": "2026-05",
        "total_target": Decimal("100000.00"),
        "min_floor": Decimal("10000.00"),
        "previous_month_floor_pct": Decimal("0.50"),
        "proposed_total": Decimal("100000.00"),
        "final_total": Decimal("100000.00"),
        "calculation_method": CALCULATION_METHOD,
        "status": "draft",
        "revision": 2,
        "source_months": (
            '[{"month":"2025-05","label":"Anterior","role":"previous_year_reference"}]'
        ),
        "warnings": '["Atentionare test"]',
        "store_count": 2,
    }
    row.update(overrides)
    return row


def scenario_rows() -> list[dict]:
    history = (
        '[{"month":"2025-05","label":"Anterior","role":"previous_year_reference",'
        '"target":40000.0,"realized":42000.0,"actual_realized":35000.0,'
        '"is_forecast":true,"forecast_factor":1.2,"attainment_pct":105.0,'
        '"weight":0.4}]'
    )
    details_1 = (
        '{"current_month":"2026-05","current_forecast":42000.0,'
        '"seasonality":{"store_years":[{"year_offset":1,"base_month":"2025-05",'
        '"target_month":"2025-06","base_value":42000.0,"target_value":50400.0,"ratio":1.2}]}}'
    )
    details_2 = (
        '{"current_month":"2026-05","current_forecast":63000.0,'
        '"seasonality":{"store_years":[{"year_offset":1,"base_month":"2025-05",'
        '"target_month":"2025-06","base_value":63000.0,"target_value":75600.0,"ratio":1.2}]}}'
    )
    return [
        {
            "site_code": "SITE01",
            "locatie": "Magazin 1",
            "firma": "Mobiup",
            "regional": "Regional A",
            "asm": "ASM 1",
            "calculated_weight": Decimal("0.40"),
            "floor_target": Decimal("10000"),
            "proposed_target": Decimal("40000"),
            "final_target": Decimal("41000"),
            "is_floor_limited": False,
            "history": history,
            "calculation_details": details_1,
            "note": "ajustat",
        },
        {
            "site_code": "SITE02",
            "locatie": "Magazin 2",
            "firma": "Mobicell",
            "regional": "Regional A",
            "asm": "ASM 2",
            "calculated_weight": Decimal("0.60"),
            "floor_target": Decimal("12000"),
            "proposed_target": Decimal("60000"),
            "final_target": Decimal("59000"),
            "is_floor_limited": True,
            "history": history,
            "calculation_details": details_2,
            "note": None,
        },
    ]


@pytest.mark.parametrize("month", ["invalid", "2026-13", "2026"])
def test_shift_month_rejects_invalid_values(month: str) -> None:
    with pytest.raises(HTTPException) as exc:
        shift_month(month, 1)
    assert exc.value.status_code == 400


def test_allocation_handles_empty_and_zero_weight_rows() -> None:
    assert allocate_with_floors([], Decimal("100")) == ([], [])

    rows = [
        {
            "calculated_weight": Decimal("0"),
            "floor_target": Decimal("0"),
            "is_floor_limited": False,
        },
        {
            "calculated_weight": Decimal("0"),
            "floor_target": Decimal("0"),
            "is_floor_limited": False,
        },
    ]
    allocated, warnings = allocate_with_floors(rows, Decimal("100"))

    assert warnings == []
    assert [row["proposed_target"] for row in allocated] == [
        Decimal("50.00"),
        Decimal("50.00"),
    ]


def test_allocation_corrects_rounding_difference() -> None:
    rows = [
        {
            "calculated_weight": Decimal("1"),
            "floor_target": Decimal("0"),
            "is_floor_limited": False,
        }
        for _ in range(3)
    ]

    allocated, warnings = allocate_with_floors(rows, Decimal("100"))

    assert warnings == []
    assert sum(row["proposed_target"] for row in allocated) == Decimal("100.00")
    assert sorted(row["proposed_target"] for row in allocated) == [
        Decimal("33.33"),
        Decimal("33.33"),
        Decimal("33.34"),
    ]


@pytest.mark.asyncio
async def test_get_context_uses_latest_target_as_fallback() -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = "2026-05"
    repo.get_target_total.side_effect = [Decimal("0"), Decimal("90000")]
    repo.get_active_cohort.return_value = [
        {"site_code": "SITE01", "regional": "Regional B"},
        {"site_code": "SITE02", "regional": "Regional A"},
    ]

    result = await service.get_context()

    assert result == {
        "latest_sales_month": "2026-05",
        "suggested_target_month": "2026-06",
        "suggested_cohort_month": "2026-05",
        "suggested_total_target": 90000.0,
        "default_min_floor": 35000.0,
        "default_previous_month_floor_pct": 0.9,
        "default_previous_month_cap_pct": 1.7,
        "default_seasonality_years": 3,
        "active_store_count": 2,
        "regionals": ["Regional A", "Regional B"],
    }
    repo.get_active_cohort.assert_awaited_once_with("2026-05", "2026-06")


@pytest.mark.asyncio
async def test_get_context_rejects_missing_sales_data() -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = None

    with pytest.raises(HTTPException) as exc:
        await service.get_context()

    assert exc.value.status_code == 404


@pytest.mark.asyncio
async def test_profitability_uses_real_pnl_margin_gross_vat_and_flags_break_even() -> None:
    service, repo = make_service()
    repo.get_profitability_inputs.return_value = {
        "pnl_months": ["2026-02", "2026-03", "2026-04"],
        "pnl_rows": [
            {"site_code": "SITE01", "category_code": "v11", "amount": Decimal("30000")},
            {"site_code": "SITE01", "category_code": "c11", "amount": Decimal("12000")},
            {"site_code": "SITE01", "category_code": "c4", "amount": Decimal("3000")},
            {"site_code": "SITE01", "category_code": "c5", "amount": Decimal("6000")},
            {"site_code": "SITE01", "category_code": "c6", "amount": Decimal("9000")},
        ],
        "forecast_run": {
            "id": 44,
            "model_name": "TimesFM + XGRegressor",
            "model_mode": "ensemble",
            "variant": "august_exact",
            "generated_at": "2026-07-29T10:00:00",
        },
        "forecast_rows": [
            {"site_code": "SITE01", "forecast_sales": Decimal("35000")},
        ],
    }
    rows = [{
        "site_code": "SITE01",
        "calculated_weight": 2.0,
        "proposed_target": 40000.0,
    }]

    summary = await service._attach_profitability(
        {"target_month": "2026-08"},
        rows,
    )

    profitability = rows[0]["profitability"]
    assert rows[0]["normalized_weight"] == 1.0
    assert profitability["salary_cost_at_90_pct"] == 11597.22
    assert profitability["operating_costs"] == 6000.0
    assert profitability["accessory_margin_pct"] == 60.0
    assert profitability["break_even_gross_sales"] == 35487.73
    assert profitability["forecast_sales"] == 35000.0
    assert profitability["anomaly_flags"] == ["FORECAST_BELOW_BREAK_EVEN"]
    assert summary["status"] == "ready"
    assert summary["pnl_store_count"] == 1
    assert summary["forecast_store_count"] == 1
    assert summary["forecast_below_break_even_count"] == 1
    assert summary["target_below_break_even_count"] == 0


@pytest.mark.asyncio
async def test_calculate_builds_forecasted_rows_and_saves_draft(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = "2026-05"
    repo.get_active_cohort.return_value = [
        {
            "site_code": "SITE01",
            "locatie": "Magazin 1",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM 1",
        },
        {
            "site_code": "SITE02",
            "locatie": "Magazin 2",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM 2",
        },
    ]
    months = ["2025-05", "2025-06", "2026-05"]
    repo.get_source_metrics.return_value = [
        {
            "site_code": site_code,
            "import_month": month,
            "target": Decimal("40000") if site_code == "SITE01" else Decimal("60000"),
            "realized": Decimal("30000") if site_code == "SITE01" else Decimal("70000"),
        }
        for site_code in ("SITE01", "SITE02")
        for month in months
    ]

    async def forecast_factor(_conn: object, month: str) -> Decimal:
        return Decimal("1.20") if month == "2026-05" else Decimal("1")

    monkeypatch.setattr(target_module, "get_forecast_factor", forecast_factor)
    repo.save_draft_scenario.return_value = 9
    service.get_scenario_detail = AsyncMock(return_value={"id": 9})  # type: ignore[method-assign]

    result = await service.calculate(
        {
            "target_month": "2026-06",
            "total_target": 100000,
            "min_floor": 10000,
            "previous_month_floor_pct": 0.5,
            "expected_revision": 2,
        }
    )

    assert result == {"id": 9}
    repo.get_active_cohort.assert_awaited_once_with("2026-05", "2026-06")
    save_call = repo.save_draft_scenario.await_args
    assert save_call is not None
    saved_scenario, saved_rows, expected_revision = save_call.args
    assert expected_revision == 2
    assert saved_scenario["cohort_month"] == "2026-05"
    assert saved_scenario["calculation_method"] == CALCULATION_METHOD
    assert any("forecastate" in warning for warning in saved_scenario["warnings"])
    assert sum(row["proposed_target"] for row in saved_rows) == Decimal("100000.00")
    may_history = saved_rows[0]["history"][-1]
    assert may_history["actual_realized"] == 30000.0
    assert may_history["realized"] == 36000.0
    assert may_history["is_forecast"] is True


@pytest.mark.asyncio
async def test_calculate_flags_extreme_seasonality_and_capped_adjustments(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = "2026-05"
    repo.get_active_cohort.return_value = [
        {
            "site_code": "SITE01",
            "locatie": "Magazin 1",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM 1",
        },
    ]
    realized_by_month = {
        "2025-05": Decimal("100000"),
        "2025-06": Decimal("300000"),
        "2026-05": Decimal("500000"),
    }
    repo.get_source_metrics.return_value = [
        {
            "site_code": "SITE01",
            "import_month": month,
            "target": Decimal("0"),
            "realized": realized_by_month.get(month, Decimal("0")),
        }
        for month in ("2023-05", "2023-06", "2024-05", "2024-06", "2025-05", "2025-06", "2026-05")
    ]
    monkeypatch.setattr(
        target_module,
        "get_forecast_factor",
        AsyncMock(return_value=Decimal("1")),
    )
    repo.save_draft_scenario.return_value = 9
    service.get_scenario_detail = AsyncMock(return_value={"id": 9})  # type: ignore[method-assign]

    await service.calculate(
        {
            "target_month": "2026-06",
            "total_target": 800000,
            "min_floor": 10000,
            "previous_month_floor_pct": 0.1,
            "expected_revision": 2,
        }
    )

    saved_rows = repo.save_draft_scenario.await_args.args[1]
    flags = saved_rows[0]["calculation_details"]["flags"]
    assert "EXTREME_SEASONALITY" in flags
    assert "SEASONALITY_CAPPED" in flags
    assert "TREND_ADJUSTMENT_CAPPED" in flags
    assert saved_rows[0]["calculation_details"]["seasonality"]["used_factor"] == 1.7
    assert saved_rows[0]["calculation_details"]["trend"]["used_adjustment"] == 1.1


@pytest.mark.asyncio
async def test_calculate_rejects_budget_above_operational_cap(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = "2026-05"
    repo.get_active_cohort.return_value = [
        {
            "site_code": "SITE01",
            "locatie": "Magazin 1",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM 1",
        },
    ]
    repo.get_source_metrics.return_value = [
        {
            "site_code": "SITE01",
            "import_month": month,
            "target": Decimal("0"),
            "realized": Decimal("100000"),
        }
        for month in ("2025-05", "2025-06", "2026-05")
    ]
    monkeypatch.setattr(
        target_module,
        "get_forecast_factor",
        AsyncMock(return_value=Decimal("1")),
    )

    with pytest.raises(HTTPException) as exc:
        await service.calculate(
            {
                "target_month": "2026-06",
                "total_target": 500000,
                "min_floor": 10000,
                "previous_month_floor_pct": 0,
                "previous_month_cap_pct": 1.7,
                "expected_revision": 2,
            }
        )

    assert exc.value.status_code == 400
    assert "depaseste cap-ul maxim calculat" in exc.value.detail
    repo.save_draft_scenario.assert_not_awaited()


@pytest.mark.asyncio
async def test_calculate_marks_new_store_when_only_current_forecast_exists(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = "2026-05"
    repo.get_active_cohort.return_value = [
        {
            "site_code": "NEW01",
            "locatie": "Magazin nou",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM 1",
        },
    ]
    repo.get_source_metrics.return_value = [
        {
            "site_code": "NEW01",
            "import_month": month,
            "target": Decimal("0"),
            "realized": Decimal("100000") if month == "2026-05" else Decimal("0"),
        }
        for month in ("2023-05", "2023-06", "2024-05", "2024-06", "2025-05", "2025-06", "2026-05")
    ]
    monkeypatch.setattr(
        target_module,
        "get_forecast_factor",
        AsyncMock(return_value=Decimal("1")),
    )
    repo.save_draft_scenario.return_value = 9
    service.get_scenario_detail = AsyncMock(return_value={"id": 9})  # type: ignore[method-assign]

    await service.calculate(
        {
            "target_month": "2026-06",
            "total_target": 100000,
            "min_floor": 10000,
            "previous_month_floor_pct": 0,
            "previous_month_cap_pct": 1.7,
            "expected_revision": 2,
        }
    )

    saved_rows = repo.save_draft_scenario.await_args.args[1]
    details = saved_rows[0]["calculation_details"]
    assert details["seasonality"]["network_factor"] == 1.0
    assert details["seasonality"]["store_factor"] is None
    assert details["seasonality"]["weights"] == {"store": 0.0, "zone": 0.6, "network": 0.4}
    assert set(details["flags"]) >= {"NEW_STORE", "LOW_HISTORY"}


@pytest.mark.asyncio
async def test_calculate_distributes_uniformly_when_all_estimates_are_zero(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = "2026-05"
    repo.get_active_cohort.return_value = [
        {
            "site_code": "SITE01",
            "locatie": "Magazin 1",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM 1",
        },
        {
            "site_code": "SITE02",
            "locatie": "Magazin 2",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM 2",
        },
    ]
    months = ("2025-05", "2025-06", "2026-05")
    repo.get_source_metrics.return_value = [
        {
            "site_code": site_code,
            "import_month": month,
            "target": Decimal("0"),
            "realized": Decimal("0"),
        }
        for site_code in ("SITE01", "SITE02")
        for month in months
    ]
    monkeypatch.setattr(
        target_module,
        "get_forecast_factor",
        AsyncMock(return_value=Decimal("1")),
    )
    repo.save_draft_scenario.return_value = 9
    service.get_scenario_detail = AsyncMock(return_value={"id": 9})  # type: ignore[method-assign]

    await service.calculate(
        {
            "target_month": "2026-06",
            "total_target": 20000,
            "min_floor": 10000,
            "previous_month_floor_pct": 0,
            "previous_month_cap_pct": 1.7,
            "expected_revision": 2,
        }
    )

    saved_scenario, saved_rows, _expected_revision = repo.save_draft_scenario.await_args.args
    assert "distribuit uniform" in saved_scenario["warnings"][-1]
    assert [row["calculated_weight"] for row in saved_rows] == [Decimal("0.5"), Decimal("0.5")]
    assert all("LOW_HISTORY" in row["flags"] for row in saved_rows)
    assert sum(row["proposed_target"] for row in saved_rows) == Decimal("20000.00")


@pytest.mark.asyncio
async def test_calculate_weakens_store_weight_when_recent_year_is_missing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = "2026-05"
    repo.get_active_cohort.return_value = [
        {
            "site_code": "SITE01",
            "locatie": "Magazin 1",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM 1",
        },
    ]
    realized_by_month = {
        "2023-05": Decimal("100000"),
        "2023-06": Decimal("110000"),
        "2024-05": Decimal("100000"),
        "2024-06": Decimal("120000"),
        "2025-05": Decimal("0"),
        "2025-06": Decimal("0"),
        "2026-05": Decimal("100000"),
    }
    repo.get_source_metrics.return_value = [
        {
            "site_code": "SITE01",
            "import_month": month,
            "target": Decimal("0"),
            "realized": realized_by_month[month],
        }
        for month in realized_by_month
    ]
    monkeypatch.setattr(
        target_module,
        "get_forecast_factor",
        AsyncMock(return_value=Decimal("1")),
    )
    repo.save_draft_scenario.return_value = 9
    service.get_scenario_detail = AsyncMock(return_value={"id": 9})  # type: ignore[method-assign]

    await service.calculate(
        {
            "target_month": "2026-06",
            "total_target": 100000,
            "min_floor": 10000,
            "previous_month_floor_pct": 0,
            "expected_revision": 2,
        }
    )

    saved_rows = repo.save_draft_scenario.await_args.args[1]
    details = saved_rows[0]["calculation_details"]
    assert "LOW_RECENT_HISTORY" in details["flags"]
    assert details["seasonality"]["weights"] == {"store": 0.3, "zone": 0.4, "network": 0.3}
    assert details["seasonality"]["store_years"][0]["ratio"] is None
    assert details["seasonality"]["store_factor"] == 1.17


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("cohort_month", "cohort", "total_target", "expected_detail"),
    [
        (None, [], 100000, "Nu exista o luna"),
        ("2026-06", [{"site_code": "S"}], 100000, "Cohorta activa"),
        ("2026-05", [], 100000, "nu are magazine active"),
        ("2026-05", [{"site_code": "S"}], 0, "Parametrii de calcul"),
    ],
)
async def test_calculate_rejects_invalid_context_or_parameters(
    cohort_month: str | None,
    cohort: list[dict],
    total_target: int,
    expected_detail: str,
) -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = cohort_month
    repo.get_active_cohort.return_value = cohort

    with pytest.raises(HTTPException) as exc:
        await service.calculate(
            {
                "target_month": "2026-06",
                "total_target": total_target,
                "cohort_month": cohort_month,
            }
        )

    assert exc.value.status_code == 400
    assert expected_detail in exc.value.detail


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("error", "expected_detail"),
    [
        (TargetScenarioFinalizedError(), "deja finalizat"),
        (TargetScenarioVersionConflict(), "alt utilizator"),
    ],
)
async def test_calculate_maps_repository_conflicts(
    monkeypatch: pytest.MonkeyPatch,
    error: Exception,
    expected_detail: str,
) -> None:
    service, repo = make_service()
    repo.get_latest_sales_month.return_value = "2026-05"
    repo.get_active_cohort.return_value = [
        {
            "site_code": "SITE01",
            "locatie": "Magazin",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM",
        }
    ]
    repo.get_source_metrics.return_value = [
        {
            "site_code": "SITE01",
            "import_month": month,
            "target": Decimal("0"),
            "realized": Decimal("100000"),
        }
        for month in ("2025-05", "2025-06", "2026-05")
    ]
    monkeypatch.setattr(
        target_module,
        "get_forecast_factor",
        AsyncMock(return_value=Decimal("1")),
    )
    repo.save_draft_scenario.side_effect = error

    with pytest.raises(HTTPException) as exc:
        await service.calculate(
            {
                "target_month": "2026-06",
                "total_target": 100000,
                "min_floor": 10000,
            }
        )

    assert exc.value.status_code == 409
    assert expected_detail in exc.value.detail


@pytest.mark.asyncio
async def test_scenario_detail_serializes_totals_and_summaries() -> None:
    service, repo = make_service()
    repo.get_scenario.return_value = scenario_header()
    repo.get_scenario_rows.return_value = scenario_rows()

    result = await service.get_scenario_detail(9)

    assert result["store_count"] == 2
    assert result["proposed_total"] == 100000.0
    assert result["final_total"] == 100000.0
    assert result["remaining_difference"] == 0.0
    assert result["pending_final_count"] == 0
    assert result["floor_limited_count"] == 1
    assert result["manual_adjustments_count"] == 2
    assert result["regional_summary"] == [
        {
            "regional": "Regional A",
            "store_count": 2,
            "floor_total": 22000.0,
            "proposed_total": 100000.0,
            "final_total": 100000.0,
            "current_month": "2026-05",
            "current_forecast_total": 105000.0,
            "proposed_growth_vs_current_pct": -4.76,
            "final_growth_vs_current_pct": -4.76,
            "last_year_base_month": "2025-05",
            "last_year_target_month": "2025-06",
            "last_year_base_total": 105000.0,
            "last_year_target_total": 126000.0,
            "last_year_growth_pct": 20.0,
        }
    ]
    assert result["source_summary"][0]["is_forecast"] is True
    assert result["source_summary"][0]["attainment_pct"] == 105.0


def test_regional_summary_falls_back_to_history_when_details_are_missing() -> None:
    service, _repo = make_service()
    rows = [
        {
            "regional": "Regional A",
            "floor_target": 10000.0,
            "proposed_target": 50000.0,
            "final_target": None,
            "calculation_details": "{}",
            "history": (
                '[{"month":"2025-06","role":"seasonality_base_y1","realized":40000.0},'
                '{"month":"2025-07","role":"seasonality_target_y1","realized":60000.0},'
                '{"month":"2026-06","role":"floor_reference","realized":45000.0}]'
            ),
        }
    ]

    result = service._regional_summary(rows)

    assert result == [
        {
            "regional": "Regional A",
            "store_count": 1,
            "floor_total": 10000.0,
            "proposed_total": 50000.0,
            "final_total": 0.0,
            "current_month": "2026-06",
            "current_forecast_total": 45000.0,
            "last_year_base_month": "2025-06",
            "last_year_target_month": "2025-07",
            "last_year_base_total": 40000.0,
            "last_year_target_total": 60000.0,
            "proposed_growth_vs_current_pct": 11.11,
            "final_growth_vs_current_pct": -100.0,
            "last_year_growth_pct": 50.0,
        }
    ]


@pytest.mark.asyncio
async def test_scenario_list_and_missing_detail() -> None:
    service, repo = make_service()
    repo.list_scenarios.return_value = [scenario_header()]

    scenarios = await service.list_scenarios()

    assert scenarios[0]["total_target"] == 100000.0
    assert scenarios[0]["source_months"][0]["month"] == "2025-05"

    repo.get_scenario.return_value = None
    with pytest.raises(HTTPException) as exc:
        await service.get_scenario_detail(999)
    assert exc.value.status_code == 404


@pytest.mark.asyncio
async def test_save_final_targets_validates_rows_and_conflicts() -> None:
    service, repo = make_service()
    duplicate_rows = [
        {"site_code": "SITE01", "final_target": 10},
        {"site_code": "SITE01", "final_target": 20},
    ]
    with pytest.raises(HTTPException) as exc:
        await service.save_final_targets(9, duplicate_rows, 2)
    assert exc.value.status_code == 400

    repo.update_final_targets.side_effect = TargetScenarioVersionConflict()
    with pytest.raises(HTTPException) as exc:
        await service.save_final_targets(
            9,
            [{"site_code": "SITE01", "final_target": 10}],
            2,
        )
    assert exc.value.status_code == 409


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("stored_scenario", "expected_status"),
    [
        (None, 404),
        ({"status": "finalized"}, 409),
        ({"status": "draft"}, 400),
    ],
)
async def test_save_final_targets_maps_partial_updates(
    stored_scenario: dict | None,
    expected_status: int,
) -> None:
    service, repo = make_service()
    repo.update_final_targets.return_value = 0
    repo.get_scenario.return_value = stored_scenario

    with pytest.raises(HTTPException) as exc:
        await service.save_final_targets(
            9,
            [{"site_code": "SITE01", "final_target": 10}],
            2,
        )

    assert exc.value.status_code == expected_status


@pytest.mark.asyncio
async def test_save_final_targets_returns_refreshed_detail() -> None:
    service, repo = make_service()
    repo.update_final_targets.return_value = 1
    service.get_scenario_detail = AsyncMock(return_value={"id": 9, "revision": 3})  # type: ignore[method-assign]

    result = await service.save_final_targets(
        9,
        [{"site_code": "SITE01", "final_target": 10}],
        2,
    )

    assert result["revision"] == 3


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("scenario", "expected_status"),
    [
        ({"calculation_method": "legacy", "pending_final_count": 0, "final_total": 100, "total_target": 100}, 409),
        ({"calculation_method": CALCULATION_METHOD, "pending_final_count": 1, "final_total": 100, "total_target": 100}, 400),
        ({"calculation_method": CALCULATION_METHOD, "pending_final_count": 0, "final_total": 99, "total_target": 100}, 400),
    ],
)
async def test_finalize_validates_scenario(
    scenario: dict,
    expected_status: int,
) -> None:
    service, _repo = make_service()
    service.get_scenario_detail = AsyncMock(return_value=scenario)  # type: ignore[method-assign]

    with pytest.raises(HTTPException) as exc:
        await service.finalize(9, 2)

    assert exc.value.status_code == expected_status


@pytest.mark.asyncio
async def test_finalize_maps_conflict_false_and_success() -> None:
    service, repo = make_service()
    valid = {
        "calculation_method": CALCULATION_METHOD,
        "pending_final_count": 0,
        "final_total": 100,
        "total_target": 100,
    }
    service.get_scenario_detail = AsyncMock(side_effect=[valid, valid, valid, {"id": 9, "status": "finalized"}])  # type: ignore[method-assign]

    repo.finalize_scenario.side_effect = TargetScenarioVersionConflict()
    with pytest.raises(HTTPException) as exc:
        await service.finalize(9, 2)
    assert exc.value.status_code == 409

    repo.finalize_scenario.side_effect = None
    repo.finalize_scenario.return_value = False
    with pytest.raises(HTTPException) as exc:
        await service.finalize(9, 2)
    assert exc.value.status_code == 409

    repo.finalize_scenario.return_value = True
    result = await service.finalize(9, 2)
    assert result == {"id": 9, "status": "finalized"}


@pytest.mark.asyncio
async def test_export_excel_contains_audit_sheets() -> None:
    service, _repo = make_service()
    detail = {
        **scenario_header(
            source_months=[
                {
                    "month": "2025-05",
                    "label": "Anterior",
                    "role": "previous_year_reference",
                }
            ],
            warnings=["Atentionare test"],
            calculation_params={"strong_weights": {"store": 0.5}},
        ),
        "rows": [
            {
                **scenario_rows()[0],
                "calculated_weight": 0.4,
                "floor_target": 10000.0,
                "proposed_target": 40000.0,
                "final_target": 41000.0,
                "history": [
                    {
                        "month": "2025-05",
                        "target": 40000.0,
                        "realized": 42000.0,
                        "actual_realized": 35000.0,
                        "attainment_pct": 105.0,
                        "is_forecast": True,
                    }
                ],
                "calculation_details": {
                    "current_month": "2026-05",
                    "current_forecast": 42000.0,
                    "seasonality": {
                        "store_years": [
                            {
                                "year_offset": 1,
                                "base_month": "2025-05",
                                "target_month": "2025-06",
                                "base_value": 42000.0,
                                "target_value": 50400.0,
                                "ratio": 1.2,
                            }
                        ]
                    },
                },
                "normalized_weight": 1.0,
                "profitability": {
                    "salary_cost_at_90_pct": 12000.0,
                    "operating_costs": 10000.0,
                    "break_even_gross_sales": 35000.0,
                    "forecast_sales": 42000.0,
                    "anomaly_flags": [],
                },
            }
        ],
        "regional_summary": [
            {
                "regional": "Regional A",
                "store_count": 1,
                "floor_total": 10000.0,
                "proposed_total": 40000.0,
                "final_total": 41000.0,
                "current_month": "2026-05",
                "current_forecast_total": 42000.0,
                "proposed_growth_vs_current_pct": -4.76,
                "final_growth_vs_current_pct": -2.38,
                "last_year_base_month": "2025-05",
                "last_year_target_month": "2025-06",
                "last_year_base_total": 42000.0,
                "last_year_target_total": 50400.0,
                "last_year_growth_pct": 20.0,
            }
        ],
        "source_summary": [
            {
                "month": "2025-05",
                "is_forecast": True,
                "forecast_factor": 1.2,
                "actual_realized": 35000.0,
                "realized": 42000.0,
            }
        ],
        "profitability_summary": {
            "status": "ready",
            "pnl_months": ["2026-02", "2026-03", "2026-04"],
            "forecast_run": {
                "id": 44,
                "model_name": "TimesFM + XGRegressor",
                "variant": "august_exact",
            },
        },
    }
    service.get_scenario_detail = AsyncMock(return_value=detail)  # type: ignore[method-assign]

    output, filename = await service.export_excel(9)
    workbook = load_workbook(BytesIO(output.getvalue()))

    assert workbook.sheetnames == [
        "Target + profitabilitate",
        "Comparație manageri",
        "Rezumat calcul",
        "Parametri",
    ]
    main = workbook["Target + profitabilitate"]
    assert main["A1"].value == "SUBTOTAL"
    assert main["E1"].value == "=SUBTOTAL(109,E3:E3)"
    assert main["G1"].value == '=IF(E1=0,0,F1/E1)'
    assert [main.cell(2, column).value for column in range(1, 21)] == [
        "Firma", "Manager", "Nume locație", "Cod locație",
        "Target 2025-05", "Realizat 2025-05", "% 2025-05",
        "Target 2025-06", "Realizat 2025-06", "% 2025-06",
        "Target 2026-05", "Realizat 2026-05", "% 2026-05",
        "Pondere calcul", "Calcul target iunie", "Propunere manager",
        "Cheltuieli salariale la 90% - P&L estimat",
        "Cheltuieli operaționale estimate", "Break-even vânzări brute",
        "Forecast iunie",
    ]
    assert main["P3"].value == 41000.0
    assert main["Q3"].value == 12000.0
    assert main["S3"].value == 35000.0
    assert main["T3"].value == 42000.0
    assert len(main.conditional_formatting) == 6
    parameter_labels = [cell.value for cell in workbook["Parametri"]["A"]]
    assert "Parametru strong_weights" in parameter_labels
    assert "Forecast 2025-05" in parameter_labels
    assert "Status surse profitabilitate" in parameter_labels
    assert filename.startswith("targete_2026-06_scenariu_9_")


@pytest.mark.asyncio
async def test_store_detail_serializes_history_agents_and_statistics() -> None:
    service, repo = make_service()
    repo.get_store_detail.return_value = {
        "scenario": {
            "site_code": "SITE01",
            "locatie": "Magazin",
            "firma": "Mobiup",
            "regional": "Regional",
            "asm": "ASM",
            "target_month": "2026-06",
            "cohort_month": "2026-05",
            "proposed_target": Decimal("50000"),
            "final_target": None,
        },
        "history": [
            {
                "import_month": "2026-04",
                "total_sales": Decimal("40000"),
                "target_value": Decimal("50000"),
                "total_quantity": 100,
                "receipt_count": 50,
                "receipt_2plus_count": 10,
                "focus_quantity": 20,
                "cartele_qty": 5,
                "active_agents": 2,
                "working_days": 20,
            },
            {
                "import_month": "2026-05",
                "total_sales": Decimal("60000"),
                "target_value": Decimal("60000"),
                "total_quantity": 120,
                "receipt_count": 60,
                "receipt_2plus_count": 15,
                "focus_quantity": 24,
                "cartele_qty": 6,
                "active_agents": 3,
                "working_days": 21,
            },
        ],
        "agents": [
            {
                "agent": "Agent",
                "total_sales": Decimal("30000"),
                "sales_share_pct": Decimal("50"),
                "total_quantity": 60,
                "receipt_count": 30,
                "receipt_2plus_count": 9,
                "focus_quantity": 12,
                "active_months_16": 10,
                "sales_16m": Decimal("300000"),
            }
        ],
    }

    result = await service.get_store_detail(9, "SITE01")

    assert result["latest"]["month"] == "2026-05"
    assert result["best_month"]["total_sales"] == 60000.0
    assert result["avg_sales_16m"] == 50000.0
    assert result["history"][0]["target_pct"] == 80.0
    assert result["agents"][0]["bon2acc_pct"] == 30.0

    repo.get_store_detail.return_value = None
    with pytest.raises(HTTPException) as exc:
        await service.get_store_detail(9, "MISSING")
    assert exc.value.status_code == 404
