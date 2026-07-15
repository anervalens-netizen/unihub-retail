from __future__ import annotations

import json
import zipfile
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import Workbook

import services.grile_monthly as grile
from services.grile_monthly import ExtractedAgentRow, StoreEntry


class AsyncAcquire:
    def __init__(self, conn: MagicMock):
        self.conn = conn

    async def __aenter__(self) -> MagicMock:
        return self.conn

    async def __aexit__(self, *args: object) -> None:
        return None


def fake_pool(conn: MagicMock) -> MagicMock:
    pool = MagicMock()
    pool.acquire.return_value = AsyncAcquire(conn)
    return pool


def entry(
    *,
    company: str = "Mobiup",
    store: str = "Park Lake",
    sheet_id: str = "sheet-1",
    site_code: str = "SITE01",
    manager: str = "Manager 1",
) -> StoreEntry:
    return StoreEntry(company, store, sheet_id, site_code, manager)


def extracted(*, status: str = "OK", error: str = "") -> ExtractedAgentRow:
    return ExtractedAgentRow(
        company="Mobiup",
        store="Park Lake",
        slot=1,
        agent="Agent Test" if status == "OK" else "",
        base_salary=2600 if status == "OK" else "",
        sales_commission=300 if status == "OK" else "",
        extra_location_commission=25 if status == "OK" else "",
        extra_hours_pay=150 if status == "OK" else "",
        bonuri=480 if status == "OK" else "",
        worked_hours=176 if status == "OK" else "",
        status=status,
        error=error,
        sheet_id="sheet-1",
        site_code="SITE01",
        error_code="" if status == "OK" else (error or "test_error"),
    )


def test_month_helpers_and_output_paths(tmp_path: Path) -> None:
    assert grile.next_ym("2026-12") == "2027-01"
    assert grile.next_ym("2026-05") == "2026-06"
    assert grile.month_slug(" Iunie 2026 / Test ") == "Iunie-2026-Test"
    assert grile.safe_filename('  <bad>|name... ') == "_bad__name"
    assert grile.safe_filename("...") == "untitled"
    assert grile.build_final_export_path(tmp_path, "Iunie 2026").name == (
        "Tabel Salarii - Iunie 2026.xlsx"
    )
    assert grile.build_archive_manifest_path(tmp_path, "Iunie 2026").name == (
        "archive-manifest-Iunie-2026.json"
    )
    assert grile.build_archive_zip_path(tmp_path, "Iunie 2026").name == (
        "Grile - Iunie 2026.zip"
    )
    assert grile.build_reset_report_path(tmp_path, "Iulie 2026").name == (
        "reset-report-Iulie-2026.json"
    )
    assert grile.build_manager_zip_path(
        tmp_path, "Iunie 2026", "Manager/Unu"
    ).name == "Grile - Iunie 2026 - Manager - Unu.zip"
    assert grile.resolve_output_path("Iunie 2026", "Park/Lake", tmp_path).name == (
        "Tabel Salarii - Iunie 2026 - TEST Park - Lake.xlsx"
    )


def test_credentials_require_existing_service_account(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    missing = tmp_path / "missing.json"
    monkeypatch.setenv("GRILE_GOOGLE_SA_FILE", str(missing))

    with pytest.raises(FileNotFoundError, match="Service account Google lipsa"):
        grile.get_credentials()


def test_credentials_load_existing_service_account(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    credentials_file = tmp_path / "service-account.json"
    credentials_file.write_text("{}")
    expected = object()
    loader = MagicMock(return_value=expected)
    monkeypatch.setenv("GRILE_GOOGLE_SA_FILE", str(credentials_file))
    monkeypatch.setattr(
        "google.oauth2.service_account.Credentials.from_service_account_file",
        loader,
    )

    assert grile.get_credentials() is expected
    loader.assert_called_once_with(str(credentials_file), scopes=grile.SCOPES)


def test_build_google_services_uses_shared_credentials(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    credentials = object()
    calls: list[tuple[str, str, object, bool]] = []

    def build(service: str, version: str, *, credentials: object, cache_discovery: bool):
        calls.append((service, version, credentials, cache_discovery))
        return f"{service}-client"

    monkeypatch.setattr(grile, "get_credentials", lambda: credentials)
    monkeypatch.setattr(grile, "build", build)

    assert grile.build_google_services() == ("sheets-client", "drive-client")
    assert calls == [
        ("sheets", "v4", credentials, False),
        ("drive", "v3", credentials, False),
    ]


def test_retry_api_retries_transient_and_wraps_terminal_errors(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    attempts = 0
    sleeps: list[float] = []

    class ApiError(Exception):
        def __init__(self, status: int):
            super().__init__(f"status {status}")
            self.resp = SimpleNamespace(status=status)

    def flaky() -> str:
        nonlocal attempts
        attempts += 1
        if attempts < 3:
            raise ApiError(503)
        return "ok"

    monkeypatch.setattr(grile.time, "sleep", sleeps.append)
    assert grile.retry_api(flaky, label="read", attempts=4, base_delay=0.5) == "ok"
    assert sleeps == [0.5, 1.0]

    with pytest.raises(grile.MonthlyIntegrityError, match="read failed") as exc_info:
        grile.retry_api(
            lambda: (_ for _ in ()).throw(ApiError(400)),
            label="read",
        )
    assert exc_info.value.code == "google_request_failed"


@pytest.mark.parametrize(
    ("status", "expected_code"),
    [(429, "google_rate_limited"), (503, "google_unavailable")],
)
def test_retry_api_exhausts_google_429_and_503_without_coercion(
    status: int,
    expected_code: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class ApiError(Exception):
        def __init__(self) -> None:
            self.resp = SimpleNamespace(status=status)

    sleeps: list[float] = []
    monkeypatch.setattr(grile.time, "sleep", sleeps.append)
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile.retry_api(
            lambda: (_ for _ in ()).throw(ApiError()),
            label="Google",
            attempts=3,
            base_delay=0.1,
        )
    assert exc_info.value.code == expected_code
    assert sleeps == [0.1, 0.2]


def test_retry_api_timeout_is_fail_closed() -> None:
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile.retry_api(
            lambda: (_ for _ in ()).throw(TimeoutError()),
            label="Google",
            attempts=1,
        )
    assert exc_info.value.code == "google_timeout"


@pytest.mark.asyncio
async def test_load_entries_normalizes_registry_and_filters() -> None:
    conn = MagicMock()
    conn.fetch = AsyncMock(
        return_value=[
            {
                "registry_key": "Mobicell/AFI Cotroceni",
                "firma": "ignored",
                "locatie": "ignored",
                "sheet_id": "sheet-1",
                "site_code": "SITE01",
                "asm": " Manager 1 ",
            },
            {
                "registry_key": None,
                "firma": "Mobiup",
                "locatie": "Park Lake",
                "sheet_id": "sheet-2",
                "site_code": "SITE02",
                "asm": "",
            },
        ]
    )
    pool = fake_pool(conn)

    entries = await grile.load_entries(pool, only="manager 1")

    assert entries == [
        StoreEntry(
            "Mobicell",
            "AFI Cotroceni",
            "sheet-1",
            "SITE01",
            "Manager 1",
        )
    ]

    with pytest.raises(RuntimeError, match="No active grile"):
        await grile.load_entries(pool, only="missing")


def test_operation_serialization_and_number_helpers() -> None:
    assert grile._operation_to_dict(None) is None
    assert grile._operation_to_dict({"id": 1, "result": '{"ok": true}'}) == {
        "id": 1,
        "result": {"ok": True},
    }
    assert grile.scalar([]) == ""
    assert grile.scalar([[]]) == ""
    assert grile.scalar([[7]]) == 7
    assert grile.to_number(12) == 12
    assert grile.to_number("1.234,50") == 1234.5
    for invalid in (None, "", "invalid", object(), float("nan"), float("inf"), -1, True):
        with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
            grile.to_number(invalid)
        assert exc_info.value.code == "invalid_numeric_value"
    assert grile.sum_scalars([{"values": [["1,5"]]}, {"values": [[2]]}]) == 3.5


@pytest.mark.asyncio
async def test_operation_state_helpers_issue_expected_updates() -> None:
    conn = MagicMock()
    conn.execute = AsyncMock(
        side_effect=[
            "UPDATE 1",
            "UPDATE 0",
            "UPDATE 1",
            "UPDATE 1",
            "UPDATE 1",
            "UPDATE 1",
        ]
    )
    conn.executemany = AsyncMock()
    conn.fetchrow = AsyncMock(return_value={"status": "completed"})
    pool = fake_pool(conn)

    await grile.attach_monthly_operation_job(pool, operation_id=1, job_id="job-1")
    assert (await grile.start_monthly_operation(pool, 1)).status == "started"
    await grile.heartbeat_monthly_operation(pool, 1)
    await grile.fail_monthly_operation(pool, 1, error_message="failed")
    await grile.ensure_reset_items(
        pool,
        operation_id=1,
        closing_month_key="2026-06",
        next_month_key="2026-07",
        entries=[entry()],
    )
    previous = await grile.get_previous_completed_reset_item(
        pool,
        closing_month_key="2026-06",
        site_code="SITE01",
    )
    await grile.mark_reset_item_running(pool, operation_id=1, site_code="SITE01")
    await grile.finish_reset_item(
        pool,
        operation_id=1,
        site_code="SITE01",
        status="completed",
    )

    assert previous == {"status": "completed"}
    records = conn.executemany.await_args.args[1]
    assert records[0][:7] == (
        1,
        "2026-06",
        "2026-07",
        "SITE01",
        "sheet-1",
        "Mobiup",
        "Park Lake",
    )


def test_validate_archive_manifest_accepts_complete_and_reports_all_failures(
    tmp_path: Path,
) -> None:
    xlsx = tmp_path / "store.xlsx"
    archive = tmp_path / "archive.zip"
    xlsx.write_bytes(b"xlsx")
    archive.write_bytes(b"zip")
    valid = {
        "registry_count": 1,
        "exported_count": 1,
        "error_count": 0,
        "zip_path": str(archive),
        "stores": [
            {
                "company": "Mobiup",
                "store": "Park Lake",
                "status": "OK",
                "xlsx_path": str(xlsx),
            }
        ],
    }
    assert grile.validate_archive_manifest(valid, 1) == (True, [])

    invalid = {
        "registry_count": 2,
        "exported_count": 0,
        "error_count": 1,
        "zip_path": str(tmp_path / "missing.zip"),
        "stores": [
            {
                "company": "Mobiup",
                "store": "Park Lake",
                "status": "ERROR",
                "xlsx_path": str(tmp_path / "missing.xlsx"),
            }
        ],
    }
    ok, errors = grile.validate_archive_manifest(invalid, 1)
    assert ok is False
    assert len(errors) == 6


def make_sheets_value_service(value_ranges: list[dict[str, Any]] | Exception):
    request = MagicMock()
    if isinstance(value_ranges, Exception):
        request.execute.side_effect = value_ranges
    else:
        request.execute.return_value = {"valueRanges": value_ranges}
    values = MagicMock()
    values.batchGet.return_value = request
    spreadsheets = MagicMock()
    spreadsheets.values.return_value = values
    service = MagicMock()
    service.spreadsheets.return_value = spreadsheets
    return service


def test_extract_store_rows_reads_two_slots_and_returns_error_row() -> None:
    values: list[dict[str, Any]] = []
    for agent in ("Agent 1", ""):
        raw = ([
            agent, 2600, 10, 20, 30, 40, 50, 25, 150, 480, 176,
        ] if agent else [""] * 11)
        values.extend({"values": [[value]]} if value != "" else {"values": []} for value in raw)

    rows = grile.extract_store_rows(make_sheets_value_service(values), entry())

    assert len(rows) == 1
    assert rows[0].agent == "Agent 1"
    assert rows[0].sales_commission == 150

    error_rows = grile.extract_store_rows(
        make_sheets_value_service(RuntimeError("Google failed")),
        entry(),
    )
    assert error_rows[0].status == "ERROR"
    assert error_rows[0].error_code == "google_request_failed"

    missing_agent_values: list[dict[str, Any]] = [
        {"values": []},
        *({"values": [[1]]} for _ in range(10)),
        *({"values": []} for _ in range(11)),
    ]
    missing_agent = grile.extract_store_rows(
        make_sheets_value_service(list(missing_agent_values)),
        entry(),
    )
    assert missing_agent[0].error_code == "missing_or_invalid_agent"


def test_finalization_coverage_rejects_unexpected_store_and_conflicts() -> None:
    expected = [entry()]
    unexpected = ExtractedAgentRow(
        **{
            **extracted().__dict__,
            "site_code": "SITE99",
            "sheet_id": "unexpected-sheet",
        }
    )
    counts = grile._validate_finalization_coverage(expected, [unexpected])
    assert counts[:4] == (1, 0, 0, 0)
    assert "unexpected_store" in counts[4]
    assert "store_not_processed" in counts[4]

    conflicting = ExtractedAgentRow(
        **{
            **extracted().__dict__,
            "store": "Conflicting Store",
        }
    )
    counts = grile._validate_finalization_coverage(expected, [conflicting])
    assert "contradictory_store_metadata" in counts[4]


@pytest.mark.asyncio
async def test_finalize_month_rejects_errors_without_official_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    entries = [entry(), entry(store="Store 2", sheet_id="sheet-2", site_code="SITE02")]
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=entries))
    monkeypatch.setattr(grile, "build_google_services", lambda: (object(), object()))
    monkeypatch.setattr(grile, "_validate_source_workbook", lambda _path: None)
    monkeypatch.setattr(
        grile,
        "extract_store_rows",
        MagicMock(side_effect=[[extracted()], [extracted(status="ERROR", error="bad")]]),
    )

    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile.finalize_month(MagicMock(), "Iunie 2026", only="test", delay=0)
    assert exc_info.value.code == "finalization_incomplete"
    assert not grile.resolve_output_path("Iunie 2026", "test", tmp_path).exists()


@pytest.mark.asyncio
async def test_finalize_month_sleeps_only_between_entries(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    entries = [entry(), entry(store="Store 2", sheet_id="sheet-2", site_code="SITE02")]
    sleeps: list[float] = []

    async def sleep(delay: float) -> None:
        sleeps.append(delay)

    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=entries))
    monkeypatch.setattr(grile, "build_google_services", lambda: (object(), object()))
    monkeypatch.setattr(grile, "_validate_source_workbook", lambda _path: None)
    monkeypatch.setattr(
        grile,
        "extract_store_rows",
        MagicMock(side_effect=[[extracted()], [ExtractedAgentRow(**{**extracted().__dict__, "site_code": "SITE02", "store": "Store 2", "sheet_id": "sheet-2"})]]),
    )
    monkeypatch.setattr(grile.asyncio, "sleep", sleep)

    await grile.finalize_month(MagicMock(), "Iunie 2026", delay=0.25)

    assert sleeps == [0.25]


@pytest.mark.asyncio
async def test_partial_workbook_never_replaces_official_artifact(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    official = grile.build_final_export_path(tmp_path, "Iunie 2026")
    official.write_bytes(b"previous-official")
    previous_hash = grile.file_sha256(official)
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[entry()]))
    monkeypatch.setattr(grile, "build_google_services", lambda: (object(), object()))
    monkeypatch.setattr(grile, "extract_store_rows", MagicMock(return_value=[extracted()]))

    def build_partial(_rows: Any, output_path: Path, _metadata: Any) -> None:
        workbook = Workbook()
        workbook.active.title = "Grila"
        workbook.save(output_path)

    monkeypatch.setattr(grile, "build_workbook", build_partial)
    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile.finalize_month(
            MagicMock(),
            "Iunie 2026",
            delay=0,
            month_key="2026-06",
        )
    assert exc_info.value.code == "workbook_structure_invalid"
    assert grile.file_sha256(official) == previous_hash


def test_export_sheet_xlsx_success_empty_and_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class Downloader:
        def __init__(self, handle, request):
            self.handle = handle
            self.request = request

        def next_chunk(self):
            self.handle.write(self.request)
            return None, True

    drive = MagicMock()
    drive.files.return_value.export_media.return_value = b"xlsx"
    monkeypatch.setattr(grile, "MediaIoBaseDownload", Downloader)
    output = tmp_path / "store.xlsx"

    result = grile.export_sheet_xlsx(drive, entry(), output)
    assert result["status"] == "OK"
    assert result["bytes"] == 4

    drive.files.return_value.export_media.return_value = b""
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile.export_sheet_xlsx(drive, entry(), tmp_path / "empty.xlsx")
    assert exc_info.value.code == "empty_source_backup"

    drive.files.return_value.export_media.side_effect = RuntimeError("denied")
    with pytest.raises(RuntimeError, match="denied"):
        grile.export_sheet_xlsx(drive, entry(), tmp_path / "failed.xlsx")


def test_archive_and_manager_zips_preserve_relative_paths(tmp_path: Path) -> None:
    archive_dir = grile.build_archive_dir(tmp_path, "Iunie 2026")
    first = archive_dir / "Mobiup" / "Store 1.xlsx"
    second = archive_dir / "Mobicell" / "Store 2.xlsx"
    first.parent.mkdir(parents=True)
    second.parent.mkdir(parents=True)
    first.write_bytes(b"one")
    second.write_bytes(b"two")
    zip_path = archive_dir / "all.zip"

    grile.create_archive_zip(zip_path, [first, second], archive_dir)
    manager_zips = grile.create_manager_zips(
        tmp_path,
        "Iunie 2026",
        [
            {"status": "OK", "manager": "Manager 1", "xlsx_path": str(first)},
            {"status": "OK", "manager": "", "xlsx_path": str(second)},
            {"status": "ERROR", "manager": "Ignored", "xlsx_path": ""},
        ],
    )

    with zipfile.ZipFile(zip_path) as archive:
        assert archive.namelist() == [
            "Mobiup/Store 1.xlsx",
            "Mobicell/Store 2.xlsx",
        ]
    assert sorted(manager_zips) == ["Manager 1", "Neatribuit"]

    summary = grile.summarize_archive_results(
        "Iunie 2026",
        2,
        [
            {"status": "OK"},
            {"status": "ERROR"},
        ],
        zip_path,
        manager_zips,
    )
    assert summary["exported_count"] == 1
    assert summary["error_count"] == 1


def test_source_workbook_must_be_complete_and_readable(tmp_path: Path) -> None:
    invalid = tmp_path / "invalid.xlsx"
    invalid.write_bytes(b"not-an-xlsx")
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._validate_source_workbook(invalid)
    assert exc_info.value.code == "source_workbook_invalid"

    partial = tmp_path / "partial.xlsx"
    workbook = Workbook()
    workbook.active.title = "Grila"
    workbook.save(partial)
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._validate_source_workbook(partial)
    assert exc_info.value.code == "source_workbook_partial"

    complete = tmp_path / "complete.xlsx"
    workbook.create_sheet("Pontaj")
    workbook.save(complete)
    grile._validate_source_workbook(complete)


def patch_verified_final_manifest(
    monkeypatch: pytest.MonkeyPatch,
    output_dir: Path,
    *,
    stores: int,
    agents: int,
    registry: list[StoreEntry] | None = None,
) -> dict[str, Any]:
    final_path = grile.build_final_export_path(output_dir, "Iunie 2026")
    final_path.parent.mkdir(parents=True, exist_ok=True)
    final_path.write_bytes(b"verified-final")
    source_entries = registry or [
        entry(
            store=f"Store {index}",
            site_code=f"SITE{index:02d}",
            sheet_id=f"sheet-{index}",
        )
        for index in range(1, stores + 1)
    ]
    manifest = grile._with_source_registry(
        grile.base_manifest(
            month="2026-06",
            operation="finalize",
            requested_by_sub="test-subject",
            expected_stores=stores,
            expected_agents=agents,
            processed_stores=stores,
            processed_agents=agents,
            control_totals={"salary_components": "1.00"},
            artifacts=[grile.relative_artifact(final_path, root=output_dir, kind="final_workbook")],
        ),
        source_entries,
    )
    monkeypatch.setattr(
        grile,
        "fetch_latest_monthly_manifest",
        AsyncMock(return_value={"manifest": manifest, "status": "verified"}),
    )
    return manifest


@pytest.mark.asyncio
async def test_archive_month_writes_valid_manifest(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    entries = [
        entry(),
        entry(
            company="Mobicell",
            store="Store 2",
            sheet_id="sheet-2",
            site_code="SITE02",
            manager="Manager 2",
        ),
    ]
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    patch_verified_final_manifest(
        monkeypatch, tmp_path, stores=2, agents=3, registry=entries
    )
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=entries))
    monkeypatch.setattr(grile, "build_google_services", lambda: (object(), object()))
    monkeypatch.setattr(grile, "_validate_source_workbook", lambda _path: None)

    def export(_drive: object, item: StoreEntry, output_path: Path) -> dict[str, Any]:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(item.sheet_id.encode())
        return {
            "company": item.company,
            "store": item.store,
            "site_code": item.site_code,
            "manager": item.manager,
            "sheet_id": item.sheet_id,
            "status": "OK",
            "xlsx_path": str(output_path),
            "bytes": output_path.stat().st_size,
            "error": "",
        }

    monkeypatch.setattr(grile, "export_sheet_xlsx", export)

    manifest_path = await grile.archive_month(
        MagicMock(),
        "Iunie 2026",
        delay=0,
        month_key="2026-06",
    )

    manifest = json.loads(manifest_path.read_text())
    assert manifest["status"] == "verified"
    assert manifest["expected"] == {"agents": 3, "stores": 2}
    assert manifest["processed"] == manifest["expected"]
    assert manifest["error_count"] == 0
    grile.validate_verified_manifest(manifest, operation="archive")
    grile.verify_artifacts(manifest, root=tmp_path)


@pytest.mark.asyncio
async def test_archive_month_rejects_incomplete_export(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    patch_verified_final_manifest(monkeypatch, tmp_path, stores=1, agents=1)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[entry()]))
    monkeypatch.setattr(grile, "build_google_services", lambda: (object(), object()))
    monkeypatch.setattr(
        grile,
        "export_sheet_xlsx",
        lambda *_args: {
            "company": "Mobiup",
            "store": "Park Lake",
            "status": "ERROR",
            "xlsx_path": "",
            "error": "failed",
        },
    )

    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile.archive_month(MagicMock(), "Iunie 2026", delay=0, month_key="2026-06")
    assert exc_info.value.code == "archive_incomplete"


@pytest.mark.asyncio
async def test_archive_month_sleeps_only_between_entries(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    entries = [entry(), entry(store="Store 2", sheet_id="sheet-2", site_code="SITE02")]
    sleeps: list[float] = []

    async def sleep(delay: float) -> None:
        sleeps.append(delay)

    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    patch_verified_final_manifest(
        monkeypatch, tmp_path, stores=2, agents=2, registry=entries
    )
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=entries))
    monkeypatch.setattr(grile, "build_google_services", lambda: (object(), object()))
    monkeypatch.setattr(grile, "_validate_source_workbook", lambda _path: None)

    def export(_drive: object, item: StoreEntry, output_path: Path) -> dict[str, Any]:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"x")
        return {
            "company": item.company,
            "store": item.store,
            "site_code": item.site_code,
            "manager": item.manager,
            "sheet_id": item.sheet_id,
            "status": "OK",
            "xlsx_path": str(output_path),
            "bytes": 1,
            "error": "",
        }

    monkeypatch.setattr(grile, "export_sheet_xlsx", export)
    monkeypatch.setattr(grile.asyncio, "sleep", sleep)

    await grile.archive_month(
        MagicMock(),
        "Iunie 2026",
        delay=0.25,
        month_key="2026-06",
    )

    assert sleeps == [0.25]


def test_reset_store_dry_run_success_and_error() -> None:
    assert grile.reset_store(None, entry(), dry_run=True)["status"] == "DRY_RUN"

    request = MagicMock()
    request.execute.return_value = {}
    values = MagicMock()
    values.batchClear.return_value = request
    spreadsheets = MagicMock()
    spreadsheets.values.return_value = values
    service = MagicMock()
    service.spreadsheets.return_value = spreadsheets
    assert grile.reset_store(service, entry(), dry_run=False)["status"] == "OK"

    request.execute.side_effect = RuntimeError("clear failed")
    result = grile.reset_store(service, entry(), dry_run=False)
    assert result["status"] == "ERROR"
    assert result["error"] == "google_request_failed"


class StatefulResetService:
    def __init__(self) -> None:
        self.state = {
            item: [[index + 1]]
            for index, item in enumerate(grile.RESET_RANGES)
        }
        self.clear_calls = 0
        self.restore_calls = 0

    def spreadsheets(self) -> "StatefulResetService":
        return self

    def values(self) -> "StatefulResetService":
        return self

    def batchGet(self, **_kwargs: Any) -> SimpleNamespace:  # noqa: N802
        return SimpleNamespace(
            execute=lambda: {
                "valueRanges": [
                    {"range": item, "majorDimension": "ROWS", "values": self.state[item]}
                    for item in grile.RESET_RANGES
                ]
            }
        )

    def batchClear(self, **_kwargs: Any) -> SimpleNamespace:  # noqa: N802
        def execute() -> dict[str, Any]:
            self.clear_calls += 1
            self.state = {item: [] for item in grile.RESET_RANGES}
            return {}

        return SimpleNamespace(execute=execute)

    def batchUpdate(self, *, body: dict[str, Any], **_kwargs: Any) -> SimpleNamespace:  # noqa: N802
        def execute() -> dict[str, Any]:
            self.restore_calls += 1
            for item in body["data"]:
                self.state[item["range"]] = item["values"]
            return {}

        return SimpleNamespace(execute=execute)


def patch_archive_prerequisite(
    monkeypatch: pytest.MonkeyPatch,
    output_dir: Path,
    *,
    approved: bool,
) -> dict[str, Any]:
    source = output_dir / "archive" / "Iunie 2026" / "Mobiup" / "Store.xlsx"
    archive = output_dir / "archive" / "Iunie 2026" / "archive.zip"
    source.parent.mkdir(parents=True, exist_ok=True)
    source.write_bytes(b"source")
    archive.write_bytes(b"archive")
    source_artifact = grile.relative_artifact(source, root=output_dir, kind="source_workbook")
    source_artifact.update({"site_code": "SITE01", "sheet_id": "sheet-1"})
    manifest = grile.base_manifest(
        month="2026-06",
        operation="archive",
        requested_by_sub="request-subject",
        expected_stores=1,
        expected_agents=1,
        processed_stores=1,
        processed_agents=1,
        control_totals={"salary_components": "1.00"},
        artifacts=[
            source_artifact,
            grile.relative_artifact(archive, root=output_dir, kind="archive_zip"),
        ],
        source_backups=[source_artifact],
    )
    status = "verified"
    if approved:
        manifest["status"] = "approved"
        manifest["approved_by_sub"] = "approval-subject"
        manifest["approved_at"] = grile.utc_now()
        manifest = grile.finalize_manifest(manifest)
        status = "approved"
    record = {"id": 31, "status": status, "manifest": manifest}
    monkeypatch.setattr(grile, "fetch_latest_monthly_manifest", AsyncMock(return_value=record))
    monkeypatch.setattr(grile, "fetch_monthly_manifest", AsyncMock(return_value=record))
    return record


@pytest.mark.asyncio
async def test_manifest_approval_reverifies_hash_and_persists_subject(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    record = patch_archive_prerequisite(monkeypatch, tmp_path, approved=False)
    record.update(
        {
            "operation": "archive",
            "manifest_sha256": record["manifest"]["manifest_sha256"],
            "closing_month": "2026-06",
            "error_count": 0,
        }
    )
    approved_record = {
        **record,
        "status": "approved",
        "approved_by_sub": "stable-approval-subject",
    }
    persist = AsyncMock(return_value=approved_record)
    monkeypatch.setattr(grile, "persist_monthly_manifest_approval", persist)

    payload = await grile.approve_monthly_manifest(
        MagicMock(),
        manifest_id=31,
        approved_by_sub="stable-approval-subject",
    )

    assert payload["approved"] is True
    assert "approved_by_sub" not in payload
    persist.assert_awaited_once()
    assert persist.await_args is not None
    assert persist.await_args.kwargs["approved_by_sub"] == "stable-approval-subject"

    source_path = grile.resolve_artifact_path(
        tmp_path,
        record["manifest"]["source_backups"][0]["path"],
    )
    source_path.write_bytes(b"tampered")
    persist.reset_mock()
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        await grile.approve_monthly_manifest(
            MagicMock(),
            manifest_id=31,
            approved_by_sub="stable-approval-subject",
        )
    assert exc_info.value.code in {"artifact_size_mismatch", "artifact_hash_mismatch"}
    persist.assert_not_awaited()


@pytest.mark.asyncio
async def test_reset_month_dry_run_writes_report(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[entry()]))
    patch_archive_prerequisite(monkeypatch, tmp_path, approved=False)
    service = StatefulResetService()
    monkeypatch.setattr(grile, "build_google_services", lambda: (service, object()))

    report_path = await grile.reset_month(
        MagicMock(),
        "Iunie 2026",
        "Iulie 2026",
        dry_run=True,
        operation_id=7,
        closing_month_key="2026-06",
        next_month_key="2026-07",
    )

    report = json.loads(report_path.read_text())
    assert report["dry_run"] is True
    assert report["processed_store_count"] == 1
    assert service.clear_calls == 0


@pytest.mark.asyncio
async def test_reset_preflight_timeout_has_zero_destructive_effects(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[entry()]))
    patch_archive_prerequisite(monkeypatch, tmp_path, approved=True)
    service = StatefulResetService()
    monkeypatch.setattr(grile, "build_google_services", lambda: (service, object()))
    monkeypatch.setattr(grile, "ensure_reset_items", AsyncMock())
    monkeypatch.setattr(
        grile,
        "_read_reset_snapshot",
        MagicMock(side_effect=grile.MonthlyIntegrityError("google_timeout", "timeout")),
    )
    clear = MagicMock()
    monkeypatch.setattr(grile, "reset_store", clear)

    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile.reset_month(
            MagicMock(),
            "Iunie 2026",
            "Iulie 2026",
            dry_run=False,
            operation_id=12,
            closing_month_key="2026-06",
            next_month_key="2026-07",
            approved_manifest_id=31,
        )
    assert exc_info.value.code == "google_timeout"
    assert service.clear_calls == 0
    clear.assert_not_called()


@pytest.mark.asyncio
async def test_reset_live_success_requires_backup_and_verifies_every_clear(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[entry()]))
    patch_archive_prerequisite(monkeypatch, tmp_path, approved=True)
    service = StatefulResetService()
    monkeypatch.setattr(grile, "build_google_services", lambda: (service, object()))
    monkeypatch.setattr(grile, "ensure_reset_items", AsyncMock())
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    monkeypatch.setattr(grile, "record_reset_item_backup", AsyncMock(return_value=True))
    monkeypatch.setattr(grile, "mark_reset_item_running", AsyncMock(return_value=True))
    finish = AsyncMock(return_value=True)
    monkeypatch.setattr(grile, "finish_reset_item", finish)

    report_path = await grile.reset_month(
        MagicMock(),
        "Iunie 2026",
        "Iulie 2026",
        dry_run=False,
        operation_id=13,
        closing_month_key="2026-06",
        next_month_key="2026-07",
        approved_manifest_id=31,
    )

    assert report_path == grile.build_reset_report_path(tmp_path, "Iulie 2026")
    assert service.clear_calls == 1
    assert all(values == [] for values in service.state.values())
    finish.assert_awaited_once()
    assert finish.await_args is not None
    assert finish.await_args.kwargs["status"] == "completed"


@pytest.mark.asyncio
async def test_reset_month_live_failure_marks_checkpoint(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[entry()]))
    patch_archive_prerequisite(monkeypatch, tmp_path, approved=True)
    service = StatefulResetService()
    original_state = dict(service.state)
    monkeypatch.setattr(grile, "build_google_services", lambda: (service, object()))
    monkeypatch.setattr(grile, "ensure_reset_items", AsyncMock())
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    monkeypatch.setattr(grile, "mark_reset_item_running", AsyncMock(return_value=True))
    monkeypatch.setattr(grile, "record_reset_item_backup", AsyncMock(return_value=True))
    rollback = AsyncMock(return_value=True)
    monkeypatch.setattr(grile, "record_reset_item_rollback", rollback)
    monkeypatch.setattr(
        grile,
        "reset_store",
        lambda *_args, **_kwargs: {
            "company": "Mobiup",
            "store": "Park Lake",
            "site_code": "SITE01",
            "sheet_id": "sheet-1",
            "status": "ERROR",
            "error": "google_unavailable",
            "ranges": [],
        },
    )

    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile.reset_month(
            MagicMock(),
            "Iunie 2026",
            "Iulie 2026",
            dry_run=False,
            operation_id=7,
            closing_month_key="2026-06",
            next_month_key="2026-07",
            approved_manifest_id=31,
        )

    assert exc_info.value.code == "rolled_back"
    assert service.state == original_state
    assert service.restore_calls == 1
    rollback.assert_awaited_once()
    assert rollback.await_args is not None
    assert rollback.await_args.kwargs["restored"] is True


@pytest.mark.asyncio
@pytest.mark.parametrize("op", ["finalize", "archive", "reset"])
async def test_run_monthly_op_dispatches_and_finishes(
    op: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    pool = object()
    monkeypatch.setattr("db.connection.get_pool", AsyncMock(return_value=pool))
    monkeypatch.setattr(
        grile,
        "start_monthly_operation",
        AsyncMock(
            return_value=grile.MonthlyOperationStartResult(
                "started",
                9,
                operation={
                    "op": op,
                    "closing_month": "2026-06",
                    "only_filter": None,
                    "dry_run": True,
                    "requested_by_sub": "test-subject",
                    "approved_manifest_id": None,
                },
            )
        ),
    )
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    execution = grile.MonthlyExecution(Path("unused"), {"status": "verified"})
    monkeypatch.setattr(grile, "_finalize_month_execution", AsyncMock(return_value=execution))
    monkeypatch.setattr(grile, "_archive_month_execution", AsyncMock(return_value=execution))
    monkeypatch.setattr(grile, "_reset_month_execution", AsyncMock(return_value=execution))
    monkeypatch.setattr(
        grile,
        "persist_manifest_result",
        AsyncMock(
            return_value={
                "id": 1,
                "operation_id": 9,
                "closing_month": "2026-06",
                "operation": op,
                "status": "verified",
                "manifest": {},
                "error_count": 0,
            }
        ),
    )
    finish = AsyncMock(return_value=True)
    monkeypatch.setattr(grile, "finish_monthly_operation", finish)

    result = await grile.run_monthly_op(operation_id=9)

    assert result["status"] == "success"
    assert result["operation_id"] == 9
    finish.assert_awaited_once()


@pytest.mark.asyncio
async def test_run_monthly_op_captures_failure_and_validates_op(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with pytest.raises(ValueError, match="Operatie necunoscuta"):
        await grile.run_monthly_op(op="invalid", month="2026-06")

    monkeypatch.setattr("db.connection.get_pool", AsyncMock(return_value=object()))
    monkeypatch.setattr(
        grile,
        "_finalize_month_execution",
        AsyncMock(side_effect=RuntimeError("finalize failed")),
    )
    result = await grile.run_monthly_op(op="finalize", month="2026-06")
    assert result["status"] == "failed"
    assert result["exit_code"] == -1
    assert result["output"] == "Operation failed: monthly_operation_failed"


@pytest.mark.asyncio
async def test_fetch_download_final_archive_missing_and_invalid(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    final_path = grile.build_final_export_path(tmp_path, "Iunie 2026")
    final_path.write_bytes(b"final")
    archive_path = grile.build_archive_zip_path(tmp_path, "Iunie 2026")
    archive_path.parent.mkdir(parents=True)
    archive_path.write_bytes(b"archive")

    content, filename, media = await grile.fetch_download("final", "2026-06")
    assert (content, filename) == (b"final", "Tabel Salarii - Iunie 2026.xlsx")
    assert "spreadsheetml" in media

    content, filename, media = await grile.fetch_download("archive", "2026-06")
    assert (content, filename, media) == (
        b"archive",
        "Arhiva Grile - Iunie 2026.zip",
        "application/zip",
    )

    archive_path.unlink()
    with pytest.raises(FileNotFoundError):
        await grile.fetch_download("archive", "2026-06")
    with pytest.raises(ValueError, match="Tip download necunoscut"):
        await grile.fetch_download("invalid", "2026-06")


@pytest.mark.asyncio
async def test_repository_delegates_and_invalid_reservation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    pool = object()
    with pytest.raises(ValueError, match="Operatie necunoscuta"):
        await grile.reserve_monthly_operation(
            pool,
            op="invalid",
            month="2026-06",
            only=None,
            dry_run=False,
            requested_by_sub="subject-1",
        )

    reservation = object()
    reserve = AsyncMock(return_value=reservation)
    finish = AsyncMock(return_value=True)
    get_manifest = AsyncMock(return_value={"id": 1})
    latest_manifest = AsyncMock(return_value={"id": 2})
    backup = AsyncMock(return_value=True)
    rollback = AsyncMock(return_value=True)
    monkeypatch.setattr(grile, "persist_monthly_operation_reservation", reserve)
    monkeypatch.setattr(grile, "persist_monthly_operation_result", finish)
    monkeypatch.setattr(grile, "fetch_monthly_manifest", get_manifest)
    monkeypatch.setattr(grile, "fetch_latest_monthly_manifest", latest_manifest)
    monkeypatch.setattr(grile, "persist_reset_item_backup", backup)
    monkeypatch.setattr(grile, "persist_reset_item_rollback", rollback)

    assert await grile.reserve_monthly_operation(
        pool,
        op="finalize",
        month="2026-06",
        only=None,
        dry_run=False,
        requested_by_sub="subject-1",
    ) is reservation
    assert await grile.finish_monthly_operation(pool, 1, result={"ok": True}) is True
    assert await grile.get_monthly_manifest(pool, 1) == {"id": 1}
    assert await grile.get_latest_monthly_manifest(pool, month="2026-06") == {"id": 2}
    assert await grile.record_reset_item_backup(
        pool,
        operation_id=1,
        site_code="SITE01",
        backup_path="backup.json",
        backup_sha256="a" * 64,
    ) is True
    assert await grile.record_reset_item_rollback(
        pool,
        operation_id=1,
        site_code="SITE01",
        restored=True,
    ) is True


def test_extract_store_rows_rejects_incomplete_invalid_empty_and_duplicate() -> None:
    incomplete = grile.extract_store_rows(make_sheets_value_service([]), entry())
    assert incomplete[0].error_code == "google_response_incomplete"

    invalid_values: list[dict[str, Any]] = []
    for raw in (["Agent 1", "invalid", 10, 20, 30, 40, 50, 25, 150, 480, 176], [""] * 11):
        invalid_values.extend(
            {"values": [[value]]} if value != "" else {"values": []}
            for value in raw
        )
    invalid = grile.extract_store_rows(make_sheets_value_service(invalid_values), entry())
    assert invalid[0].error_code == "invalid_numeric_value"

    empty_values: list[dict[str, Any]] = [{"values": []} for _ in range(22)]
    empty = grile.extract_store_rows(make_sheets_value_service(empty_values), entry())
    assert empty[0].error_code == "store_has_no_agent"

    duplicate_values: list[dict[str, Any]] = []
    for _slot in range(2):
        duplicate_values.extend(
            {"values": [[value]]}
            for value in ["Agent 1", 2600, 10, 20, 30, 40, 50, 25, 150, 480, 176]
        )
    duplicate = grile.extract_store_rows(
        make_sheets_value_service(duplicate_values),
        entry(),
    )
    assert [row.error_code for row in duplicate] == ["", "duplicate_agent"]


def test_workbook_and_coverage_defensive_validation(tmp_path: Path) -> None:
    output = tmp_path / "with-error-row.xlsx"
    grile.build_workbook(
        [extracted(), extracted(status="ERROR", error="invalid_numeric_value")],
        output,
        {},
    )
    assert output.exists()

    duplicate_entry = entry(store="Duplicate", sheet_id="sheet-2")
    duplicate_agent = ExtractedAgentRow(**{**extracted().__dict__, "slot": 2})
    coverage = grile._validate_finalization_coverage(
        [entry(), duplicate_entry],
        [extracted(), duplicate_agent],
    )
    assert "duplicate_registry_entry" in coverage[4]
    assert "duplicate_agent" in coverage[4]

    coverage_mismatch = tmp_path / "coverage-mismatch.xlsx"
    grile.build_workbook([extracted()], coverage_mismatch, {})
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._validate_final_workbook(coverage_mismatch, expected_agents=2)
    assert exc_info.value.code == "workbook_coverage_incomplete"

    audit_invalid = tmp_path / "audit-invalid.xlsx"
    workbook = Workbook()
    workbook.active.title = "Mobiup"
    workbook.active.append(grile.HEADERS)
    workbook.active.append([1])
    workbook.create_sheet("Mobicell").append(grile.HEADERS)
    workbook.create_sheet("Audit").append(grile.AUDIT_HEADERS)
    workbook["Audit"].append([None] * 11 + ["ERROR"])
    workbook.save(audit_invalid)
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._validate_final_workbook(audit_invalid, expected_agents=1)
    assert exc_info.value.code == "workbook_audit_invalid"

    unreadable = tmp_path / "unreadable.xlsx"
    unreadable.write_bytes(b"invalid")
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._validate_final_workbook(unreadable, expected_agents=0)
    assert exc_info.value.code == "workbook_invalid"


def test_staging_and_atomic_file_promotion_restore_previous(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    stale = tmp_path / ".staging" / "finalize-1"
    stale.mkdir(parents=True)
    (stale / "stale").write_text("old")
    assert grile._staging_dir("finalize", 1) == stale
    assert not (stale / "stale").exists()

    destination = tmp_path / "official.xlsx"
    destination.write_bytes(b"previous")
    staged = tmp_path / "staged.xlsx"
    staged.write_bytes(b"replacement")
    grile._promote_file(staged, destination)
    assert destination.read_bytes() == b"replacement"

    staged_same_revision = tmp_path / "staged-again.xlsx"
    staged_same_revision.write_bytes(b"new")
    previous_revision = tmp_path / ".revisions" / (
        f"{destination.name}.{grile.file_sha256(destination)[:16]}"
    )
    previous_revision.parent.mkdir(parents=True, exist_ok=True)
    previous_revision.write_bytes(b"already-revisioned")
    grile._promote_file(staged_same_revision, destination)
    assert destination.read_bytes() == b"new"

    staged_failure = tmp_path / "staged-failure.xlsx"
    staged_failure.write_bytes(b"failure")
    original_replace = grile.os.replace

    def fail_staged(source: Path | str, target: Path | str) -> None:
        if Path(source) == staged_failure:
            raise OSError("promotion failed")
        original_replace(source, target)

    monkeypatch.setattr(grile.os, "replace", fail_staged)
    with pytest.raises(OSError, match="promotion failed"):
        grile._promote_file(staged_failure, destination)
    assert destination.read_bytes() == b"new"


def test_archive_zip_and_directory_promotion_fail_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    empty_zip = tmp_path / "empty.zip"
    with zipfile.ZipFile(empty_zip, "w"):
        pass
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._validate_archive_zip(empty_zip, expected_files=1)
    assert exc_info.value.code == "archive_coverage_incomplete"

    invalid_zip = tmp_path / "invalid.zip"
    invalid_zip.write_bytes(b"invalid")
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._validate_archive_zip(invalid_zip, expected_files=1)
    assert exc_info.value.code == "archive_invalid"

    class CorruptArchive:
        def __enter__(self) -> "CorruptArchive":
            return self

        def __exit__(self, *_args: object) -> None:
            return None

        def infolist(self) -> list[SimpleNamespace]:
            return [SimpleNamespace(is_dir=lambda: False, filename="one.xlsx")]

        def testzip(self) -> str:
            return "one.xlsx"

    monkeypatch.setattr(grile.zipfile, "ZipFile", lambda *_args, **_kwargs: CorruptArchive())
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._validate_archive_zip(tmp_path / "corrupt.zip", expected_files=1)
    assert exc_info.value.code == "archive_corrupt"

    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    destination = tmp_path / "archive" / "Iunie-2026"
    destination.mkdir(parents=True)
    (destination / "old").write_text("previous")
    staged = tmp_path / "staged-archive"
    staged.mkdir()
    (staged / "new").write_text("replacement")
    grile._promote_directory(staged, destination)
    assert (destination / "new").exists()

    staged_failure = tmp_path / "staged-archive-failure"
    staged_failure.mkdir()
    (staged_failure / "bad").write_text("failure")
    original_replace = grile.os.replace

    def fail_staged(source: Path | str, target: Path | str) -> None:
        if Path(source) == staged_failure:
            raise OSError("directory promotion failed")
        original_replace(source, target)

    monkeypatch.setattr(grile.os, "replace", fail_staged)
    with pytest.raises(OSError, match="directory promotion failed"):
        grile._promote_directory(staged_failure, destination)
    assert (destination / "new").exists()


@pytest.mark.asyncio
async def test_archive_requires_full_verified_finalization(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile._archive_month_execution(
            MagicMock(),
            "Iunie 2026",
            month_key="2026-06",
            requested_by_sub="subject-1",
            operation_id=1,
            only="one-store",
        )
    assert exc_info.value.code == "partial_archive_forbidden"

    monkeypatch.setattr(grile, "fetch_latest_monthly_manifest", AsyncMock(return_value=None))
    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile._archive_month_execution(
            MagicMock(),
            "Iunie 2026",
            month_key="2026-06",
            requested_by_sub="subject-1",
            operation_id=1,
        )
    assert exc_info.value.code == "verified_finalization_missing"

    patch_verified_final_manifest(monkeypatch, tmp_path, stores=1, agents=1)
    duplicate = [entry(), entry(store="Duplicate")]
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=duplicate))
    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile._archive_month_execution(
            MagicMock(),
            "Iunie 2026",
            month_key="2026-06",
            requested_by_sub="subject-1",
            operation_id=1,
        )
    assert exc_info.value.code == "registry_changed_or_duplicate_after_finalization"


@pytest.mark.asyncio
async def test_archive_rejects_same_count_registry_replacement(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    finalized_entries = [
        entry(),
        entry(store="Store 2", site_code="SITE02", sheet_id="sheet-2"),
    ]
    current_entries = [
        entry(),
        entry(store="Store 3", site_code="SITE03", sheet_id="sheet-3"),
    ]
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    patch_verified_final_manifest(
        monkeypatch,
        tmp_path,
        stores=2,
        agents=2,
        registry=finalized_entries,
    )
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=current_entries))

    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile._archive_month_execution(
            MagicMock(),
            "Iunie 2026",
            month_key="2026-06",
            requested_by_sub="subject-1",
            operation_id=1,
        )

    assert exc_info.value.code == "registry_changed_or_duplicate_after_finalization"


@pytest.mark.asyncio
async def test_manifest_approval_rejects_missing_wrong_state_and_race(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    fetch = AsyncMock(return_value=None)
    monkeypatch.setattr(grile, "fetch_monthly_manifest", fetch)
    with pytest.raises(FileNotFoundError):
        await grile.approve_monthly_manifest(
            MagicMock(), manifest_id=1, approved_by_sub="subject-1"
        )

    fetch.return_value = {"operation": "finalize", "status": "verified", "manifest": {}}
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        await grile.approve_monthly_manifest(
            MagicMock(), manifest_id=1, approved_by_sub="subject-1"
        )
    assert exc_info.value.code == "manifest_not_approvable"

    fetch.return_value = {"operation": "archive", "status": "verified", "manifest": None}
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        await grile.approve_monthly_manifest(
            MagicMock(), manifest_id=1, approved_by_sub="subject-1"
        )
    assert exc_info.value.code == "manifest_invalid"

    record = patch_archive_prerequisite(monkeypatch, tmp_path, approved=False)
    record.update(
        {
            "operation": "archive",
            "manifest_sha256": record["manifest"]["manifest_sha256"],
            "closing_month": "2026-06",
        }
    )
    fetch = AsyncMock(side_effect=[record, {**record, "status": "approved"}])
    monkeypatch.setattr(grile, "fetch_monthly_manifest", fetch)
    monkeypatch.setattr(
        grile,
        "persist_monthly_manifest_approval",
        AsyncMock(return_value=None),
    )
    replay = await grile.approve_monthly_manifest(
        MagicMock(), manifest_id=1, approved_by_sub="subject-1"
    )
    assert replay["approved"] is False

    fetch = AsyncMock(side_effect=[record, record])
    monkeypatch.setattr(grile, "fetch_monthly_manifest", fetch)
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        await grile.approve_monthly_manifest(
            MagicMock(), manifest_id=1, approved_by_sub="subject-1"
        )
    assert exc_info.value.code == "manifest_approval_race"


def test_reset_snapshot_restore_and_verification_guards() -> None:
    service = StatefulResetService()
    incomplete = MagicMock()
    incomplete.spreadsheets.return_value.values.return_value.batchGet.return_value.execute.return_value = {
        "valueRanges": []
    }
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._read_reset_snapshot(incomplete, entry())
    assert exc_info.value.code == "backup_response_incomplete"

    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._restore_reset_snapshot(service, entry(), {})
    assert exc_info.value.code == "backup_invalid"

    original = grile._read_reset_snapshot(service, entry())
    service.state[grile.RESET_RANGES[0]] = [[999]]
    service.batchUpdate = lambda **_kwargs: SimpleNamespace(execute=lambda: {})  # type: ignore[method-assign]
    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._restore_reset_snapshot(service, entry(), original)
    assert exc_info.value.code == "rollback_verification_failed"

    with pytest.raises(grile.MonthlyIntegrityError) as exc_info:
        grile._verify_reset_cleared(service, entry())
    assert exc_info.value.code == "reset_verification_failed"


@pytest.mark.asyncio
async def test_reset_rejects_missing_approval_partial_and_registry_drift(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile._reset_month_execution(
            MagicMock(),
            "Iunie 2026",
            "Iulie 2026",
            closing_month_key="2026-06",
            next_month_key="2026-07",
            requested_by_sub="subject-1",
            operation_id=None,
            approved_manifest_id=None,
            dry_run=False,
        )
    assert exc_info.value.code == "approved_manifest_required"

    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile._reset_month_execution(
            MagicMock(),
            "Iunie 2026",
            "Iulie 2026",
            closing_month_key="2026-06",
            next_month_key="2026-07",
            requested_by_sub="subject-1",
            operation_id=1,
            approved_manifest_id=31,
            only="one-store",
            dry_run=False,
        )
    assert exc_info.value.code == "partial_live_reset_forbidden"

    monkeypatch.setattr(grile, "fetch_latest_monthly_manifest", AsyncMock(return_value=None))
    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile._reset_month_execution(
            MagicMock(),
            "Iunie 2026",
            "Iulie 2026",
            closing_month_key="2026-06",
            next_month_key="2026-07",
            requested_by_sub="subject-1",
            operation_id=1,
            approved_manifest_id=None,
            dry_run=True,
        )
    assert exc_info.value.code == "verified_archive_required"

    patch_archive_prerequisite(monkeypatch, tmp_path, approved=False)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[]))
    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile._reset_month_execution(
            MagicMock(),
            "Iunie 2026",
            "Iulie 2026",
            closing_month_key="2026-06",
            next_month_key="2026-07",
            requested_by_sub="subject-1",
            operation_id=1,
            approved_manifest_id=None,
            dry_run=True,
        )
    assert exc_info.value.code == "registry_or_archive_coverage_changed"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("failure", "expected_code"),
    [
        ("backup", "backup_checkpoint_failed"),
        ("claim", "rolled_back"),
        ("finish", "rolled_back"),
    ],
)
async def test_live_reset_checkpoint_failures_are_fail_closed(
    failure: str,
    expected_code: str,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[entry()]))
    patch_archive_prerequisite(monkeypatch, tmp_path, approved=True)
    service = StatefulResetService()
    monkeypatch.setattr(grile, "build_google_services", lambda: (service, object()))
    monkeypatch.setattr(grile, "ensure_reset_items", AsyncMock())
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    monkeypatch.setattr(
        grile,
        "record_reset_item_backup",
        AsyncMock(return_value=failure != "backup"),
    )
    monkeypatch.setattr(
        grile,
        "mark_reset_item_running",
        AsyncMock(return_value=failure != "claim"),
    )
    monkeypatch.setattr(
        grile,
        "finish_reset_item",
        AsyncMock(return_value=failure != "finish"),
    )
    monkeypatch.setattr(grile, "record_reset_item_rollback", AsyncMock(return_value=True))

    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile.reset_month(
            MagicMock(),
            "Iunie 2026",
            "Iulie 2026",
            dry_run=False,
            operation_id=13,
            closing_month_key="2026-06",
            next_month_key="2026-07",
            approved_manifest_id=31,
        )
    assert exc_info.value.code == expected_code
    if failure == "backup":
        assert service.clear_calls == 0
    else:
        assert service.state[grile.RESET_RANGES[0]] == [[1]]


@pytest.mark.asyncio
async def test_live_reset_rollback_failure_is_uncertain(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[entry()]))
    patch_archive_prerequisite(monkeypatch, tmp_path, approved=True)
    service = StatefulResetService()
    monkeypatch.setattr(grile, "build_google_services", lambda: (service, object()))
    monkeypatch.setattr(grile, "ensure_reset_items", AsyncMock())
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    monkeypatch.setattr(grile, "record_reset_item_backup", AsyncMock(return_value=True))
    monkeypatch.setattr(grile, "mark_reset_item_running", AsyncMock(return_value=True))
    monkeypatch.setattr(grile, "finish_reset_item", AsyncMock(return_value=False))
    monkeypatch.setattr(
        grile,
        "_restore_reset_snapshot",
        MagicMock(side_effect=RuntimeError("restore failed")),
    )
    monkeypatch.setattr(grile, "record_reset_item_rollback", AsyncMock(return_value=False))

    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile.reset_month(
            MagicMock(),
            "Iunie 2026",
            "Iulie 2026",
            dry_run=False,
            operation_id=13,
            closing_month_key="2026-06",
            next_month_key="2026-07",
            approved_manifest_id=31,
        )
    assert exc_info.value.code == "uncertain"


@pytest.mark.asyncio
async def test_live_reset_output_failure_restores_all_google_values(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(grile, "OUTPUTS_DIR", tmp_path)
    monkeypatch.setattr(grile, "load_entries", AsyncMock(return_value=[entry()]))
    patch_archive_prerequisite(monkeypatch, tmp_path, approved=True)
    service = StatefulResetService()
    original_state = dict(service.state)
    monkeypatch.setattr(grile, "build_google_services", lambda: (service, object()))
    monkeypatch.setattr(grile, "ensure_reset_items", AsyncMock())
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    monkeypatch.setattr(grile, "record_reset_item_backup", AsyncMock(return_value=True))
    monkeypatch.setattr(grile, "mark_reset_item_running", AsyncMock(return_value=True))
    monkeypatch.setattr(grile, "finish_reset_item", AsyncMock(return_value=True))
    monkeypatch.setattr(grile, "record_reset_item_rollback", AsyncMock(return_value=True))
    monkeypatch.setattr(
        grile,
        "_promote_file",
        MagicMock(side_effect=OSError("report promotion failed")),
    )

    with pytest.raises(grile.MonthlyManifestError) as exc_info:
        await grile.reset_month(
            MagicMock(),
            "Iunie 2026",
            "Iulie 2026",
            dry_run=False,
            operation_id=13,
            closing_month_key="2026-06",
            next_month_key="2026-07",
            approved_manifest_id=31,
        )
    assert exc_info.value.code == "rolled_back"
    assert service.state == original_state
    assert service.restore_calls == 1


@pytest.mark.asyncio
async def test_run_monthly_op_persisted_state_guards(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("db.connection.get_pool", AsyncMock(return_value=object()))
    start = AsyncMock()
    monkeypatch.setattr(grile, "start_monthly_operation", start)

    start.return_value = grile.MonthlyOperationStartResult(
        "already_completed",
        1,
        result={"status": "success"},
    )
    replay = await grile.run_monthly_op(operation_id=1)
    assert replay["idempotent_replay"] is True

    start.return_value = grile.MonthlyOperationStartResult(
        "already_failed",
        2,
        operation={"op": "archive", "closing_month": "2026-06", "dry_run": False},
    )
    failed = await grile.run_monthly_op(operation_id=2)
    assert failed["status"] == "failed"
    assert failed["operation_status"] == "failed"

    start.return_value = grile.MonthlyOperationStartResult("started", 3, operation=None)
    with pytest.raises(RuntimeError, match="no persisted state"):
        await grile.run_monthly_op(operation_id=3)

    base_operation = {
        "op": "finalize",
        "closing_month": "2026-06",
        "only_filter": None,
        "dry_run": False,
        "requested_by_sub": "subject-1",
        "approved_manifest_id": None,
    }
    start.return_value = grile.MonthlyOperationStartResult(
        "started", 4, operation={**base_operation, "requested_by_sub": ""}
    )
    with pytest.raises(RuntimeError, match="no OIDC subject"):
        await grile.run_monthly_op(operation_id=4)

    start.return_value = grile.MonthlyOperationStartResult(
        "started", 5, operation={**base_operation, "op": "invalid"}
    )
    with pytest.raises(ValueError, match="Operatie necunoscuta"):
        await grile.run_monthly_op(operation_id=5)

    start.return_value = grile.MonthlyOperationStartResult(
        "started", 6, operation={**base_operation, "closing_month": None}
    )
    with pytest.raises(ValueError, match="month is required"):
        await grile.run_monthly_op(operation_id=6)


@pytest.mark.asyncio
async def test_run_monthly_op_rejects_partial_official_operation(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("db.connection.get_pool", AsyncMock(return_value=object()))
    monkeypatch.setattr(
        grile,
        "start_monthly_operation",
        AsyncMock(
            return_value=grile.MonthlyOperationStartResult(
                "started",
                7,
                operation={
                    "op": "finalize",
                    "closing_month": "2026-06",
                    "only_filter": "one-store",
                    "dry_run": False,
                    "requested_by_sub": "subject-1",
                    "approved_manifest_id": None,
                },
            )
        ),
    )
    persist = AsyncMock(return_value={"id": 1})
    finish = AsyncMock(return_value=True)
    monkeypatch.setattr(grile, "persist_manifest_result", persist)
    monkeypatch.setattr(grile, "finish_monthly_operation", finish)

    result = await grile.run_monthly_op(operation_id=7)

    assert result["status"] == "failed"
    assert result["output"].endswith("partial_official_operation_forbidden")
    persist.assert_awaited_once()
    finish.assert_awaited_once()


@pytest.mark.asyncio
@pytest.mark.parametrize("manifest_failure", [True, False])
async def test_run_monthly_op_persists_structured_failures(
    manifest_failure: bool,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("db.connection.get_pool", AsyncMock(return_value=object()))
    monkeypatch.setattr(
        grile,
        "start_monthly_operation",
        AsyncMock(
            return_value=grile.MonthlyOperationStartResult(
                "started",
                8,
                operation={
                    "op": "finalize",
                    "closing_month": "2026-06",
                    "only_filter": None,
                    "dry_run": False,
                    "requested_by_sub": "subject-1",
                    "approved_manifest_id": None,
                },
            )
        ),
    )
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    failed_manifest = grile.base_manifest(
        month="2026-06",
        operation="finalize",
        requested_by_sub="subject-1",
        expected_stores=0,
        expected_agents=0,
        processed_stores=0,
        processed_agents=0,
        control_totals={},
        artifacts=[],
        errors=["finalization_incomplete"],
        status="failed",
    )
    failure: Exception
    if manifest_failure:
        failure = grile.MonthlyManifestError(
            "finalization_incomplete", "failed", failed_manifest
        )
    else:
        failure = RuntimeError("unexpected failure")
    monkeypatch.setattr(
        grile,
        "_finalize_month_execution",
        AsyncMock(side_effect=failure),
    )
    persist = AsyncMock(return_value={"id": 1, "manifest": failed_manifest})
    monkeypatch.setattr(grile, "persist_manifest_result", persist)
    monkeypatch.setattr(grile, "finish_monthly_operation", AsyncMock(return_value=True))

    result = await grile.run_monthly_op(operation_id=8)

    assert result["status"] == "failed"
    assert result["exit_code"] == -1
    persist.assert_awaited_once()


@pytest.mark.asyncio
async def test_run_monthly_op_consumes_approved_manifest_atomically(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("db.connection.get_pool", AsyncMock(return_value=object()))
    monkeypatch.setattr(
        grile,
        "start_monthly_operation",
        AsyncMock(
            return_value=grile.MonthlyOperationStartResult(
                "started",
                9,
                operation={
                    "op": "reset",
                    "closing_month": "2026-06",
                    "only_filter": None,
                    "dry_run": False,
                    "requested_by_sub": "subject-1",
                    "approved_manifest_id": 31,
                },
            )
        ),
    )
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    execution = grile.MonthlyExecution(Path("report.json"), {"status": "verified"})
    monkeypatch.setattr(grile, "_reset_month_execution", AsyncMock(return_value=execution))
    monkeypatch.setattr(
        grile,
        "persist_manifest_result",
        AsyncMock(return_value={"id": 2, "status": "verified", "manifest": {}}),
    )
    approved_manifest = {"status": "approved", "manifest_sha256": "b" * 64}
    monkeypatch.setattr(
        grile,
        "fetch_monthly_manifest",
        AsyncMock(
            return_value={
                "manifest": approved_manifest,
                "manifest_sha256": approved_manifest["manifest_sha256"],
            }
        ),
    )
    consume = AsyncMock(
        return_value={
            "id": 41,
            "operation_id": 9,
            "closing_month": "2026-06",
            "operation": "reset",
            "status": "verified",
            "manifest": execution.manifest,
            "error_count": 0,
        }
    )
    monkeypatch.setattr(grile, "persist_reset_success", consume)

    result = await grile.run_monthly_op(operation_id=9)

    assert result["status"] == "success"
    consume.assert_awaited_once()
    assert consume.await_args is not None
    assert consume.await_args.kwargs["reset_manifest"] == execution.manifest
    consumed = consume.await_args.kwargs["consumed_manifest"]
    assert consumed["status"] == "consumed"


@pytest.mark.asyncio
async def test_run_monthly_op_commit_failure_rolls_back_google_effects(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("db.connection.get_pool", AsyncMock(return_value=object()))
    monkeypatch.setattr(
        grile,
        "start_monthly_operation",
        AsyncMock(
            return_value=grile.MonthlyOperationStartResult(
                "started",
                9,
                operation={
                    "op": "reset",
                    "closing_month": "2026-06",
                    "only_filter": None,
                    "dry_run": False,
                    "requested_by_sub": "subject-1",
                    "approved_manifest_id": 31,
                },
            )
        ),
    )
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    rollback_manifest = grile.base_manifest(
        month="2026-06",
        operation="reset",
        requested_by_sub="subject-1",
        expected_stores=1,
        expected_agents=1,
        processed_stores=0,
        processed_agents=0,
        control_totals={},
        artifacts=[],
        errors=["reset_commit_failed", "rollback_verified"],
        status="rolled_back",
    )
    rollback = AsyncMock(return_value=rollback_manifest)
    execution = grile.MonthlyExecution(
        Path("report.json"),
        {"status": "verified"},
        rollback=rollback,
    )
    monkeypatch.setattr(grile, "_reset_month_execution", AsyncMock(return_value=execution))
    approved_manifest = {"status": "approved", "manifest_sha256": "b" * 64}
    monkeypatch.setattr(
        grile,
        "fetch_monthly_manifest",
        AsyncMock(
            return_value={
                "manifest": approved_manifest,
                "manifest_sha256": approved_manifest["manifest_sha256"],
            }
        ),
    )
    monkeypatch.setattr(
        grile,
        "persist_reset_success",
        AsyncMock(side_effect=RuntimeError("commit failed")),
    )
    monkeypatch.setattr(
        grile,
        "persist_manifest_result",
        AsyncMock(
            return_value={
                "id": 41,
                "operation_id": 9,
                "closing_month": "2026-06",
                "operation": "reset",
                "status": "rolled_back",
                "manifest": rollback_manifest,
                "error_count": 2,
            }
        ),
    )
    monkeypatch.setattr(grile, "finish_monthly_operation", AsyncMock(return_value=True))

    result = await grile.run_monthly_op(operation_id=9)

    rollback.assert_awaited_once()
    assert result["status"] == "failed"
    assert result["output"].endswith("rolled_back")


@pytest.mark.asyncio
async def test_run_monthly_op_reports_lost_completion_lease(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("db.connection.get_pool", AsyncMock(return_value=object()))
    monkeypatch.setattr(
        grile,
        "start_monthly_operation",
        AsyncMock(
            return_value=grile.MonthlyOperationStartResult(
                "started",
                10,
                operation={
                    "op": "finalize",
                    "closing_month": "2026-06",
                    "only_filter": None,
                    "dry_run": False,
                    "requested_by_sub": "subject-1",
                    "approved_manifest_id": None,
                },
            )
        ),
    )
    monkeypatch.setattr(grile, "heartbeat_monthly_operation", AsyncMock())
    execution = grile.MonthlyExecution(Path("final.xlsx"), {"status": "verified"})
    monkeypatch.setattr(grile, "_finalize_month_execution", AsyncMock(return_value=execution))
    monkeypatch.setattr(grile, "persist_manifest_result", AsyncMock(return_value=None))
    monkeypatch.setattr(grile, "finish_monthly_operation", AsyncMock(return_value=False))

    result = await grile.run_monthly_op(operation_id=10)

    assert result["status"] == "failed"
    assert result["output"].endswith("operation_lease_lost")
