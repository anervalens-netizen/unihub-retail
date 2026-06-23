from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from services.exports import (
    COMPARISON_LEVELS,
    ExportValidationError,
    ExportsService,
)


class FakeRepo:
    def __init__(self) -> None:
        self.report_calls: list[dict[str, Any]] = []
        self.daily_comparison_calls: list[dict[str, Any]] = []
        self.daily_evolution_rows: list[dict[str, Any]] = []

    async def fetch_report_rows(self, **kwargs: Any) -> list[dict[str, Any]]:
        self.report_calls.append(kwargs)
        period = kwargs.get("period")
        if period == "month":
            return [
                row(period_key=month, total_sales=amount)
                for month, amount in [("2026-05", "1000"), ("2026-06", "1500")]
            ]
        if period == "day":
            return [
                row(period_key="2026-06-01", total_sales="500", total_quantity=5, total_receipts=2),
                row(period_key="2026-06-02", total_sales="700", total_quantity=7, total_receipts=3),
            ]
        return [row(total_sales="2500", total_quantity=25, total_receipts=10)]

    async def fetch_daily_evolution_rows(self, **_kwargs: Any) -> list[dict[str, Any]]:
        return self.daily_evolution_rows

    async def fetch_daily_comparison_rows(self, **kwargs: Any) -> list[dict[str, Any]]:
        self.daily_comparison_calls.append(kwargs)
        return [
            row(import_month="2026-05", day_of_month=1, total_sales="100", total_quantity=10, total_receipts=5),
            row(import_month="2026-06", day_of_month=1, total_sales="150", total_quantity=15, total_receipts=6),
        ]


def row(**overrides: Any) -> dict[str, Any]:
    defaults: dict[str, Any] = {
        "agent": "Agent 1",
        "site_code": "S001",
        "locatie": "Magazin 1",
        "firma": "Firma 1",
        "regional": "RM 1",
        "asm": "ASM 1",
        "period_key": "2026-06",
        "import_month": "2026-06",
        "day_of_month": 1,
        "total_sales": Decimal("1000"),
        "total_quantity": 10,
        "total_receipts": 5,
        "receipt_2plus_count": 2,
        "focus_quantity": 1,
        "target": Decimal("2000"),
        "working_days": 2,
        "store_count": 1,
        "agent_count": 1,
    }
    defaults.update(overrides)
    for key in ("total_sales", "target"):
        defaults[key] = Decimal(str(defaults[key]))
    return defaults


@pytest.mark.asyncio
async def test_preview_builds_monthly_and_daily_columns_without_db() -> None:
    repo = FakeRepo()
    service = ExportsService(repo)  # type: ignore[arg-type]

    result = await service.preview(
        {
            "dataset": "agents",
            "months": ["2026-06", "2026-05"],
            "dimensions": ["agent", "site_code"],
            "metrics": ["total_sales", "target_progress_pct", "avg_receipt_value"],
            "monthly_metrics": ["total_sales"],
            "daily_metrics": ["total_sales"],
            "preview_limit": 10,
        }
    )

    keys = [column["key"] for column in result["columns"]]
    assert keys[:5] == ["agent", "site_code", "total_sales", "target_progress_pct", "avg_receipt_value"]
    assert "month:2026-05:total_sales" in keys
    assert "month:2026-06:total_sales" in keys
    assert "day:2026-06-01:total_sales" in keys
    assert result["total_rows"] == 1
    assert result["rows"][0]["total_sales"] == 2500.0
    assert result["rows"][0]["target_progress_pct"] == 125.0
    assert result["rows"][0]["avg_receipt_value"] == 250.0
    assert repo.report_calls[0]["months"] == ["2026-05", "2026-06"]


@pytest.mark.asyncio
async def test_preview_rejects_invalid_or_too_wide_requests() -> None:
    service = ExportsService(FakeRepo())  # type: ignore[arg-type]

    with pytest.raises(ExportValidationError, match="Dataset invalid"):
        await service.preview({"dataset": "bad", "months": ["2026-06"]})

    with pytest.raises(ExportValidationError, match="Evolutia zilnica este limitata"):
        await service.preview(
            {
                "dataset": "stores",
                "months": ["2026-03", "2026-04", "2026-05", "2026-06"],
                "daily_metrics": ["total_sales"],
            }
        )

    with pytest.raises(ExportValidationError, match="Selectie invalida pentru metrici"):
        await service.preview({"dataset": "stores", "months": ["2026-06"], "metrics": ["unknown"]})


@pytest.mark.asyncio
async def test_daily_comparison_preview_computes_delta_rows() -> None:
    repo = FakeRepo()
    service = ExportsService(repo)  # type: ignore[arg-type]

    result = await service.preview(
        {
            "export_mode": "daily_comparison",
            "dataset": "stores",
            "months": ["2026-05", "2026-06"],
            "daily_metrics": ["total_sales"],
            "comparison_levels": ["general"],
            "preview_limit": 2,
        }
    )

    keys = [column["key"] for column in result["columns"]]
    assert keys == [
        "day_of_month",
        "2026-05:total_sales",
        "2026-06:total_sales",
        "delta:total_sales",
        "delta_pct:total_sales",
    ]
    assert result["total_rows"] == 31
    assert result["truncated"] is True
    assert result["rows"][0]["day_of_month"] == 1
    assert result["rows"][0]["2026-05:total_sales"] == 100.0
    assert result["rows"][0]["2026-06:total_sales"] == 150.0
    assert result["rows"][0]["delta:total_sales"] == 50.0
    assert result["rows"][0]["delta_pct:total_sales"] == 50.0
    assert repo.daily_comparison_calls[0]["level"] == "general"


@pytest.mark.asyncio
async def test_build_xlsx_sanitizes_filename_and_writes_config_sheet() -> None:
    service = ExportsService(FakeRepo())  # type: ignore[arg-type]

    content, filename = await service.build_xlsx(
        {
            "dataset": "stores",
            "months": ["2026-06"],
            "dimensions": ["site_code", "locatie"],
            "metrics": ["total_sales"],
            "filename": "../raport retail final",
        }
    )

    assert filename == "raport_retail_final.xlsx"
    wb = load_workbook(BytesIO(content), read_only=True)
    assert wb.sheetnames == ["Raport", "Configuratie"]
    config = wb["Configuratie"]
    assert config["A2"].value == "Dataset"
    assert config["B2"].value == "Magazine"


def test_catalog_exposes_datasets_metrics_and_comparison_levels() -> None:
    catalog = ExportsService(FakeRepo()).catalog()  # type: ignore[arg-type]

    assert {item["key"] for item in catalog["datasets"]} == {
        "agents",
        "stores",
        "regionals",
        "asms",
    }
    assert {item["key"] for item in catalog["comparison_levels"]} == set(
        COMPARISON_LEVELS
    )
    assert any(item["key"] == "target_progress_pct" for item in catalog["metrics"])


@pytest.mark.asyncio
async def test_report_validation_limits_months_and_daily_width() -> None:
    service = ExportsService(FakeRepo())  # type: ignore[arg-type]

    with pytest.raises(ExportValidationError, match="cel putin o luna"):
        await service.build_report({"dataset": "stores", "months": []})

    with pytest.raises(ExportValidationError, match="maxim 24 luni"):
        await service.build_report(
            {
                "dataset": "stores",
                "months": [f"2024-{index:02d}" for index in range(1, 26)],
            }
        )

    with pytest.raises(ExportValidationError, match="Prea multe coloane"):
        await service.build_report(
            {
                "dataset": "stores",
                "months": ["2026-04", "2026-05", "2026-06"],
                "daily_metrics": [
                    "total_sales",
                    "total_quantity",
                    "total_receipts",
                ],
            }
        )


@pytest.mark.asyncio
async def test_preview_limit_truncates_regular_report() -> None:
    class TwoRowsRepo(FakeRepo):
        async def fetch_report_rows(self, **kwargs: Any) -> list[dict[str, Any]]:
            if kwargs.get("period"):
                return []
            return [
                row(site_code="S001", locatie="A"),
                row(site_code="S002", locatie="B"),
            ]

    result = await ExportsService(TwoRowsRepo()).preview(  # type: ignore[arg-type]
        {
            "dataset": "stores",
            "months": ["2026-06"],
            "preview_limit": 1,
        }
    )

    assert result["total_rows"] == 2
    assert len(result["rows"]) == 1
    assert result["truncated"] is True


@pytest.mark.asyncio
async def test_build_xlsx_adds_daily_evolution_sheet_with_deltas() -> None:
    repo = FakeRepo()
    repo.daily_evolution_rows = [
        row(
            import_month="2026-05",
            day_of_month=1,
            total_sales=100,
            total_quantity=10,
            total_receipts=5,
            receipt_2plus_count=1,
        ),
        row(
            import_month="2026-06",
            day_of_month=1,
            total_sales=150,
            total_quantity=15,
            total_receipts=6,
            receipt_2plus_count=3,
        ),
    ]
    service = ExportsService(repo)  # type: ignore[arg-type]

    content, _filename = await service.build_xlsx(
        {
            "dataset": "stores",
            "months": ["2026-05", "2026-06"],
            "dimensions": ["site_code"],
            "metrics": ["total_sales"],
            "daily_metrics": ["total_sales", "proc_bon2acc"],
        }
    )

    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == ["Raport", "Configuratie", "Evolutie zilnica"]
    sheet = workbook["Evolutie zilnica"]
    headers = [cell.value for cell in sheet[1]]
    assert "Delta Vanzari" in headers
    assert "Delta % Vanzari" in headers
    assert "Delta pp Bon2Acc %" in headers
    assert sheet["A2"].value == 1
    assert len(sheet._charts) == 1


@pytest.mark.asyncio
async def test_daily_comparison_xlsx_writes_requested_levels_and_charts() -> None:
    repo = FakeRepo()
    service = ExportsService(repo)  # type: ignore[arg-type]

    content, filename = await service.build_xlsx(
        {
            "export_mode": "daily_comparison",
            "months": ["2026-05", "2026-06"],
            "daily_metrics": ["total_sales", "proc_bon2acc"],
            "comparison_levels": ["general", "stores"],
            "include_closed_stores": True,
            "filename": "comparatie / iunie",
        }
    )

    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == ["General", "Magazine", "Configuratie"]
    assert len(workbook["General"]._charts) == 1
    assert len(workbook["Magazine"]._charts) == 1
    assert workbook["Configuratie"]["B6"].value == "Da"
    assert filename == "comparatie___iunie.xlsx"
    assert [call["level"] for call in repo.daily_comparison_calls] == [
        "general",
        "stores",
    ]


@pytest.mark.asyncio
async def test_daily_comparison_validation_and_percent_delta() -> None:
    service = ExportsService(FakeRepo())  # type: ignore[arg-type]

    with pytest.raises(ExportValidationError, match="cel putin o luna"):
        await service.preview({"export_mode": "daily_comparison", "months": []})

    with pytest.raises(ExportValidationError, match="maxim 6 luni"):
        await service.preview(
            {
                "export_mode": "daily_comparison",
                "months": [f"2026-{month:02d}" for month in range(1, 8)],
            }
        )

    with pytest.raises(ExportValidationError, match="maxim 4 metrici"):
        await service.preview(
            {
                "export_mode": "daily_comparison",
                "months": ["2026-06"],
                "daily_metrics": [
                    "total_sales",
                    "total_quantity",
                    "total_receipts",
                    "avg_receipt_value",
                    "proc_bon2acc",
                ],
            }
        )

    result = await service.preview(
        {
            "export_mode": "daily_comparison",
            "months": ["2026-05", "2026-06"],
            "daily_metrics": ["proc_bon2acc"],
            "comparison_levels": ["general"],
            "preview_limit": 1,
        }
    )
    assert result["rows"][0]["delta_pp:proc_bon2acc"] == -6.67


def test_daily_comparison_table_handles_missing_and_invalid_days() -> None:
    service = ExportsService(FakeRepo())  # type: ignore[arg-type]
    table = service._daily_comparison_table(
        level="general",
        months=["invalid", "2026-02"],
        metrics=["total_sales"],
        records=[
            row(import_month="2026-02", day_of_month=0),
            row(import_month="2026-02", day_of_month=1, total_sales=0),
        ],
    )

    assert table["max_day"] == 28
    assert table["rows"][0]["invalid:total_sales"] is None
    assert table["rows"][0]["delta:total_sales"] is None
    assert table["rows"][0]["delta_pct:total_sales"] is None

    empty_general = service._daily_comparison_table(
        level="general",
        months=["bad"],
        metrics=["total_sales"],
        records=[],
    )
    assert empty_general["max_day"] == 31
    assert len(empty_general["rows"]) == 31
    assert service._max_days_for_months(["2026-12"]) == 31


def test_filters_keys_metrics_filename_and_number_formats() -> None:
    service = ExportsService(FakeRepo())  # type: ignore[arg-type]

    assert service._normalize_filters(
        {
            "firma": ["Mobiup", ""],
            "regional": "RM 1",
            "unknown": "ignored",
            "asm": None,
        }
    ) == {"firma": ["Mobiup"], "regional": ["RM 1"]}
    assert service._valid_keys(None, {"a"}, ["a"], "test") == ["a"]
    assert service._ratio(Decimal("10"), 0) is None
    assert service._pct(Decimal("10"), Decimal("0")) is None
    assert service._json_value(Decimal("1.25")) == 1.25
    assert service._safe_filename("...") == "export_retail.xlsx"
    assert service._excel_number_format("currency") == "#,##0.00"
    assert service._excel_number_format("percent") == "0.00"
    assert service._excel_number_format("integer") == "0"
    assert service._excel_number_format("text") is None


def test_attach_period_metrics_skips_unknown_row_and_sheet_helpers() -> None:
    service = ExportsService(FakeRepo())  # type: ignore[arg-type]
    rows = {("S001",): {"site_code": "S001"}}
    service._attach_period_metrics(
        rows,
        [
            row(site_code="S999", period_key="2026-06"),
            row(site_code="S001", period_key="2026-06", total_sales=123),
        ],
        "stores",
        ["total_sales"],
        period_prefix="month",
    )
    assert rows[("S001",)]["month:2026-06:total_sales"] == Decimal("123")

    workbook = Workbook()
    sheet = workbook.active
    service._write_table_sheet(
        sheet,
        [
            {"key": "sales", "label": "Sales", "type": "currency"},
            {"key": "name", "label": "Name", "type": "text"},
        ],
        [{"sales": 10, "name": "Store"}],
        header_fill="FFFFFF",
    )
    assert sheet["A2"].number_format == "#,##0.00"
    assert sheet.freeze_panes == "A2"

    service._add_daily_comparison_chart(
        sheet,
        months=[],
        metric="total_sales",
        max_row=1,
        first_data_col=2,
    )
    assert sheet._charts == []


def test_daily_evolution_sheet_skips_empty_metrics_and_missing_values() -> None:
    service = ExportsService(FakeRepo())  # type: ignore[arg-type]
    workbook = Workbook()
    service._add_daily_evolution_sheet(
        workbook,
        months=["2026-05", "2026-06"],
        metrics=[],
        records=[],
    )
    assert workbook.sheetnames == ["Sheet"]

    service._add_daily_evolution_sheet(
        workbook,
        months=["2026-05", "2026-06"],
        metrics=["total_sales"],
        records=[
            row(import_month="2026-05", day_of_month=0),
            row(import_month="2026-05", day_of_month=32, total_sales=100),
        ],
    )
    sheet = workbook["Evolutie zilnica"]
    assert sheet["B2"].value is None
    assert sheet["D2"].value is None
    assert sheet.max_row == 33
