from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import pytest
from openpyxl import Workbook

import services.importer as importer_module
import services.sales_import_parsing as sales_import_parsing
import services.spreadsheet_readers as spreadsheet_readers
from services.importer import (
    SALES_COLUMNS,
    detect_month,
    is_month_final,
    load_sales_dataframe,
    load_targets_dataframe,
    normalize_firma,
)
from services.sales_generation import SalesAnomalyClassification, SalesPolicyValidationError


def sales_workbook(rows: list[dict]) -> bytes:
    output = BytesIO()
    pd.DataFrame(rows, columns=SALES_COLUMNS).to_excel(output, index=False)
    return output.getvalue()


def test_sales_import_has_bounded_headroom_for_growing_erp_worksheet() -> None:
    limits = sales_import_parsing.SALES_IMPORT_SPREADSHEET_LIMITS

    assert limits.max_member_bytes == 128 * 1024 * 1024
    assert limits.max_member_bytes <= limits.max_uncompressed_bytes


def sales_row(**overrides: object) -> dict:
    row: dict[str, object] = {
        "Data": "01.07.2099",
        "SiteCode": " SITE01 ",
        "ItemCode": " ITEM01 ",
        "ItemName": " Produs ",
        "Cantitate": 2,
        "Brand": " Brand ",
        "Pret": 10.125,
        "Valoare": 20.255,
        "Locatie": " Magazin ",
        "Firma": " mobiup ",
        "ASM": " Manager ",
        "Regional": " Regional ",
        "Nr": " BON-A1 ",
        "Categorie": " Accesorii ",
        "SubCategorie": " Test ",
        "Agent": " Agent ",
    }
    row.update(overrides)
    return row


def test_load_sales_dataframe_normalizes_and_flags_rows() -> None:
    content = sales_workbook(
        [
            sales_row(),
            sales_row(
                Data="02.07.2099",
                SiteCode="SITE02",
                Nr="BON2",
                Firma="MOBICELL",
                Cantitate=-1,
                Pret=0,
                Valoare=0,
                Brand=None,
                Categorie=None,
                SubCategorie=None,
            ),
        ]
    )

    frame = load_sales_dataframe(content)

    assert detect_month(frame) == "2099-07"
    assert list(frame["SiteCode"]) == ["SITE01", "SITE02"]
    assert list(frame["Firma"]) == ["Mobiup", "MobiCell"]
    assert list(frame["Nr"]) == ["BON-A1", "BON2"]
    assert list(frame["Pret"]) == [10.125, 0.0]
    assert list(frame["Valoare"]) == [20.255, 0.0]
    assert list(frame["is_cartela"]) == [False, True]
    assert list(frame["is_return"]) == [False, True]
    assert frame.loc[0, "Categorie"] == "Accesorii"
    assert frame.loc[1, "Categorie"] is None


def test_load_sales_dataframe_rejects_missing_columns() -> None:
    output = BytesIO()
    pd.DataFrame([{"Data": "01.07.2099"}]).to_excel(output, index=False)

    with pytest.raises(ValueError, match="Lipsesc coloane obligatorii"):
        load_sales_dataframe(output.getvalue())


@pytest.mark.parametrize(
    ("column", "value", "message"),
    [
        ("Pret", None, "Pret"),
        ("Valoare", "invalid", "Valoare"),
        ("Cantitate", 1.5, "Cantitate"),
    ],
)
def test_load_sales_dataframe_rejects_invalid_numeric_values(
    column: str,
    value: object,
    message: str,
) -> None:
    content = sales_workbook([sales_row(**{column: value})])

    with pytest.raises(ValueError, match=message):
        load_sales_dataframe(content)


def test_load_sales_dataframe_preserves_identical_sales_rows() -> None:
    row = sales_row()

    frame = load_sales_dataframe(sales_workbook([row, row]))

    assert len(frame) == 2
    assert frame.loc[0, SALES_COLUMNS].equals(frame.loc[1, SALES_COLUMNS])
    assert frame["Cantitate"].sum() == 4


def test_sales_loader_parses_the_worksheet_once(monkeypatch: pytest.MonkeyPatch) -> None:
    original = spreadsheet_readers.pd.read_excel
    parse_calls: list[dict[str, object]] = []

    def tracking_read_excel(*args: object, **kwargs: object) -> pd.DataFrame:
        parse_calls.append(dict(kwargs))
        return original(*args, **kwargs)

    monkeypatch.setattr(spreadsheet_readers.pd, "read_excel", tracking_read_excel)

    load_sales_dataframe(sales_workbook([sales_row()]))

    assert parse_calls == [{"sheet_name": 0, "header": None, "engine": "openpyxl"}]


def test_sales_loader_keeps_finite_parser_resources_for_manifest() -> None:
    frame = load_sales_dataframe(sales_workbook([sales_row()]))

    resources = frame.attrs["parser_resource_stats"]
    assert resources["parser"] == "sales"
    assert resources["format"] == "xlsx"
    assert resources["rows"] == 1
    assert int(resources["compressed_bytes"]) > 0
    assert int(resources["expanded_bytes"]) > 0
    assert float(resources["parse_seconds"]) >= 0
    assert int(resources["peak_rss_bytes"]) > 0


def test_load_sales_dataframe_rejects_duplicate_raw_excel_headers() -> None:
    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(SALES_COLUMNS)
    sheet.cell(row=1, column=3, value="SiteCode")
    workbook.save(output)

    with pytest.raises(SalesPolicyValidationError, match="antete duplicate") as exc_info:
        load_sales_dataframe(output.getvalue())

    assert exc_info.value.anomalies[0]["code"] == "duplicate_headers"
    assert (
        exc_info.value.anomalies[0]["classification"]
        == SalesAnomalyClassification.STRUCTURAL_CONTRADICTION.value
    )


def test_load_sales_dataframe_rejects_missing_required_identifier() -> None:
    with pytest.raises(SalesPolicyValidationError, match="identificatori obligatorii") as exc_info:
        load_sales_dataframe(sales_workbook([sales_row(Agent=None)]))

    assert exc_info.value.anomalies[0]["code"] == "missing_required_identifiers"
    assert exc_info.value.anomalies[0]["blocking"] is True


def test_load_sales_dataframe_ignores_missing_identifier_on_excluded_row() -> None:
    frame = load_sales_dataframe(
        sales_workbook(
            [
                sales_row(),
                sales_row(
                    SiteCode="TR-IGNORED",
                    ASM="-",
                    Agent=None,
                    Nr=None,
                ),
            ]
        )
    )

    assert list(frame["SiteCode"]) == ["SITE01", "TR-IGNORED"]


def test_load_sales_dataframe_rejects_conflicting_store_metadata() -> None:
    with pytest.raises(SalesPolicyValidationError, match="contradictorii") as exc_info:
        load_sales_dataframe(
            sales_workbook(
                [sales_row(), sales_row(Nr="BON2", Locatie="Alt magazin")]
            )
        )

    anomaly = exc_info.value.anomalies[0]
    assert anomaly["code"] == "contradictory_store_metadata"
    assert anomaly["classification"] == SalesAnomalyClassification.STRUCTURAL_CONTRADICTION.value

def test_detect_month_rejects_mixed_months() -> None:
    frame = pd.DataFrame({"Data": [date(2099, 7, 1), date(2099, 8, 1)]})

    with pytest.raises(ValueError, match="mai multe luni"):
        detect_month(frame)


def test_company_normalization_and_month_finality() -> None:
    assert normalize_firma(" MOBIUP ") == "Mobiup"
    assert normalize_firma("mobicell") == "MobiCell"
    assert normalize_firma("Alta Firma") == "Alta Firma"
    assert is_month_final("2000-01") is True
    assert is_month_final("9999-12") is False


def test_month_finality_uses_bucharest_business_midnight() -> None:
    class FixedClock:
        def now(self) -> datetime:
            return datetime(2026, 4, 1, 0, 30, tzinfo=ZoneInfo("Europe/Bucharest"))

    assert is_month_final("2026-03", clock=FixedClock()) is True
    assert is_month_final("2026-04", clock=FixedClock()) is False


def write_targets_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["", 2098, None, None, 2099])
    sheet.append(["SiteCode", "TG L01", "TG L02", "TG invalid", "TG L03"])
    sheet.append(["SITE01", 100.125, 200, 999, 300])
    sheet.append(["", 400, 500, 999, 600])
    sheet.append(["SITE02", None, 250.555, 999, None])
    workbook.save(path)


def test_load_targets_dataframe_extracts_years_and_skips_empty_values(
    tmp_path: Path,
) -> None:
    source = tmp_path / "targets.xlsx"
    write_targets_workbook(source)

    targets = load_targets_dataframe(source)

    assert targets == [
        {
            "site_code": "SITE01",
            "import_month": "2098-01",
            "target_value": Decimal("100.12"),
        },
        {
            "site_code": "SITE01",
            "import_month": "2098-02",
            "target_value": Decimal("200.00"),
        },
        {
            "site_code": "SITE01",
            "import_month": "2099-03",
            "target_value": Decimal("300.00"),
        },
        {
            "site_code": "SITE02",
            "import_month": "2098-02",
            "target_value": Decimal("250.56"),
        },
    ]


def test_targets_loader_parses_the_worksheet_once(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    source = tmp_path / "targets.xlsx"
    write_targets_workbook(source)
    original = importer_module.pd.ExcelFile
    parse_calls: list[dict[str, object]] = []

    class TrackingExcelFile:
        def __init__(self, *args: object, **kwargs: object) -> None:
            self._inner = original(*args, **kwargs)

        def __enter__(self) -> "TrackingExcelFile":
            self._inner.__enter__()
            return self

        def __exit__(self, *args: object) -> None:
            self._inner.__exit__(*args)

        def parse(self, *args: object, **kwargs: object) -> pd.DataFrame:
            parse_calls.append(dict(kwargs))
            return self._inner.parse(*args, **kwargs)

    monkeypatch.setattr(importer_module.pd, "ExcelFile", TrackingExcelFile)

    load_targets_dataframe(source)

    assert parse_calls == [{"header": None}]


def test_targets_loader_runs_structural_preflight(tmp_path: Path) -> None:
    source = tmp_path / "targets.xlsx"
    source.write_bytes(b"not-an-xlsx")

    with pytest.raises(ValueError, match="semnătură"):
        load_targets_dataframe(source)
