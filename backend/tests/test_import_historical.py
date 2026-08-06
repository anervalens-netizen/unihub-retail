from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

import scripts.import_historical as historical
import scripts.import_historical_monthly_sales as historical_monthly
import scripts.import_annual_summary as historical_annual


LEGACY_COLUMNS = historical.OLD_COLUMNS


def legacy_workbook(path: Path, **overrides: object) -> None:
    row: dict[str, object] = {
        "Data": "01.07.2099",
        "SiteCode": "SITE01",
        "ItemCode": "ITEM01",
        "ItemName": "Produs",
        "Cantitate": 2,
        "Brand": "Brand",
        "Pret": 10.125,
        "Valoare": 20.255,
        "Locatie": "Magazin",
        "Firma": "mobiup",
        "ASM": "Manager",
        "Regional": "Regional",
        "Nr": "BON1",
    }
    row.update(overrides)
    pd.DataFrame([row], columns=LEGACY_COLUMNS).to_excel(
        path, sheet_name="MobiUp_MobiCell", index=False
    )


def test_historical_workbook_is_validated_without_database_dependencies(tmp_path: Path) -> None:
    source = tmp_path / "legacy.xlsx"
    legacy_workbook(source)
    before = source.read_bytes()

    report = historical.validate_historical_file(source)
    loaded = historical.load_historical_df(source)

    assert report.import_month == "2099-07"
    assert report.rows_in_file == 1
    assert report.rows_without_valid_asm == 0
    assert report.stores == 1
    assert report.parser_resources["format"] == "xlsx"
    assert report.parser_resources["rows"] == 1
    assert float(report.parser_resources["parse_seconds"]) >= 0
    assert int(report.parser_resources["peak_rss_bytes"]) > 0
    assert list(loaded.columns) == historical.OLD_COLUMNS
    assert "Categorie" not in loaded.columns
    assert "is_cartela" not in loaded.columns
    assert loaded.loc[0, "Firma"] == "Mobiup"
    assert loaded.loc[0, "Data"].isoformat() == "2099-07-01"
    assert source.read_bytes() == before


def test_validation_rejects_bad_source_instead_of_coercing_to_zero(tmp_path: Path) -> None:
    source = tmp_path / "invalid.xlsx"
    legacy_workbook(source, Valoare="not-a-number")

    with pytest.raises(ValueError, match="Valoare"):
        historical.load_historical_df(source)


def test_validation_rejects_non_workbook_before_pandas_parse(tmp_path: Path) -> None:
    source = tmp_path / "invalid.xlsx"
    source.write_bytes(b"not-an-xlsx")

    with pytest.raises(ValueError, match="semnătură"):
        historical.load_historical_df(source)


def test_validation_only_does_not_create_output_files(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    source = input_dir / "2024" / "legacy.xlsx"
    source.parent.mkdir(parents=True)
    legacy_workbook(source)

    reports = historical.process_files(input_dir)

    assert len(reports) == 1
    assert sorted(path.relative_to(input_dir) for path in input_dir.rglob("*")) == [
        Path("2024"),
        Path("2024/legacy.xlsx"),
    ]


def test_historical_validator_has_no_output_or_apply_surface() -> None:
    source = Path(historical.__file__).read_text(encoding="utf-8")

    for forbidden in (
        "--output-dir",
        "output_dir",
        "write_converted_file",
        "to_excel",
        "to_csv",
        "--apply",
        "asyncpg",
        "DATABASE_URL",
        "reserve_snapshot",
        "import_sales_dataframe",
    ):
        assert forbidden not in source


def test_monthly_history_parser_preflights_and_streams_one_open(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    source = tmp_path / "monthly.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([])
    sheet.append([])
    sheet.append(["A2022", "MobiUp", historical_monthly.VALUE_TYPE, "Manager", "Store", *range(1, 13)])
    workbook.save(source)
    monkeypatch.setattr(historical_monthly, "SOURCE_FILES", [source])

    records, duplicates = historical_monthly.load_source_files()

    assert duplicates == []
    assert len(records) == 1
    assert next(iter(records.values())).values[12] == 12
    output = capsys.readouterr().out
    assert 'parser_resources={"cells":' in output
    assert '"expanded_bytes":' in output
    assert '"parser":"historical_monthly_sales"' in output


def test_annual_history_parser_preflights_and_streams_all_sheets(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    source = tmp_path / "annual.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, (firma, _kind) in historical_annual.SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([firma, None, None, "Store", 0, 30, 20])
    workbook.save(source)
    monkeypatch.setattr(historical_annual, "SUMMARY_FILE", source)

    summary = historical_annual.load_summary()

    assert summary[("MobiUp", "STORE")][2022]["value"] == 20
    assert summary[("MobiCell", "STORE")][2023]["qty"] == 30
    output = capsys.readouterr().out
    assert output.startswith('parser_resources={"cells":')
    assert '"expanded_bytes":' in output
    assert '"parser":"historical_annual_sales"' in output


def test_historical_cli_emits_parser_resource_evidence(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    source = tmp_path / "2024" / "legacy.xlsx"
    source.parent.mkdir(parents=True)
    legacy_workbook(source)

    assert historical.main(tmp_path) == 0

    output = capsys.readouterr().out
    assert 'parser_resources={"cells":' in output
    assert '"expanded_bytes":' in output
    assert '"parser":"historical_sales"' in output
