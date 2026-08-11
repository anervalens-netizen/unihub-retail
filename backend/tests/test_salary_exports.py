from __future__ import annotations

import hashlib
import json
import os
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any, cast
from unittest.mock import AsyncMock, MagicMock

import asyncpg
import pytest
from openpyxl import load_workbook

import repositories.export_operations as export_operations_module
import worker
from db.connection import get_pool
from repositories.export_operations import ExportOperationsRepository
from services.salary_exports import SalaryExportsService
from services.exports import ExportValidationError


def test_salary_export_request_is_canonical_and_site_scope_dominates() -> None:
    normalized, kind = SalaryExportsService.validate_request(
        {
            "export_kind": "agents",
            "company_name": " Mobicell ",
            "regional": " Manager ",
            "site_code": [" B, Nord ", "B, Nord", " C "],
            "q": " Ana ",
        }
    )

    assert kind == "salary_agents"
    assert normalized == {
        "export_kind": "agents",
        "company_name": None,
        "site_code": ["B, Nord", "C"],
        "regional": None,
        "asm": None,
        "year": None,
        "month": None,
        "q": "Ana",
    }


@pytest.mark.asyncio
async def test_salary_agent_artifact_attests_true_rows_and_excludes_private_ids() -> None:
    service = SalaryExportsService(MagicMock())
    service.salary_service = cast(Any, SimpleNamespace(
        get_agents_summary=AsyncMock(
            return_value={
                "items": [
                    {
                        "person_id": "sp1_" + "a" * 64,
                        "full_name": '=WEBSERVICE("https://example.invalid")',
                        "company_name": "Mobiup",
                        "locatie": "B, Nord",
                        "month_count": 2,
                        "avg_month_count": 1,
                        "avg_salary": 2500,
                        "total_salary": 3500,
                    }
                ],
                "total": 1,
            }
        )
    ))

    artifact = await service.build_xlsx_artifact(
        {"export_kind": "agents", "site_code": ["B, Nord"]}
    )
    try:
        content = b"".join(artifact.iter_chunks())
        assert artifact.row_count == 1
        assert artifact.cell_count == 34
        assert artifact.sha256 == hashlib.sha256(content).hexdigest()
        assert artifact.size == len(content)
        assert artifact.peak_rss_bytes is not None and artifact.peak_rss_bytes > 0

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
        try:
            report = workbook["Raport salarii"]
            assert report.max_row == 2
            assert report.max_column == 7
            assert report["A2"].value == '\'=WEBSERVICE("https://example.invalid")'
            assert "sp1_" not in " ".join(
                str(cell.value or "")
                for row in report.iter_rows()
                for cell in row
            )
            config = {
                str(row[0].value): row[1].value
                for row in workbook["Configuratie"].iter_rows(min_row=2)
            }
            assert config["Magazine"] == '["B, Nord"]'
            assert config["Randuri"] == 1
        finally:
            workbook.close()
    finally:
        artifact.close()


@pytest.mark.asyncio
async def test_every_salary_export_kind_has_the_same_server_row_limit() -> None:
    service = SalaryExportsService(MagicMock())
    service.salary_service = cast(Any, SimpleNamespace(
        get_trend=AsyncMock(return_value=[{}] * 5_001),
    ))

    with pytest.raises(ExportValidationError, match="5000"):
        await service.build_xlsx_artifact({"export_kind": "monthly_trend"})


def test_salary_export_migration_keeps_server_attestation_immutable() -> None:
    sql = (
        Path(__file__).parents[1]
        / "db"
        / "migrations"
        / "065_salary_export_evidence.sql"
    ).read_text(encoding="utf-8")

    assert "NOT (request_payload ? 'row_count')" in sql
    assert "OLD.row_count IS DISTINCT FROM NEW.row_count" in sql
    assert "export_operations_completed_salary_row_count_check" in sql
    assert "^salary-export:[1-9][0-9]*$" in sql
    assert "^salary/[0-9a-f]{32}\\.xlsx$" in sql


def test_salary_export_authority_is_column_scoped_and_rls_fenced() -> None:
    sql = (
        Path(__file__).parents[1]
        / "db"
        / "migrations"
        / "066_salary_export_authority.sql"
    ).read_text(encoding="utf-8")

    assert "must be provisioned before migration 066" in sql
    assert "GRANT SELECT (\n    id, year, month, full_name, person_id" in sql
    salary_grant = sql.split("GRANT SELECT (", 1)[1].split(
        ") ON TABLE salary_records", 1
    )[0]
    assert "created_at" not in salary_grant
    assert "ALTER TABLE export_operations ENABLE ROW LEVEL SECURITY" in sql
    assert "export_operations_operations_read" in sql
    assert "export_operations_salary_read" in sql
    assert "REVOKE ALL ON SCHEMA salary_private" in sql


@pytest.mark.asyncio
async def test_salary_worker_claim_is_db_fenced_to_salary_kinds(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    repo = SimpleNamespace(
        claim=AsyncMock(return_value=None),
        get=AsyncMock(return_value=None),
    )
    monkeypatch.setattr(
        export_operations_module,
        "ExportOperationsRepository",
        lambda _pool: repo,
    )

    result = await worker.build_salary_export_background(
        {"db_pool": MagicMock()},
        17,
    )

    assert result == {"operation_id": 17, "status": "not_found"}
    repo.claim.assert_awaited_once()
    assert repo.claim.await_args is not None
    assert repo.claim.await_args.args == (17,)
    assert repo.claim.await_args.kwargs["lease_seconds"] == 300
    assert repo.claim.await_args.kwargs["allowed_kinds"] == (
        "salary_store_summary",
        "salary_monthly_trend",
        "salary_agents",
    )


@pytest.mark.asyncio
@pytest.mark.skipif(
    os.getenv("UNIHUB_TEST_DATABASE") != "1",
    reason="requires isolated PostgreSQL",
)
async def test_salary_export_row_count_is_worker_written_and_immutable_in_postgres() -> None:
    pool = await get_pool()
    repo = ExportOperationsRepository(pool)
    request = {
        "export_kind": "agents",
        "company_name": None,
        "site_code": ["B, Nord"],
        "regional": None,
        "asm": None,
        "year": None,
        "month": None,
        "q": None,
    }
    operation = await repo.reserve(
        kind="salary_agents",
        request_payload=request,
        request_sha256=hashlib.sha256(
            json.dumps(request, sort_keys=True, separators=(",", ":")).encode()
        ).hexdigest(),
        requested_by_sub="salary-export-test-owner",
        job_prefix="salary-export",
    )
    claimed = await repo.claim(
        int(operation["id"]),
        execution_owner="salary-export-test-worker",
        lease_seconds=300,
        allowed_kinds=(
            "salary_store_summary",
            "salary_monthly_trend",
            "salary_agents",
        ),
    )
    assert claimed is not None
    assert await repo.complete(
        int(operation["id"]),
        execution_owner="salary-export-test-worker",
        execution_epoch=int(claimed["execution_epoch"]),
        artifact_key=f"salary/{'a' * 32}.xlsx",
        artifact_sha256="b" * 64,
        artifact_size=42,
        peak_rss_bytes=1024,
        build_seconds=0.1,
        cell_count=7,
        row_count=1,
        download_filename="salary.xlsx",
        ttl_seconds=3600,
    )
    completed = await repo.get(int(operation["id"]))
    assert completed is not None and completed["row_count"] == 1

    async with pool.acquire() as connection:
        with pytest.raises(asyncpg.RaiseError, match="download claim only"):
            await connection.execute(
                "UPDATE export_operations SET row_count = 2 WHERE id = $1",
                int(operation["id"]),
            )
