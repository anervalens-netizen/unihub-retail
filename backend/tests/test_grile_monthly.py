from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import AsyncMock, Mock

from openpyxl import load_workbook
import pytest

from services.spreadsheet_safety import TrustedFormula
from services.grile_monthly import (
    ExtractedAgentRow,
    RESET_RANGES,
    RESET_RANGES_V3,
    StoreEntry,
    build_store_export_path,
    build_workbook,
    make_output_row,
    ro_month_label,
    safe_filename,
    validate_archive_manifest,
)


def test_reset_ranges_clear_manual_pontaj_without_total_column():
    joined = " ".join(RESET_RANGES)

    assert "Pontaj!C8:AG31" in RESET_RANGES
    assert "Pontaj!AH" not in joined
    assert "Grila!B32:F46" in RESET_RANGES
    assert "Grila!B32:F37" not in RESET_RANGES
    assert "Grila!G10" not in joined
    assert "Grila!G24" not in joined


def test_v3_reset_ranges_include_agent3_and_shifted_supplemental_section():
    joined = " ".join(RESET_RANGES_V3)
    assert "Grila!D36" in RESET_RANGES_V3
    assert "Grila!Z5:Z50" in RESET_RANGES_V3
    assert "Grila!AA5:AC50" in RESET_RANGES_V3
    assert "Grila!B46:F60" in RESET_RANGES_V3
    assert "Grila!F40:F42" in RESET_RANGES_V3
    assert "Pontaj!AH" not in joined


def test_ro_month_label_and_safe_filename():
    assert ro_month_label("2026-06") == "Iunie 2026"
    assert safe_filename("Mobiup/Bad:Name") == "Mobiup - Bad_Name"


def test_build_store_export_path_uses_company_folder(tmp_path: Path):
    entry = StoreEntry(
        company="Mobiup",
        store="Park Lake",
        sheet_id="sheet-1",
        site_code="PARKLAKE",
        manager="Andrei Stancu",
    )

    assert build_store_export_path(tmp_path, "Iunie 2026", entry) == (
        tmp_path / "archive" / "Iunie 2026" / "Mobiup" / "Park Lake.xlsx"
    )


def test_make_output_row_uses_new_salary_formulas():
    row = ExtractedAgentRow(
        company="Mobiup",
        store="Park Lake",
        slot=1,
        agent="Agent Test",
        base_salary=2600,
        sales_commission=300,
        extra_location_commission=25,
        extra_hours_pay=150,
        bonuri=480,
        worked_hours=176,
        status="OK",
        error="",
        sheet_id="sheet-1",
    )

    output = make_output_row(row, nr=1, metadata={"Manager": "Andrei Stancu"})

    assert output[:6] == [1, "Andrei Stancu", "Park Lake", "Agent Test", 2600, 300]
    assert output[7] == 25
    assert output[9] == 150
    assert isinstance(output[10], TrustedFormula)
    assert isinstance(output[11], TrustedFormula)
    assert output[10].expression == "=SUM(E2:J2,M2)"
    assert output[11].expression == "=K2-M2"
    assert output[12] == 480


def test_build_workbook_creates_company_sheets_and_audit(tmp_path: Path):
    rows = [
        ExtractedAgentRow("Mobiup", "Park Lake", 1, "A1", 2600, 300, 0, 150, 480, 176, "OK", "", "s1"),
        ExtractedAgentRow("Mobicell", "AFI Cotroceni", 1, "B1", 2600, 100, 0, 0, 480, 165, "OK", "", "s2"),
    ]
    output = tmp_path / "Tabel Salarii - Test.xlsx"

    build_workbook(
        rows,
        output,
        metadata_by_company_store={
            ("Mobiup", "Park Lake"): {"Manager": "Andrei Stancu"},
            ("Mobicell", "AFI Cotroceni"): {"Manager": "Andrei Stancu"},
        },
    )

    wb = load_workbook(output, data_only=False)
    assert wb.sheetnames == ["Mobiup", "Mobicell", "Audit"]
    assert wb["Mobiup"]["B2"].value == "Andrei Stancu"
    assert wb["Mobiup"]["C2"].value == "Park Lake"
    assert wb["Mobiup"]["F1"].value == "Comision vanzare"
    assert wb["Mobiup"]["K2"].value == "=SUM(E2:J2,M2)"
    assert wb["Audit"]["L2"].value == "OK"


def test_validate_archive_manifest_rejects_missing_zip(tmp_path: Path):
    xlsx = tmp_path / "store.xlsx"
    xlsx.write_bytes(b"x")
    manifest = {
        "registry_count": 1,
        "exported_count": 1,
        "error_count": 0,
        "zip_path": str(tmp_path / "missing.zip"),
        "stores": [{"company": "Mobiup", "store": "Park Lake", "status": "OK", "xlsx_path": str(xlsx)}],
    }

    ok, errors = validate_archive_manifest(manifest, expected_count=1)

    assert not ok
    assert any("missing or empty archive zip" in error for error in errors)


@pytest.mark.asyncio
async def test_monthly_facade_delegates_google_and_registry(monkeypatch):
    import services.grile_monthly as subject

    marker = object()
    entry = StoreEntry("Mobiup", "Store", "sheet", "SITE", "Manager")
    monkeypatch.setattr(subject.google_api, "service_account_file", Mock(return_value=marker))
    monkeypatch.setattr(subject.google_api, "credentials", Mock(return_value=marker))
    monkeypatch.setattr(subject.google_api, "build_services", Mock(return_value=(marker, marker)))
    monkeypatch.setattr(subject.google_api, "is_transient", Mock(return_value=True))
    monkeypatch.setattr(subject.google_api, "error_code", Mock(return_value="google_error"))
    monkeypatch.setattr(subject.google_api, "retry", Mock(return_value=marker))
    monkeypatch.setattr(subject.google_api, "request", AsyncMock(return_value=marker))
    monkeypatch.setattr(subject.registry, "company_from_values", Mock(return_value="Mobiup"))
    monkeypatch.setattr(subject.registry, "store_from_values", Mock(return_value="Store"))
    monkeypatch.setattr(subject.registry, "load_entries", AsyncMock(return_value=[entry]))

    assert subject._sa_file() is marker
    assert subject.get_credentials() is marker
    assert subject.build_google_services() == (marker, marker)
    assert subject._is_transient(RuntimeError()) is True
    assert subject._google_error_code(RuntimeError()) == "google_error"
    assert subject.retry_api(lambda: None, label="read") is marker
    assert await subject._google_request(marker, "read", {}, label="read") is marker
    assert subject._company_from_values("key", None) == "Mobiup"
    assert subject._store_from_values("key", None) == "Store"
    assert await subject.load_entries(marker, "SITE", month="2026-06") == [entry]

    monkeypatch.setattr(subject, "persisted_operation_to_dict", Mock(return_value={"id": 1}))
    monkeypatch.setattr(subject, "safe_persisted_result", Mock(return_value={"ok": True}))
    assert subject._operation_to_dict(marker) == {"id": 1}
    assert subject._safe_operation_result({"result": {}}) == {"ok": True}
    with pytest.raises(ValueError, match="Operatie necunoscuta"):
        await subject.reserve_monthly_operation(marker, op="unknown")


@pytest.mark.asyncio
async def test_monthly_facade_delegates_repository(monkeypatch):
    import services.grile_monthly as subject

    marker = object()
    repository_calls = {
        "persist_monthly_operation_job": True,
        "persist_monthly_operation_heartbeat": True,
        "fetch_monthly_execution_lease": marker,
        "persist_cancelled_uncertain": True,
        "fetch_monthly_manifest": {"id": 2},
        "fetch_latest_monthly_manifest": {"id": 3},
        "fetch_previous_completed_reset_item": marker,
        "persist_reset_item_backup": True,
    }
    for name, result in repository_calls.items():
        monkeypatch.setattr(subject, name, AsyncMock(return_value=result))
    assert await subject.attach_monthly_operation_job(marker) is True
    assert await subject.heartbeat_monthly_operation(marker) is True
    assert await subject.get_monthly_execution_lease(marker) is marker
    assert await subject.mark_monthly_operation_cancelled_uncertain(marker) is True
    assert await subject.get_monthly_manifest(marker, 2) == {"id": 2}
    assert await subject.get_latest_monthly_manifest(marker, month="2026-06") == {"id": 3}
    assert await subject.get_previous_completed_reset_item(marker) is marker
    assert await subject.get_previous_completed_reset_item(
        marker, closing_month_key="2026-06"
    ) is marker
    assert await subject.record_reset_item_backup(marker) is True

    lease_runner = AsyncMock(return_value=marker)
    monkeypatch.setattr(subject.monthly_lease, "run_with_lease", lease_runner)
    assert await subject._run_with_monthly_lease(
        marker,
        1,
        execution_owner="owner",
        execution_epoch=2,
        operation=AsyncMock(),
    ) is marker


def test_monthly_facade_characterizes_parse_success_and_failures(monkeypatch):
    import services.grile_monthly as subject

    entry = StoreEntry("Mobiup", "Store", "sheet", "SITE", "Manager")
    parsed_row = ExtractedAgentRow(
        "Mobiup", "Store", 1, "Agent", 1, 2, 3, 4, 5, 6, "OK", "", "sheet"
    )
    monkeypatch.setattr(subject, "value_ranges_for_entry", Mock(return_value=["A1"]))
    monkeypatch.setattr(subject, "parse_store_rows", Mock(return_value=[parsed_row]))
    monkeypatch.setattr(subject, "_error_row", Mock(return_value=parsed_row))
    assert subject.extract_store_rows(entry=entry, sheets_svc=None, value_ranges=[{}]) == [
        parsed_row
    ]
    assert subject.extract_store_rows(entry=entry, sheets_svc=None, value_ranges=[]) == [
        parsed_row
    ]

    monkeypatch.setattr(subject, "parse_store_rows", Mock(side_effect=RuntimeError("bad")))
    monkeypatch.setattr(subject, "_google_error_code", Mock(return_value="bad_read"))
    assert subject.extract_store_rows(entry=entry, sheets_svc=None, value_ranges=[{}]) == [
        parsed_row
    ]

    service = Mock()
    service.spreadsheets.return_value.values.return_value.batchGet.return_value.execute.return_value = {
        "valueRanges": [{}]
    }
    monkeypatch.setattr(subject, "retry_api", lambda fn, **_kwargs: fn())
    assert subject._read_store_value_ranges(service, entry, ["A1"]) == [{}]
    service.spreadsheets.return_value.values.return_value.batchGet.return_value.execute.return_value = {}
    with pytest.raises(subject.MonthlyIntegrityError, match="incomplete"):
        subject._read_store_value_ranges(service, entry, ["A1"])


def test_monthly_facade_delegates_artifact_operations(monkeypatch, tmp_path: Path):
    import services.grile_monthly as subject

    marker = {"ok": True}
    path = tmp_path / "artifact"
    monkeypatch.setattr(subject.artifacts, "validate_final_workbook", Mock())
    monkeypatch.setattr(subject.artifacts, "staging_dir", Mock(return_value=path))
    monkeypatch.setattr(subject.artifacts, "promote_file", Mock())
    monkeypatch.setattr(subject.archive_artifacts, "export_sheet_xlsx", Mock(return_value=marker))
    monkeypatch.setattr(subject.archive_artifacts, "write_exported_xlsx", Mock(return_value=marker))
    monkeypatch.setattr(subject.archive_artifacts, "create_archive_zip", Mock())
    monkeypatch.setattr(subject.archive_artifacts, "create_manager_zips", Mock(return_value={"m": path}))
    monkeypatch.setattr(subject.archive_artifacts, "summarize_archive_results", Mock(return_value=marker))
    monkeypatch.setattr(subject.archive_artifacts, "validate_archive_zip", Mock())
    monkeypatch.setattr(subject.archive_artifacts, "validate_source_workbook", Mock())
    monkeypatch.setattr(subject.archive_artifacts, "future_artifact", Mock(return_value=marker))
    monkeypatch.setattr(subject.archive_artifacts, "promote_directory", Mock())

    subject._validate_final_workbook(path)
    assert subject._staging_dir("archive", 1) == path
    subject._promote_file(path, path)
    assert subject.export_sheet_xlsx(None, "sheet", path) == marker
    assert subject.write_exported_xlsx(path, b"data") == marker
    subject.create_archive_zip(path, [path], tmp_path)
    assert subject.create_manager_zips([]) == {"m": path}
    assert subject.summarize_archive_results([]) == marker
    subject._validate_archive_zip(path)
    subject._validate_source_workbook(path)
    assert subject._future_artifact(path, role="archive") == marker
    subject._promote_directory(path, path, manifest={})


@pytest.mark.asyncio
async def test_monthly_facade_delegates_execution(monkeypatch):
    import services.grile_monthly as subject

    marker = object()
    async_names = (
        "finalize_execution",
        "finalize_month",
        "archive_execution",
        "archive_month",
        "approve_manifest",
    )
    for name in async_names:
        monkeypatch.setattr(subject.execution_adapters, name, AsyncMock(return_value=marker))
    monkeypatch.setattr(subject.execution_adapters, "public_manifest_payload", Mock(return_value={"id": 1}))

    assert await subject._finalize_month_execution(marker) is marker
    assert await subject.finalize_month(marker) is marker
    assert await subject._archive_month_execution(marker) is marker
    assert await subject.archive_month(marker) is marker
    assert subject.public_manifest_payload({}) == {"id": 1}
    now = datetime(2026, 1, 1, tzinfo=timezone.utc)
    assert subject._public_timestamp(now) == now.isoformat()
    assert subject._public_timestamp("unchanged") == "unchanged"
    assert await subject.approve_monthly_manifest(marker) is marker


@pytest.mark.asyncio
async def test_monthly_facade_delegates_reset_adapters(monkeypatch):
    import services.grile_monthly as subject

    marker = object()
    entry = StoreEntry("Mobiup", "Store", "sheet", "SITE", "Manager")
    monkeypatch.setattr(subject.reset_adapters, "read_snapshot", Mock(return_value={"v": 1}))
    monkeypatch.setattr(subject.reset_adapters, "read_snapshot_async", AsyncMock(return_value={"v": 1}))
    monkeypatch.setattr(subject.reset_adapters, "restore_snapshot", Mock())
    monkeypatch.setattr(subject.reset_adapters, "restore_snapshot_async", AsyncMock())
    monkeypatch.setattr(subject.reset_adapters, "verify_cleared", Mock())
    monkeypatch.setattr(subject.reset_adapters, "verify_cleared_async", AsyncMock())
    monkeypatch.setattr(subject.reset_adapters, "reset_store", Mock(return_value={"ok": True}))
    monkeypatch.setattr(subject.reset_adapters, "rollback_sync", AsyncMock(return_value=True))
    monkeypatch.setattr(subject.reset_adapters, "rollback_adapter", AsyncMock(return_value=True))

    async def run_cancel_safe(thunk):
        return await thunk()

    monkeypatch.setattr(subject.reset_adapters, "cancel_safe", run_cancel_safe)
    assert subject._read_reset_snapshot(marker, entry) == {"v": 1}
    assert await subject._read_reset_snapshot_async(marker, entry) == {"v": 1}
    subject._restore_reset_snapshot(marker, entry, {})
    await subject._restore_reset_snapshot_async(marker, entry, {})
    subject._verify_reset_cleared(marker, entry)
    await subject._verify_reset_cleared_async(marker, entry)
    assert subject.reset_store(marker, entry, dry_run=True) == {"ok": True}
    assert await subject._rollback_reset_entries(marker) is True
    assert await subject._rollback_reset_entries_cancel_safe(marker) is True
    assert await subject._rollback_reset_entries_adapter(marker) is True
    assert await subject._rollback_reset_entries_adapter_cancel_safe(marker) is True


@pytest.mark.asyncio
async def test_monthly_facade_delegates_reset_execution_and_reconciliation(monkeypatch):
    import services.grile_monthly as subject

    marker = object()
    entry = StoreEntry("Mobiup", "Store", "sheet", "SITE", "Manager")
    for name in ("reset_execution", "reset_month", "reconcile", "fetch_download"):
        monkeypatch.setattr(subject.execution_adapters, name, AsyncMock(return_value=marker))
    monkeypatch.setattr(subject.reconciler, "reconciliation_entry", Mock(return_value=entry))
    monkeypatch.setattr(subject.reconciler, "read_reset_backup", Mock(return_value={"v": 1}))
    monkeypatch.setattr(subject.reconciler, "snapshot_is_cleared", Mock(return_value=True))

    assert await subject._reset_month_execution(marker) is marker
    assert await subject.reset_month(marker) is marker
    assert subject._reconciliation_entry({}) == entry
    assert subject._read_reset_backup({}) == {"v": 1}
    assert subject._snapshot_is_cleared({}) is True
    assert await subject.reconcile_monthly_operations(marker, marker) is marker
    assert await subject.fetch_download("archive", "2026-06") is marker
